from __future__ import annotations

import io
import asyncio
import base64
import gzip
import hashlib
import html
import json
import logging
import os
import re
import secrets
import tempfile
import time
import uuid
import zipfile
import sys
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from xml.etree import ElementTree

import bcrypt
from asyncpg.exceptions import UniqueViolationError
from cryptography.fernet import Fernet, InvalidToken
from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding, rsa
from dotenv import load_dotenv
import httpx
from fastapi import APIRouter, BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import segno
from pydantic import BaseModel, EmailStr, Field
from starlette.background import BackgroundTask
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware

try:
    from .postgres_database import PostgresDatabase
    from .app_version import version_payload
    from .identity_integrity import (
        normalize_nim,
        student_identity_conflict_query,
        student_identity_values,
    )
    from .program_scope import record_matches_program_scope, resolve_program_identifiers
    from .storage_policy import (
        DRIVE_LOCAL_RETENTION_DAYS,
        DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY,
        local_copy_is_expired,
        next_drive_retry_at,
        portable_storage_path_from_local_path,
        resolve_storage_local_path,
        retry_is_due,
        sync_attempt_day,
    )
    from .user_activity import aggregate_user_activity, classify_activity
    from .user_notifications import finalize_notifications, notification_event
    from .youtube_urls import normalize_youtube_url
    from .rps_parser import RPSPdfDependencyError, RPSPdfParseError, parse_rps_document
except ImportError:  # Supports `uvicorn server:app` from the backend directory.
    from postgres_database import PostgresDatabase
    from app_version import version_payload
    from identity_integrity import (
        normalize_nim,
        student_identity_conflict_query,
        student_identity_values,
    )
    from program_scope import record_matches_program_scope, resolve_program_identifiers
    from storage_policy import (
        DRIVE_LOCAL_RETENTION_DAYS,
        DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY,
        local_copy_is_expired,
        next_drive_retry_at,
        portable_storage_path_from_local_path,
        resolve_storage_local_path,
        retry_is_due,
        sync_attempt_day,
    )
    from user_activity import aggregate_user_activity, classify_activity
    from user_notifications import finalize_notifications, notification_event
    from youtube_urls import normalize_youtube_url
    from rps_parser import RPSPdfDependencyError, RPSPdfParseError, parse_rps_document


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# Router modules historically use top-level imports such as
# ``from routers.feeder import ...``. Keep those imports working when the
# application is started from the repository root as ``backend.server:app``
# as well as from ``backend`` as ``server:app``.
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

db = PostgresDatabase(os.environ["DATABASE_URL"])

app = FastAPI(title="SIAKAD & E-Learning Dosen API")
api_router = APIRouter(prefix="/api")

from routers.akademik import router as akademik_router
from routers.krs_khs import router as krs_khs_router
from routers.keuangan import ensure_default_finance_components, router as keuangan_router
from routers.master_data import router as master_data_router
from routers.master_data import (
    DEFAULT_JABATAN_AKADEMIK,
    DEFAULT_UNIT_ORGANISASI,
    _active_pejabat,
    _recommended_rombel_name,
)
from routers.kurikulum import router as kurikulum_router
from routers.feeder import router as feeder_router
from routers.user_access import (
    build_effective_user_access,
    normalize_base_role,
    rebuild_user_position_access,
    router as user_access_router,
    user_has_access_role,
    user_is_admin_or_access_role,
    user_is_program_manager,
)
from routers.sk_mengajar import router as sk_mengajar_router
from routers.sk_jabatan import router as sk_jabatan_router
from routers.pmb import router as pmb_router

app.include_router(akademik_router)
app.include_router(krs_khs_router)
app.include_router(keuangan_router)
app.include_router(master_data_router)
app.include_router(kurikulum_router)
app.include_router(feeder_router)
app.include_router(user_access_router)
app.include_router(sk_mengajar_router)
app.include_router(sk_jabatan_router)
app.include_router(pmb_router)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)
_whatsapp_send_lock = asyncio.Lock()
_material_creation_lock = asyncio.Lock()
_drive_sync_attempt_lock = asyncio.Lock()
_user_activity_cleanup_lock = asyncio.Lock()

_settings_cache: Dict[str, Any] = {}
_settings_cache_times: Dict[str, float] = {}
_SETTINGS_CACHE_TTL = 30.0
_auth_cache: Dict[str, Any] = {}
_AUTH_CACHE_TTL = max(1.0, float(os.environ.get("AUTH_CACHE_TTL", "5")))
_AUTH_CACHE_MAX_ENTRIES = max(128, int(os.environ.get("AUTH_CACHE_MAX_ENTRIES", "2048")))
_auth_cache_lock = asyncio.Lock()
_class_scope_cache: Dict[str, Any] = {}
_CLASS_SCOPE_CACHE_TTL = max(1.0, float(os.environ.get("CLASS_SCOPE_CACHE_TTL", "5")))
_CLASS_SCOPE_CACHE_MAX_ENTRIES = max(128, int(os.environ.get("CLASS_SCOPE_CACHE_MAX_ENTRIES", "2048")))
_class_scope_cache_lock = asyncio.Lock()
_user_activity_tasks: set[asyncio.Task] = set()
_oidc_discovery_cache: Dict[str, Any] = {}
_oidc_discovery_cached_at = 0.0
_OIDC_DISCOVERY_CACHE_TTL = 300.0
_oidc_runtime_settings: Dict[str, Any] = {}
_last_user_activity_cleanup_at = 0.0
USER_ACTIVITY_TIMEZONE = ZoneInfo(os.environ.get("USER_ACTIVITY_TIMEZONE", "Asia/Jakarta"))
USER_ACTIVITY_RETENTION_DAYS = max(
    30,
    int(os.environ.get("USER_ACTIVITY_RETENTION_DAYS", "180")),
)
NOTIFICATION_LOOKBACK_DAYS = max(
    7,
    min(int(os.environ.get("NOTIFICATION_LOOKBACK_DAYS", "30")), 90),
)
NOTIFICATION_READ_RETENTION_DAYS = max(
    NOTIFICATION_LOOKBACK_DAYS,
    int(os.environ.get("NOTIFICATION_READ_RETENTION_DAYS", "180")),
)


def env_enabled(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def environment_oidc_settings() -> Dict[str, Any]:
    discovery_url = os.environ.get("OIDC_DISCOVERY_URL", "").strip()
    client_id = os.environ.get("OIDC_CLIENT_ID", "").strip()
    redirect_uri = os.environ.get("OIDC_REDIRECT_URI", "").strip()
    frontend_url = os.environ.get("OIDC_FRONTEND_URL", os.environ.get("APP_URL", "http://127.0.0.1:3000")).strip().rstrip("/")
    return {
        "enabled": env_enabled("OIDC_ENABLED", bool(discovery_url and client_id and redirect_uri)),
        "discovery_url": discovery_url,
        "issuer": os.environ.get("OIDC_ISSUER", "").strip().rstrip("/"),
        "client_id": client_id,
        "client_secret": os.environ.get("OIDC_CLIENT_SECRET", "").strip(),
        "redirect_uri": redirect_uri,
        "frontend_url": frontend_url,
        "scopes": os.environ.get("OIDC_SCOPES", "openid profile email roles").strip(),
        "local_login_enabled": env_enabled("OIDC_LOCAL_LOGIN_ENABLED", True),
    }


def oidc_settings() -> Dict[str, Any]:
    return {**environment_oidc_settings(), **_oidc_runtime_settings}


def clear_oidc_discovery_cache() -> None:
    global _oidc_discovery_cache, _oidc_discovery_cached_at
    _oidc_discovery_cache = {}
    _oidc_discovery_cached_at = 0.0


def oidc_require_settings() -> Dict[str, Any]:
    settings = oidc_settings()
    missing = [key for key in ["discovery_url", "client_id", "redirect_uri", "frontend_url"] if not settings.get(key)]
    if not settings["enabled"] or missing:
        raise HTTPException(status_code=503, detail="Login SCI-ID belum dikonfigurasi")
    return settings


def _get_cached_settings(key: str) -> Optional[Dict[str, Any]]:
    now = time.time()
    if key in _settings_cache and (now - _settings_cache_times.get(key, 0)) < _SETTINGS_CACHE_TTL:
        return _settings_cache[key]
    return None


def _set_cached_settings(key: str, value: Dict[str, Any]) -> None:
    _settings_cache[key] = value
    _settings_cache_times[key] = time.time()


def cache_authenticated_user(token: str, session: Dict[str, Any], user: Dict[str, Any]) -> None:
    now = time.monotonic()
    if len(_auth_cache) >= _AUTH_CACHE_MAX_ENTRIES:
        expired = [
            key
            for key, value in _auth_cache.items()
            if now - value[0] >= _AUTH_CACHE_TTL
        ]
        for key in expired:
            _auth_cache.pop(key, None)
        while len(_auth_cache) >= _AUTH_CACHE_MAX_ENTRIES:
            oldest = min(_auth_cache, key=lambda key: _auth_cache[key][0])
            _auth_cache.pop(oldest, None)
    _auth_cache[token] = (now, session, user)


def _invalidate_settings_cache(key: str = "") -> None:
    if key:
        _settings_cache.pop(key, None)
        _settings_cache_times.pop(key, None)
    else:
        _settings_cache.clear()
        _settings_cache_times.clear()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def should_log_user_activity(method: str, path: str) -> bool:
    clean_method = str(method or "").upper()
    clean_path = str(path or "")
    if clean_method in {"OPTIONS", "HEAD"}:
        return False
    ignored = {
        "/api/version",
        "/api/settings/public",
        "/api/auth/me",
        "/api/auth/sso/config",
        "/api/user-activity",
        "/api/notifications",
    }
    return (
        clean_path not in ignored
        and not clean_path.startswith("/api/user-activity/")
        and not clean_path.startswith("/api/notifications/")
    )


def request_client_ip(request: Optional[Request]) -> str:
    if request is None:
        return ""
    forwarded = str(request.headers.get("x-forwarded-for") or "").split(",", 1)[0].strip()
    return (forwarded or str(request.client.host if request.client else ""))[:64]


async def cleanup_old_user_activity_logs() -> None:
    global _last_user_activity_cleanup_at
    current_monotonic = time.monotonic()
    if current_monotonic - _last_user_activity_cleanup_at < 3600:
        return
    async with _user_activity_cleanup_lock:
        current_monotonic = time.monotonic()
        if current_monotonic - _last_user_activity_cleanup_at < 3600:
            return
        cutoff = (
            datetime.now(timezone.utc) - timedelta(days=USER_ACTIVITY_RETENTION_DAYS)
        ).isoformat()
        await db.user_activity_logs.delete_many({"created_at": {"$lt": cutoff}})
        _last_user_activity_cleanup_at = current_monotonic


async def record_user_activity(
    user: Dict[str, Any],
    method: str,
    path: str,
    status_code: int,
    duration_ms: int = 0,
    request: Optional[Request] = None,
    action_override: str = "",
) -> None:
    if not user or not user.get("id"):
        return
    classification = classify_activity(method, path)
    if action_override:
        classification = {
            **classification,
            "action": action_override,
            "activity_label": classify_activity(
                "POST",
                f"/api/auth/{action_override}",
            )["activity_label"],
        }
    await db.user_activity_logs.insert_one(
        {
            "id": new_id(),
            "user_id": user["id"],
            "user_name": str(user.get("name") or user.get("username") or "Pengguna")[:160],
            "user_role": str(user.get("role") or "unknown")[:32],
            **classification,
            "method": str(method or "GET").upper()[:10],
            "path": str(path or "")[:300],
            "status_code": int(status_code or 0),
            "success": 200 <= int(status_code or 0) < 400,
            "duration_ms": max(0, int(duration_ms or 0)),
            "client_ip": request_client_ip(request),
            "user_agent": (
                str(request.headers.get("user-agent") or "")[:300]
                if request is not None
                else ""
            ),
            "created_at": now_iso(),
        }
    )
    await cleanup_old_user_activity_logs()


async def safe_record_user_activity(*args: Any, **kwargs: Any) -> None:
    try:
        await record_user_activity(*args, **kwargs)
    except Exception as exc:
        logger.warning("Aktivitas pengguna gagal dicatat: %s", exc)


def queue_user_activity(*args: Any, **kwargs: Any) -> None:
    """Persist analytics after the HTTP response has been handed off.

    Activity analytics are useful but never part of the response contract. On
    a small host, waiting for an extra JSONB INSERT after every API call made
    the login fan-out noticeably slower and occupied the limited DB pool.
    """
    task = asyncio.create_task(safe_record_user_activity(*args, **kwargs))
    _user_activity_tasks.add(task)
    task.add_done_callback(_user_activity_tasks.discard)


def base64url_decode(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + ("=" * (-len(value) % 4)))


def base64url_encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


async def oidc_discovery(force: bool = False) -> Dict[str, Any]:
    global _oidc_discovery_cache, _oidc_discovery_cached_at
    settings = oidc_require_settings()
    if not force and _oidc_discovery_cache and time.time() - _oidc_discovery_cached_at < _OIDC_DISCOVERY_CACHE_TTL:
        return _oidc_discovery_cache
    try:
        async with httpx.AsyncClient(timeout=10) as http:
            response = await http.get(settings["discovery_url"], headers={"Accept": "application/json"})
            response.raise_for_status()
            metadata = response.json()
    except (httpx.HTTPError, ValueError) as exc:
        logger.warning("OIDC discovery SCI-ID gagal: %s", exc)
        raise HTTPException(status_code=503, detail="SCI-ID sedang tidak dapat dihubungi") from exc
    required = ["issuer", "authorization_endpoint", "token_endpoint", "jwks_uri", "end_session_endpoint"]
    if not isinstance(metadata, dict) or any(not metadata.get(key) for key in required):
        raise HTTPException(status_code=503, detail="Metadata OIDC SCI-ID tidak lengkap")
    expected_issuer = settings["issuer"] or metadata["issuer"]
    if metadata["issuer"].rstrip("/") != expected_issuer.rstrip("/"):
        raise HTTPException(status_code=503, detail="Issuer OIDC SCI-ID tidak sesuai konfigurasi")
    _oidc_discovery_cache = metadata
    _oidc_discovery_cached_at = time.time()
    return metadata


def oidc_roles(claims: Dict[str, Any], client_id: str) -> List[str]:
    roles: set[str] = set()
    realm_access = claims.get("realm_access")
    if isinstance(realm_access, dict) and isinstance(realm_access.get("roles"), list):
        roles.update(str(role) for role in realm_access["roles"])
    resource_access = claims.get("resource_access")
    client_access = resource_access.get(client_id) if isinstance(resource_access, dict) else None
    if isinstance(client_access, dict) and isinstance(client_access.get("roles"), list):
        roles.update(str(role) for role in client_access["roles"])
    return sorted(roles)


def oidc_application_role(claims: Dict[str, Any], client_id: str) -> str:
    roles = set(oidc_roles(claims, client_id))
    if "super_admin" in roles:
        return "admin"
    if roles.intersection({"tendik", "staff", "staf", "pegawai"}):
        return "staff"
    if "dosen" in roles:
        return "lecturer"
    if "mahasiswa" in roles:
        return "student"
    raise HTTPException(status_code=403, detail="Akun SCI-ID belum memiliki role tendik, dosen, mahasiswa, atau super_admin")


async def validate_oidc_id_token(id_token: str, metadata: Dict[str, Any], expected_nonce: str) -> Dict[str, Any]:
    settings = oidc_require_settings()
    parts = id_token.split(".")
    if len(parts) != 3:
        raise HTTPException(status_code=401, detail="Format ID token SCI-ID tidak valid")
    try:
        header = json.loads(base64url_decode(parts[0]))
        claims = json.loads(base64url_decode(parts[1]))
        signature = base64url_decode(parts[2])
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=401, detail="ID token SCI-ID tidak dapat dibaca") from exc
    if header.get("alg") != "RS256" or not header.get("kid"):
        raise HTTPException(status_code=401, detail="Algoritma ID token SCI-ID tidak diizinkan")
    try:
        async with httpx.AsyncClient(timeout=10) as http:
            jwks_response = await http.get(metadata["jwks_uri"], headers={"Accept": "application/json"})
            jwks_response.raise_for_status()
            jwks = jwks_response.json()
    except (httpx.HTTPError, ValueError) as exc:
        raise HTTPException(status_code=503, detail="Kunci verifikasi SCI-ID tidak dapat dimuat") from exc
    key = next((item for item in jwks.get("keys", []) if item.get("kid") == header["kid"] and item.get("kty") == "RSA"), None)
    if not key or not key.get("n") or not key.get("e"):
        raise HTTPException(status_code=401, detail="Kunci ID token SCI-ID tidak ditemukan")
    try:
        public_key = rsa.RSAPublicNumbers(
            int.from_bytes(base64url_decode(key["e"]), "big"),
            int.from_bytes(base64url_decode(key["n"]), "big"),
        ).public_key()
        public_key.verify(signature, f"{parts[0]}.{parts[1]}".encode("ascii"), padding.PKCS1v15(), hashes.SHA256())
    except (InvalidSignature, ValueError) as exc:
        raise HTTPException(status_code=401, detail="Signature ID token SCI-ID tidak valid") from exc
    now = int(time.time())
    issuer = settings["issuer"] or metadata["issuer"]
    audiences = claims.get("aud") if isinstance(claims.get("aud"), list) else [claims.get("aud")]
    if claims.get("iss", "").rstrip("/") != issuer.rstrip("/"):
        raise HTTPException(status_code=401, detail="Issuer ID token SCI-ID tidak sesuai")
    if settings["client_id"] not in audiences:
        raise HTTPException(status_code=401, detail="Audience ID token SCI-ID tidak sesuai")
    if claims.get("azp") and claims["azp"] != settings["client_id"]:
        raise HTTPException(status_code=401, detail="Authorized party ID token SCI-ID tidak sesuai")
    if not isinstance(claims.get("exp"), (int, float)) or now - 30 >= int(claims["exp"]):
        raise HTTPException(status_code=401, detail="ID token SCI-ID sudah kedaluwarsa")
    if isinstance(claims.get("nbf"), (int, float)) and now + 30 < int(claims["nbf"]):
        raise HTTPException(status_code=401, detail="ID token SCI-ID belum berlaku")
    if isinstance(claims.get("iat"), (int, float)) and int(claims["iat"]) > now + 60:
        raise HTTPException(status_code=401, detail="Waktu penerbitan ID token SCI-ID tidak valid")
    if not claims.get("nonce") or not secrets.compare_digest(str(claims["nonce"]), expected_nonce):
        raise HTTPException(status_code=401, detail="Nonce login SCI-ID tidak sesuai")
    if not str(claims.get("sub", "")).strip():
        raise HTTPException(status_code=401, detail="Subject ID token SCI-ID tidak valid")
    return claims


def oidc_frontend_redirect(**params: str) -> str:
    base = oidc_require_settings()["frontend_url"]
    clean = {key: value for key, value in params.items() if value}
    return f"{base}/?{urlencode(clean)}" if clean else base


async def provision_oidc_user(claims: Dict[str, Any]) -> Dict[str, Any]:
    settings = oidc_require_settings()
    issuer = settings["issuer"] or str(claims["iss"])
    subject = str(claims["sub"])
    role = oidc_application_role(claims, settings["client_id"])
    username = str(claims.get("preferred_username") or f"sso-{subject[:12]}").strip().lower()
    email = str(claims.get("email") or f"sso-{subject}@local.invalid").strip().lower()
    name = str(claims.get("name") or username).strip()
    linked = await db.users.find_one({"sso_issuer": issuer, "sso_subject": subject}, {"_id": 0})
    if not linked:
        matches = await db.users.find({"$or": [{"email": email}, {"username": username}]}, {"_id": 0}).to_list(3)
        unique_matches = {item["id"]: item for item in matches}
        if len(unique_matches) > 1:
            raise HTTPException(status_code=409, detail="Email dan username SCI-ID terhubung ke akun lokal yang berbeda")
        linked = next(iter(unique_matches.values()), None)
    role_claims = oidc_roles(claims, settings["client_id"])
    synced = {
        "role": role,
        "name": name,
        "sso_issuer": issuer,
        "sso_subject": subject,
        "sso_roles": role_claims,
        "auth_source": "sso",
        "status": "active",
        "last_login_at": now_iso(),
        "sso_synced_at": now_iso(),
    }
    employee_id = str(claims.get("employee_id") or claims.get("nidn") or claims.get("nip") or "").strip()
    nim = str(claims.get("nim") or "").strip()
    if employee_id:
        synced["employee_id"] = employee_id
    if nim:
        synced["nim"] = nim
    if linked:
        await db.users.update_one({"id": linked["id"]}, {"$set": synced})
        user = await db.users.find_one({"id": linked["id"]}, {"_id": 0})
    else:
        candidate = username
        if await db.users.find_one({"username": candidate}, {"_id": 0, "id": 1}):
            candidate = f"{candidate}-{subject[:6]}"
        doc = {
            "id": new_id(),
            "username": candidate,
            "email": email,
            "whatsapp": str(claims.get("phone_number") or ""),
            "password_hash": hash_password(secrets.token_urlsafe(32)),
            "class_ids": [],
            "created_at": now_iso(),
            **synced,
        }
        await db.users.insert_one(doc)
        user = doc
    return public_doc(user)


def parse_iso_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        return None


def normalize_optional_datetime(value: str, field_label: str) -> str:
    if not value:
        return ""
    parsed = parse_iso_datetime(value)
    if not parsed:
        raise HTTPException(status_code=400, detail=f"{field_label} tidak valid")
    return parsed.isoformat()


def assignment_is_published(assignment: Dict[str, Any], reference: Optional[datetime] = None) -> bool:
    published_at = parse_iso_datetime(assignment.get("published_at", ""))
    if not published_at:
        return True
    return published_at <= (reference or datetime.now(timezone.utc))


def assignment_publish_status(assignment: Dict[str, Any]) -> str:
    return "published" if assignment_is_published(assignment) else "scheduled"


def new_id() -> str:
    return str(uuid.uuid4())


def clean_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())[:10]


def public_doc(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not doc:
        return None
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    doc.pop("local_path", None)
    doc.pop("drive_error", None)
    doc.pop("drive_hierarchy", None)
    if doc.get("role") == "student":
        doc["physical_document_status"] = physical_document_status_payload(doc)
    return doc


PHYSICAL_DOCUMENT_TYPES: Dict[str, str] = {
    "ijazah": "Ijazah",
    "transkip": "Transkrip Nilai",
    "ktp": "KTP",
    "kk": "Kartu Keluarga (KK)",
    "akte": "Akta Kelahiran",
    "kip_k": "KIP-K",
    "surat_keterangan": "Surat Keterangan",
}
PHYSICAL_DOCUMENT_TYPE_ALIASES = {
    "transkrip": "transkip",
    "kip-k": "kip_k",
    "kipkuliah": "kip_k",
    "surat-keterangan": "surat_keterangan",
}
PHYSICAL_DOCUMENT_ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}
PHYSICAL_DOCUMENT_MAX_FILE_MB = max(
    1.0,
    float(os.environ.get("PHYSICAL_DOCUMENT_MAX_FILE_MB", "10")),
)


def normalize_physical_document_type(value: str) -> str:
    normalized = re.sub(r"\s+", "_", str(value or "").strip().lower())
    normalized = PHYSICAL_DOCUMENT_TYPE_ALIASES.get(normalized, normalized)
    if normalized not in PHYSICAL_DOCUMENT_TYPES:
        raise HTTPException(status_code=400, detail="Jenis dokumen fisik tidak dikenal")
    return normalized


def physical_document_status_payload(user: Dict[str, Any]) -> Dict[str, Any]:
    raw_documents = user.get("physical_documents") or {}
    documents: Dict[str, Any] = {}
    for document_type, label in PHYSICAL_DOCUMENT_TYPES.items():
        item = raw_documents.get(document_type)
        if isinstance(item, dict) and (item.get("file_id") or item.get("id")):
            documents[document_type] = {
                **enrich_file_urls(dict(item)),
                "document_type": document_type,
                "label": label,
            }
    completed_count = len(documents)
    return {
        "required": [
            {"type": document_type, "label": label}
            for document_type, label in PHYSICAL_DOCUMENT_TYPES.items()
        ],
        "documents": documents,
        "completed_count": completed_count,
        "total_count": len(PHYSICAL_DOCUMENT_TYPES),
        "is_complete": completed_count == len(PHYSICAL_DOCUMENT_TYPES),
        "angkatan": str(user.get("angkatan") or user.get("academic_year") or "").strip(),
    }


STORAGE_ROOT = ROOT_DIR / "storage" / "E-Learning Dosen"
PMB_STORAGE_ROOT = ROOT_DIR / "storage" / "pmb"
DEFAULT_SUBMISSION_MAX_FILE_MB = 5
STORAGE_POLICY_TIMEZONE = ZoneInfo(os.environ.get("STORAGE_POLICY_TIMEZONE", "Asia/Jakarta"))
STORAGE_MAINTENANCE_INTERVAL_SECONDS = max(
    60,
    int(os.environ.get("STORAGE_MAINTENANCE_INTERVAL_SECONDS", "300")),
)
_storage_maintenance_scheduler_task: Optional[asyncio.Task] = None


def resolved_stored_file_path(file_doc: Dict[str, Any]) -> Optional[Path]:
    """Return a safe local path, preferring the portable storage_path field."""
    storage_path = str(file_doc.get("storage_path") or "")
    if not storage_path:
        storage_path = portable_storage_path_from_local_path(
            str(file_doc.get("local_path") or "")
        )
    portable_path = resolve_storage_local_path(
        ROOT_DIR / "storage",
        storage_path,
    )
    if portable_path and portable_path.exists() and portable_path.is_file():
        return portable_path

    local_path = str(file_doc.get("local_path") or "").strip()
    if not local_path:
        return None
    candidate = Path(local_path).resolve()
    storage_roots = (STORAGE_ROOT.resolve(), PMB_STORAGE_ROOT.resolve())
    if any(root == candidate or root in candidate.parents for root in storage_roots):
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def safe_path_segment(value: str) -> str:
    cleaned = re.sub(r"[\\/]+", "-", str(value or "").strip())
    cleaned = re.sub(r"[^A-Za-z0-9._ \-()]+", "-", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-")
    return cleaned[:100] or "Tanpa Nama"


def build_local_file_path(
    hierarchy: List[str], student_nim: str, student_name: str, file_token: str, original_filename: str
) -> tuple[Path, str, str]:
    safe_hierarchy = [safe_path_segment(item) for item in hierarchy]
    student_folder = safe_path_segment(f"{student_nim or 'NO-NIM'} - {student_name}")
    safe_filename = safe_path_segment(original_filename or "submission.bin")
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    final_filename = f"{timestamp}_{file_token[:8]}_{safe_filename}"
    folder = STORAGE_ROOT.joinpath(*safe_hierarchy, student_folder)
    absolute_path = folder / final_filename
    relative_folder = str(Path("E-Learning Dosen").joinpath(*safe_hierarchy, student_folder))
    relative_file = str(Path(relative_folder) / final_filename)
    return absolute_path, relative_folder, relative_file


def local_file_urls(file_id: str) -> Dict[str, str]:
    return {
        "file_url": f"/api/files/{file_id}/download",
        "preview_url": f"/api/files/{file_id}/preview",
        "inline_url": f"/api/files/{file_id}/inline",
    }


async def multipart_uploads(request: Request, fields: List[str]) -> tuple[Any, List[UploadFile]]:
    form = await request.form()
    uploads: List[UploadFile] = []
    for field in fields:
        for item in form.getlist(field):
            if getattr(item, "filename", "") and hasattr(item, "read"):
                uploads.append(item)
    return form, uploads


def assignment_max_file_size_mb(assignment: Dict[str, Any]) -> float:
    raw_value = assignment.get("max_file_size_mb", assignment.get("max_submission_size_mb", DEFAULT_SUBMISSION_MAX_FILE_MB))
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        value = DEFAULT_SUBMISSION_MAX_FILE_MB
    return value if value > 0 else DEFAULT_SUBMISSION_MAX_FILE_MB


async def validate_upload_file_sizes(uploads: List[UploadFile], max_mb: float, label: str) -> None:
    max_bytes = int(max_mb * 1024 * 1024)
    for upload in uploads:
        file_size = 0
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            file_size += len(chunk)
            if file_size > max_bytes:
                raise HTTPException(status_code=400, detail=f"Ukuran file {upload.filename or label} maksimal {max_mb:g} MB")
        await upload.seek(0)


def enrich_file_urls(file_doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not file_doc:
        return file_doc
    file_id = file_doc.get("file_id") or file_doc.get("id")
    if file_id:
        file_doc.update(local_file_urls(file_id))
    return file_doc


def enrich_submission_file_urls(submission: Dict[str, Any]) -> Dict[str, Any]:
    if isinstance(submission.get("file"), dict):
        submission["file"] = enrich_file_urls(submission["file"])
    if isinstance(submission.get("files"), list):
        submission["files"] = [enrich_file_urls(item) if isinstance(item, dict) else item for item in submission["files"]]
    return submission


def mask_secret(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 6:
        return "******"
    return f"{value[:3]}***{value[-3:]}"


def normalize_phone(value: str) -> str:
    cleaned = re.sub(r"\D+", "", value or "")
    if cleaned.startswith("0"):
        cleaned = "62" + cleaned[1:]
    return cleaned


def normalize_http_base_url(value: str) -> str:
    base_url = (value or "").strip()
    if not base_url:
        return ""
    if not re.match(r"^https?://", base_url, flags=re.IGNORECASE):
        base_url = f"http://{base_url}"
    base_url = base_url.rstrip("/")
    if base_url.lower().endswith("/api"):
        base_url = base_url[:-4].rstrip("/")
    return base_url


def normalize_app_url(value: str) -> str:
    app_url = (value or "").strip()
    if not app_url or "domain-aplikasi" in app_url.lower():
        app_url = os.environ.get("APP_URL", "").strip()
    if not app_url:
        return ""
    if not re.match(r"^https?://", app_url, flags=re.IGNORECASE):
        if app_url.startswith(("localhost", "127.", "10.", "192.168.")) or re.match(r"^\d{1,3}(\.\d{1,3}){3}", app_url):
            app_url = f"http://{app_url}"
        else:
            app_url = f"https://{app_url}"
    return app_url.rstrip("/")


def build_password_reset_link(app_url: str, identifier: str) -> str:
    base_url = normalize_app_url(app_url)
    if not base_url:
        return "Buka halaman aplikasi dan pilih Lupa Password"
    parts = urlsplit(base_url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({"reset": "password", "identifier": identifier})
    path = parts.path or "/"
    return urlunsplit((parts.scheme, parts.netloc, path, urlencode(query), parts.fragment))


def build_app_fragment_link(app_url: str, fragment: str = "") -> str:
    base_url = normalize_app_url(app_url)
    if not base_url:
        return "Buka halaman aplikasi"
    parts = urlsplit(base_url)
    return urlunsplit((parts.scheme, parts.netloc, parts.path or "/", parts.query, fragment.lstrip("#")))


def waha_chat_id(phone_or_chat_id: str) -> str:
    raw = (phone_or_chat_id or "").strip()
    if raw.endswith(("@c.us", "@g.us", "@newsletter")):
        return raw
    phone = normalize_phone(raw)
    if not phone:
        raise RuntimeError("Nomor tujuan WhatsApp kosong")
    return f"{phone}@c.us"


def waha_headers(api_key: str = "") -> Dict[str, str]:
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    api_key = (api_key or "").strip()
    if api_key:
        headers["X-Api-Key"] = api_key
    return headers


def response_excerpt(response: httpx.Response) -> str:
    try:
        data = response.json()
        return json.dumps(data, ensure_ascii=False)[:600]
    except Exception:
        return response.text[:600]


def file_extension(file_doc: Dict[str, Any], path: Optional[Path] = None) -> str:
    source = file_doc.get("file_name") or file_doc.get("storage_path") or ""
    if path:
        source = source or path.name
    return source.rsplit(".", 1)[-1].lower() if "." in source else ""


def preview_kind(file_doc: Dict[str, Any], path: Path) -> str:
    mime = (file_doc.get("mime_type") or "").lower()
    ext = file_extension(file_doc, path)
    if mime == "application/pdf" or ext == "pdf":
        return "pdf"
    if mime.startswith("image/") or ext in {"png", "jpg", "jpeg", "webp", "gif"}:
        return "image"
    if ext in {"docx"}:
        return "docx"
    if ext in {"xlsx", "xlsm"}:
        return "xlsx"
    if ext in {"txt", "md", "csv", "json"} or mime.startswith("text/"):
        return "text"
    return "unsupported"


def html_panel(title: str, body: str) -> str:
    return f"<section class=\"doc-preview-section\"><h3>{html.escape(title)}</h3>{body}</section>"


def preview_docx_html(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml_data = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml_data)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    blocks: List[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        line = "".join(texts).strip()
        if line:
            blocks.append(f"<p>{html.escape(line)}</p>")
    if not blocks:
        blocks.append("<p class=\"doc-preview-muted\">Dokumen tidak memiliki teks yang bisa diekstrak.</p>")
    return html_panel(path.name, "".join(blocks[:1200]))


def preview_xlsx_html(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: List[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            rows_html: List[str] = []
            for row_index, row in enumerate(sheet.iter_rows(max_row=80, max_col=20, values_only=True), start=1):
                cells = ["" if value is None else str(value) for value in row]
                if not any(cells):
                    continue
                cell_tag = "th" if row_index == 1 else "td"
                row_html = "".join(f"<{cell_tag}>{html.escape(value)}</{cell_tag}>" for value in cells)
                rows_html.append(f"<tr>{row_html}</tr>")
            table = "<table>" + "".join(rows_html) + "</table>" if rows_html else "<p class=\"doc-preview-muted\">Sheet kosong.</p>"
            sections.append(f"<h3>{html.escape(sheet.title)}</h3>{table}")
    finally:
        workbook.close()
    return "<section class=\"doc-preview-section\">" + "".join(sections) + "</section>"


def preview_text_html(path: Path) -> str:
    content = path.read_text(encoding="utf-8", errors="replace")[:120000]
    return html_panel(path.name, f"<pre>{html.escape(content)}</pre>")


async def fetch_waha_session_status(client: httpx.AsyncClient, base_url: str, session: str, headers: Dict[str, str]) -> Optional[Dict[str, Any]]:
    response = await client.get(f"{base_url}/api/sessions/{session}", headers=headers)
    if response.status_code == 404:
        return None
    if response.status_code >= 400:
        raise RuntimeError(f"WAHA session check gagal ({response.status_code}): {response_excerpt(response)}")
    return response.json()


async def ensure_waha_session_working(client: httpx.AsyncClient, base_url: str, session: str, headers: Dict[str, str]) -> None:
    session_doc = await fetch_waha_session_status(client, base_url, session, headers)
    if session_doc is None:
        raise RuntimeError(f"Sesi WAHA '{session}' tidak ditemukan. Buat/start session di WAHA lalu scan QR.")
    status = str(session_doc.get("status") or "").upper()
    if status and status != "WORKING":
        raise RuntimeError(f"Sesi WAHA '{session}' belum WORKING (status: {status}). Start session dan scan QR di dashboard WAHA.")


def generate_otp() -> str:
    return "".join(str(secrets.randbelow(10)) for _ in range(6))


def local_reset_otp_enabled() -> bool:
    return os.environ.get("ALLOW_LOCAL_RESET_OTP", "").lower() in {"1", "true", "yes", "on"}


def identity_query(identifier: str) -> Dict[str, Any]:
    raw = (identifier or "").strip()
    lowered = raw.lower()
    normalized_phone = normalize_phone(raw)
    candidates: List[Dict[str, str]] = [
        {"email": lowered},
        {"username": lowered},
        {"employee_id": raw},
        {"nim": raw.upper()},
        {"nim": raw},
        {"whatsapp": raw},
    ]
    if normalized_phone and normalized_phone != raw:
        candidates.append({"whatsapp": normalized_phone})
    return {"$or": candidates}


async def find_unique_identity_user(
    identifier: str,
    *,
    ambiguous_status: int = 409,
    ambiguous_detail: str = "Identitas terhubung ke lebih dari satu akun. Gunakan email unik atau hubungi admin kampus.",
) -> Optional[Dict[str, Any]]:
    users = await db.users.find(identity_query(identifier), {"_id": 0}).to_list(2)
    if len(users) > 1:
        raise HTTPException(status_code=ambiguous_status, detail=ambiguous_detail)
    return users[0] if users else None


async def ensure_unique_identity_index(field: str) -> None:
    try:
        await db.users.create_index(field, unique=True, sparse=True)
    except UniqueViolationError:
        logger.error(
            "Unique index %s belum dapat diaktifkan karena data lama duplikat; "
            "fallback non-unique sementara dipasang. Jalankan audit/repair identitas.",
            field,
        )
        await db.users.create_index(field, unique=False, sparse=True)


DEFAULT_GRADE_PREDICATES = [
    {"label": "A", "min_score": 85, "max_score": 100},
    {"label": "B", "min_score": 70, "max_score": 84.99},
    {"label": "C", "min_score": 60, "max_score": 69.99},
    {"label": "D", "min_score": 50, "max_score": 59.99},
    {"label": "E", "min_score": 0, "max_score": 49.99},
]

DEFAULT_GRADE_WEIGHTS = {
    "tugas": 25.0,
    "uts": 35.0,
    "uas": 40.0,
}
GRADE_WEIGHT_COMPONENTS = tuple(DEFAULT_GRADE_WEIGHTS.keys())

CLASS_STATUS_ACTIVE = "active"
CLASS_STATUS_ENDED = "ended"
CLASS_STATUS_FINALIZED = "finalized"
CLASS_STATUS_ARCHIVED = "archived"
CLASS_STATUS_DELETED = "deleted"
CLASS_STATUSES_READ_ONLY = {CLASS_STATUS_FINALIZED, CLASS_STATUS_ARCHIVED}


def normalize_assessment_category(value: Any) -> str:
    category = str(value or "tugas").strip().lower()
    aliases = {
        "assignment": "tugas",
        "task": "tugas",
        "midterm": "uts",
        "final": "uas",
        "final_exam": "uas",
    }
    category = aliases.get(category, category)
    if category not in GRADE_WEIGHT_COMPONENTS:
        raise HTTPException(status_code=400, detail="Komponen nilai harus Tugas, UTS, atau UAS")
    return category


def grade_weights_from_document(value: Any) -> Dict[str, float]:
    if not isinstance(value, dict):
        return dict(DEFAULT_GRADE_WEIGHTS)
    return {
        component: round(float(value.get(component, DEFAULT_GRADE_WEIGHTS[component])), 2)
        for component in GRADE_WEIGHT_COMPONENTS
    }


def validate_grade_weights(value: Dict[str, Any]) -> Dict[str, float]:
    cleaned: Dict[str, float] = {}
    for component in GRADE_WEIGHT_COMPONENTS:
        try:
            weight = float(value.get(component, 0))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"Bobot {component.upper()} harus berupa angka") from exc
        if weight < 0 or weight > 100:
            raise HTTPException(status_code=400, detail=f"Bobot {component.upper()} harus berada di antara 0 dan 100")
        cleaned[component] = round(weight, 2)
    if abs(sum(cleaned.values()) - 100) > 0.01:
        raise HTTPException(status_code=400, detail="Total bobot Tugas, UTS, dan UAS harus tepat 100%")
    return cleaned


def default_whatsapp_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "provider": "disabled",
        "app_url": os.environ.get("APP_URL", ""),
        "fonnte_token": "",
        "fonnte_url": "https://api.fonnte.com/send",
        "waha_base_url": "",
        "waha_api_key": "",
        "waha_session": "default",
        "send_delay_seconds": 3,
        "typing_simulation_seconds": 30,
        "otp_template": "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}",
        "assignment_template": "Halo {name}, ada tugas baru: {title}. Kelas: {class_name}. Deadline: {deadline}. Link: {link}",
        "grade_template": "Halo {name}, tugas {title} sudah dinilai. Nilai: {grade} ({predicate}). Feedback: {feedback}. Link: {link}",
        "revision_template": "Halo {name}, tugas {title} perlu revisi. Catatan: {revision_note}. Link: {link}",
    }


DEFAULT_INTEGRATION_SETTINGS: Dict[str, Any] = {
    "sekolah": {
        "label": "API Data Sekolah Indonesia",
        "provider": "apiindonesia.id",
        "enabled": False,
        "base_url": "https://use.apiindonesia.id",
        "api_key": "",
    },
}


async def get_integration_settings(mask: bool = True) -> Dict[str, Any]:
    """Mengambil konfigurasi seluruh integrasi sistem (sekolah, dan integrasi lain ke depan).

    Disimpan di koleksi `integration_settings` dengan struktur:
    {"id": "main", "integrations": {"<nama>": {config...}}}
    """
    cached = _get_cached_settings("integration_settings")
    if cached is None:
        doc = await db.integration_settings.find_one({"id": "main"}, {"_id": 0})
        stored = (doc or {}).get("integrations", {})
        merged: Dict[str, Any] = {}
        for name, default in DEFAULT_INTEGRATION_SETTINGS.items():
            current = stored.get(name, {})
            if not isinstance(current, dict):
                current = {}
            merged[name] = {**default, **{k: v for k, v in current.items() if v is not None}}
        cached = {"id": "main", "integrations": merged}
        _set_cached_settings("integration_settings", cached)
    result = json.loads(json.dumps(cached))
    if mask:
        for cfg in result.get("integrations", {}).values():
            if isinstance(cfg, dict) and cfg.get("api_key"):
                cfg["api_key_masked"] = mask_secret(cfg["api_key"])
                cfg.pop("api_key", None)
    return result


def default_email_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "enabled": False,
        "smtp_host": os.environ.get("SMTP_HOST", ""),
        "smtp_port": int(os.environ.get("SMTP_PORT", "587")),
        "smtp_user": os.environ.get("SMTP_USER", ""),
        "smtp_password": os.environ.get("SMTP_PASSWORD", ""),
        "smtp_use_tls": os.environ.get("SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"},
        "from_name": os.environ.get("SMTP_FROM_NAME", "E-Learning Dosen"),
        "from_email": os.environ.get("SMTP_FROM_EMAIL", ""),
    }


async def get_email_settings(mask: bool = False) -> Dict[str, Any]:
    cached = _get_cached_settings("email_settings")
    if cached is None:
        settings = await db.email_settings.find_one({"id": "main"}, {"_id": 0})
        cached = {**default_email_settings(), **(settings or {})}
        _set_cached_settings("email_settings", cached)
    if mask:
        settings = cached.copy()
        settings["smtp_password_masked"] = mask_secret(settings.get("smtp_password", ""))
        settings.pop("smtp_password", None)
        return settings
    return cached


async def send_email_message(to_email: str, subject: str, html_body: str) -> Dict[str, Any]:
    settings = await get_email_settings(mask=False)
    if not settings.get("enabled"):
        return {"ok": False, "error": "Email belum diaktifkan"}
    host = settings.get("smtp_host", "")
    port = int(settings.get("smtp_port", 587))
    user = settings.get("smtp_user", "")
    password = settings.get("smtp_password", "")
    use_tls = settings.get("smtp_use_tls", True)
    from_name = settings.get("from_name", "E-Learning Dosen")
    from_email = settings.get("from_email", "") or user
    if not host or not from_email:
        return {"ok": False, "error": "SMTP host atau email pengirim belum dikonfigurasi"}
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(html_body, "html", "utf-8"))
        if port == 465:
            server = smtplib.SMTP_SSL(host, port, timeout=15)
        else:
            server = smtplib.SMTP(host, port, timeout=15)
            server.ehlo()
            if use_tls:
                try:
                    server.starttls()
                    server.ehlo()
                except smtplib.SMTPNotSupportedError:
                    pass
        auth_capable = server.has_extn("AUTH")
        if user and password and auth_capable:
            server.login(user, password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()
        return {"ok": True}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def default_app_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "app_name": "SIAKAD ONE",
        "meta_description": "Sistem Informasi Akademik terpadu untuk mengelola pembelajaran, presensi, penilaian, dan layanan akademik perguruan tinggi.",
        "campus_name": "POLITEKNIK SCI",
        "campus_code": "POLTEK-SCI",
        "institution_type": "Politeknik",
        "accreditation": "Unggul",
        "accreditation_sk": "SK BAN-PT No. 1024/SK/BAN-PT/Akred/PT/2024",
        "campus_motto": "Unggul, Berkarakter, Berbasis Industri & Teknologi",
        "campus_phone": "(021) 789-0123",
        "campus_whatsapp": "0812-3456-7890",
        "campus_email": "info@politekniksci.ac.id",
        "campus_website": "https://politekniksci.ac.id",
        "campus_address": "Jl. Pendidikan Raya No. 45, Kompleks Akademik SCI, Jakarta",
        "program_name": "Program Studi",
        "lecturer_name": "Syahrul Anwar, M.Kom",
        "lecturer_email": "syahrul@politekniksci.ac.id",
        "app_logo_url": "",
        "campus_logo_url": "",
        "rector_name": "Prof. Dr. Ir. H. Ahmad Dahlan, M.T.",
        "rector_nidn": "0012056801",
        "vice_rector_1": "Dr. Eng. Rina Wati, M.Eng.",
        "head_of_baak": "Drs. Budi Santoso, M.Si.",
        "head_of_lppm": "Dr. Ir. Hendra Wijaya, M.T.",
        "kop_letterhead": "KEMENTERIAN PENDIDIKAN, KEBUDAYAAN, RISET, DAN TEKNOLOGI\nPOLITEKNIK SCI JAKARTA",
        "active_academic_year": "2025/2026",
        "active_semester": "Genap",
        "min_attendance_percentage": 75,
    }


async def get_app_settings_cached() -> Dict[str, Any]:
    cached = _get_cached_settings("app_settings")
    if cached is None:
        cached = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or default_app_settings()
        _set_cached_settings("app_settings", cached)
    return cached


def validate_grade_predicates(predicates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    cleaned = []
    for item in predicates:
        label = str(item.get("label", "")).strip().upper()
        min_score = float(item.get("min_score", 0))
        max_score = float(item.get("max_score", 0))
        if not label:
            raise HTTPException(status_code=400, detail="Label predikat wajib diisi")
        if min_score < 0 or max_score > 100 or min_score > max_score:
            raise HTTPException(status_code=400, detail="Range predikat harus 0-100 dan minimum <= maksimum")
        cleaned.append({"label": label, "min_score": min_score, "max_score": max_score})
    ordered = sorted(cleaned, key=lambda item: item["min_score"])
    for idx in range(len(ordered) - 1):
        if ordered[idx + 1]["min_score"] <= ordered[idx]["max_score"]:
            raise HTTPException(status_code=400, detail="Range predikat tidak boleh tumpang tindih")
    return cleaned


async def get_grade_predicates_for_class(class_id: str = "") -> List[Dict[str, Any]]:
    if class_id:
        class_doc = await db.grade_predicates.find_one({"class_id": class_id}, {"_id": 0})
        if class_doc:
            return class_doc.get("predicates", DEFAULT_GRADE_PREDICATES)
    global_doc = await db.grade_predicates.find_one({"class_id": ""}, {"_id": 0})
    return global_doc.get("predicates", DEFAULT_GRADE_PREDICATES) if global_doc else DEFAULT_GRADE_PREDICATES


async def calculate_grade_predicate(score: float, class_id: str = "") -> str:
    predicates = await get_grade_predicates_for_class(class_id)
    for item in predicates:
        if float(item["min_score"]) <= float(score) <= float(item["max_score"]):
            return item["label"]
    return "-"


async def get_whatsapp_settings(mask: bool = False) -> Dict[str, Any]:
    cached = _get_cached_settings("whatsapp_settings")
    if cached is None:
        settings = await db.whatsapp_settings.find_one({"id": "main"}, {"_id": 0})
        cached = {**default_whatsapp_settings(), **(settings or {})}
        _set_cached_settings("whatsapp_settings", cached)
    if mask:
        settings = cached.copy()
        settings["fonnte_token_masked"] = mask_secret(settings.get("fonnte_token", ""))
        settings["waha_api_key_masked"] = mask_secret(settings.get("waha_api_key", ""))
        settings.pop("fonnte_token", None)
        settings.pop("waha_api_key", None)
        return settings
    return cached


class SafeTemplateContext(dict):
    def __missing__(self, key: str) -> str:
        return "{" + key + "}"


def format_message_template(template: str, context: Dict[str, Any]) -> str:
    cleaned = {key: "" if value is None else str(value) for key, value in context.items()}
    try:
        return template.format_map(SafeTemplateContext(cleaned))
    except Exception:
        return template


def format_message_datetime(value: str) -> str:
    if not value:
        return "-"
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return value


def whatsapp_send_delay_seconds(settings: Dict[str, Any]) -> int:
    try:
        delay = int(settings.get("send_delay_seconds", 3))
    except Exception:
        delay = 3
    return max(0, min(delay, 300))


async def wait_for_whatsapp_send_window(delay_seconds: int) -> None:
    if delay_seconds <= 0:
        return
    gate = await db.whatsapp_runtime.find_one({"id": "send_gate"}, {"_id": 0})
    last_attempt = (gate or {}).get("last_attempt_at", "")
    if last_attempt:
        try:
            last_at = datetime.fromisoformat(last_attempt.replace("Z", "+00:00"))
            wait_seconds = delay_seconds - (datetime.now(timezone.utc) - last_at).total_seconds()
            if wait_seconds > 0:
                await asyncio.sleep(wait_seconds)
        except Exception:
            pass
    await db.whatsapp_runtime.update_one({"id": "send_gate"}, {"$set": {"last_attempt_at": now_iso()}}, upsert=True)


async def enqueue_whatsapp_message(to: str, message: str, message_type: str, ref_id: str = "") -> Dict[str, Any]:
    settings = await get_whatsapp_settings(mask=False)
    provider = settings.get("provider", "disabled")
    status = "pending_config" if provider == "disabled" else "pending"
    doc = {
        "id": new_id(),
        "to": normalize_phone(to),
        "message": message,
        "message_type": message_type,
        "ref_id": ref_id,
        "provider": provider,
        "status": status,
        "response": "",
        "error": "Gateway belum dikonfigurasi" if status == "pending_config" else "",
        "created_at": now_iso(),
        "sent_at": "",
    }
    await db.whatsapp_messages.insert_one(doc)
    return public_doc(doc.copy())


def public_whatsapp_delivery_status(message: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not message:
        return {"message_id": "", "status": "", "provider": "", "created_at": "", "sent_at": "", "error": ""}
    status = message.get("status", "")
    return {
        "message_id": message.get("id", ""),
        "status": status,
        "provider": message.get("provider", ""),
        "created_at": message.get("created_at", ""),
        "sent_at": message.get("sent_at", ""),
        "error": message.get("error", "") if status in {"failed", "pending_config"} else "",
    }


def forgot_password_response_message(delivery: Dict[str, Any]) -> str:
    status = delivery.get("status", "")
    provider = delivery.get("provider", "")
    if status == "pending":
        provider_label = f" via {provider}" if provider and provider != "disabled" else ""
        return f"Permintaan reset password diproses. OTP masuk antrian WhatsApp{provider_label}. Tunggu sampai status berubah menjadi terkirim."
    if status == "sent":
        return "Permintaan reset password diproses. OTP sudah dikirim via WhatsApp."
    if status == "pending_config":
        return "Permintaan reset password diproses. OTP dibuat, tetapi gateway WhatsApp belum aktif. Pesan masuk antrian konfigurasi."
    if status == "no_whatsapp":
        return "Permintaan reset password diproses, tetapi nomor WhatsApp belum terdaftar."
    return "Permintaan reset password diproses."


async def simulate_whatsapp_typing(
    client: httpx.AsyncClient,
    provider: str,
    to: str,
    settings: Dict[str, Any],
    typing_seconds: int,
) -> None:
    if typing_seconds <= 0:
        return
    try:
        if provider == "fonnte":
            token = settings.get("fonnte_token", "")
            if token:
                await client.post(
                    settings.get("fonnte_url") or "https://api.fonnte.com/send",
                    data={"target": to, "message": "...", "delay": str(typing_seconds), "typing": "true"},
                    headers={"Authorization": token},
                    timeout=10,
                )
        elif provider == "waha":
            base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
            api_key = settings.get("waha_api_key", "")
            session = (settings.get("waha_session") or "default").strip() or "default"
            if base_url:
                headers = waha_headers(api_key)
                chat_id = waha_chat_id(to)
                for _ in range(typing_seconds // 5):
                    await client.post(
                        f"{base_url}/api/sendPresence",
                        json={"session": session, "chatId": chat_id, "presence": "typing"},
                        headers=headers,
                        timeout=10,
                    )
                    await asyncio.sleep(5)
                remaining = typing_seconds % 5
                if remaining > 0:
                    await asyncio.sleep(remaining)
                return
        await asyncio.sleep(typing_seconds)
    except Exception:
        logger.debug("Typing simulation gagal, melanjutkan kirim pesan", exc_info=True)
        await asyncio.sleep(typing_seconds)


async def send_whatsapp_message(message_id: str) -> None:
    msg = await db.whatsapp_messages.find_one({"id": message_id}, {"_id": 0})
    if not msg:
        return
    provider = "disabled"
    async with _whatsapp_send_lock:
        settings = await get_whatsapp_settings(mask=False)
        provider = settings.get("provider", "disabled")
        if provider == "disabled":
            await db.whatsapp_messages.update_one(
                {"id": message_id}, {"$set": {"status": "pending_config", "error": "Gateway belum dikonfigurasi"}}
            )
            return
        try:
            await wait_for_whatsapp_send_window(whatsapp_send_delay_seconds(settings))
            typing_seconds = max(0, min(int(settings.get("typing_simulation_seconds", 30)), 120))
            async with httpx.AsyncClient(timeout=20) as client:
                await simulate_whatsapp_typing(client, provider, msg["to"], settings, typing_seconds)
                if provider == "fonnte":
                    token = settings.get("fonnte_token", "")
                    if not token:
                        raise RuntimeError("Token Fonnte belum diisi")
                    response = await client.post(
                        settings.get("fonnte_url") or "https://api.fonnte.com/send",
                        data={"target": msg["to"], "message": msg["message"]},
                        headers={"Authorization": token},
                    )
                elif provider == "waha":
                    base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
                    api_key = settings.get("waha_api_key", "")
                    session = (settings.get("waha_session") or "default").strip() or "default"
                    if not base_url:
                        raise RuntimeError("WAHA Base URL belum diisi")
                    headers = waha_headers(api_key)
                    await ensure_waha_session_working(client, base_url, session, headers)
                    response = await client.post(
                        f"{base_url}/api/sendText",
                        json={"session": session, "chatId": waha_chat_id(msg["to"]), "text": msg["message"]},
                        headers=headers,
                    )
                else:
                    raise RuntimeError("Provider WhatsApp tidak dikenal")
            if response.status_code >= 400:
                raise RuntimeError(f"Gateway {provider} gagal ({response.status_code}): {response_excerpt(response)}")
            await db.whatsapp_messages.update_one(
                {"id": message_id},
                {"$set": {"status": "sent", "response": response.text[:2000], "sent_at": now_iso(), "provider": provider, "error": ""}},
            )
        except Exception as exc:
            await db.whatsapp_messages.update_one(
                {"id": message_id},
                {"$set": {"status": "failed", "error": str(exc), "provider": provider}},
            )


async def queue_student_whatsapp_message(
    student: Dict[str, Any],
    message: str,
    message_type: str,
    ref_id: str,
    background_tasks: BackgroundTasks,
) -> Optional[Dict[str, Any]]:
    if not normalize_phone(student.get("whatsapp", "")):
        return None
    queued = await enqueue_whatsapp_message(student.get("whatsapp", ""), message, message_type, ref_id)
    if queued.get("status") == "pending":
        background_tasks.add_task(send_whatsapp_message, queued["id"])
    return queued


def unique_ids(values: List[str]) -> List[str]:
    seen = set()
    result = []
    for value in values or []:
        item = str(value or "").strip()
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


async def get_manageable_class(class_id: str, active_only: bool = False, user: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    query: Dict[str, Any] = {"id": class_id}
    if active_only:
        query["status"] = "active"
    else:
        query["status"] = {"$ne": "deleted"}
    if user and user.get("role") == "lecturer":
        query["lecturer_id"] = user["id"]
    class_doc = await db.classes.find_one(query, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas aktif tidak ditemukan" if active_only else "Kelas tidak ditemukan")
    return class_doc


def class_status_label(status: str) -> str:
    return {
        CLASS_STATUS_ACTIVE: "Aktif",
        CLASS_STATUS_ENDED: "Berakhir",
        CLASS_STATUS_FINALIZED: "Nilai difinalisasi",
        CLASS_STATUS_ARCHIVED: "Arsip",
        CLASS_STATUS_DELETED: "Dihapus",
    }.get(status, "Tidak diketahui")


def class_is_read_only(class_doc: Dict[str, Any]) -> bool:
    return class_doc.get("status") in CLASS_STATUSES_READ_ONLY


def class_allows_learning(class_doc: Dict[str, Any]) -> bool:
    return class_doc.get("status") == CLASS_STATUS_ACTIVE


def class_allows_grading(class_doc: Dict[str, Any]) -> bool:
    return class_doc.get("status") in {CLASS_STATUS_ACTIVE, CLASS_STATUS_ENDED}


async def require_class_mutation_access(class_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    class_doc = await require_class_access(class_id, user)
    if not class_allows_learning(class_doc):
        raise HTTPException(
            status_code=409,
            detail=f"Kelas {class_status_label(class_doc.get('status', ''))} bersifat read-only. Buat kelas baru untuk semester berikutnya.",
        )
    return class_doc


async def require_class_grading_access(class_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    class_doc = await require_class_access(class_id, user)
    if not class_allows_grading(class_doc):
        raise HTTPException(
            status_code=409,
            detail=f"Penilaian tidak dapat diubah karena kelas berstatus {class_status_label(class_doc.get('status', ''))}.",
        )
    return class_doc


async def get_active_student(student_id: str) -> Dict[str, Any]:
    student = await db.users.find_one({"id": student_id, "role": "student"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Mahasiswa tidak ditemukan")
    if student.get("status", "active") != "active":
        raise HTTPException(status_code=400, detail="Hanya mahasiswa aktif yang bisa diproses")
    return student


async def add_student_to_class_record(class_doc: Dict[str, Any], student: Dict[str, Any], actor_id: str) -> Dict[str, Any]:
    class_id = class_doc["id"]
    student_id = student["id"]
    already_joined = class_id in student.get("class_ids", [])
    await db.users.update_one({"id": student_id}, {"$addToSet": {"class_ids": class_id}})
    await db.classes.update_one({"id": class_id}, {"$addToSet": {"student_ids": student_id}})
    await db.enrollment_requests.update_many(
        {"class_id": class_id, "student_id": student_id, "status": {"$in": ["pending", "invited"]}},
        {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": actor_id}},
    )
    return {
        "student_id": student_id,
        "student_name": student.get("name", ""),
        "status": "already_joined" if already_joined else "approved",
    }


async def invite_student_to_class_record(
    class_doc: Dict[str, Any],
    student: Dict[str, Any],
    background_tasks: BackgroundTasks,
    actor_id: str,
) -> Dict[str, Any]:
    class_id = class_doc["id"]
    if class_id in student.get("class_ids", []):
        return {
            "student_id": student["id"],
            "student_name": student.get("name", ""),
            "status": "already_joined",
            "delivery_status": "",
            "message_id": "",
        }

    now = now_iso()
    existing = await db.enrollment_requests.find_one(
        {"class_id": class_id, "student_id": student["id"], "status": {"$in": ["pending", "invited"]}}, {"_id": 0}
    )
    request_doc = {
        "id": existing.get("id") if existing else new_id(),
        "class_id": class_doc["id"],
        "class_name": class_doc.get("name", ""),
        "class_code": class_doc.get("class_code", ""),
        "lecturer_id": class_doc.get("lecturer_id", ""),
        "lecturer_name": class_doc.get("lecturer_name", ""),
        "student_id": student["id"],
        "student_name": student.get("name", ""),
        "student_nim": student.get("nim", ""),
        "student_email": student.get("email", ""),
        "status": existing.get("status", "invited") if existing else "invited",
        "requested_at": existing.get("requested_at", now) if existing else now,
        "invited_at": now,
        "invited_by": actor_id,
    }

    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""))
    message = format_message_template(
        "Halo {name}, Anda diundang bergabung ke kelas {class_name}. Kode kelas: {class_code}. Login ke aplikasi lalu masukkan kode kelas tersebut: {link}",
        {
            "name": student.get("name", ""),
            "nim": student.get("nim", ""),
            "class_name": class_doc.get("name", ""),
            "course_name": class_doc.get("course_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "link": link,
        },
    )
    queued = await queue_student_whatsapp_message(student, message, "invite_kelas", request_doc["id"], background_tasks)
    request_doc["delivery_status"] = queued.get("status", "") if queued else "no_whatsapp"
    request_doc["message_id"] = queued.get("id", "") if queued else ""
    if existing:
        await db.enrollment_requests.update_one({"id": existing["id"]}, {"$set": request_doc})
    else:
        await db.enrollment_requests.insert_one(request_doc)
    return public_doc(request_doc.copy())


async def active_students_for_class(class_id: str, student_ids: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    conditions: List[Dict[str, Any]] = [{"class_ids": class_id}]
    if student_ids:
        conditions.append({"id": {"$in": student_ids}})
    students = await db.users.find(
        {
            "role": "student",
            "status": "active",
            "whatsapp": {"$exists": True, "$ne": ""},
            "$or": conditions,
        },
        {"_id": 0},
    ).to_list(2000)
    deduped: Dict[str, Dict[str, Any]] = {}
    for student in students:
        if normalize_phone(student.get("whatsapp", "")):
            deduped[student["id"]] = student
    return list(deduped.values())


async def notify_new_assignment_whatsapp(doc: Dict[str, Any], class_doc: Dict[str, Any], background_tasks: BackgroundTasks) -> None:
    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""), f"assignment-{doc['id']}")
    students = await active_students_for_class(doc["class_id"], class_doc.get("student_ids", []))
    template = settings.get("assignment_template") or default_whatsapp_settings()["assignment_template"]
    for student in students:
        message = format_message_template(
            template,
            {
                "name": student.get("name", ""),
                "nim": student.get("nim", ""),
                "title": doc.get("title", ""),
                "class_name": doc.get("class_name", ""),
                "course_name": doc.get("course_name", ""),
                "deadline": format_message_datetime(doc.get("deadline", "")),
                "description": doc.get("description", ""),
                "link": link,
            },
        )
        await queue_student_whatsapp_message(student, message, "tugas_baru", doc["id"], background_tasks)


async def create_assignment_publication_reminders(doc: Dict[str, Any], class_doc: Dict[str, Any]) -> None:
    for student_id in class_doc.get("student_ids", []):
        await db.reminder_logs.insert_one(
            {
                "id": new_id(),
                "assignment_id": doc["id"],
                "student_id": student_id,
                "reminder_type": "tugas_baru",
                "sent_at": now_iso(),
                "status": "in_app",
                "response": "Reminder tampil di aplikasi",
            }
        )


async def send_assignment_publication_notifications(
    doc: Dict[str, Any], class_doc: Dict[str, Any], background_tasks: BackgroundTasks
) -> None:
    await create_assignment_publication_reminders(doc, class_doc)
    await notify_new_assignment_whatsapp(doc, class_doc, background_tasks)


async def dispatch_due_assignment_notifications(background_tasks: BackgroundTasks) -> None:
    query = {
        "is_active": True,
        "published_at": {"$nin": ["", None]},
        "$or": [{"published_notification_sent_at": {"$exists": False}}, {"published_notification_sent_at": ""}],
    }
    scheduled = await db.assignments.find(query, {"_id": 0}).to_list(1000)
    for assignment in scheduled:
        if not assignment_is_published(assignment):
            continue
        update = await db.assignments.update_one(
            {
                "id": assignment["id"],
                "$or": [{"published_notification_sent_at": {"$exists": False}}, {"published_notification_sent_at": ""}],
            },
            {"$set": {"published_notification_sent_at": now_iso()}},
        )
        if update.modified_count:
            class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
            await send_assignment_publication_notifications(assignment, class_doc, background_tasks)


async def notify_submission_status_whatsapp(
    submission: Dict[str, Any],
    assignment: Dict[str, Any],
    message_type: str,
    background_tasks: BackgroundTasks,
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    student = await db.users.find_one({"id": submission.get("student_id")}, {"_id": 0})
    if not student:
        return
    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""), f"assignment-{submission.get('assignment_id', '')}")
    defaults = default_whatsapp_settings()
    template_key = "revision_template" if message_type == "revisi_tugas" else "grade_template"
    context = {
        "name": student.get("name", submission.get("student_name", "")),
        "nim": student.get("nim", submission.get("student_nim", "")),
        "title": assignment.get("title") or submission.get("assignment_title", ""),
        "class_name": assignment.get("class_name", ""),
        "course_name": assignment.get("course_name", ""),
        "grade": submission.get("grade", ""),
        "predicate": submission.get("grade_predicate", ""),
        "feedback": submission.get("feedback", ""),
        "revision_note": submission.get("revision_note", ""),
        "link": link,
    }
    context.update(extra or {})
    message = format_message_template(settings.get(template_key) or defaults[template_key], context)
    await queue_student_whatsapp_message(student, message, message_type, submission.get("id", ""), background_tasks)


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, password_hash: str) -> bool:
    return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))


async def find_user(user_id: str, include_password: bool = True) -> Dict[str, Any]:
    projection: Dict[str, int] = {"_id": 0}
    if not include_password:
        projection["password_hash"] = 0
    user = await db.users.find_one({"id": user_id}, projection)
    if not user:
        raise HTTPException(status_code=401, detail="Sesi tidak valid")
    if user.get("status", "active") != "active":
        raise HTTPException(status_code=403, detail="Akun tidak aktif")
    return user


async def get_current_user(
    request: Request,
    authorization: Optional[str] = Header(None),
) -> Dict[str, Any]:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token diperlukan")
    token = authorization.replace("Bearer ", "", 1).strip()
    cached = _auth_cache.get(token)
    if cached and (time.monotonic() - cached[0]) < _AUTH_CACHE_TTL:
        session, user = cached[1], cached[2]
    else:
        async with _auth_cache_lock:
            cached = _auth_cache.get(token)
            if cached and (time.monotonic() - cached[0]) < _AUTH_CACHE_TTL:
                session, user = cached[1], cached[2]
            else:
                session = await db.sessions.find_one({"token": token}, {"_id": 0})
                if not session:
                    raise HTTPException(status_code=401, detail="Sesi tidak ditemukan")
                user = await db.users.find_one(
                    {"id": session["user_id"]},
                    {"_id": 0, "password_hash": 0},
                )
                if not user:
                    applicant = await db.pmb_applicants.find_one(
                        {"id": session["user_id"]},
                        {"_id": 0, "password_hash": 0},
                    )
                    if applicant:
                        applicant["role"] = "camaba"
                        user = applicant
                if not user:
                    raise HTTPException(status_code=401, detail="Sesi tidak valid")
                if user.get("role") != "camaba":
                    user["role"] = normalize_base_role(user.get("role"))
                if user.get("role") != "camaba" and user.get("status", "active") != "active":
                    raise HTTPException(status_code=403, detail="Akun tidak aktif")
                cache_authenticated_user(token, session, user)
    if user.get("role") != "camaba":
        user["role"] = normalize_base_role(user.get("role"))
    request.state.current_user = user
    request.state.current_session = session
    return user


async def require_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") not in {"admin", "lecturer"}:
        raise HTTPException(status_code=403, detail="Hanya dosen atau admin kampus")
    return user


async def require_admin_or_academic_operator(
    user: Dict[str, Any] = Depends(get_current_user),
) -> Dict[str, Any]:
    if not user_is_admin_or_access_role(user, "academic_operator"):
        raise HTTPException(status_code=403, detail="Hanya admin kampus atau operator akademik")
    return user


async def require_lecturer_or_academic_manager(
    user: Dict[str, Any] = Depends(get_current_user),
) -> Dict[str, Any]:
    if user.get("role") == "lecturer" or user_is_admin_or_access_role(user, "academic_operator"):
        return user
    raise HTTPException(status_code=403, detail="Akses hanya untuk dosen, admin kampus, atau operator akademik")


async def require_student_records_reader(
    user: Dict[str, Any] = Depends(get_current_user),
) -> Dict[str, Any]:
    if user.get("role") == "lecturer" or user_is_admin_or_access_role(
        user,
        "academic_operator",
        "finance_officer",
    ):
        return user
    raise HTTPException(status_code=403, detail="Akses data mahasiswa tidak diizinkan")


async def require_admin_or_operational_staff(
    user: Dict[str, Any] = Depends(get_current_user),
) -> Dict[str, Any]:
    if not user_is_admin_or_access_role(user, "academic_operator", "finance_officer"):
        raise HTTPException(status_code=403, detail="Akses hanya untuk admin atau staf operasional yang ditugaskan")
    return user


async def require_campus_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Hanya admin kampus")
    return user


def is_campus_admin(user: Dict[str, Any]) -> bool:
    return normalize_base_role(user.get("role")) == "admin"


def is_academic_operator(user: Dict[str, Any]) -> bool:
    return user_is_admin_or_access_role(user, "academic_operator")


CALENDAR_EVENT_CATEGORIES = {
    "academic",
    "registration",
    "krs",
    "exam",
    "graduation",
    "holiday",
    "finance",
    "campus",
}
CALENDAR_EVENT_AUDIENCES = {"all", "student", "lecturer", "staff"}
CALENDAR_EVENT_STATUSES = {"draft", "published", "archived"}
ACADEMIC_DEADLINE_DEFINITIONS = {
    "curriculum_setup": {
        "title": "Deadline Setting Kurikulum",
        "target_role": "kaprodi",
        "target_label": "Kaprodi",
        "description": "Batas waktu penyiapan dan penetapan kurikulum oleh Ketua Program Studi.",
    },
    "rps_submission": {
        "title": "Deadline Pengisian RPS",
        "target_role": "lecturer",
        "target_label": "Dosen",
        "description": "Batas waktu dosen melengkapi Rencana Pembelajaran Semester.",
    },
    "grade_entry": {
        "title": "Deadline Pengisian Nilai",
        "target_role": "lecturer",
        "target_label": "Dosen",
        "description": "Batas waktu dosen menyelesaikan pengisian nilai mahasiswa.",
    },
}


def can_manage_academic_calendar(user: Dict[str, Any]) -> bool:
    """Administrasi kalender berada pada Admin Kampus dan operator akademik."""
    return is_campus_admin(user) or "academic_operator" in (user.get("access_roles") or [])


def is_kaprodi_user(user: Dict[str, Any]) -> bool:
    return user_is_program_manager(user)


def academic_deadline_visible_to_user(deadline_type: str, user: Dict[str, Any]) -> bool:
    definition = ACADEMIC_DEADLINE_DEFINITIONS.get(deadline_type)
    if not definition:
        return False
    if can_manage_academic_calendar(user):
        return True
    if str(user.get("role") or "").lower() not in {"lecturer", "dosen"}:
        return False
    if definition["target_role"] == "lecturer":
        return True
    return definition["target_role"] == "kaprodi" and is_kaprodi_user(user)


def default_academic_deadlines() -> Dict[str, Dict[str, Any]]:
    return {
        deadline_type: {
            "deadline_type": deadline_type,
            **definition,
            "enabled": False,
            "deadline_at": "",
        }
        for deadline_type, definition in ACADEMIC_DEADLINE_DEFINITIONS.items()
    }


def academic_deadline_settings_payload(document: Dict[str, Any]) -> Dict[str, Any]:
    deadlines = default_academic_deadlines()
    stored_deadlines = document.get("deadlines") or {}
    for deadline_type, item in deadlines.items():
        stored = stored_deadlines.get(deadline_type) or {}
        item["enabled"] = bool(stored.get("enabled", False))
        item["deadline_at"] = str(stored.get("deadline_at") or "")
    return {
        "id": document.get("id", ""),
        "academic_year_id": str(document.get("academic_year_id") or ""),
        "deadlines": deadlines,
        "updated_at": document.get("updated_at", ""),
        "updated_by": document.get("updated_by", ""),
    }


def validate_academic_deadline_settings(
    payload: AcademicDeadlineSettingsInput,
) -> Dict[str, Any]:
    unknown_types = sorted(set(payload.deadlines) - set(ACADEMIC_DEADLINE_DEFINITIONS))
    if unknown_types:
        raise HTTPException(
            status_code=400,
            detail=f"Jenis deadline akademik tidak valid: {', '.join(unknown_types)}",
        )

    normalized = default_academic_deadlines()
    for deadline_type, item in normalized.items():
        submitted = payload.deadlines.get(deadline_type)
        if not submitted:
            continue
        deadline_at = normalize_optional_datetime(
            submitted.deadline_at,
            item["title"],
        )
        if submitted.enabled and not deadline_at:
            raise HTTPException(
                status_code=400,
                detail=f"Tanggal {item['title']} wajib diisi saat switch diaktifkan",
            )
        item["enabled"] = bool(submitted.enabled)
        item["deadline_at"] = deadline_at

    return {
        "academic_year_id": str(payload.academic_year_id or "").strip(),
        "deadlines": normalized,
    }


def academic_deadline_event_payload(
    deadline_type: str,
    item: Dict[str, Any],
    academic_year_id: str,
) -> Dict[str, Any]:
    definition = ACADEMIC_DEADLINE_DEFINITIONS[deadline_type]
    return {
        "id": f"academic-deadline-{academic_year_id or 'global'}-{deadline_type}",
        "source": "academic_deadline",
        "type": "academic_deadline",
        "deadline_type": deadline_type,
        "category": "academic",
        "title": definition["title"],
        "date": item.get("deadline_at", ""),
        "end_at": "",
        "all_day": False,
        "academic_year_id": academic_year_id,
        "audience": "lecturer",
        "target_role": definition["target_role"],
        "target_label": definition["target_label"],
        "description": definition["description"],
        "enabled": bool(item.get("enabled", False)),
    }


def user_calendar_prodi_ids(user: Dict[str, Any]) -> set[str]:
    values = {
        str(user.get(key) or "").strip()
        for key in ("prodi_id", "program_id", "kaprodi_prodi_id")
    }
    values.update(
        str(item or "").strip()
        for item in (user.get("access_scope_prodi_ids") or [])
    )
    return values - {""}


def calendar_event_visible_to_user(event: Dict[str, Any], user: Dict[str, Any]) -> bool:
    """Filter publikasi kalender berdasarkan sasaran peran dan prodi."""
    if event.get("status") != "published":
        return False
    if is_campus_admin(user) or can_manage_academic_calendar(user):
        return True

    audience = str(event.get("audience") or "all")
    role = str(user.get("role") or "")
    if audience == "student" and role != "student":
        return False
    if audience == "lecturer" and role not in {"lecturer", "admin"}:
        return False
    if audience == "staff" and role != "staff":
        return False

    target_prodi_ids = {
        str(item or "").strip()
        for item in (event.get("target_prodi_ids") or [])
    } - {""}
    return not target_prodi_ids or bool(target_prodi_ids & user_calendar_prodi_ids(user))


def calendar_event_payload(event: Dict[str, Any]) -> Dict[str, Any]:
    """Bentuk event institusi yang konsisten dengan endpoint kalender lama."""
    return {
        "id": f"academic-{event['id']}",
        "event_id": event["id"],
        "source": "academic_calendar",
        "type": "academic",
        "category": event.get("category", "academic"),
        "title": event.get("title", "Kegiatan akademik"),
        "date": event.get("start_at", ""),
        "end_at": event.get("end_at", ""),
        "all_day": bool(event.get("all_day", True)),
        "academic_year_id": event.get("academic_year_id", ""),
        "audience": event.get("audience", "all"),
        "target_prodi_ids": event.get("target_prodi_ids", []),
        "description": event.get("description", ""),
        "location": event.get("location", ""),
        "link": event.get("link", ""),
    }


def validate_academic_calendar_event(payload: AcademicCalendarEventInput) -> Dict[str, Any]:
    category = str(payload.category or "academic").strip().lower()
    audience = str(payload.audience or "all").strip().lower()
    status = str(payload.status or "published").strip().lower()
    if category not in CALENDAR_EVENT_CATEGORIES:
        raise HTTPException(status_code=400, detail="Kategori kegiatan kalender tidak valid")
    if audience not in CALENDAR_EVENT_AUDIENCES:
        raise HTTPException(status_code=400, detail="Sasaran publikasi kalender tidak valid")
    if status not in CALENDAR_EVENT_STATUSES:
        raise HTTPException(status_code=400, detail="Status kalender harus draft, published, atau archived")

    start_at = normalize_optional_datetime(payload.start_at, "Tanggal mulai")
    end_at = normalize_optional_datetime(payload.end_at, "Tanggal selesai")
    if end_at and parse_iso_datetime(end_at) < parse_iso_datetime(start_at):
        raise HTTPException(status_code=400, detail="Tanggal selesai tidak boleh sebelum tanggal mulai")

    return {
        "title": payload.title.strip(),
        "category": category,
        "start_at": start_at,
        "end_at": end_at,
        "all_day": bool(payload.all_day),
        "academic_year_id": str(payload.academic_year_id or "").strip(),
        "audience": audience,
        "target_prodi_ids": sorted({
            str(item or "").strip()
            for item in payload.target_prodi_ids
        } - {""}),
        "description": payload.description.strip(),
        "location": payload.location.strip(),
        "link": payload.link.strip(),
        "status": status,
    }


async def lecturer_class_ids(user: Dict[str, Any], include_deleted: bool = False) -> List[str]:
    if user.get("role") == "student":
        return list(user.get("class_ids", []))

    cache_key = f"{user.get('id', '')}:{int(include_deleted)}"
    cached = _class_scope_cache.get(cache_key)
    if cached and (time.monotonic() - cached[0]) < _CLASS_SCOPE_CACHE_TTL:
        return list(cached[1])

    # Several startup endpoints need the same lecturer scope concurrently.
    # Serialize only the cache fill so the first request does the query and
    # the other requests reuse its result instead of hitting PostgreSQL again.
    async with _class_scope_cache_lock:
        cached = _class_scope_cache.get(cache_key)
        if cached and (time.monotonic() - cached[0]) < _CLASS_SCOPE_CACHE_TTL:
            return list(cached[1])

        result: List[str]
        if is_campus_admin(user) or user_has_access_role(user, "academic_operator"):
            query = {} if include_deleted else {"status": {"$ne": "deleted"}}
            docs = await db.classes.find(query, {"_id": 0, "id": 1}).to_list(5000)
            result = [item["id"] for item in docs]
        else:
            target_ids = list({user.get("id", ""), user.get("username", ""), user.get("nidn", ""), user.get("employee_id", "")} - {""})
            class_ids = set()
            for tid in target_ids:
                docs = await db.classes.find({"lecturer_id": tid}, {"_id": 0, "id": 1, "status": 1}).to_list(5000)
                for d in docs:
                    if include_deleted or d.get("status") != "deleted":
                        class_ids.add(d["id"])
                docs_nidn = await db.classes.find({"lecturer_nidn": tid}, {"_id": 0, "id": 1, "status": 1}).to_list(5000)
                for d in docs_nidn:
                    if include_deleted or d.get("status") != "deleted":
                        class_ids.add(d["id"])
            result = list(class_ids)
        if len(_class_scope_cache) >= _CLASS_SCOPE_CACHE_MAX_ENTRIES:
            expired = [
                key
                for key, value in _class_scope_cache.items()
                if time.monotonic() - value[0] >= _CLASS_SCOPE_CACHE_TTL
            ]
            for key in expired:
                _class_scope_cache.pop(key, None)
            while len(_class_scope_cache) >= _CLASS_SCOPE_CACHE_MAX_ENTRIES:
                oldest = min(_class_scope_cache, key=lambda key: _class_scope_cache[key][0])
                _class_scope_cache.pop(oldest, None)
        _class_scope_cache[cache_key] = (time.monotonic(), result)
        return list(result)


async def require_class_access(class_id: str, user: Dict[str, Any], active_only: bool = False) -> Dict[str, Any]:
    class_doc = await db.classes.find_one({"id": class_id}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    if class_doc.get("status") == "deleted":
        raise HTTPException(status_code=404, detail="Kelas telah dihapus")
    if active_only and class_doc.get("status") != "active":
        raise HTTPException(status_code=409, detail="Kelas tidak aktif")

    if is_campus_admin(user):
        return class_doc

    if user.get("role") == "lecturer":
        target_ids = {user.get("id", ""), user.get("username", ""), user.get("nidn", ""), user.get("employee_id", "")} - {""}
        c_lid = str(class_doc.get("lecturer_id") or "").strip()
        c_nidn = str(class_doc.get("lecturer_nidn") or "").strip()
        c_name = str(class_doc.get("lecturer_name") or "").strip().upper()
        u_name = str(user.get("name") or "").strip().upper()
        if c_lid in target_ids or c_nidn in target_ids or (u_name and u_name in c_name):
            return class_doc
        raise HTTPException(status_code=404, detail="Bukan kelas yang Anda kelola")

    if user.get("role") == "student":
        student_identifiers = {user.get("id", ""), user.get("username", ""), user.get("nim", "")} - {""}
        enrolled_students = set(class_doc.get("student_ids", []))
        if class_id in user.get("class_ids", []) or bool(student_identifiers & enrolled_students):
            return class_doc
        raise HTTPException(status_code=404, detail="Anda tidak terdaftar di kelas ini")

    return class_doc


async def require_course_access(course_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    course = await db.courses.find_one({"id": course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    if user.get("role") == "lecturer":
        owned_class = await db.classes.find_one(
            {"course_id": course_id, "lecturer_id": user["id"], "status": {"$ne": "deleted"}},
            {"_id": 0, "id": 1},
        )
        if not owned_class:
            raise HTTPException(status_code=404, detail="Mata kuliah bukan bagian dari kelas yang Anda kelola")
    return course


async def require_submission_access(submission_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    submission = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    if not submission:
        if submission_id.startswith("draft_sub_"):
            remainder = submission_id[len("draft_sub_"):]
            rparts = remainder.rsplit("_", 1)
            if len(rparts) == 2:
                assignment_id, student_id = rparts[0], rparts[1]
                assignment = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
                student = await db.users.find_one({"$or": [{"id": student_id}, {"username": student_id}, {"nim": student_id}]}, {"_id": 0})
                if assignment:
                    class_id = assignment.get("class_id", "")
                    await require_class_grading_access(class_id, user)
                    student_name = student.get("name") if student else f"Mahasiswa ({student_id})"
                    student_nim = student.get("nim") or student.get("username") if student else student_id
                    doc = {
                        "id": submission_id,
                        "assignment_id": assignment_id,
                        "assignment_title": assignment.get("title", "Tugas"),
                        "student_id": student.get("id", student_id),
                        "student_name": student_name,
                        "student_nim": student_nim,
                        "class_id": class_id,
                        "status": "Dinilai",
                        "review_status": "graded",
                        "submitted_at": now_iso(),
                        "created_at": now_iso(),
                    }
                    await db.submissions.insert_one(doc)
                    return doc
        raise HTTPException(status_code=404, detail="Submission tidak ditemukan")
    await require_class_grading_access(submission.get("class_id", ""), user)
    return submission


def chat_conversation_id(first_user_id: str, second_user_id: str) -> str:
    participants = sorted([first_user_id, second_user_id])
    return hashlib.sha256(":".join(participants).encode("utf-8")).hexdigest()


def chat_user_payload(user: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": user.get("id", ""),
        "role": user.get("role", ""),
        "username": user.get("username", ""),
        "nim": user.get("nim", ""),
        "name": user.get("name", ""),
        "email": user.get("email", ""),
    }


def chat_contact_matches_query(user: Dict[str, Any], query: str) -> bool:
    """Match chat contacts by a partial, case-insensitive name or identifier."""
    normalized_query = str(query or "").strip().casefold()
    if not normalized_query:
        return True
    return any(
        normalized_query in str(user.get(field) or "").casefold()
        for field in ("name", "username", "email")
    )


async def chat_contact_payload(user: Dict[str, Any]) -> Dict[str, Any]:
    payload = chat_user_payload(user)
    if user.get("role") != "admin":
        return payload
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0, "lecturer_name": 1, "lecturer_email": 1}) or {}
    configured_name = str(settings.get("lecturer_name", "")).strip()
    configured_email = str(settings.get("lecturer_email", "")).strip().lower()
    applies_to_user = not configured_email or configured_email == str(user.get("email", "")).strip().lower()
    if applies_to_user and configured_name and configured_name.lower() != "dosen admin":
        payload["name"] = configured_name
    elif str(payload.get("name", "")).strip().lower() == "dosen admin":
        payload["name"] = "Dosen Pengampu"
    return payload


async def mark_chat_read(user_id: str, contact_id: str) -> None:
    read_at = now_iso()
    await db.chat_read_receipts.update_one(
        {"user_id": user_id, "contact_id": contact_id},
        {"$set": {"last_read_at": read_at, "updated_at": read_at}},
        upsert=True,
    )


async def chat_unread_count(contact_id: str, viewer: Dict[str, Any]) -> int:
    if chat_connections.is_viewing(viewer["id"], contact_id):
        return 0
    receipt = await db.chat_read_receipts.find_one(
        {"user_id": viewer["id"], "contact_id": contact_id},
        {"_id": 0, "last_read_at": 1},
    )
    query: Dict[str, Any] = {
        "conversation_id": chat_conversation_id(viewer["id"], contact_id),
        "recipient_id": viewer["id"],
    }
    last_read_at = str((receipt or {}).get("last_read_at") or "").strip()
    if last_read_at:
        query["created_at"] = {"$gt": last_read_at}
    return await db.chat_messages.count_documents(query)


async def chat_contact_view_payload(
    contact: Dict[str, Any], viewer: Dict[str, Any]
) -> Dict[str, Any]:
    """Build a contact row with live presence and unread message count."""
    viewing_chat = chat_connections.is_viewing(viewer["id"], contact["id"])
    unread_count = await chat_unread_count(contact["id"], viewer)
    return {
        **await chat_contact_payload(contact),
        "online": chat_connections.is_online(contact["id"]),
        "viewing_chat": viewing_chat,
        "unread_count": unread_count,
    }


class ChatConnectionManager:
    def __init__(self) -> None:
        self.connections: Dict[str, set[WebSocket]] = {}
        self.viewing: Dict[WebSocket, str] = {}

    def is_online(self, user_id: str) -> bool:
        return bool(self.connections.get(user_id))

    def is_viewing(self, viewer_id: str, target_user_id: str) -> bool:
        return any(self.viewing.get(socket) == target_user_id for socket in self.connections.get(viewer_id, set()))

    async def connect(self, user_id: str, websocket: WebSocket) -> bool:
        await websocket.accept()
        self.connections.setdefault(user_id, set()).add(websocket)
        try:
            await websocket.send_json({"type": "presence_snapshot", "online_user_ids": list(self.connections.keys())})
            await self.broadcast({"type": "presence", "user_id": user_id, "online": True})
            return True
        except Exception:
            sockets = self.connections.get(user_id, set())
            sockets.discard(websocket)
            if not sockets:
                self.connections.pop(user_id, None)
            return False

    async def disconnect(self, user_id: str, websocket: WebSocket) -> None:
        viewed_user_id = self.viewing.pop(websocket, "")
        sockets = self.connections.get(user_id, set())
        sockets.discard(websocket)
        if not sockets:
            self.connections.pop(user_id, None)
        if viewed_user_id:
            await self.send_to_users(
                {viewed_user_id},
                {"type": "chat_focus", "user_id": user_id, "viewing": self.is_viewing(user_id, viewed_user_id)},
            )
        await self.broadcast({"type": "presence", "user_id": user_id, "online": self.is_online(user_id)})

    async def set_viewing(self, user_id: str, websocket: WebSocket, target_user_id: str) -> None:
        previous = self.viewing.get(websocket, "")
        self.viewing[websocket] = target_user_id
        if previous and previous != target_user_id:
            await self.send_to_users(
                {previous},
                {"type": "chat_focus", "user_id": user_id, "viewing": self.is_viewing(user_id, previous)},
            )
        if target_user_id:
            await self.send_to_users(
                {target_user_id},
                {"type": "chat_focus", "user_id": user_id, "viewing": True},
            )

    async def send_to_users(self, user_ids: set[str], payload: Dict[str, Any]) -> None:
        for user_id in user_ids:
            for socket in list(self.connections.get(user_id, set())):
                try:
                    await socket.send_json(payload)
                except Exception:
                    await self.disconnect(user_id, socket)

    async def broadcast(self, payload: Dict[str, Any]) -> None:
        await self.send_to_users(set(self.connections.keys()), payload)


chat_connections = ChatConnectionManager()


class LoginInput(BaseModel):
    identifier: str = ""
    email: str = ""
    password: str = Field(min_length=3)


class SsoExchangeInput(BaseModel):
    ticket: str = Field(min_length=20)


class RegisterStudentInput(BaseModel):
    username: str = ""
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(min_length=3)


class ForgotPasswordInput(BaseModel):
    identifier: str


class ChangePasswordInput(BaseModel):
    current_password: str
    new_password: str = Field(min_length=3)


class ProfileInput(BaseModel):
    name: str = Field(min_length=1)
    username: str = Field(min_length=1)
    email: EmailStr
    whatsapp: str = ""
    employee_id: Optional[str] = None
    nip: Optional[str] = None
    nidn: Optional[str] = None
    nik: Optional[str] = None
    gelar: Optional[str] = None
    gelar_depan: Optional[str] = None
    gelar_belakang: Optional[str] = None
    prodi_id: Optional[str] = None
    prodi_name: Optional[str] = None
    prodi_kode: Optional[str] = None
    homebase: Optional[str] = None
    gender: Optional[str] = None
    agama: Optional[str] = None
    tempat_lahir: Optional[str] = None
    tanggal_lahir: Optional[str] = None
    alamat: Optional[str] = None
    kota: Optional[str] = None
    provinsi: Optional[str] = None
    kode_pos: Optional[str] = None
    spesialisasi: Optional[str] = None
    jabatan: Optional[str] = None
    status_kepegawaian: Optional[str] = None
    nim: Optional[str] = None
    nisn: Optional[str] = None
    angkatan: Optional[str] = None
    academic_year: Optional[str] = None
    semester: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_job: Optional[str] = None
    parent_address: Optional[str] = None
    nuptk: Optional[str] = None
    nrsd: Optional[str] = None
    nama_panggilan: Optional[str] = None
    kewarganegaraan: Optional[str] = None
    rt: Optional[str] = None
    rw: Optional[str] = None
    dusun: Optional[str] = None
    kelurahan: Optional[str] = None
    kecamatan: Optional[str] = None
    kode_wilayah: Optional[str] = None
    jenis_tinggal_id: Optional[str] = None
    jenis_tinggal: Optional[str] = None
    transportasi_id: Optional[str] = None
    transportasi: Optional[str] = None
    asal_sekolah: Optional[str] = None
    status_sipil: Optional[str] = None
    no_kk: Optional[str] = None
    npwp: Optional[str] = None
    no_kip: Optional[str] = None
    no_kps: Optional[str] = None
    kebutuhan_khusus: Optional[str] = None
    tinggi_badan: Optional[str] = None
    berat_badan: Optional[str] = None
    semester_masuk: Optional[str] = None
    tanggal_masuk: Optional[str] = None
    jenis_pendaftaran_id: Optional[str] = None
    jenis_pendaftaran: Optional[str] = None
    jalur_masuk_id: Optional[str] = None
    jalur_masuk: Optional[str] = None
    jenis_pembiayaan_id: Optional[str] = None
    jenis_pembiayaan: Optional[str] = None
    status_mahasiswa_id: Optional[str] = None
    feeder_student_id: Optional[str] = None
    feeder_registration_id: Optional[str] = None
    foto_url: Optional[str] = None
    jabatan_dikti_id: Optional[str] = None
    jabatan_kode: Optional[str] = None
    jenjang_pendidikan: Optional[str] = None
    ikatan_kerja: Optional[str] = None
    status_pegawai: Optional[str] = None
    status_kerja: Optional[str] = None
    jenis_pegawai: Optional[str] = None
    pangkat_golongan_id: Optional[str] = None
    pangkat_golongan: Optional[str] = None
    no_sk: Optional[str] = None
    unit_organisasi_id: Optional[str] = None
    institusi_induk: Optional[str] = None
    status_dosen: Optional[str] = None
    status_dosen_id: Optional[str] = None
    tanggal_mulai_mengajar: Optional[str] = None
    orang_tua: Optional[Dict[str, Any]] = None
    registration: Optional[Dict[str, Any]] = None
    pddikti_ids: Optional[Dict[str, Any]] = None


class ResetPasswordOtpInput(BaseModel):
    identifier: str
    otp: str
    new_password: str = Field(min_length=3)


class JoinClassInput(BaseModel):
    class_code: str
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(min_length=3)


class ProgramInput(BaseModel):
    code: str
    name: str
    description: str = ""


class CourseInput(BaseModel):
    program_id: str = ""
    code: str
    name: str
    credits: int = 3
    description: str = ""


class ClassInput(BaseModel):
    academic_year: str
    semester: str
    course_id: str
    name: str
    schedule: str = ""


class ClassDuplicateInput(BaseModel):
    academic_year: str = Field(min_length=1)
    semester: str = Field(min_length=1)
    name: str = Field(min_length=1)
    schedule: str = ""
    confirmation: str = ""


class StudentInput(BaseModel):
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    nik: str = ""
    nisn: str = ""
    gender: str = "L"
    agama: str = "Islam"
    tempat_lahir: str = ""
    tanggal_lahir: str = ""
    alamat: str = ""
    kota: str = ""
    provinsi: str = ""
    kode_pos: str = ""
    class_id: Optional[str] = ""
    prodi_id: Optional[str] = None
    angkatan: Optional[str] = "2024"
    dosen_wali_id: Optional[str] = None
    parent_name: str = ""
    parent_phone: str = ""
    parent_job: str = ""
    parent_address: str = ""
    parent_email: str = ""
    parent_rt: str = ""
    parent_rw: str = ""
    parent_kota: str = ""
    parent_provinsi: str = ""
    parent_kode_pos: str = ""
    parent_negara: str = ""
    kewarganegaraan: str = ""
    rt: str = ""
    rw: str = ""
    dusun: str = ""
    kelurahan: str = ""
    kecamatan: str = ""
    kode_wilayah: str = ""
    jenis_tinggal_id: str = ""
    jenis_tinggal: str = ""
    transportasi_id: str = ""
    transportasi: str = ""
    asal_sekolah: str = ""
    status_sipil: str = ""
    no_kk: str = ""
    npwp: str = ""
    no_kip: str = ""
    no_kps: str = ""
    kebutuhan_khusus: str = ""
    tinggi_badan: str = ""
    berat_badan: str = ""
    semester_masuk: str = ""
    tanggal_masuk: str = ""
    jenis_pendaftaran_id: str = ""
    jenis_pendaftaran: str = ""
    jalur_masuk_id: str = ""
    jalur_masuk: str = ""
    jenis_pembiayaan_id: str = ""
    jenis_pembiayaan: str = ""
    status_mahasiswa_id: str = ""
    feeder_student_id: str = ""
    feeder_registration_id: str = ""
    foto_url: str = ""
    orang_tua: Dict[str, Any] = Field(default_factory=dict)
    registration: Dict[str, Any] = Field(default_factory=dict)
    pddikti_ids: Dict[str, Any] = Field(default_factory=dict)
    status: str = "active"
    password: str = "Mahasiswa1231!"


class StudentUpdateInput(BaseModel):
    nim: str
    name: str
    email: EmailStr
    whatsapp: Optional[str] = None
    nik: Optional[str] = None
    nisn: Optional[str] = None
    gender: Optional[str] = None
    agama: Optional[str] = None
    tempat_lahir: Optional[str] = None
    tanggal_lahir: Optional[str] = None
    alamat: Optional[str] = None
    kota: Optional[str] = None
    provinsi: Optional[str] = None
    kode_pos: Optional[str] = None
    prodi_id: Optional[str] = None
    angkatan: Optional[str] = None
    dosen_wali_id: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_job: Optional[str] = None
    parent_address: Optional[str] = None
    parent_email: Optional[str] = None
    parent_rt: Optional[str] = None
    parent_rw: Optional[str] = None
    parent_kota: Optional[str] = None
    parent_provinsi: Optional[str] = None
    parent_kode_pos: Optional[str] = None
    parent_negara: Optional[str] = None
    kewarganegaraan: Optional[str] = None
    rt: Optional[str] = None
    rw: Optional[str] = None
    dusun: Optional[str] = None
    kelurahan: Optional[str] = None
    kecamatan: Optional[str] = None
    kode_wilayah: Optional[str] = None
    jenis_tinggal_id: Optional[str] = None
    jenis_tinggal: Optional[str] = None
    transportasi_id: Optional[str] = None
    transportasi: Optional[str] = None
    asal_sekolah: Optional[str] = None
    status_sipil: Optional[str] = None
    no_kk: Optional[str] = None
    npwp: Optional[str] = None
    no_kip: Optional[str] = None
    no_kps: Optional[str] = None
    kebutuhan_khusus: Optional[str] = None
    tinggi_badan: Optional[str] = None
    berat_badan: Optional[str] = None
    semester_masuk: Optional[str] = None
    tanggal_masuk: Optional[str] = None
    jenis_pendaftaran_id: Optional[str] = None
    jenis_pendaftaran: Optional[str] = None
    jalur_masuk_id: Optional[str] = None
    jalur_masuk: Optional[str] = None
    jenis_pembiayaan_id: Optional[str] = None
    jenis_pembiayaan: Optional[str] = None
    status_mahasiswa_id: Optional[str] = None
    feeder_student_id: Optional[str] = None
    feeder_registration_id: Optional[str] = None
    foto_url: Optional[str] = None
    orang_tua: Optional[Dict[str, Any]] = None
    registration: Optional[Dict[str, Any]] = None
    pddikti_ids: Optional[Dict[str, Any]] = None
    status: Optional[str] = None


class LecturerInput(BaseModel):
    employee_id: str = ""
    nip: str = ""
    nidn: str = ""
    nuptk: str = ""
    nrsd: str = ""
    nik: str = ""
    username: str = Field(min_length=3)
    name: str = Field(min_length=1)
    gelar: str = ""
    gelar_depan: str = ""
    gelar_belakang: str = ""
    nama_panggilan: str = ""
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(default="Dosen123!", min_length=6)
    status: str = "active"
    prodi_id: str = ""
    homebase: str = ""
    gender: str = ""
    agama: str = ""
    tempat_lahir: str = ""
    tanggal_lahir: str = ""
    alamat: str = ""
    kota: str = ""
    provinsi: str = ""
    kode_pos: str = ""
    kewarganegaraan: str = ""
    rt: str = ""
    rw: str = ""
    dusun: str = ""
    kelurahan: str = ""
    kecamatan: str = ""
    kode_wilayah: str = ""
    jabatan_akademik: str = ""
    jabatan_dikti_id: str = ""
    jabatan_kode: str = ""
    keilmuan: str = ""
    pendidikan_terakhir: str = ""
    jenjang_pendidikan: str = ""
    ikatan_kerja: str = ""
    status_pegawai: str = ""
    status_kerja: str = ""
    jenis_pegawai: str = ""
    pangkat_golongan_id: str = ""
    pangkat_golongan: str = ""
    no_sk: str = ""
    unit_organisasi_id: str = ""
    institusi_induk: str = ""
    status_dosen: str = ""
    status_dosen_id: str = ""
    tanggal_masuk: str = ""
    tanggal_mulai_mengajar: str = ""
    foto_url: str = ""


class StaffInput(BaseModel):
    employee_id: str = ""
    nip: str = ""
    nik: str = ""
    nuptk: str = ""
    username: str = Field(min_length=3)
    name: str = Field(min_length=1)
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(default="Tendik123!", min_length=6)
    status: str = "active"
    jabatan_id: str = ""
    jabatan: str = ""
    unit_organisasi: str = ""
    unit_organisasi_id: str = ""
    jenis_pegawai: str = ""
    status_pegawai: str = ""
    status_kerja: str = ""
    no_sk: str = ""
    tanggal_masuk: str = ""
    alamat: str = ""
    kota: str = ""
    provinsi: str = ""
    foto_url: str = ""


class StaffUpdateInput(StaffInput):
    password: Optional[str] = Field(default=None, min_length=6)


async def resolve_staff_master_references(jabatan_id: str, unit_organisasi_id: str) -> Dict[str, str]:
    """Resolve ID master Tendik menjadi snapshot nama yang konsisten."""
    jabatan_id = jabatan_id.strip()
    unit_organisasi_id = unit_organisasi_id.strip()
    if not jabatan_id:
        raise HTTPException(status_code=400, detail="Jabatan / fungsi wajib dipilih dari master")
    if not unit_organisasi_id:
        raise HTTPException(status_code=400, detail="Unit organisasi wajib dipilih dari master")

    jabatan = await db.jabatan_akademik.find_one({"id": jabatan_id}, {"_id": 0})
    if not jabatan:
        jabatan = next((item for item in DEFAULT_JABATAN_AKADEMIK if item["id"] == jabatan_id), None)
    if not jabatan or jabatan.get("status") == "inactive":
        raise HTTPException(status_code=400, detail="Jabatan / fungsi tidak ditemukan atau sudah nonaktif")

    unit = await db.unit_organisasi.find_one({"id": unit_organisasi_id}, {"_id": 0})
    if not unit:
        unit = next((item for item in DEFAULT_UNIT_ORGANISASI if item["id"] == unit_organisasi_id), None)
    if not unit or unit.get("status") == "inactive":
        raise HTTPException(status_code=400, detail="Unit organisasi tidak ditemukan atau sudah nonaktif")

    return {
        "jabatan_id": jabatan["id"],
        "jabatan": str(jabatan.get("nama") or "").strip(),
        "jabatan_kode": str(jabatan.get("kode") or "").strip(),
        "unit_organisasi_id": unit["id"],
        "unit_organisasi": str(unit.get("nama") or "").strip(),
    }


class LecturerUpdateInput(BaseModel):
    employee_id: str = ""
    nip: str = ""
    nidn: str = ""
    nuptk: str = ""
    nrsd: str = ""
    nik: str = ""
    username: str = Field(min_length=3)
    name: str = Field(min_length=1)
    gelar: str = ""
    gelar_depan: str = ""
    gelar_belakang: str = ""
    nama_panggilan: str = ""
    email: EmailStr
    whatsapp: str = ""
    status: str = "active"
    prodi_id: str = ""
    homebase: str = ""
    gender: str = ""
    agama: str = ""
    tempat_lahir: str = ""
    tanggal_lahir: str = ""
    alamat: str = ""
    kota: str = ""
    provinsi: str = ""
    kode_pos: str = ""
    kewarganegaraan: str = ""
    rt: str = ""
    rw: str = ""
    dusun: str = ""
    kelurahan: str = ""
    kecamatan: str = ""
    kode_wilayah: str = ""
    jabatan_akademik: str = ""
    jabatan_dikti_id: str = ""
    jabatan_kode: str = ""
    keilmuan: str = ""
    pendidikan_terakhir: str = ""
    jenjang_pendidikan: str = ""
    ikatan_kerja: str = ""
    status_pegawai: str = ""
    status_kerja: str = ""
    jenis_pegawai: str = ""
    pangkat_golongan_id: str = ""
    pangkat_golongan: str = ""
    no_sk: str = ""
    unit_organisasi_id: str = ""
    institusi_induk: str = ""
    status_dosen: str = ""
    status_dosen_id: str = ""
    tanggal_masuk: str = ""
    tanggal_mulai_mengajar: str = ""
    foto_url: str = ""


class MaterialInput(BaseModel):
    class_id: str
    title: str
    description: str = ""
    meeting: str = "Pertemuan 1"
    file_url: str = ""
    video_url: str = ""
    meeting_type: str = "offline"
    meeting_url: str = ""
    is_active: bool = True
    locked_until: str = ""
    rps_meeting_number: Optional[int] = None


class GoogleMeetInput(BaseModel):
    class_id: str
    title: str = ""


class CommentInput(BaseModel):
    material_id: str
    content: str
    parent_id: str = ""


class JoinRequestInput(BaseModel):
    class_code: str


class StudentIdsInput(BaseModel):
    student_ids: List[str] = Field(default_factory=list)


class RubricItem(BaseModel):
    criterion: str
    weight: float


class AssignmentInput(BaseModel):
    class_id: str
    title: str
    description: str
    attachment_link: str = ""
    deadline: str
    published_at: str = ""
    tolerance_hours: int = 0
    allowed_formats: List[str] = Field(default_factory=lambda: ["pdf", "docx", "xlsx", "zip", "png", "jpg"])
    max_file_size_mb: float = Field(default=DEFAULT_SUBMISSION_MAX_FILE_MB, gt=0)
    rubric: List[RubricItem] = Field(default_factory=list)
    assignment_type: str = "individu"
    assessment_category: str = "tugas"
    allow_revision: bool = True
    is_active: bool = True
    is_practicum: bool = False
    practicum_goal: str = ""
    practicum_tools: str = ""
    practicum_steps: List[str] = Field(default_factory=list)
    required_screenshot: bool = False
    late_penalty_per_day: float = 0
    close_after_deadline: bool = False
    material_id: str = ""


class AcademicCalendarEventInput(BaseModel):
    """Kegiatan institusi yang menjadi rujukan kalender akademik kampus."""

    title: str = Field(min_length=3, max_length=180)
    category: str = "academic"
    start_at: str
    end_at: str = ""
    all_day: bool = True
    academic_year_id: str = ""
    audience: str = "all"
    target_prodi_ids: List[str] = Field(default_factory=list)
    description: str = Field(default="", max_length=2000)
    location: str = Field(default="", max_length=240)
    link: str = Field(default="", max_length=1000)
    status: str = "published"


class AcademicDeadlineItemInput(BaseModel):
    enabled: bool = False
    deadline_at: str = ""


class AcademicDeadlineSettingsInput(BaseModel):
    academic_year_id: str = ""
    deadlines: Dict[str, AcademicDeadlineItemInput] = Field(default_factory=dict)


class GradeItem(BaseModel):
    criterion: str
    weight: float
    score: float = Field(..., ge=0, le=100)


class GradeInput(BaseModel):
    rubric_scores: List[GradeItem]
    feedback: str = ""
    revision_note: str = ""
    status: str = "Dinilai"


class GradePredicateItem(BaseModel):
    label: str
    min_score: float
    max_score: float


class GradePredicateInput(BaseModel):
    class_id: str = ""
    predicates: List[GradePredicateItem]


class GradeWeightsInput(BaseModel):
    tugas: float = Field(default=25, ge=0, le=100)
    uts: float = Field(default=35, ge=0, le=100)
    uas: float = Field(default=40, ge=0, le=100)


class FinalizationInput(BaseModel):
    confirmation: str = Field(default="")


class CleanDataInput(BaseModel):
    confirmation: str = ""
    confirmation_label: str = ""


class ReminderInput(BaseModel):
    assignment_id: str
    student_id: str = ""
    reminder_type: str = "manual"
    message: str = ""


class BulkGradeItem(BaseModel):
    submission_id: str
    score: float = Field(..., ge=0, le=100)
    feedback: str = ""
    revision_note: str = ""


class BulkGradeInput(BaseModel):
    grades: List[BulkGradeItem]


class ResetPasswordInput(BaseModel):
    password: str = ""


class AppSettingsInput(BaseModel):
    app_name: str = "SIAKAD ONE"
    meta_description: str = Field(
        default="Sistem Informasi Akademik terpadu untuk mengelola pembelajaran, presensi, penilaian, dan layanan akademik perguruan tinggi.",
        max_length=320,
    )
    campus_name: str = ""
    campus_code: Optional[str] = ""
    institution_type: Optional[str] = "Politeknik"
    accreditation: Optional[str] = "Unggul"
    accreditation_sk: Optional[str] = ""
    campus_motto: Optional[str] = ""
    campus_phone: Optional[str] = ""
    campus_whatsapp: Optional[str] = ""
    campus_email: Optional[str] = ""
    campus_website: Optional[str] = ""
    campus_address: str = ""
    program_name: str = ""
    lecturer_name: str = ""
    lecturer_email: str = ""
    app_logo_url: str = ""
    campus_logo_url: str = ""
    rector_name: Optional[str] = ""
    rector_nidn: Optional[str] = ""
    vice_rector_1: Optional[str] = ""
    head_of_baak: Optional[str] = ""
    head_of_lppm: Optional[str] = ""
    kop_letterhead: Optional[str] = ""
    active_academic_year: str = ""
    active_semester: str = ""
    min_attendance_percentage: Optional[int] = 75


class GoogleDriveSettingsInput(BaseModel):
    enabled: bool = True
    root_folder_id: str = ""
    root_folder_name: str = "E-Learning Dosen"
    require_upload: bool = False
    lecturer_folder_sharing_enabled: bool = False
    lecturer_folder_role: str = "reader"
    google_meet_enabled: bool = False
    google_workspace_delegated_user: str = ""
    service_account_json: str = ""
    clear_service_account: bool = False


class DatabaseBackupSettingsInput(BaseModel):
    enabled: bool = False
    frequency: str = "daily"
    run_time: str = "02:00"
    weekly_day: int = Field(default=0, ge=0, le=6)
    retention_count: int = Field(default=14, ge=1, le=90)
    upload_to_drive: bool = True
    keep_local: bool = True


class WhatsAppSettingsInput(BaseModel):
    provider: str = "disabled"
    app_url: str = ""
    fonnte_token: str = ""
    fonnte_url: str = "https://api.fonnte.com/send"
    waha_base_url: str = ""
    waha_api_key: str = ""
    waha_session: str = "default"
    send_delay_seconds: int = Field(default=3, ge=0, le=300)
    typing_simulation_seconds: int = Field(default=30, ge=0, le=120)
    otp_template: str = "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}"
    assignment_template: str = default_whatsapp_settings()["assignment_template"]
    grade_template: str = default_whatsapp_settings()["grade_template"]
    revision_template: str = default_whatsapp_settings()["revision_template"]


class EmailSettingsInput(BaseModel):
    enabled: bool = False
    smtp_host: str = ""
    smtp_port: int = Field(default=587, ge=1, le=65535)
    smtp_user: str = ""
    smtp_password: str = ""
    smtp_use_tls: bool = True
    from_name: str = "E-Learning Dosen"
    from_email: str = ""


class OidcSettingsInput(BaseModel):
    enabled: bool = True
    discovery_url: str = ""
    issuer: str = ""
    client_id: str = ""
    client_secret: str = ""
    redirect_uri: str = ""
    frontend_url: str = ""
    scopes: str = "openid profile email roles"
    local_login_enabled: bool = True
    clear_client_secret: bool = False


async def seed_data() -> None:
    existing_admin = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if existing_admin:
        return

    admin_id = new_id()
    await db.users.insert_one(
        {
            "id": admin_id,
            "role": "admin",
            "username": "dosenadmin",
            "name": "Dosen Admin",
            "email": "dosen@demo.id",
            "whatsapp": "628000000001",
            "password_hash": hash_password("Dosen123!"),
            "status": "active",
            "created_at": now_iso(),
            "last_login_at": "",
        }
    )

    program_id = new_id()
    course_id = new_id()
    class_id = new_id()
    student_id = new_id()
    assignment_id = new_id()
    material_id = new_id()

    await db.programs.insert_one(
        {
            "id": program_id,
            "code": "IF",
            "name": "Teknik Informatika",
            "description": "Program studi demo untuk kelas e-learning.",
            "status": "active",
            "created_at": now_iso(),
        }
    )
    await db.courses.insert_one(
        {
            "id": course_id,
            "program_id": program_id,
            "program_name": "Teknik Informatika",
            "code": "IF401",
            "name": "Pemrograman Web Lanjut",
            "credits": 3,
            "description": "Mata kuliah praktis untuk membangun aplikasi web modern.",
            "dosen_utama_id": admin_id,
            "dosen_utama_nama": "Dosen Admin",
            "status": "active",
            "created_at": now_iso(),
        }
    )
    await db.classes.insert_one(
        {
            "id": class_id,
            "academic_year": "2025/2026",
            "semester": "Ganjil",
            "course_id": course_id,
            "course_name": "Pemrograman Web Lanjut",
            "name": "IF-4A",
            "schedule": "Selasa 09.00",
            "class_code": "WEB4A1",
            "lecturer_id": admin_id,
            "lecturer_name": "Dosen Admin",
            "status": "active",
            "student_ids": [student_id],
            "created_at": now_iso(),
        }
    )
    await db.users.insert_one(
        {
            "id": student_id,
            "role": "student",
            "username": "alya",
            "nim": "230001001",
            "name": "Alya Pratama",
            "email": "alya@demo.id",
            "whatsapp": "628123456789",
            "password_hash": hash_password("Mahasiswa1231!"),
            "status": "active",
            "class_ids": [class_id],
            "created_at": now_iso(),
            "last_login_at": "",
        }
    )
    deadline = (datetime.now(timezone.utc) + timedelta(days=3)).isoformat()
    await db.assignments.insert_one(
        {
            "id": assignment_id,
            "class_id": class_id,
            "course_id": course_id,
            "course_name": "Pemrograman Web Lanjut",
            "class_name": "IF-4A",
            "title": "Praktikum CRUD API",
            "description": "Buat endpoint CRUD sederhana dan unggah laporan singkat.",
            "deadline": deadline,
            "published_at": "",
            "tolerance_hours": 6,
            "allowed_formats": ["pdf", "docx", "zip"],
            "rubric": [
                {"criterion": "Ketepatan jawaban", "weight": 40},
                {"criterion": "Kerapian laporan", "weight": 20},
                {"criterion": "Kreativitas", "weight": 20},
                {"criterion": "Ketepatan waktu", "weight": 20},
            ],
            "assignment_type": "individu",
            "allow_revision": True,
            "is_active": True,
            "is_practicum": True,
            "practicum_goal": "Mahasiswa memahami struktur API CRUD.",
            "practicum_tools": "Python, FastAPI, PostgreSQL",
            "practicum_steps": ["Rancang endpoint", "Uji dengan curl", "Tulis laporan"],
            "required_screenshot": True,
            "late_penalty_per_day": 5,
            "close_after_deadline": False,
            "material_id": material_id,
            "created_at": now_iso(),
            "created_by": admin_id,
            "lecturer_id": admin_id,
        }
    )
    await db.materials.insert_one(
        {
            "id": material_id,
            "class_id": class_id,
            "title": "Arsitektur Aplikasi Full-stack",
            "description": "Materi pembuka tentang frontend, backend, dan database.",
            "meeting": "Pertemuan 1",
            "file_url": "https://drive.google.com/",
            "video_url": "",
            "is_active": True,
            "locked_until": "",
            "created_at": now_iso(),
            "created_by": admin_id,
            "lecturer_id": admin_id,
        }
    )
    logger.info("Seed data e-learning dibuat")


async def ensure_program_course_links() -> None:
    program = await db.programs.find_one({}, {"_id": 0})
    if not program:
        settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or {}
        program = {
            "id": new_id(),
            "code": "PRODI",
            "name": settings.get("program_name") or "Program Studi",
            "description": "",
            "status": "active",
            "created_at": now_iso(),
        }
        await db.programs.insert_one(program)
    await db.courses.update_many(
        {"$or": [{"program_id": {"$exists": False}}, {"program_id": ""}, {"program_id": None}]},
        {"$set": {"program_id": program["id"], "program_name": program["name"]}},
    )
    async for course in db.courses.find({"program_id": {"$exists": True}}, {"_id": 0, "id": 1, "program_id": 1}):
        linked = await db.programs.find_one({"id": course.get("program_id")}, {"_id": 0})
        if linked:
            await db.courses.update_one({"id": course["id"]}, {"$set": {"program_name": linked.get("name", "")}})


async def ensure_multi_lecturer_schema() -> None:
    """Backfill ownership so existing single-lecturer installations remain usable."""
    fallback = await db.users.find_one({"role": "admin", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not fallback:
        return
    fallback_id = fallback["id"]
    fallback_name = fallback.get("name", "Admin Kampus")
    async for class_doc in db.classes.find(
        {"$or": [{"lecturer_id": {"$exists": False}}, {"lecturer_id": ""}, {"lecturer_id": None}]},
        {"_id": 0, "id": 1, "created_by": 1},
    ):
        owner_id = class_doc.get("created_by") or fallback_id
        owner = await db.users.find_one({"id": owner_id, "role": {"$in": ["admin", "lecturer"]}}, {"_id": 0}) or fallback
        await db.classes.update_one(
            {"id": class_doc["id"]},
            {"$set": {"lecturer_id": owner["id"], "lecturer_name": owner.get("name", fallback_name)}},
        )
    async for class_doc in db.classes.find({}, {"_id": 0, "id": 1, "lecturer_id": 1, "lecturer_name": 1}):
        owner_id = class_doc.get("lecturer_id") or fallback_id
        owner_name = class_doc.get("lecturer_name") or fallback_name
        class_id = class_doc["id"]
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many(
                {"class_id": class_id, "$or": [{"lecturer_id": {"$exists": False}}, {"lecturer_id": ""}, {"lecturer_id": None}]},
                {"$set": {"lecturer_id": owner_id, "lecturer_name": owner_name}},
            )


async def ensure_class_lifecycle_schema() -> None:
    """Backfill lifecycle fields for classes created before finalization support."""
    async for class_doc in db.classes.find(
        {
            "status": {"$in": [CLASS_STATUS_ENDED, CLASS_STATUS_FINALIZED, CLASS_STATUS_ARCHIVED]},
            "$or": [
                {"grade_weights_snapshot": {"$exists": False}},
                {"grade_weights_snapshot_customized": {"$exists": False}},
            ],
        },
        {"_id": 0, "id": 1, "course_id": 1, "grade_weights_snapshot": 1},
    ):
        course = await db.courses.find_one({"id": class_doc.get("course_id")}, {"_id": 0, "grade_weights": 1}) or {}
        await db.classes.update_one(
            {"id": class_doc["id"]},
            {
                "$set": {
                    "grade_weights_snapshot": grade_weights_from_document(
                        class_doc.get("grade_weights_snapshot") or course.get("grade_weights")
                    ),
                    "grade_weights_snapshot_customized": isinstance(course.get("grade_weights"), dict),
                    "lifecycle_migrated_at": now_iso(),
                }
            },
        )


def env_flag(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_enabled(name: str, default: bool = True) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() not in {"0", "false", "no", "n", "off", "disabled"}


def env_service_account_configured() -> bool:
    return bool(
        os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        or os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        or os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    )


def drive_key_file_path() -> Path:
    configured = os.environ.get("GOOGLE_DRIVE_CONFIG_KEY_FILE", "").strip()
    return Path(configured).expanduser() if configured else ROOT_DIR / ".drive_config.key"


def normalize_fernet_key(value: str) -> bytes:
    raw = value.strip().encode()
    try:
        Fernet(raw)
        return raw
    except Exception:
        return base64.urlsafe_b64encode(hashlib.sha256(raw).digest())


def get_drive_config_key() -> bytes:
    env_key = os.environ.get("GOOGLE_DRIVE_CONFIG_KEY", "").strip()
    if env_key:
        return normalize_fernet_key(env_key)
    key_path = drive_key_file_path()
    if key_path.exists():
        return key_path.read_bytes().strip()
    key = Fernet.generate_key()
    key_path.write_bytes(key)
    os.chmod(key_path, 0o600)
    return key


def encrypt_secret(value: str) -> str:
    return Fernet(get_drive_config_key()).encrypt(value.encode("utf-8")).decode("utf-8")


def decrypt_secret(value: str) -> str:
    if not value:
        return ""
    try:
        return Fernet(get_drive_config_key()).decrypt(value.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        logger.error("Kunci enkripsi aplikasi tidak cocok dengan credential tersimpan")
        return ""


def validate_oidc_url(value: str, label: str, required: bool = True) -> str:
    normalized = (value or "").strip().rstrip("/")
    if not normalized and not required:
        return ""
    parsed = urlsplit(normalized)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=400, detail=f"{label} harus berupa URL HTTP/HTTPS yang valid")
    host = (parsed.hostname or "").lower()
    local_http = host in {"localhost", "127.0.0.1", "::1"} or host.startswith(("10.", "192.168."))
    if parsed.scheme != "https" and not local_http:
        raise HTTPException(status_code=400, detail=f"{label} non-local wajib menggunakan HTTPS")
    return normalized


def oidc_admin_view(settings: Dict[str, Any], source: str = "environment", metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    secret_configured = bool(settings.get("client_secret"))
    return {
        "enabled": bool(settings.get("enabled")),
        "provider": "SCI-ID",
        "discovery_url": settings.get("discovery_url", ""),
        "issuer": settings.get("issuer", ""),
        "client_id": settings.get("client_id", ""),
        "client_secret_configured": secret_configured,
        "client_secret_masked": "••••••••••••" if secret_configured else "",
        "redirect_uri": settings.get("redirect_uri", ""),
        "frontend_url": settings.get("frontend_url", ""),
        "scopes": settings.get("scopes", "openid profile email roles"),
        "local_login_enabled": bool(settings.get("local_login_enabled", True)),
        "source": source,
        "updated_at": (metadata or {}).get("updated_at", ""),
        "updated_by": (metadata or {}).get("updated_by", ""),
    }


async def load_oidc_runtime_settings() -> Dict[str, Any]:
    doc = await db.oidc_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    _oidc_runtime_settings.clear()
    if doc:
        _oidc_runtime_settings.update(
            {
                "enabled": bool(doc.get("enabled")),
                "discovery_url": str(doc.get("discovery_url", "")).strip(),
                "issuer": str(doc.get("issuer", "")).strip().rstrip("/"),
                "client_id": str(doc.get("client_id", "")).strip(),
                "client_secret": decrypt_secret(str(doc.get("client_secret_encrypted", ""))),
                "redirect_uri": str(doc.get("redirect_uri", "")).strip(),
                "frontend_url": str(doc.get("frontend_url", "")).strip().rstrip("/"),
                "scopes": str(doc.get("scopes", "openid profile email roles")).strip(),
                "local_login_enabled": bool(doc.get("local_login_enabled", True)),
            }
        )
    clear_oidc_discovery_cache()
    return oidc_admin_view(oidc_settings(), "admin_ui" if doc else "environment", doc)


def normalize_service_account_payload(value: str) -> Dict[str, Any]:
    raw = (value or "").strip()
    if not raw:
        return {}
    if not raw.startswith("{"):
        raw = base64.b64decode(raw).decode("utf-8")
    info = json.loads(raw)
    if not isinstance(info, dict) or info.get("type") != "service_account" or not info.get("client_email") or not info.get("private_key"):
        raise ValueError("JSON service account tidak valid")
    return info


async def get_google_drive_settings(mask: bool = True) -> Dict[str, Any]:
    cached = _get_cached_settings("google_drive_settings")
    if cached is not None:
        if mask:
            return cached.copy()
        return cached
    doc = await db.google_drive_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    credential = ""
    credential_source = ""
    service_account_email = doc.get("service_account_email", "")
    if doc.get("service_account_json_encrypted"):
        credential = decrypt_secret(doc.get("service_account_json_encrypted", ""))
        credential_source = "admin_ui" if credential else ""
    env_has_credential = env_service_account_configured()
    if not credential and env_has_credential:
        credential_source = "environment"
        service_account_email = service_account_email or service_account_email_from_env()
    settings = {
        "id": "main",
        "enabled": bool(doc.get("enabled", env_enabled("GOOGLE_DRIVE_ENABLED", True))),
        "root_folder_id": doc.get("root_folder_id", os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID", "")),
        "root_folder_name": safe_path_segment(doc.get("root_folder_name", os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_NAME", "E-Learning Dosen"))),
        "require_upload": bool(doc.get("require_upload", env_flag("GOOGLE_DRIVE_REQUIRE_UPLOAD", False))),
        "lecturer_folder_sharing_enabled": bool(doc.get("lecturer_folder_sharing_enabled", False)),
        "lecturer_folder_role": doc.get("lecturer_folder_role", "reader") if doc.get("lecturer_folder_role") in {"reader", "writer"} else "reader",
        "google_meet_enabled": bool(doc.get("google_meet_enabled", env_enabled("GOOGLE_MEET_ENABLED", False))),
        "google_workspace_delegated_user": str(doc.get("google_workspace_delegated_user") or os.environ.get("GOOGLE_WORKSPACE_DELEGATED_USER", "")).strip().lower(),
        "service_account_configured": bool(credential or env_has_credential),
        "service_account_source": credential_source,
        "service_account_email": service_account_email,
        "updated_at": doc.get("updated_at", ""),
        "updated_by": doc.get("updated_by", ""),
    }
    settings["google_meet_ready"] = bool(
        settings["google_meet_enabled"]
        and settings["service_account_configured"]
        and settings["google_workspace_delegated_user"]
    )
    settings["google_meet_scope"] = "https://www.googleapis.com/auth/meetings.space.created"
    if not mask:
        settings["service_account_json"] = credential
    _set_cached_settings("google_drive_settings", settings)
    if mask:
        return settings.copy()
    return settings


def service_account_email_from_env() -> str:
    try:
        json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
        if json_payload:
            return json.loads(json_payload).get("client_email", "")
        if json_payload_b64:
            return json.loads(base64.b64decode(json_payload_b64).decode("utf-8")).get("client_email", "")
        if file_path and Path(file_path).exists():
            return json.loads(Path(file_path).read_text()).get("client_email", "")
    except Exception:
        return ""
    return ""


def google_drive_upload_enabled(settings: Dict[str, Any]) -> bool:
    return bool(settings.get("enabled") and settings.get("service_account_configured"))


def google_drive_upload_required(settings: Dict[str, Any]) -> bool:
    return bool(settings.get("require_upload"))


def google_drive_error_message(exc: Exception) -> str:
    text = str(exc)
    if "storageQuotaExceeded" in text or "Service Accounts do not have storage quota" in text:
        return (
            "Google Drive menolak upload file karena service account tidak punya kuota penyimpanan. "
            "Gunakan Shared Drive atau folder Drive yang mendukung upload service account, lalu jalankan retry."
        )
    if "insufficientFilePermissions" in text or "The user does not have sufficient permissions" in text:
        return "Service account belum punya izin upload ke folder Google Drive. Bagikan folder dengan akses Editor."
    if "File not found" in text or "notFound" in text:
        return "Folder Google Drive tidak ditemukan atau belum dibagikan ke service account."
    return text[:500]


async def storage_status_summary() -> Dict[str, Any]:
    settings = await get_google_drive_settings(mask=True)
    drive_enabled = google_drive_upload_enabled(settings)
    drive_sync_counts = {
        "pending": await db.stored_files.count_documents({"drive_sync_status": "pending"}),
        "synced": await db.stored_files.count_documents({"drive_sync_status": "synced"}),
        "failed": await db.stored_files.count_documents({"drive_sync_status": "failed"}),
        "not_configured": await db.stored_files.count_documents({"drive_sync_status": "not_configured"}),
    }
    return {
        "storage_mode": "google_drive" if drive_enabled else "server_local",
        "drive_configured": bool(settings.get("service_account_configured")),
        "drive_enabled": drive_enabled,
        "drive_required": google_drive_upload_required(settings),
        "drive_root_folder_id_configured": bool(settings.get("root_folder_id")),
        "drive_root_folder_name": settings.get("root_folder_name", "E-Learning Dosen"),
        "drive_service_account_email": settings.get("service_account_email", ""),
        "drive_service_account_source": settings.get("service_account_source", ""),
        "drive_sync": drive_sync_counts,
    }


def get_drive_service(settings: Optional[Dict[str, Any]] = None):
    scopes = ["https://www.googleapis.com/auth/drive"]
    settings = settings or {}
    if settings.get("service_account_json"):
        credentials = service_account.Credentials.from_service_account_info(json.loads(settings["service_account_json"]), scopes=scopes)
        return build("drive", "v3", credentials=credentials, cache_discovery=False)
    json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
    file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    if json_payload:
        credentials = service_account.Credentials.from_service_account_info(json.loads(json_payload), scopes=scopes)
    elif json_payload_b64:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(base64.b64decode(json_payload_b64).decode("utf-8")),
            scopes=scopes,
        )
    elif file_path:
        credentials = service_account.Credentials.from_service_account_file(file_path, scopes=scopes)
    else:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON, GOOGLE_SERVICE_ACCOUNT_JSON_B64, atau GOOGLE_SERVICE_ACCOUNT_FILE belum dikonfigurasi")
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


BACKUP_ROOT = STORAGE_ROOT / "Database Backups"
BACKUP_TIMEZONE = ZoneInfo(os.environ.get("BACKUP_TIMEZONE", "Asia/Jakarta"))
_database_backup_scheduler_task: Optional[asyncio.Task] = None


async def reconcile_local_storage_paths() -> Dict[str, int]:
    """Rebase machine-specific file paths after a database/storage restore."""
    rebased = 0
    missing = 0
    async for file_doc in db.stored_files.find(
        {},
        {"_id": 0, "id": 1, "storage_path": 1, "local_path": 1, "local_available": 1},
    ):
        storage_path = str(file_doc.get("storage_path") or "")
        if not storage_path:
            storage_path = portable_storage_path_from_local_path(
                str(file_doc.get("local_path") or "")
            )
        portable_path = resolve_storage_local_path(
            ROOT_DIR / "storage",
            storage_path,
        )
        if not portable_path:
            continue
        available = portable_path.exists() and portable_path.is_file()
        updates: Dict[str, Any] = {}
        if str(file_doc.get("storage_path") or "") != storage_path:
            updates["storage_path"] = storage_path
        if str(file_doc.get("local_path") or "") != str(portable_path):
            updates["local_path"] = str(portable_path)
        if bool(file_doc.get("local_available")) != available:
            updates["local_available"] = available
        if updates:
            updates["updated_at"] = now_iso()
            await db.stored_files.update_one({"id": file_doc["id"]}, {"$set": updates})
            rebased += 1
        if not available:
            missing += 1

    async for backup_doc in db.database_backups.find(
        {},
        {"_id": 0, "id": 1, "file_name": 1, "local_path": 1, "local_available": 1},
    ):
        file_name = Path(str(backup_doc.get("file_name") or "")).name
        if not file_name:
            continue
        portable_path = BACKUP_ROOT / file_name
        available = portable_path.exists() and portable_path.is_file()
        updates = {}
        if str(backup_doc.get("local_path") or "") != str(portable_path):
            updates["local_path"] = str(portable_path)
        if bool(backup_doc.get("local_available")) != available:
            updates["local_available"] = available
        if updates:
            await db.database_backups.update_one(
                {"id": backup_doc["id"]},
                {"$set": updates},
            )
            rebased += 1
        if not available:
            missing += 1
    return {"rebased": rebased, "missing": missing}


def default_database_backup_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "enabled": False,
        "frequency": "daily",
        "run_time": "02:00",
        "weekly_day": 0,
        "retention_count": 14,
        "upload_to_drive": True,
        "keep_local": True,
        "timezone": str(BACKUP_TIMEZONE),
    }


async def get_database_backup_settings() -> Dict[str, Any]:
    doc = await db.database_backup_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    return {**default_database_backup_settings(), **doc, "running": bool(doc.get("running"))}


def next_database_backup_at(settings: Dict[str, Any], after: Optional[datetime] = None) -> str:
    current = (after or datetime.now(timezone.utc)).astimezone(BACKUP_TIMEZONE)
    try:
        hour, minute = [int(item) for item in str(settings.get("run_time", "02:00")).split(":", 1)]
    except (TypeError, ValueError):
        hour, minute = 2, 0
    candidate = current.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if settings.get("frequency") == "weekly":
        target_day = int(settings.get("weekly_day", 0))
        candidate += timedelta(days=(target_day - candidate.weekday()) % 7)
        if candidate <= current:
            candidate += timedelta(days=7)
    elif candidate <= current:
        candidate += timedelta(days=1)
    return candidate.astimezone(timezone.utc).isoformat()


async def database_backup_payload() -> bytes:
    collection_names = sorted(
        name for name in await db.list_collection_names() if not name.startswith("system.")
    )
    collections: Dict[str, List[Dict[str, Any]]] = {}
    for name in collection_names:
        documents: List[Dict[str, Any]] = []
        async for document in db[name].find({}):
            documents.append(document)
        collections[name] = documents
    payload = {
        "format": "nugaslagi-postgresql-jsonb",
        "version": 2,
        "database": db.name,
        "created_at": now_iso(),
        "collections": collections,
    }
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return gzip.compress(raw, compresslevel=6)


def upload_database_backup_to_drive_sync(
    content: bytes, file_name: str, settings: Dict[str, Any]
) -> Dict[str, str]:
    service = get_drive_service(settings)
    root_folder_id = settings.get("root_folder_id", "")
    if not root_folder_id:
        root_folder_id = drive_find_or_create_folder(
            service,
            safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen"),
            None,
        )
    backup_folder_id = drive_find_or_create_folder(
        service, "Database Backups", root_folder_id
    )
    media = MediaIoBaseUpload(
        io.BytesIO(content), mimetype="application/gzip", resumable=True
    )
    uploaded = (
        service.files()
        .create(
            body={"name": file_name, "parents": [backup_folder_id]},
            media_body=media,
            fields="id,name,webViewLink",
            supportsAllDrives=True,
        )
        .execute()
    )
    return {
        "drive_file_id": uploaded.get("id", ""),
        "drive_file_url": uploaded.get("webViewLink", ""),
    }


def delete_database_backup_from_drive_sync(file_id: str, settings: Dict[str, Any]) -> None:
    if file_id:
        get_drive_service(settings).files().delete(
            fileId=file_id, supportsAllDrives=True
        ).execute()


async def enforce_database_backup_retention(settings: Dict[str, Any]) -> None:
    retention = max(1, int(settings.get("retention_count", 14)))
    expired = await db.database_backups.find(
        {"status": {"$in": ["completed", "completed_with_warning"]}}, {"_id": 0}
    ).sort("created_at", -1).skip(retention).to_list(500)
    drive_settings = await get_google_drive_settings(mask=False)
    for item in expired:
        local_path = item.get("local_path", "")
        if local_path:
            path = Path(local_path)
            if path.exists() and BACKUP_ROOT.resolve() in path.resolve().parents:
                path.unlink(missing_ok=True)
        if item.get("drive_file_id") and google_drive_upload_enabled(drive_settings):
            try:
                await asyncio.to_thread(
                    delete_database_backup_from_drive_sync,
                    item["drive_file_id"],
                    drive_settings,
                )
            except Exception as exc:
                logger.warning("Backup lama gagal dihapus dari Drive: %s", exc)
        await db.database_backups.delete_one({"id": item["id"]})


async def create_database_backup(trigger: str, user_id: str = "system") -> Dict[str, Any]:
    backup_id = str(uuid.uuid4())
    created_at = now_iso()
    file_name = f"nugaslagi-db-{datetime.now(BACKUP_TIMEZONE).strftime('%Y%m%d-%H%M%S')}.json.gz"
    record = {
        "id": backup_id,
        "file_name": file_name,
        "trigger": trigger,
        "status": "running",
        "created_at": created_at,
        "created_by": user_id,
        "size": 0,
        "local_available": False,
        "drive_file_id": "",
        "drive_file_url": "",
        "error": "",
    }
    await db.database_backups.insert_one(record.copy())
    settings = await get_database_backup_settings()
    try:
        content = await database_backup_payload()
        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
        local_path = BACKUP_ROOT / file_name
        local_path.write_bytes(content)
        update: Dict[str, Any] = {
            "status": "completed",
            "completed_at": now_iso(),
            "size": len(content),
            "local_path": str(local_path),
            "local_available": True,
        }
        if settings.get("upload_to_drive"):
            drive_settings = await get_google_drive_settings(mask=False)
            if not google_drive_upload_enabled(drive_settings):
                update.update(
                    status="completed_with_warning",
                    error="Backup lokal berhasil, tetapi Google Drive belum aktif atau belum dikonfigurasi.",
                )
            else:
                try:
                    drive_result = await asyncio.to_thread(
                        upload_database_backup_to_drive_sync,
                        content,
                        file_name,
                        drive_settings,
                    )
                    update.update(drive_result)
                except Exception as exc:
                    update.update(
                        status="completed_with_warning",
                        error=f"Backup lokal berhasil, upload Drive gagal: {google_drive_error_message(exc)}",
                    )
        if not settings.get("keep_local") and update.get("drive_file_id"):
            local_path.unlink(missing_ok=True)
            update.update(local_path="", local_available=False)
        await db.database_backups.update_one({"id": backup_id}, {"$set": update})
        await enforce_database_backup_retention(settings)
    except Exception as exc:
        logger.exception("Backup database gagal: %s", exc)
        await db.database_backups.update_one(
            {"id": backup_id},
            {"$set": {"status": "failed", "completed_at": now_iso(), "error": str(exc)[:500]}},
        )
    return await db.database_backups.find_one({"id": backup_id}, {"_id": 0})


async def database_backup_scheduler() -> None:
    while True:
        try:
            settings = await get_database_backup_settings()
            if settings.get("enabled"):
                now = now_iso()
                next_run = settings.get("next_run_at") or next_database_backup_at(settings)
                if not settings.get("next_run_at"):
                    await db.database_backup_settings.update_one(
                        {"id": "main"}, {"$set": {"next_run_at": next_run}}, upsert=True
                    )
                if next_run <= now:
                    lock_until = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
                    claimed = await db.database_backup_settings.update_one(
                        {
                            "id": "main",
                            "enabled": True,
                            "next_run_at": {"$lte": now},
                            "$or": [
                                {"auto_lock_until": {"$exists": False}},
                                {"auto_lock_until": {"$lt": now}},
                            ],
                        },
                        {"$set": {"running": True, "auto_lock_until": lock_until}},
                    )
                    if claimed.modified_count:
                        await create_database_backup("automatic", "system")
                        latest = await get_database_backup_settings()
                        await db.database_backup_settings.update_one(
                            {"id": "main"},
                            {"$set": {
                                "running": False,
                                "last_run_at": now_iso(),
                                "next_run_at": next_database_backup_at(latest),
                            }, "$unset": {"auto_lock_until": ""}},
                        )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.exception("Scheduler backup database bermasalah: %s", exc)
        await asyncio.sleep(60)


def get_meet_access_token(settings: Dict[str, Any], delegated_user: str = "") -> str:
    """Get a user-authorized token for Meet using Workspace domain-wide delegation."""
    credential_payload = settings.get("service_account_json", "")
    scopes = ["https://www.googleapis.com/auth/meetings.space.created"]
    if credential_payload:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(credential_payload), scopes=scopes
        )
    else:
        json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
        if json_payload:
            credentials = service_account.Credentials.from_service_account_info(json.loads(json_payload), scopes=scopes)
        elif json_payload_b64:
            credentials = service_account.Credentials.from_service_account_info(
                json.loads(base64.b64decode(json_payload_b64).decode("utf-8")), scopes=scopes
            )
        elif file_path:
            credentials = service_account.Credentials.from_service_account_file(file_path, scopes=scopes)
        else:
            raise RuntimeError("Service account Google belum dikonfigurasi")
    effective_user = str(delegated_user or settings.get("google_workspace_delegated_user", "")).strip()
    if effective_user:
        credentials = credentials.with_subject(effective_user)
    credentials.refresh(GoogleAuthRequest())
    if not credentials.token:
        raise RuntimeError("Token Google Meet tidak berhasil dibuat")
    return credentials.token


def google_meet_delegation_error(exc: Exception) -> bool:
    """Return whether Google rejected the delegated Workspace user/token grant."""
    details = " ".join(str(item) for item in getattr(exc, "args", ()) if item)
    message = f"{exc} {details}".lower()
    return "unauthorized_client" in message or (
        "invalid_grant" in message
        and ("invalid email" in message or "not a valid email" in message)
    )


def google_meet_error_message(exc: Exception) -> str:
    if google_meet_delegation_error(exc):
        return (
            "Google menolak delegasi akun Workspace. Pastikan Domain-wide Delegation "
            "menggunakan Client ID numerik service account (bukan email), scope "
            "https://www.googleapis.com/auth/meetings.space.created sudah diotorisasi, "
            "dan akun penyelenggara adalah pengguna pada domain Google Workspace yang sama."
        )
    return str(exc)[:500]


def create_google_meet_space_sync(settings: Dict[str, Any], delegated_user: str = "") -> Dict[str, str]:
    if not settings.get("service_account_configured"):
        raise RuntimeError("Service account Google belum dikonfigurasi")
    effective_user = str(delegated_user or settings.get("google_workspace_delegated_user", "")).strip()
    if not effective_user:
        raise RuntimeError(
            "Isi Google Workspace user untuk Meet dan aktifkan Domain-wide Delegation pada service account"
        )
    token = get_meet_access_token(settings, effective_user)
    response = httpx.post(
        "https://meet.googleapis.com/v2/spaces",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={},
        timeout=20,
    )
    if response.status_code >= 400:
        detail = ""
        try:
            detail = response.json().get("error", {}).get("message", "")
        except ValueError:
            detail = response.text
        raise RuntimeError(f"Google Meet menolak pembuatan ruang: {detail[:400]}")
    payload = response.json()
    meeting_uri = str(payload.get("meetingUri", "")).strip()
    meeting_code = str(payload.get("meetingCode", "")).strip()
    if not meeting_uri:
        raise RuntimeError("Google Meet tidak mengembalikan link pertemuan")
    return {"meeting_url": meeting_uri, "meeting_code": meeting_code, "meeting_space_name": str(payload.get("name", ""))}


def create_google_meet_for_app_user_sync(settings: Dict[str, Any], user_email: str = "") -> Dict[str, Any]:
    """Create a Meet as the app user, falling back to the verified default Workspace user."""
    requested_user = str(user_email or "").strip().lower()
    default_user = str(settings.get("google_workspace_delegated_user", "")).strip().lower()
    organizer = requested_user or default_user
    try:
        meet = create_google_meet_space_sync(settings, organizer)
        return {
            **meet,
            "organizer_email": organizer,
            "organizer_fallback_used": False,
        }
    except Exception as exc:
        if not (
            default_user
            and requested_user
            and default_user != requested_user
            and google_meet_delegation_error(exc)
        ):
            raise
        logger.warning(
            "Delegasi Google Meet untuk %s ditolak; memakai akun Workspace default %s",
            requested_user,
            default_user,
        )
        meet = create_google_meet_space_sync(settings, default_user)
        return {
            **meet,
            "organizer_email": default_user,
            "organizer_fallback_used": True,
        }


def drive_query_literal(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def drive_find_or_create_folder(service, name: str, parent_id: Optional[str]) -> str:
    folder_name = safe_path_segment(name or "Tanpa Nama")
    safe_name = drive_query_literal(folder_name)
    parent_clause = f" and '{parent_id}' in parents" if parent_id else " and 'root' in parents"
    query = (
        "mimeType='application/vnd.google-apps.folder' "
        f"and name='{safe_name}' and trashed=false{parent_clause}"
    )
    result = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            corpora="allDrives",
            fields="files(id,name)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )
    files = result.get("files", [])
    if files:
        return files[0]["id"]
    metadata = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        metadata["parents"] = [parent_id]
    folder = service.files().create(body=metadata, fields="id", supportsAllDrives=True).execute()
    return folder["id"]


def lecturer_drive_folder_name(lecturer: Dict[str, Any]) -> str:
    existing_name = str(lecturer.get("drive_folder_name", "")).strip()
    if existing_name:
        return safe_path_segment(existing_name)
    lecturer_code = str(lecturer.get("employee_id") or lecturer.get("id") or "DOSEN").strip()
    lecturer_name = str(lecturer.get("name") or "Dosen").strip()
    return safe_path_segment(f"{lecturer_code} - {lecturer_name}")


def drive_ensure_user_permission(service, folder_id: str, email: str, role: str) -> Dict[str, str]:
    clean_email = (email or "").strip().lower()
    clean_role = role if role in {"reader", "writer"} else "reader"
    if not clean_email:
        return {"drive_access_status": "missing_email", "drive_permission_id": ""}
    permissions = (
        service.permissions()
        .list(
            fileId=folder_id,
            fields="permissions(id,type,emailAddress,role)",
            supportsAllDrives=True,
        )
        .execute()
        .get("permissions", [])
    )
    existing = next(
        (
            item
            for item in permissions
            if item.get("type") == "user" and str(item.get("emailAddress", "")).lower() == clean_email
        ),
        None,
    )
    if existing:
        permission_id = existing.get("id", "")
        if existing.get("role") != clean_role:
            service.permissions().update(
                fileId=folder_id,
                permissionId=permission_id,
                body={"role": clean_role},
                supportsAllDrives=True,
            ).execute()
    else:
        created = (
            service.permissions()
            .create(
                fileId=folder_id,
                body={"type": "user", "role": clean_role, "emailAddress": clean_email},
                fields="id",
                sendNotificationEmail=False,
                supportsAllDrives=True,
            )
            .execute()
        )
        permission_id = created.get("id", "")
    return {
        "drive_access_status": "shared",
        "drive_permission_id": permission_id,
        "drive_permission_email": clean_email,
        "drive_permission_role": clean_role,
    }


def drive_revoke_user_permission(service, folder_id: str, permission_id: str) -> None:
    if folder_id and permission_id:
        service.permissions().delete(
            fileId=folder_id,
            permissionId=permission_id,
            supportsAllDrives=True,
        ).execute()


def drive_prepare_lecturer_folder(
    settings: Dict[str, Any], lecturer: Dict[str, Any]
) -> Dict[str, str]:
    service = get_drive_service(settings)
    root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
    root_folder_id = settings.get("root_folder_id") or os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID")
    if not root_folder_id:
        root_folder_id = drive_find_or_create_folder(service, root_folder_name, None)
    lecturers_folder_id = drive_find_or_create_folder(service, "Dosen", root_folder_id)
    folder_name = lecturer_drive_folder_name(lecturer)
    folder_id = drive_find_or_create_folder(service, folder_name, lecturers_folder_id)
    result = {
        "drive_folder_id": folder_id,
        "drive_folder_name": folder_name,
        "drive_access_status": "app_only",
        "drive_permission_id": "",
        "drive_permission_email": "",
        "drive_permission_role": "",
        "drive_access_error": "",
    }
    if settings.get("lecturer_folder_sharing_enabled") and lecturer.get("status", "active") == "active":
        try:
            result.update(
                drive_ensure_user_permission(
                    service,
                    folder_id,
                    lecturer.get("email", ""),
                    settings.get("lecturer_folder_role", "reader"),
                )
            )
        except Exception as exc:
            result.update(
                {
                    "drive_access_status": "share_failed",
                    "drive_access_error": google_drive_error_message(exc),
                }
            )
    return result


async def reconcile_lecturer_drive_access(lecturer_id: str) -> None:
    lecturer = await db.users.find_one({"id": lecturer_id, "role": "lecturer"}, {"_id": 0})
    if not lecturer:
        return
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        return
    old_folder_id = str(lecturer.get("drive_folder_id", ""))
    old_permission_id = str(lecturer.get("drive_permission_id", ""))
    should_share = bool(
        settings.get("lecturer_folder_sharing_enabled") and lecturer.get("status", "active") == "active"
    )
    if old_folder_id and old_permission_id and not should_share:
        try:
            await asyncio.to_thread(
                drive_revoke_user_permission,
                get_drive_service(settings),
                old_folder_id,
                old_permission_id,
            )
        except Exception as exc:
            logger.warning("Gagal mencabut akses Drive dosen %s: %s", lecturer_id, exc)
    try:
        drive_doc = await asyncio.to_thread(drive_prepare_lecturer_folder, settings, lecturer)
    except Exception as exc:
        logger.exception("Gagal menyiapkan folder Drive dosen %s: %s", lecturer_id, exc)
        await db.users.update_one(
            {"id": lecturer_id},
            {"$set": {"drive_access_status": "provision_failed", "drive_access_error": google_drive_error_message(exc)}},
        )
        return
    await db.users.update_one(
        {"id": lecturer_id},
        {"$set": {**drive_doc, "drive_access_updated_at": now_iso()}},
    )


async def reconcile_all_lecturer_drive_access() -> None:
    lecturer_ids = await db.users.distinct("id", {"role": "lecturer", "status": {"$ne": "deleted"}})
    for lecturer_id in lecturer_ids:
        await reconcile_lecturer_drive_access(str(lecturer_id))


def upload_to_drive(
    temp_path: str,
    filename: str,
    mime_type: str,
    hierarchy: List[str],
    settings: Dict[str, Any],
    lecturer_email: str = "",
) -> Dict[str, Any]:
    service = get_drive_service(settings)
    root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
    parent_id = settings.get("root_folder_id") or os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID")
    path_parts = [root_folder_name]
    if not parent_id:
        parent_id = drive_find_or_create_folder(service, root_folder_name, None)
    lecturer_folder_id = ""
    for folder_index, folder_name in enumerate(hierarchy):
        clean_folder_name = safe_path_segment(folder_name or "Tanpa Nama")
        parent_id = drive_find_or_create_folder(service, clean_folder_name, parent_id)
        path_parts.append(clean_folder_name)
        if folder_index == 1 and hierarchy[0] == "Dosen":
            lecturer_folder_id = parent_id
    media = MediaFileUpload(temp_path, mimetype=mime_type or "application/octet-stream", resumable=True)
    metadata = {"name": safe_path_segment(filename), "parents": [parent_id]}
    uploaded = (
        service.files()
        .create(
            body=metadata,
            media_body=media,
            fields="id,name,webViewLink,mimeType,size,parents",
            supportsAllDrives=True,
        )
        .execute()
    )
    result = {
        "drive_file_id": uploaded.get("id", ""),
        "drive_file_name": uploaded.get("name", filename),
        "drive_file_url": uploaded.get("webViewLink", ""),
        "drive_mime_type": uploaded.get("mimeType", mime_type),
        "drive_folder_id": parent_id,
        "drive_folder_path": str(Path(*path_parts)),
        "drive_uploaded_at": now_iso(),
    }
    if lecturer_folder_id:
        result.update(
            {
                "drive_lecturer_folder_id": lecturer_folder_id,
                "drive_access_status": "app_only",
                "drive_permission_id": "",
                "drive_permission_email": "",
                "drive_permission_role": "",
                "drive_access_error": "",
            }
        )
        if settings.get("lecturer_folder_sharing_enabled"):
            try:
                result.update(
                    drive_ensure_user_permission(
                        service,
                        lecturer_folder_id,
                        lecturer_email,
                        settings.get("lecturer_folder_role", "reader"),
                    )
                )
            except Exception as exc:
                result.update(
                    {
                        "drive_access_status": "share_failed",
                        "drive_access_error": google_drive_error_message(exc),
                    }
                )
    return result


def delete_drive_file_sync(file_id: str, settings: Dict[str, Any]) -> None:
    if not file_id:
        return
    try:
        get_drive_service(settings).files().delete(
            fileId=file_id,
            supportsAllDrives=True,
        ).execute()
    except Exception as exc:
        if int(getattr(getattr(exc, "resp", None), "status", 0) or 0) == 404:
            return
        raise


def download_drive_file_to_temp_sync(
    file_id: str,
    file_name: str,
    settings: Dict[str, Any],
) -> Path:
    suffix = Path(safe_path_segment(file_name or "attachment.bin")).suffix[:20]
    temporary = tempfile.NamedTemporaryFile(
        prefix="nugas-drive-",
        suffix=suffix,
        delete=False,
    )
    temporary_path = Path(temporary.name)
    try:
        request = get_drive_service(settings).files().get_media(
            fileId=file_id,
            supportsAllDrives=True,
        )
        downloader = MediaIoBaseDownload(temporary, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        temporary.close()
        return temporary_path
    except Exception:
        temporary.close()
        temporary_path.unlink(missing_ok=True)
        raise


def remove_temporary_file(path: Path) -> None:
    path.unlink(missing_ok=True)


async def refresh_embedded_file_references(file_id: str) -> None:
    updated = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not updated:
        return
    public_file = enrich_file_urls(public_doc(updated.copy()))
    if updated.get("record_type") == "physical_document":
        document_type = str(updated.get("document_type") or "").strip()
        student_id = str(updated.get("student_id") or updated.get("uploaded_by") or "").strip()
        if document_type in PHYSICAL_DOCUMENT_TYPES and student_id:
            public_file.update(
                {
                    "document_type": document_type,
                    "document_label": PHYSICAL_DOCUMENT_TYPES[document_type],
                    "student_id": student_id,
                    "student_nim": str(updated.get("student_nim") or ""),
                    "student_name": str(updated.get("student_name") or ""),
                    "angkatan": str(updated.get("angkatan") or ""),
                }
            )
            await db.users.update_one(
                {
                    "id": student_id,
                    f"physical_documents.{document_type}.file_id": file_id,
                },
                {
                    "$set": {
                        f"physical_documents.{document_type}": public_file,
                        "physical_documents_updated_at": now_iso(),
                    }
                },
            )
    await db.submissions.update_many(
        {"file.file_id": file_id},
        {"$set": {"file": public_file}},
    )
    await db.submissions.update_many(
        {"files.file_id": file_id},
        {"$set": {"files.$[item]": public_file}},
        array_filters=[{"item.file_id": file_id}],
    )
    await db.assignments.update_many(
        {"attachments.file_id": file_id},
        {"$set": {"attachments.$[item]": public_file}},
        array_filters=[{"item.file_id": file_id}],
    )
    await db.materials.update_many(
        {"attachment.file_id": file_id},
        {"$set": {"attachment": public_file, "file_url": public_file.get("file_url", "")}},
    )
    await db.comments.update_many(
        {"attachment.file_id": file_id},
        {"$set": {"attachment": public_file}},
    )
    await db.chat_messages.update_many(
        {"attachment.file_id": file_id},
        {"$set": {"attachment": public_file}},
    )


async def claim_drive_sync_attempt(file_id: str) -> Optional[int]:
    """Claim one of the five daily synchronization attempts for a file."""
    now = datetime.now(timezone.utc)
    attempt_day = sync_attempt_day(now, STORAGE_POLICY_TIMEZONE)
    async with _drive_sync_attempt_lock:
        file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
        if not file_doc:
            return None
        previous_attempts = (
            int(file_doc.get("drive_sync_attempts_today", 0) or 0)
            if file_doc.get("drive_sync_attempt_date") == attempt_day
            else 0
        )
        if previous_attempts >= DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY:
            await db.stored_files.update_one(
                {"id": file_id},
                {
                    "$set": {
                        "drive_sync_status": "failed",
                        "drive_next_retry_at": next_drive_retry_at(
                            previous_attempts,
                            now,
                            STORAGE_POLICY_TIMEZONE,
                        ),
                        "updated_at": now_iso(),
                    }
                },
            )
            return None
        attempts_today = previous_attempts + 1
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    "drive_sync_attempt_date": attempt_day,
                    "drive_sync_attempts_today": attempts_today,
                    "drive_sync_attempts_total": int(file_doc.get("drive_sync_attempts_total", 0) or 0) + 1,
                    "drive_last_attempt_at": now_iso(),
                    "drive_next_retry_at": (now + timedelta(minutes=15)).isoformat(),
                    "drive_sync_status": "pending",
                    "updated_at": now_iso(),
                }
            },
        )
        return attempts_today


async def mark_drive_sync_failure(file_id: str, error: str, attempts_today: int) -> None:
    await db.stored_files.update_one(
        {"id": file_id},
        {
            "$set": {
                "upload_status": "drive_upload_failed",
                "drive_sync_status": "failed",
                "drive_error": error[:500],
                "drive_next_retry_at": next_drive_retry_at(
                    attempts_today,
                    datetime.now(timezone.utc),
                    STORAGE_POLICY_TIMEZONE,
                ),
                "updated_at": now_iso(),
            }
        },
    )
    await refresh_embedded_file_references(file_id)


async def sync_stored_file_to_drive(file_id: str) -> None:
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        return
    if file_doc.get("drive_file_id") and file_doc.get("drive_file_url"):
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    "storage_provider": "google_drive",
                    "upload_status": "uploaded_to_drive",
                    "drive_sync_status": "synced",
                    "drive_next_retry_at": "",
                    "updated_at": now_iso(),
                },
                "$unset": {"drive_error": ""},
            },
        )
        await refresh_embedded_file_references(file_id)
        return
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    "drive_sync_status": "not_configured",
                    "drive_next_retry_at": (
                        datetime.now(timezone.utc) + timedelta(hours=1)
                    ).isoformat(),
                    "updated_at": now_iso(),
                }
            },
        )
        await refresh_embedded_file_references(file_id)
        return
    attempts_today = await claim_drive_sync_attempt(file_id)
    if attempts_today is None:
        await refresh_embedded_file_references(file_id)
        return
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0}) or file_doc
    local_path = resolved_stored_file_path(file_doc)
    if not local_path:
        await mark_drive_sync_failure(
            file_id,
            "File lokal tidak ditemukan untuk sinkron Google Drive.",
            attempts_today,
        )
        return
    if str(file_doc.get("local_path") or "") != str(local_path):
        await db.stored_files.update_one(
            {"id": file_id},
            {"$set": {"local_path": str(local_path), "local_available": True}},
        )
    try:
        drive_doc = await asyncio.to_thread(
            upload_to_drive,
            str(local_path),
            local_path.name,
            file_doc.get("mime_type") or "application/octet-stream",
            file_doc.get("drive_hierarchy") or [],
            settings,
            file_doc.get("lecturer_email", ""),
        )
        if not await db.stored_files.find_one({"id": file_id}, {"_id": 0, "id": 1}):
            orphan_doc = {
                "id": file_id,
                "file_name": file_doc.get("file_name", ""),
                "drive_file_id": drive_doc.get("drive_file_id", ""),
            }
            try:
                await asyncio.to_thread(
                    delete_drive_file_sync,
                    drive_doc.get("drive_file_id", ""),
                    settings,
                )
            except Exception as delete_exc:
                await queue_drive_file_deletion(
                    orphan_doc,
                    google_drive_error_message(delete_exc),
                )
            return
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    **drive_doc,
                    "storage_provider": "google_drive",
                    "upload_status": "uploaded_to_drive",
                    "drive_sync_status": "synced",
                    "drive_next_retry_at": "",
                    "updated_at": now_iso(),
                },
                "$unset": {"drive_error": ""},
            },
        )
        if file_doc.get("lecturer_id") and drive_doc.get("drive_lecturer_folder_id"):
            await db.users.update_one(
                {"id": file_doc["lecturer_id"], "role": "lecturer"},
                {
                    "$set": {
                        "drive_folder_id": drive_doc.get("drive_lecturer_folder_id", ""),
                        "drive_folder_name": file_doc.get("lecturer_folder_name", ""),
                        "drive_access_status": drive_doc.get("drive_access_status", "app_only"),
                        "drive_permission_id": drive_doc.get("drive_permission_id", ""),
                        "drive_permission_email": drive_doc.get("drive_permission_email", ""),
                        "drive_permission_role": drive_doc.get("drive_permission_role", ""),
                        "drive_access_error": drive_doc.get("drive_access_error", ""),
                        "drive_access_updated_at": now_iso(),
                    }
                },
            )
        await refresh_embedded_file_references(file_id)
    except Exception as exc:
        logger.exception("Sinkron Google Drive background gagal untuk %s: %s", file_doc.get("file_name", file_id), exc)
        drive_error = google_drive_error_message(exc)
        await mark_drive_sync_failure(file_id, drive_error, attempts_today)


async def save_uploaded_file_record(
    upload: UploadFile,
    hierarchy: List[str],
    owner_code: str,
    owner_name: str,
    uploaded_by: str,
    submission_id: str = "",
    assignment_id: str = "",
    record_type: str = "submission",
    sync_drive: bool = True,
    background_tasks: Optional[BackgroundTasks] = None,
    async_drive: bool = False,
    lecturer_id: str = "",
) -> Dict[str, Any]:
    drive_settings = await get_google_drive_settings(mask=False) if sync_drive else {}
    if sync_drive and google_drive_upload_required(drive_settings) and not google_drive_upload_enabled(drive_settings):
        raise HTTPException(
            status_code=503,
            detail="Google Drive belum dikonfigurasi. Isi konfigurasi Google Drive di menu admin terlebih dahulu.",
        )

    lecturer = {}
    if lecturer_id:
        lecturer = await db.users.find_one(
            {"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}},
            {"_id": 0, "id": 1, "employee_id": 1, "name": 1, "email": 1, "drive_folder_name": 1},
        ) or {}
    lecturer_folder = lecturer_drive_folder_name(lecturer) if lecturer else ""
    scoped_hierarchy = ["Dosen", lecturer_folder, *hierarchy] if lecturer_folder else list(hierarchy)

    file_id = new_id()
    original_filename = upload.filename or "upload.bin"
    mime_type = upload.content_type or "application/octet-stream"
    local_path, folder_path, storage_path = build_local_file_path(
        scoped_hierarchy,
        owner_code,
        owner_name,
        file_id,
        original_filename,
    )
    local_path.parent.mkdir(parents=True, exist_ok=True)
    size = 0
    with local_path.open("wb") as output:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            output.write(chunk)
    await upload.close()
    owner_folder = safe_path_segment(f"{owner_code or 'NO-ID'} - {owner_name}")
    drive_hierarchy = [*scoped_hierarchy, owner_folder]
    file_doc = {
        "id": file_id,
        "file_name": original_filename,
        "file_id": file_id,
        **local_file_urls(file_id),
        "mime_type": mime_type,
        "size": size,
        "folder_path": folder_path,
        "storage_path": storage_path,
        "storage_provider": "server_local",
        "local_path": str(local_path),
        "local_available": True,
        "uploaded_by": uploaded_by,
        "uploaded_at": now_iso(),
        "submission_id": submission_id,
        "assignment_id": assignment_id,
        "record_type": record_type,
        "lecturer_id": lecturer.get("id", lecturer_id),
        "lecturer_name": lecturer.get("name", ""),
        "lecturer_code": lecturer.get("employee_id", ""),
        "lecturer_email": lecturer.get("email", ""),
        "lecturer_folder_name": lecturer_folder,
        "upload_status": "stored_on_server",
        "drive_sync_status": "not_configured",
        "drive_sync_attempt_date": "",
        "drive_sync_attempts_today": 0,
        "drive_sync_attempts_total": 0,
        "drive_last_attempt_at": "",
        "drive_next_retry_at": "",
        "drive_hierarchy": drive_hierarchy,
    }

    if sync_drive and google_drive_upload_enabled(drive_settings):
        if async_drive and background_tasks:
            file_doc.update({"drive_sync_status": "pending"})
            await db.stored_files.insert_one(file_doc)
            background_tasks.add_task(sync_stored_file_to_drive, file_id)
            return public_doc(file_doc.copy())
        try:
            drive_doc = await asyncio.to_thread(
                upload_to_drive,
                str(local_path),
                local_path.name,
                mime_type,
                drive_hierarchy,
                drive_settings,
                lecturer.get("email", ""),
            )
            file_doc.update(
                {
                    **drive_doc,
                    "storage_provider": "google_drive",
                    "upload_status": "uploaded_to_drive",
                    "drive_sync_status": "synced",
                }
            )
        except Exception as exc:
            logger.exception("Upload Google Drive gagal untuk %s: %s", original_filename, exc)
            if google_drive_upload_required(drive_settings):
                local_path.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=503,
                    detail="Upload ke Google Drive gagal. File belum disimpan sebagai submission.",
                )
            file_doc.update(
                {
                    "upload_status": "drive_upload_failed",
                    "drive_sync_status": "failed",
                    "drive_error": str(exc)[:500],
                }
            )

    await db.stored_files.insert_one(file_doc)
    return public_doc(file_doc.copy())


async def enrich_class_payload(class_doc: Dict[str, Any]) -> Dict[str, Any]:
    course = await db.courses.find_one({"id": class_doc.get("course_id")}, {"_id": 0})
    if course:
        class_doc["course_code"] = course.get("code", class_doc.get("course_code", ""))
        class_doc["course_name"] = course.get("name", class_doc.get("course_name", ""))
        course_program_id = course.get("program_id") or course.get("prodi_id") or ""
        class_doc["program_id"] = course_program_id or class_doc.get("program_id", "")
        if course.get("program_name"):
            class_doc["program_name"] = course.get("program_name")
        elif course_program_id:
            program = await db.programs.find_one({"id": course_program_id}, {"_id": 0, "nama": 1})
            if program:
                class_doc["program_name"] = program.get("nama", class_doc.get("program_name", ""))
        class_doc["sks"] = course.get("sks", class_doc.get("sks", ""))
        class_doc["semester_paket"] = course.get("semester_paket", course.get("semester", class_doc.get("semester_paket", "")))
    class_doc["student_count"] = len(class_doc.get("student_ids", []))
    class_doc["status_label"] = class_status_label(class_doc.get("status", CLASS_STATUS_ACTIVE))
    class_doc["read_only"] = class_is_read_only(class_doc)
    class_doc["allows_learning"] = class_allows_learning(class_doc)
    class_doc["allows_grading"] = class_allows_grading(class_doc)
    try:
        rps_complete, _ = await class_rps_complete(class_doc.get("id", ""))
        class_doc["rps_complete"] = rps_complete
        rps_doc = await db.rps.find_one({"class_id": class_doc.get("id")}, {"_id": 0, "id": 1})
        class_doc["rps_exists"] = rps_doc is not None
    except Exception:
        class_doc["rps_complete"] = False
        class_doc["rps_exists"] = False
    return class_doc


async def enrich_course_payload(course_doc: Dict[str, Any]) -> Dict[str, Any]:
    program = await db.programs.find_one({"id": course_doc.get("program_id")}, {"_id": 0})
    course_doc["program_name"] = program.get("name", course_doc.get("program_name", "")) if program else course_doc.get("program_name", "")
    course_doc["has_dosen_pengampu"] = bool(str(course_doc.get("dosen_utama_id") or "").strip())
    customized = isinstance(course_doc.get("grade_weights"), dict)
    course_doc["grade_weights"] = grade_weights_from_document(course_doc.get("grade_weights"))
    course_doc["grade_weights_customized"] = customized
    return course_doc


async def normalize_new_class_name(course: Dict[str, Any], value: str) -> str:
    """Expand numeric rombel names to the Feeder-friendly prodi prefix format."""
    raw_name = str(value or "").strip()
    if not re.fullmatch(r"\d{1,2}", raw_name):
        return raw_name

    program_id = course.get("program_id") or course.get("prodi_id") or ""
    program = None
    programs_collection = getattr(db, "programs", None)
    if program_id and programs_collection is not None:
        program = await programs_collection.find_one(
            {"id": program_id},
            {"_id": 0, "nama": 1, "name": 1, "kode": 1, "code": 1},
        )
    prodi_kode = (
        (program.get("kode") or program.get("code", ""))
        if program
        else course.get("prodi_kode") or course.get("program_code") or ""
    )
    prodi_nama = (
        (program.get("nama") or program.get("name", ""))
        if program
        else course.get("program_name") or course.get("prodi_name") or ""
    )
    if not prodi_kode and not prodi_nama:
        return raw_name
    return _recommended_rombel_name(prodi_kode, prodi_nama, int(raw_name))


async def require_course_lecturer(course: Dict[str, Any]) -> tuple[str, str]:
    """Return the assigned lecturer or stop class creation with a clear error."""
    lecturer_id = str(course.get("dosen_utama_id") or "").strip()
    if not lecturer_id:
        raise HTTPException(
            status_code=409,
            detail="Dosen pengampu mata kuliah belum ditetapkan. Tetapkan dosen pengampu terlebih dahulu.",
        )

    lecturer = await db.users.find_one(
        {
            "id": lecturer_id,
            "role": {"$nin": ["student", "Mahasiswa", "mahasiswa"]},
            "status": {"$ne": "deleted"},
        },
        {"_id": 0, "name": 1},
    )
    if not lecturer:
        raise HTTPException(
            status_code=409,
            detail="Dosen pengampu mata kuliah tidak ditemukan atau sudah tidak aktif.",
        )

    lecturer_name = str(course.get("dosen_utama_nama") or lecturer.get("name") or "").strip()
    if not lecturer_name:
        raise HTTPException(
            status_code=409,
            detail="Nama dosen pengampu mata kuliah belum lengkap. Perbarui penugasan dosen terlebih dahulu.",
        )
    return lecturer_id, lecturer_name


async def enrich_material_payload(material_doc: Dict[str, Any]) -> Dict[str, Any]:
    if isinstance(material_doc.get("attachment"), dict):
        material_doc["attachment"] = enrich_file_urls(material_doc["attachment"])
    class_doc = await db.classes.find_one({"id": material_doc.get("class_id")}, {"_id": 0})
    class_doc = await enrich_class_payload(class_doc) if class_doc else {}
    linked_assignment = await db.assignments.find_one(
        {"material_id": material_doc.get("id")},
        {"_id": 0, "class_name": 1, "course_id": 1, "course_name": 1},
    ) or {}
    material_doc.update(
        {
            "class_name": class_doc.get("name") or linked_assignment.get("class_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "class_status": class_doc.get("status", ""),
            "class_status_label": class_doc.get("status_label") or class_status_label(class_doc.get("status", "")),
            "class_read_only": bool(class_doc.get("read_only", class_is_read_only(class_doc))),
            "class_allows_learning": bool(class_doc.get("allows_learning", class_allows_learning(class_doc))),
            "class_allows_grading": bool(class_doc.get("allows_grading", class_allows_grading(class_doc))),
            "course_id": class_doc.get("course_id") or linked_assignment.get("course_id", ""),
            "course_name": class_doc.get("course_name") or linked_assignment.get("course_name", ""),
            "academic_year": class_doc.get("academic_year", ""),
            "semester": class_doc.get("semester", ""),
        }
    )
    return material_doc


async def enrich_materials_batch(materials: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not materials:
        return []
    class_ids = list(set(m.get("class_id", "") for m in materials if m.get("class_id")))
    material_ids = [m.get("id", "") for m in materials if m.get("id")]
    classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0}).to_list(1000) if class_ids else []
    classes_map: Dict[str, Dict[str, Any]] = {c["id"]: c for c in classes}
    course_ids = list(set(c.get("course_id", "") for c in classes if c.get("course_id")))
    courses = await db.courses.find({"id": {"$in": course_ids}}, {"_id": 0}).to_list(1000) if course_ids else []
    courses_map: Dict[str, Dict[str, Any]] = {c["id"]: c for c in courses}
    for class_doc in classes_map.values():
        course = courses_map.get(class_doc.get("course_id", ""))
        if course:
            class_doc["course_name"] = course.get("name", class_doc.get("course_name", ""))
            course_program_id = course.get("program_id") or course.get("prodi_id") or ""
            class_doc["program_id"] = course_program_id or class_doc.get("program_id", "")
            if course.get("program_name"):
                class_doc["program_name"] = course.get("program_name")
            elif course_program_id:
                program = await db.programs.find_one({"id": course_program_id}, {"_id": 0, "nama": 1})
                if program:
                    class_doc["program_name"] = program.get("nama", class_doc.get("program_name", ""))
        class_doc["student_count"] = len(class_doc.get("student_ids", []))
    linked_assignments = await db.assignments.find(
        {"material_id": {"$in": material_ids}},
        {"_id": 0, "material_id": 1, "class_name": 1, "course_id": 1, "course_name": 1},
    ).to_list(1000) if material_ids else []
    assignments_by_material: Dict[str, Dict[str, Any]] = {}
    for a in linked_assignments:
        mid = a.get("material_id", "")
        if mid and mid not in assignments_by_material:
            assignments_by_material[mid] = a
    for material in materials:
        if isinstance(material.get("attachment"), dict):
            material["attachment"] = enrich_file_urls(material["attachment"])
        class_doc = classes_map.get(material.get("class_id"), {})
        linked = assignments_by_material.get(material.get("id", ""), {})
        material.update({
            "class_name": class_doc.get("name") or linked.get("class_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "class_status": class_doc.get("status", ""),
            "class_status_label": class_status_label(class_doc.get("status", "")),
            "class_read_only": class_is_read_only(class_doc),
            "class_allows_learning": class_allows_learning(class_doc),
            "class_allows_grading": class_allows_grading(class_doc),
            "course_id": class_doc.get("course_id") or linked.get("course_id", ""),
            "course_name": class_doc.get("course_name") or linked.get("course_name", ""),
            "academic_year": class_doc.get("academic_year", ""),
            "semester": class_doc.get("semester", ""),
        })
    return materials


async def material_meeting_label_map(class_ids: Optional[List[str]] = None) -> Dict[str, str]:
    query: Dict[str, Any] = {}
    if class_ids is not None:
        query = {"class_id": {"$in": list(set(class_ids))}}
    material_docs = (
        await db.materials.find(query, {"_id": 0, "id": 1, "class_id": 1})
        .sort([("created_at", 1), ("id", 1)])
        .to_list(10000)
    )
    counts: Dict[str, int] = {}
    labels: Dict[str, str] = {}
    for material in material_docs:
        class_id = material.get("class_id", "")
        counts[class_id] = counts.get(class_id, 0) + 1
        labels[material["id"]] = f"Pertemuan {counts[class_id]}"
    return labels


def normalized_material_payload(payload: MaterialInput) -> Dict[str, Any]:
    doc = payload.model_dump()
    meeting_type = str(doc.get("meeting_type") or "offline").strip().lower()
    if meeting_type not in {"offline", "online"}:
        raise HTTPException(status_code=400, detail="Metode pertemuan harus offline atau online")
    meeting_url = str(doc.get("meeting_url") or "").strip()
    if meeting_url and not re.fullmatch(r"https://meet\.google\.com/[a-z0-9-]+", meeting_url, re.IGNORECASE):
        raise HTTPException(status_code=400, detail="Link Google Meet tidak valid")
    doc["meeting_type"] = meeting_type
    doc["meeting_url"] = meeting_url if meeting_type == "online" else ""
    try:
        doc["video_url"] = normalize_youtube_url(str(doc.get("video_url") or ""))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return doc


@api_router.get("/")
async def root():
    return {
        "message": "E-Learning Dosen API aktif",
        "version": version_payload(),
        **await storage_status_summary(),
    }


@api_router.get("/version")
async def app_version():
    schema_versions: list[str] = []
    try:
        rows = await db.pool.fetch("SELECT version FROM app_schema_migrations ORDER BY version")
        schema_versions = [str(row["version"]) for row in rows]
    except Exception:
        # Version information should stay available while the database is starting.
        schema_versions = []
    return version_payload(schema_versions=schema_versions)


@api_router.get("/auth/sso/config")
async def sso_config():
    settings = oidc_settings()
    login_url = ""
    if settings["enabled"] and settings["redirect_uri"]:
        callback = urlsplit(settings["redirect_uri"])
        login_url = f"{callback.scheme}://{callback.netloc}/api/auth/sso/login"
    local_enabled = bool(settings["local_login_enabled"])
    # Jika SSO tidak aktif, login lokal wajib aktif agar form login selalu muncul
    if not settings["enabled"]:
        local_enabled = True
    return {
        "enabled": bool(settings["enabled"]),
        "provider": "SCI-ID",
        "login_url": login_url,
        "local_login_enabled": local_enabled,
    }


@api_router.get("/auth/sso/login")
async def sso_login():
    settings = oidc_require_settings()
    metadata = await oidc_discovery()
    state = secrets.token_urlsafe(48)
    nonce = secrets.token_urlsafe(48)
    verifier = secrets.token_urlsafe(64)
    challenge = base64url_encode(hashlib.sha256(verifier.encode("ascii")).digest())
    await db.oidc_flows.insert_one(
        {
            "id": new_id(),
            "state_hash": hashlib.sha256(state.encode("utf-8")).hexdigest(),
            "nonce": nonce,
            "code_verifier": verifier,
            "created_at": now_iso(),
            "expires_at": datetime.now(timezone.utc) + timedelta(minutes=5),
        }
    )
    query = urlencode(
        {
            "client_id": settings["client_id"],
            "redirect_uri": settings["redirect_uri"],
            "response_type": "code",
            "scope": settings["scopes"],
            "state": state,
            "nonce": nonce,
            "code_challenge": challenge,
            "code_challenge_method": "S256",
        }
    )
    return RedirectResponse(f"{metadata['authorization_endpoint']}?{query}", status_code=307)


@api_router.get("/auth/sso/callback")
async def sso_callback(
    code: str = "",
    state: str = "",
    error: str = "",
    error_description: str = "",
):
    try:
        oidc_require_settings()
        if not state:
            raise HTTPException(status_code=400, detail="State login SCI-ID tidak tersedia")
        flow = await db.oidc_flows.find_one_and_delete(
            {"state_hash": hashlib.sha256(state.encode("utf-8")).hexdigest()},
            {"_id": 0},
        )
        if not flow:
            raise HTTPException(status_code=400, detail="State login SCI-ID tidak valid atau sudah digunakan")
        expires_at = flow.get("expires_at")
        if isinstance(expires_at, datetime):
            expires_at = expires_at.replace(tzinfo=timezone.utc) if expires_at.tzinfo is None else expires_at.astimezone(timezone.utc)
            if expires_at <= datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="Permintaan login SCI-ID sudah kedaluwarsa")
        if error:
            raise HTTPException(status_code=401, detail=error_description or f"SCI-ID menolak login: {error}")
        if not code:
            raise HTTPException(status_code=400, detail="Authorization code SCI-ID tidak tersedia")
        settings = oidc_require_settings()
        metadata = await oidc_discovery()
        token_payload = {
            "grant_type": "authorization_code",
            "client_id": settings["client_id"],
            "redirect_uri": settings["redirect_uri"],
            "code": code,
            "code_verifier": flow["code_verifier"],
        }
        if settings["client_secret"]:
            token_payload["client_secret"] = settings["client_secret"]
        try:
            async with httpx.AsyncClient(timeout=10) as http:
                token_response = await http.post(
                    metadata["token_endpoint"],
                    data=token_payload,
                    headers={"Accept": "application/json"},
                )
                token_response.raise_for_status()
                token_data = token_response.json()
        except (httpx.HTTPError, ValueError) as exc:
            logger.warning("Penukaran authorization code SCI-ID gagal: %s", type(exc).__name__)
            raise HTTPException(status_code=401, detail="Authorization code SCI-ID tidak dapat ditukar") from exc
        id_token = token_data.get("id_token") if isinstance(token_data, dict) else None
        if not isinstance(id_token, str) or not id_token:
            raise HTTPException(status_code=401, detail="SCI-ID tidak mengembalikan ID token")
        claims = await validate_oidc_id_token(id_token, metadata, flow["nonce"])
        user = await provision_oidc_user(claims)
        ticket = secrets.token_urlsafe(48)
        await db.oidc_login_tickets.insert_one(
            {
                "ticket_hash": hashlib.sha256(ticket.encode("utf-8")).hexdigest(),
                "user_id": user["id"],
                "sso_subject": claims["sub"],
                "created_at": now_iso(),
                "expires_at": datetime.now(timezone.utc) + timedelta(minutes=1),
            }
        )
        return RedirectResponse(oidc_frontend_redirect(sso_ticket=ticket), status_code=303)
    except HTTPException as exc:
        logger.warning("Callback SCI-ID ditolak: %s", exc.detail)
        return RedirectResponse(oidc_frontend_redirect(sso_error=str(exc.detail)), status_code=303)
    except Exception:
        logger.exception("Callback SCI-ID gagal")
        return RedirectResponse(oidc_frontend_redirect(sso_error="Login SCI-ID gagal diproses"), status_code=303)


EFFECTIVE_ACCESS_RESPONSE_FIELDS = (
    "access_mode",
    "template_id",
    "template_name",
    "base_permissions",
    "custom_permissions",
    "effective_permissions",
    "position_accesses",
    "access_roles",
    "access_scope_prodi_ids",
)


async def public_user_with_effective_access(user: Dict[str, Any]) -> Dict[str, Any]:
    """Return a login-safe user payload with the same access data as /auth/me."""
    response_user = public_doc(dict(user)) or {}
    if normalize_base_role(response_user.get("role")) not in {"admin", "lecturer", "student", "staff"}:
        return response_user
    try:
        effective_access = await build_effective_user_access(db, user)
        for field in EFFECTIVE_ACCESS_RESPONSE_FIELDS:
            if field in effective_access:
                response_user[field] = effective_access[field]
    except Exception:
        # Authentication must remain available while malformed legacy access
        # data is reported and repaired. The frontend fails closed when the
        # effective matrix is absent.
        logger.exception("Gagal memuat hak akses efektif untuk user %s", response_user.get("id"))
    return response_user


@api_router.post("/auth/sso/exchange")
async def sso_exchange(payload: SsoExchangeInput):
    ticket = await db.oidc_login_tickets.find_one_and_delete(
        {"ticket_hash": hashlib.sha256(payload.ticket.encode("utf-8")).hexdigest()},
        {"_id": 0},
    )
    if not ticket:
        raise HTTPException(status_code=401, detail="Tiket login SCI-ID tidak valid atau sudah digunakan")
    expires_at = ticket.get("expires_at")
    if isinstance(expires_at, datetime):
        expires_at = expires_at.replace(tzinfo=timezone.utc) if expires_at.tzinfo is None else expires_at.astimezone(timezone.utc)
        if expires_at <= datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Tiket login SCI-ID sudah kedaluwarsa")
    user = await find_user(ticket["user_id"])
    user, show_physical_documents_reminder = await auth_user_physical_document_reminder(user)
    token = new_id() + new_id()
    await db.sessions.insert_one(
        {
            "token": token,
            "user_id": user["id"],
            "auth_source": "sso",
            "sso_subject": ticket.get("sso_subject", ""),
            "created_at": now_iso(),
        }
    )
    queue_user_activity(
        user,
        "POST",
        "/api/auth/sso/exchange",
        200,
        action_override="login",
    )
    response_user = await public_user_with_effective_access(user)
    response_user["show_physical_documents_reminder"] = show_physical_documents_reminder
    return {"token": token, "user": response_user}


async def try_camaba_login(identifier: str, password: str) -> Optional[Dict[str, Any]]:
    """Login calon mahasiswa via nomor registrasi pada form login terpadu.

    Hanya nomor registrasi yang diterima sebagai identitas camaba. Email/WhatsApp
    camaba tidak dikenali untuk login — arahkan ke nomor registrasi. Akun yang
    sudah di-convert ke mahasiswa wajib login memakai NIM/Email/WhatsApp.
    """
    reg_number = identifier.strip().upper()
    applicant = await db.pmb_applicants.find_one(
        {"registration_number": reg_number}, {"_id": 0}
    )
    if not applicant:
        hint = await db.pmb_applicants.find_one(
            {"$or": [{"email": identifier}, {"whatsapp": identifier.strip()}]},
            {"_id": 0},
        )
        if hint and not hint.get("is_converted_to_student"):
            raise HTTPException(
                status_code=401,
                detail="Calon mahasiswa login menggunakan Nomor Registrasi (contoh: PMB20260001). "
                "Email/No WhatsApp baru berlaku setelah menjadi mahasiswa.",
            )
        return None
    if applicant.get("is_converted_to_student"):
        raise HTTPException(
            status_code=403,
            detail="Akun ini sudah menjadi mahasiswa. Silakan login menggunakan NIM, Email, atau WhatsApp.",
        )
    if not applicant.get("password_hash") or not verify_password(password, applicant["password_hash"]):
        raise HTTPException(status_code=401, detail="Nomor Registrasi atau password salah")
    token = new_id() + new_id()
    await db.sessions.insert_one({
        "token": token,
        "user_id": applicant["id"],
        "created_at": now_iso(),
    })
    await db.pmb_applicants.update_one(
        {"id": applicant["id"]},
        {"$set": {"last_login_at": now_iso()}},
    )
    camaba = public_doc(dict(applicant))
    camaba["role"] = "camaba"
    return {"token": token, "user": camaba}


async def auth_user_physical_document_reminder(
    user: Dict[str, Any],
) -> tuple[Dict[str, Any], bool]:
    """Mark and expose the one-time physical-document reminder for students."""
    show_reminder = bool(
        user.get("role") == "student"
        and not user.get("physical_documents_reminder_seen_at")
        and not physical_document_status_payload(user).get("is_complete")
    )
    if show_reminder:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"physical_documents_reminder_seen_at": now_iso()}},
        )
    return user, show_reminder


@api_router.post("/auth/login")
async def login(payload: LoginInput):
    local_login_enabled = bool(oidc_settings()["local_login_enabled"])
    if not oidc_settings()["enabled"]:
        local_login_enabled = True
    if not local_login_enabled:
        raise HTTPException(status_code=403, detail="Login lokal dinonaktifkan. Gunakan SCI-ID.")
    identifier = (payload.identifier or payload.email or "").strip().lower()
    if not identifier:
        raise HTTPException(status_code=400, detail="Username, NIM, nomor HP, atau email diperlukan")
    user = await find_unique_identity_user(identifier)
    if not user:
        camaba_result = await try_camaba_login(identifier, payload.password)
        if camaba_result:
            return camaba_result
        raise HTTPException(status_code=401, detail="Identitas login atau password salah")
    if not user.get("password_hash") or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Identitas login atau password salah")
    if user.get("status", "active") != "active":
        raise HTTPException(status_code=403, detail="Akun tidak aktif. Hubungi admin kampus.")
    user, show_physical_documents_reminder = await auth_user_physical_document_reminder(user)
    token = new_id() + new_id()
    await db.sessions.insert_one({"token": token, "user_id": user["id"], "created_at": now_iso()})
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login_at": now_iso()}})
    queue_user_activity(
        user,
        "POST",
        "/api/auth/login",
        200,
        action_override="login",
    )
    user = await public_user_with_effective_access(user)
    user["show_physical_documents_reminder"] = show_physical_documents_reminder
    return {"token": token, "user": user}


@api_router.post("/auth/register-student")
async def register_student(payload: RegisterStudentInput):
    identity = student_identity_values(payload.email, payload.nim, payload.username, payload.whatsapp)
    email = identity["email"]
    username = identity["username"]
    nim = identity["nim"]
    whatsapp = payload.whatsapp.strip()
    existing = await db.users.find_one(
        student_identity_conflict_query(email, nim, username, whatsapp),
        {"_id": 0},
    )
    if existing:
        raise HTTPException(status_code=409, detail="Email, username, NIM, atau WhatsApp sudah terdaftar")
    student_id = new_id()
    doc = {
        "id": student_id,
        "role": "student",
        "username": username,
        "nim": nim,
        "name": payload.name,
        "email": email,
        "whatsapp": whatsapp,
        "password_hash": hash_password(payload.password),
        "status": "active",
        "class_ids": [],
        "created_at": now_iso(),
        "last_login_at": "",
    }
    await db.users.insert_one(doc)
    return await login(LoginInput(identifier=email, password=payload.password))


@api_router.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordInput, background_tasks: BackgroundTasks):
    identifier = payload.identifier.strip().lower()
    try:
        user = await find_unique_identity_user(identifier)
    except HTTPException as exc:
        if exc.status_code != 409:
            raise
        logger.warning("Reset password diabaikan karena identitas ambigu: %s", identifier)
        user = None
    if not user:
        return {
            "ok": True,
            "message": "Jika akun ditemukan, permintaan reset password akan diproses.",
            "otp_delivery": {"status": "not_found", "message_id": ""},
            "email_delivery": {"status": "not_found", "error": ""},
        }
    response: Dict[str, Any] = {
        "ok": True,
        "message": "Jika akun ditemukan, permintaan reset password akan diproses.",
    }
    request_doc = {
        "id": new_id(),
        "identifier": identifier,
        "user_id": user["id"],
        "status": "otp_created",
        "requested_at": now_iso(),
    }
    otp = generate_otp()
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    reset_id = new_id()
    settings = await get_whatsapp_settings(mask=False)
    email_settings = await get_email_settings(mask=False)
    reset_link = build_password_reset_link(settings.get("app_url") or os.environ.get("APP_URL", ""), identifier)
    await db.password_reset_otps.insert_one(
        {
            "id": reset_id,
            "user_id": user["id"],
            "identifier": identifier,
            "otp": otp,
            "status": "active",
            "expires_at": expires_at,
            "created_at": now_iso(),
            "used_at": "",
        }
    )
    request_doc.update({"otp_id": reset_id, "expires_at": expires_at})
    email_delivery: Dict[str, Any] = {"status": "not_attempted", "error": ""}
    if email_settings.get("enabled") and user.get("email"):
        html_body = (
            f"<h3>Reset Password</h3>"
            f"<p>Halo {user.get('name', '')},</p>"
            f"<p>Kode OTP reset password Anda: <strong>{otp}</strong></p>"
            f"<p>Kode ini berlaku selama 10 menit.</p>"
            f"<p>Link reset: <a href='{reset_link}'>{reset_link}</a></p>"
            f"<p>Jika Anda tidak meminta reset password, abaikan email ini.</p>"
        )
        email_result = await send_email_message(user["email"], "Kode OTP Reset Password - E-Learning Dosen", html_body)
        email_delivery = {
            "status": "sent" if email_result["ok"] else "failed",
            "error": email_result.get("error", "") if not email_result["ok"] else "",
        }
        if email_result["ok"]:
            await db.email_messages.insert_one({
                "id": new_id(),
                "to": user["email"],
                "subject": "Kode OTP Reset Password - E-Learning Dosen",
                "message_type": "password_reset_otp",
                "ref_id": reset_id,
                "status": "sent",
                "created_at": now_iso(),
                "sent_at": now_iso(),
            })
        else:
            await db.email_messages.insert_one({
                "id": new_id(),
                "to": user["email"],
                "subject": "Kode OTP Reset Password - E-Learning Dosen",
                "message_type": "password_reset_otp",
                "ref_id": reset_id,
                "status": "failed",
                "error": email_result.get("error", ""),
                "created_at": now_iso(),
            })
    elif not user.get("email"):
        email_delivery = {"status": "no_email", "error": ""}
    response["email_delivery"] = email_delivery
    if user.get("whatsapp"):
        template = settings.get("otp_template") or "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}"
        message = template.format(code=otp, minutes=10, link=reset_link, name=user.get("name", ""))
        queued = await enqueue_whatsapp_message(user.get("whatsapp", ""), message, "password_reset_otp", reset_id)
        request_doc.update({"delivery_status": queued.get("status", ""), "message_id": queued.get("id", "")})
        delivery = public_whatsapp_delivery_status(queued)
        response["otp_delivery"] = delivery
        if queued.get("status") == "pending":
            background_tasks.add_task(send_whatsapp_message, queued["id"])
        response["message"] = forgot_password_response_message(delivery)
    else:
        request_doc.update({"delivery_status": "no_whatsapp", "message_id": ""})
        delivery = {"message_id": "", "status": "no_whatsapp", "provider": "", "created_at": now_iso(), "sent_at": "", "error": ""}
        response["otp_delivery"] = delivery
        if email_delivery.get("status") == "sent":
            response["message"] = "OTP sudah dikirim ke email Anda."
        elif email_delivery.get("status") == "failed":
            response["message"] = "Permintaan reset password diproses, tetapi gagal mengirim email."
        else:
            response["message"] = forgot_password_response_message(delivery)
    if local_reset_otp_enabled() and request_doc.get("delivery_status", "") in {"pending_config", "no_whatsapp", ""}:
        response["local_otp_available"] = True
    await db.password_reset_requests.insert_one(request_doc)
    return response


@api_router.get("/auth/forgot-password/messages/{message_id}")
async def forgot_password_message_status(message_id: str):
    message = await db.whatsapp_messages.find_one({"id": message_id, "message_type": "password_reset_otp"}, {"_id": 0})
    if not message:
        raise HTTPException(status_code=404, detail="Status antrean OTP tidak ditemukan")
    return public_whatsapp_delivery_status(message)


@api_router.post("/auth/reset-password-otp")
async def reset_password_otp(payload: ResetPasswordOtpInput):
    identifier = payload.identifier.strip().lower()
    user = await find_unique_identity_user(
        identifier,
        ambiguous_status=400,
        ambiguous_detail="OTP atau akun tidak valid",
    )
    if not user:
        raise HTTPException(status_code=400, detail="OTP atau akun tidak valid")
    otp_doc = await db.password_reset_otps.find_one(
        {"user_id": user["id"], "otp": payload.otp.strip(), "status": "active"}, {"_id": 0}
    )
    if not otp_doc:
        raise HTTPException(status_code=400, detail="OTP tidak valid")
    if datetime.fromisoformat(otp_doc["expires_at"].replace("Z", "+00:00")) < datetime.now(timezone.utc):
        await db.password_reset_otps.update_one({"id": otp_doc["id"]}, {"$set": {"status": "expired"}})
        raise HTTPException(status_code=400, detail="OTP sudah kedaluwarsa")
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": hash_password(payload.new_password), "password_changed_at": now_iso()}})
    await db.password_reset_otps.update_one({"id": otp_doc["id"]}, {"$set": {"status": "used", "used_at": now_iso()}})
    return {"ok": True}


@api_router.post("/auth/change-password")
async def change_password(payload: ChangePasswordInput, user: Dict[str, Any] = Depends(get_current_user)):
    full_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not full_user or not verify_password(payload.current_password, full_user["password_hash"]):
        raise HTTPException(status_code=400, detail="Password lama salah")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": hash_password(payload.new_password), "password_changed_at": now_iso()}},
    )
    return {"ok": True}


@api_router.post("/auth/join-class")
async def join_class(payload: JoinClassInput):
    """Legacy registration endpoint kept for compatibility.

    It no longer grants class membership directly. New and existing students
    are signed in without the class link and receive a pending enrollment
    request that must be approved by the lecturer/admin.
    """
    class_code = clean_code(payload.class_code)
    class_doc = await db.classes.find_one({"class_code": class_code, "status": "active"}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kode kelas tidak ditemukan")
    identity = student_identity_values(payload.email, payload.nim, payload.nim, payload.whatsapp)
    existing = await db.users.find_one({"email": identity["email"]}, {"_id": 0})
    if existing:
        if existing.get("role") != "student":
            raise HTTPException(status_code=409, detail="Email sudah digunakan oleh akun non-mahasiswa")
        if normalize_nim(existing.get("nim")) != identity["nim"]:
            raise HTTPException(
                status_code=409,
                detail="Email sudah terdaftar dengan NIM berbeda. Gunakan identitas akun yang sudah terdaftar.",
            )
        if not verify_password(payload.password, existing.get("password_hash", "")):
            raise HTTPException(status_code=401, detail="Password akun mahasiswa salah")
        student_id = existing["id"]
    else:
        conflict = await db.users.find_one(
            student_identity_conflict_query(
                identity["email"],
                identity["nim"],
                identity["username"],
                identity["whatsapp"],
            ),
            {"_id": 0, "id": 1},
        )
        if conflict:
            raise HTTPException(status_code=409, detail="Email, username, NIM, atau WhatsApp sudah terdaftar")
        student_id = new_id()
        await db.users.insert_one(
            {
                "id": student_id,
                "role": "student",
                "username": identity["username"],
                "nim": identity["nim"],
                "name": payload.name,
                "email": identity["email"],
                "whatsapp": identity["whatsapp"],
                "password_hash": hash_password(payload.password),
                "status": "active",
                "class_ids": [],
                "created_at": now_iso(),
                "last_login_at": "",
            }
        )
    already_joined = bool(existing and class_doc["id"] in existing.get("class_ids", []))
    if already_joined:
        login_response = await login(LoginInput(identifier=payload.email, password=payload.password))
        login_response["enrollment_status"] = "approved"
        return login_response
    request_doc = await db.enrollment_requests.find_one(
        {"class_id": class_doc["id"], "student_id": student_id, "status": "pending"},
        {"_id": 0},
    )
    if not request_doc:
        request_doc = {
            "id": new_id(),
            "class_id": class_doc["id"],
            "class_name": class_doc["name"],
            "class_code": class_doc.get("class_code", ""),
            "lecturer_id": class_doc.get("lecturer_id", ""),
            "lecturer_name": class_doc.get("lecturer_name", ""),
            "student_id": student_id,
            "student_name": (existing or {}).get("name") or payload.name,
            "student_nim": (existing or {}).get("nim") or identity["nim"],
            "student_email": identity["email"],
            "status": "pending",
            "requested_at": now_iso(),
            "source": "legacy_join_class",
        }
        await db.enrollment_requests.insert_one(request_doc)
    login_response = await login(LoginInput(identifier=payload.email, password=payload.password))
    login_response["enrollment_status"] = "pending"
    login_response["enrollment_request_id"] = request_doc["id"]
    return login_response


@api_router.get("/auth/me")
async def me(user: Dict[str, Any] = Depends(get_current_user)):
    return await public_user_with_effective_access(user)


@api_router.get("/auth/physical-documents")
async def list_physical_documents(user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Fitur ini hanya tersedia untuk mahasiswa")
    fresh_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0}) or user
    return {
        **physical_document_status_payload(fresh_user),
        "max_file_size_mb": PHYSICAL_DOCUMENT_MAX_FILE_MB,
        "allowed_extensions": sorted(PHYSICAL_DOCUMENT_ALLOWED_EXTENSIONS),
    }


@api_router.post("/auth/physical-documents")
async def upload_physical_document(
    background_tasks: BackgroundTasks,
    document_type: str = Form(...),
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(get_current_user),
):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Fitur ini hanya tersedia untuk mahasiswa")
    normalized_type = normalize_physical_document_type(document_type)
    if not file.filename:
        raise HTTPException(status_code=400, detail="File dokumen fisik wajib dipilih")
    extension = Path(file.filename).suffix.lower()
    if extension not in PHYSICAL_DOCUMENT_ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Format dokumen harus PDF, JPG, JPEG, PNG, atau WEBP",
        )
    await validate_upload_file_sizes(
        [file],
        PHYSICAL_DOCUMENT_MAX_FILE_MB,
        PHYSICAL_DOCUMENT_TYPES[normalized_type],
    )

    current_user = await db.users.find_one({"id": user["id"]}, {"_id": 0}) or user
    old_document = (current_user.get("physical_documents") or {}).get(normalized_type) or {}
    angkatan = str(
        current_user.get("angkatan")
        or current_user.get("academic_year")
        or "Tanpa Angkatan"
    ).strip()
    file_doc = await save_uploaded_file_record(
        file,
        ["Kelengkapan Data Fisik", angkatan],
        str(current_user.get("nim") or current_user.get("username") or current_user["id"]),
        str(current_user.get("name") or "Mahasiswa"),
        current_user["id"],
        record_type="physical_document",
        sync_drive=True,
        background_tasks=background_tasks,
        async_drive=True,
    )
    file_doc.update(
        {
            "document_type": normalized_type,
            "document_label": PHYSICAL_DOCUMENT_TYPES[normalized_type],
            "student_id": current_user["id"],
            "student_nim": str(current_user.get("nim") or current_user.get("username") or ""),
            "student_name": str(current_user.get("name") or "Mahasiswa"),
            "angkatan": angkatan,
        }
    )
    await db.stored_files.update_one(
        {"id": file_doc["id"]},
        {
            "$set": {
                "document_type": normalized_type,
                "document_label": PHYSICAL_DOCUMENT_TYPES[normalized_type],
                "student_id": current_user["id"],
                "student_nim": file_doc["student_nim"],
                "student_name": file_doc["student_name"],
                "angkatan": angkatan,
            }
        },
    )
    await db.users.update_one(
        {"id": current_user["id"]},
        {
            "$set": {
                f"physical_documents.{normalized_type}": file_doc,
                "physical_documents_updated_at": now_iso(),
            }
        },
    )
    old_file_id = str(old_document.get("file_id") or old_document.get("id") or "").strip()
    if old_file_id and old_file_id != file_doc["id"]:
        await delete_stored_files({"id": old_file_id})
    fresh_user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0}) or current_user
    return {
        "ok": True,
        "document_type": normalized_type,
        "document": file_doc,
        "status": physical_document_status_payload(fresh_user),
        "user": public_doc(fresh_user),
    }


@api_router.delete("/auth/physical-documents/{document_type}")
async def delete_physical_document(
    document_type: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Fitur ini hanya tersedia untuk mahasiswa")
    normalized_type = normalize_physical_document_type(document_type)
    current_user = await db.users.find_one({"id": user["id"]}, {"_id": 0}) or user
    old_document = (current_user.get("physical_documents") or {}).get(normalized_type) or {}
    await db.users.update_one(
        {"id": user["id"]},
        {
            "$unset": {f"physical_documents.{normalized_type}": ""},
            "$set": {"physical_documents_updated_at": now_iso()},
        },
    )
    old_file_id = str(old_document.get("file_id") or old_document.get("id") or "").strip()
    if old_file_id:
        await delete_stored_files({"id": old_file_id})
    fresh_user = await db.users.find_one({"id": user["id"]}, {"_id": 0}) or user
    return {"ok": True, "status": physical_document_status_payload(fresh_user), "user": public_doc(fresh_user)}


@api_router.put("/auth/me")
async def update_profile(payload: ProfileInput, user: Dict[str, Any] = Depends(get_current_user)):
    name = payload.name.strip()
    username = payload.username.strip().lower()
    email = str(payload.email).strip().lower()
    whatsapp = payload.whatsapp.strip()
    if not name or not username:
        raise HTTPException(status_code=400, detail="Nama dan username wajib diisi")
    current_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "whatsapp": 1})
    current_whatsapp = (current_user or {}).get("whatsapp", "")
    candidates: List[Dict[str, Any]] = [{"email": email}, {"username": username}]
    if whatsapp and normalize_phone(whatsapp) != normalize_phone(current_whatsapp):
        candidates.append({"whatsapp": whatsapp})
    existing = await db.users.find_one(
        {"id": {"$ne": user["id"]}, "$or": candidates},
        {"_id": 0, "id": 1},
    )
    if existing:
        raise HTTPException(status_code=409, detail="Email, username, atau WhatsApp sudah digunakan akun lain")
    set_dict = {
        "name": name,
        "username": username,
        "email": email,
        "whatsapp": whatsapp,
        "profile_updated_at": now_iso(),
    }
    extra_fields = [
        "employee_id", "nip", "nidn", "nik", "gelar", "gelar_depan", "gelar_belakang",
        "prodi_id", "prodi_name", "prodi_kode", "homebase", "gender", "agama",
        "tempat_lahir", "tanggal_lahir", "alamat", "kota", "provinsi", "kode_pos",
        "spesialisasi", "jabatan", "status_kepegawaian", "nim", "nisn", "angkatan",
        "academic_year", "semester", "parent_name", "parent_phone", "parent_job", "parent_address",
        "nuptk", "nrsd", "nama_panggilan", "kewarganegaraan", "rt", "rw", "dusun",
        "kelurahan", "kecamatan", "kode_wilayah", "jenis_tinggal_id", "jenis_tinggal",
        "transportasi_id", "transportasi", "asal_sekolah", "status_sipil", "no_kk", "npwp",
        "no_kip", "no_kps", "kebutuhan_khusus", "tinggi_badan", "berat_badan", "semester_masuk",
        "tanggal_masuk", "jenis_pendaftaran_id", "jenis_pendaftaran", "jalur_masuk_id", "jalur_masuk",
        "jenis_pembiayaan_id", "jenis_pembiayaan", "status_mahasiswa_id", "feeder_student_id",
        "feeder_registration_id", "foto_url", "orang_tua", "registration", "pddikti_ids",
        "jabatan_dikti_id", "jabatan_kode", "jenjang_pendidikan", "ikatan_kerja", "status_pegawai",
        "status_kerja", "jenis_pegawai", "pangkat_golongan_id", "pangkat_golongan", "no_sk",
        "unit_organisasi_id", "institusi_induk", "status_dosen", "status_dosen_id", "tanggal_mulai_mengajar"
    ]
    for field in extra_fields:
        val = getattr(payload, field, None)
        if val is not None:
            set_dict[field] = val

    await db.users.update_one(
        {"id": user["id"]},
        {"$set": set_dict},
    )
    if user.get("role") in {"admin", "lecturer"} and name != user.get("name"):
        await db.classes.update_many({"lecturer_id": user["id"]}, {"$set": {"lecturer_name": name}})
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many({"lecturer_id": user["id"]}, {"$set": {"lecturer_name": name}})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    return public_doc(updated)


@api_router.post("/auth/avatar")
async def upload_user_avatar(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(get_current_user),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File foto profil tidak valid")

    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Format foto harus berupa JPG, PNG, WEBP, atau GIF")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran foto maksimal 5 MB")

    file_token = secrets.token_hex(8)
    safe_name = safe_path_segment(file.filename) or f"avatar{ext}"
    filename = f"avatar_{user['id']}_{file_token[:8]}_{safe_name}"

    avatar_dir = STORAGE_ROOT / "Avatars"
    avatar_dir.mkdir(parents=True, exist_ok=True)
    file_path = avatar_dir / filename
    file_path.write_bytes(content)

    file_id = f"avatar-{user['id']}-{file_token[:8]}"
    file_doc = {
        "id": file_id,
        "record_type": "avatar",
        "owner_user_id": user["id"],
        "file_name": safe_name,
        "original_name": file.filename,
        "mime_type": file.content_type or f"image/{ext.replace('.', '')}",
        "size": len(content),
        "storage_path": portable_storage_path_from_local_path(str(file_path)),
        "local_path": str(file_path),
        "local_available": True,
        "created_at": now_iso(),
    }
    await db.stored_files.update_one({"id": file_id}, {"$set": file_doc}, upsert=True)

    avatar_url = f"/api/files/{file_id}/inline"
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"avatar_url": avatar_url, "avatar_file_id": file_id}},
    )

    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    return public_doc(updated_user)


@api_router.delete("/auth/avatar")
async def remove_user_avatar(
    user: Dict[str, Any] = Depends(get_current_user),
):
    await db.users.update_one(
        {"id": user["id"]},
        {"$unset": {"avatar_url": "", "avatar_file_id": ""}},
    )
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    return public_doc(updated_user)


@api_router.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    logout_url = ""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "", 1).strip()
        _auth_cache.pop(token, None)
        session = await db.sessions.find_one_and_delete({"token": token}, {"_id": 0})
        if session:
            logout_user = await db.users.find_one(
                {"id": session.get("user_id", "")},
                {"_id": 0, "password_hash": 0},
            )
            if logout_user:
                queue_user_activity(
                    logout_user,
                    "POST",
                    "/api/auth/logout",
                    200,
                    action_override="logout",
                )
        if session and session.get("auth_source") == "sso" and oidc_settings()["enabled"]:
            try:
                settings = oidc_require_settings()
                metadata = await oidc_discovery()
                logout_url = f"{metadata['end_session_endpoint']}?{urlencode({'client_id': settings['client_id'], 'post_logout_redirect_uri': settings['frontend_url']})}"
            except HTTPException:
                logout_url = ""
    return {"ok": True, "logout_url": logout_url}


@api_router.get("/password-reset-requests")
async def list_password_reset_requests(_: Dict[str, Any] = Depends(require_admin)):
    return await db.password_reset_requests.find({}, {"_id": 0}).sort("requested_at", -1).to_list(1000)


@api_router.get("/lecturers")
async def list_lecturers(user: Dict[str, Any] = Depends(get_current_user)):
    lecturers = await db.users.find(
        {"role": {"$in": ["lecturer", "admin", "dosen"]}, "status": {"$ne": "deleted"}},
        {"_id": 0, "password_hash": 0}
    ).sort("name", 1).to_list(2000)
    if not lecturers:
        lecturers = await db.users.find(
            {"role": {"$nin": ["student", "staff", "tendik", "staf", "pegawai"]}},
            {"_id": 0, "password_hash": 0}
        ).sort("name", 1).to_list(2000)
    structural_scope = await active_program_manager_scope_values(user)
    if user.get("role") != "admin" and (
        user_is_program_manager(user) or structural_scope
    ):
        scope_values = await resolved_program_scope_values(user, structural_scope)
        lecturers = [
            lecturer
            for lecturer in lecturers
            if record_matches_program_scope(
                lecturer,
                scope_values,
                fields=(
                    "prodi_id", "prodi_kode", "program_id", "program_code",
                    "prodi_name", "program_name", "nama_prodi", "homebase",
                ),
            )
        ]
    counts = await db.classes.aggregate(
        [
            {"$match": {"status": {"$ne": "deleted"}, "lecturer_id": {"$ne": ""}}},
            {"$group": {"_id": "$lecturer_id", "total": {"$sum": 1}}},
        ]
    ).to_list(2000)
    class_counts = {item["_id"]: item["total"] for item in counts}
    storage_counts = await db.stored_files.aggregate(
        [
            {"$match": {"lecturer_id": {"$nin": ["", None]}}},
            {
                "$group": {
                    "_id": "$lecturer_id",
                    "file_count": {"$sum": 1},
                    "storage_bytes": {"$sum": {"$ifNull": ["$size", 0]}},
                    "drive_synced_count": {
                        "$sum": {"$cond": [{"$eq": ["$drive_sync_status", "synced"]}, 1, 0]}
                    },
                    "drive_failed_count": {
                        "$sum": {"$cond": [{"$eq": ["$drive_sync_status", "failed"]}, 1, 0]}
                    },
                }
            },
        ]
    ).to_list(2000)
    lecturer_storage = {item["_id"]: item for item in storage_counts}
    for lecturer in lecturers:
        lecturer["class_count"] = class_counts.get(lecturer["id"], 0)
        storage = lecturer_storage.get(lecturer["id"], {})
        lecturer["storage_file_count"] = storage.get("file_count", 0)
        lecturer["storage_bytes"] = storage.get("storage_bytes", 0)
        lecturer["drive_synced_count"] = storage.get("drive_synced_count", 0)
        lecturer["drive_failed_count"] = storage.get("drive_failed_count", 0)
    return lecturers


@api_router.get("/academic-position-candidates")
async def list_academic_position_candidates(
    _: Dict[str, Any] = Depends(require_admin_or_academic_operator),
):
    """Daftar dosen dan tendik yang dapat ditunjuk pada Jabatan Akademik.

    Endpoint ini sengaja terpisah dari ``/lecturers`` agar Tendik tersedia di
    halaman penugasan jabatan tanpa ikut muncul sebagai Dosen pada modul lain.
    """
    candidates = await db.users.find(
        {
            "role": {
                "$in": [
                    "admin", "lecturer", "dosen",
                    "staff", "tendik", "staf", "pegawai",
                ]
            },
            "status": {"$ne": "deleted"},
        },
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(4000)
    for item in candidates:
        item["role"] = normalize_base_role(item.get("role"))
        item["candidate_type"] = "tendik" if item["role"] == "staff" else "dosen"
    return candidates


@api_router.get("/staff")
async def list_staff(_: Dict[str, Any] = Depends(require_campus_admin)):
    staff = await db.users.find(
        {
            "role": {"$in": ["staff", "tendik", "staf", "pegawai"]},
            "status": {"$ne": "deleted"},
        },
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(2000)
    for item in staff:
        item["role"] = "staff"
    return staff


@api_router.post("/staff")
async def create_staff(
    payload: StaffInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status tendik harus active atau inactive")
    master_fields = await resolve_staff_master_references(
        payload.jabatan_id,
        payload.unit_organisasi_id,
    )
    username = payload.username.strip().lower()
    email = str(payload.email).strip().lower()
    duplicate = await db.users.find_one(
        {"$or": [{"username": username}, {"email": email}]},
        {"_id": 0, "id": 1},
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email tendik sudah digunakan")
    doc = {
        "id": new_id(),
        "role": "staff",
        "class_ids": [],
        "access_roles": [],
        **payload.model_dump(),
        **master_fields,
        "username": username,
        "email": email,
        "password_hash": hash_password(payload.password),
        "created_at": now_iso(),
        "last_login_at": "",
    }
    doc.pop("password", None)
    await db.users.insert_one(doc)
    return public_doc(doc)


@api_router.put("/staff/{staff_id}")
async def update_staff(
    staff_id: str,
    payload: StaffUpdateInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    existing = await db.users.find_one(
        {
            "id": staff_id,
            "role": {"$in": ["staff", "tendik", "staf", "pegawai"]},
            "status": {"$ne": "deleted"},
        },
        {"_id": 0},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Tendik tidak ditemukan")
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status tendik harus active atau inactive")
    master_fields = await resolve_staff_master_references(
        payload.jabatan_id,
        payload.unit_organisasi_id,
    )
    username = payload.username.strip().lower()
    email = str(payload.email).strip().lower()
    duplicate = await db.users.find_one(
        {
            "id": {"$ne": staff_id},
            "$or": [{"username": username}, {"email": email}],
        },
        {"_id": 0, "id": 1},
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email tendik sudah digunakan")
    update = payload.model_dump(exclude={"password"})
    update.update({
        **master_fields,
        "username": username,
        "email": email,
        "role": "staff",
        "updated_at": now_iso(),
    })
    if payload.password:
        update["password_hash"] = hash_password(payload.password)
    await db.users.update_one({"id": staff_id}, {"$set": update})
    if payload.status != existing.get("status"):
        await db.sessions.delete_many({"user_id": staff_id})
    return public_doc(await db.users.find_one({"id": staff_id}, {"_id": 0}))


@api_router.post("/staff/{staff_id}/reset-password")
async def reset_staff_password(
    staff_id: str,
    payload: ResetPasswordInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    staff = await db.users.find_one(
        {
            "id": staff_id,
            "role": {"$in": ["staff", "tendik", "staf", "pegawai"]},
            "status": {"$ne": "deleted"},
        },
        {"_id": 0, "id": 1},
    )
    if not staff:
        raise HTTPException(status_code=404, detail="Tendik tidak ditemukan")
    password = payload.password.strip() or "Tendik123!"
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password tendik minimal 6 karakter")
    await db.users.update_one(
        {"id": staff_id},
        {"$set": {"password_hash": hash_password(password), "password_reset_at": now_iso()}},
    )
    await db.sessions.delete_many({"user_id": staff_id})
    return {"ok": True, "temporary_password": password}


@api_router.delete("/staff/{staff_id}")
async def delete_staff(
    staff_id: str,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    result = await db.users.update_one(
        {
            "id": staff_id,
            "role": {"$in": ["staff", "tendik", "staf", "pegawai"]},
        },
        {"$set": {"role": "staff", "status": "deleted", "deleted_at": now_iso()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tendik tidak ditemukan")
    await db.sessions.delete_many({"user_id": staff_id})
    return {"ok": True}


@api_router.get("/jabatan-assignments")
@api_router.get("/jabatan-assignments/")
async def list_jabatan_assignments_api_alias(request: Request):
    from routers.master_data import list_jabatan_assignments
    return await list_jabatan_assignments(db=request.app.state.db)


@api_router.post("/jabatan-assignments")
@api_router.post("/jabatan-assignments/")
@api_router.put("/jabatan-assignments")
@api_router.put("/jabatan-assignments/")
async def save_jabatan_assignment_api_alias(request: Request):
    from routers.master_data import save_jabatan_assignment, JabatanAssignmentInput
    body_json = await request.json()
    input_data = JabatanAssignmentInput(**body_json)
    return await save_jabatan_assignment(body=input_data, db=request.app.state.db)


@api_router.post("/lecturers")
async def create_lecturer(
    payload: LecturerInput,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status dosen harus active atau inactive")
    username = payload.username.strip().lower()
    email = payload.email.lower()
    duplicate = await db.users.find_one({"$or": [{"username": username}, {"email": email}]}, {"_id": 0, "id": 1})
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email dosen sudah digunakan")
    doc = {
        "id": new_id(),
        "role": "lecturer",
        **payload.model_dump(),
        "username": username,
        "email": email,
        "password_hash": hash_password(payload.password),
        "created_at": now_iso(),
        "last_login_at": "",
    }
    doc.pop("password", None)
    await db.users.insert_one(doc)
    background_tasks.add_task(reconcile_lecturer_drive_access, doc["id"])
    return public_doc(doc)


@api_router.put("/lecturers/{lecturer_id}")
async def update_lecturer(
    lecturer_id: str,
    payload: LecturerUpdateInput,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    existing = await db.users.find_one({"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status dosen harus active atau inactive")
    username = payload.username.strip().lower()
    email = payload.email.lower()
    duplicate = await db.users.find_one(
        {"id": {"$ne": lecturer_id}, "$or": [{"username": username}, {"email": email}]}, {"_id": 0, "id": 1}
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email dosen sudah digunakan")
    update = payload.model_dump()
    update.update({"username": username, "email": email, "updated_at": now_iso()})
    await db.users.update_one({"id": lecturer_id}, {"$set": update})
    if payload.name != existing.get("name"):
        await db.classes.update_many({"lecturer_id": lecturer_id}, {"$set": {"lecturer_name": payload.name}})
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many({"lecturer_id": lecturer_id}, {"$set": {"lecturer_name": payload.name}})
    background_tasks.add_task(reconcile_lecturer_drive_access, lecturer_id)
    return public_doc(await db.users.find_one({"id": lecturer_id}, {"_id": 0}))


@api_router.post("/lecturers/{lecturer_id}/reset-password")
async def reset_lecturer_password(
    lecturer_id: str, payload: ResetPasswordInput, _: Dict[str, Any] = Depends(require_campus_admin)
):
    lecturer = await db.users.find_one({"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not lecturer:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    password = payload.password.strip() or "Dosen123!"
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password dosen minimal 6 karakter")
    await db.users.update_one(
        {"id": lecturer_id}, {"$set": {"password_hash": hash_password(password), "password_reset_at": now_iso()}}
    )
    await db.sessions.delete_many({"user_id": lecturer_id})
    return {"ok": True, "temporary_password": password}


@api_router.delete("/lecturers/{lecturer_id}")
async def delete_lecturer(
    lecturer_id: str,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    active_classes = await db.classes.count_documents({"lecturer_id": lecturer_id, "status": "active"})
    if active_classes:
        raise HTTPException(status_code=400, detail="Dosen masih memiliki kelas aktif. Arsipkan atau pindahkan kelas terlebih dahulu.")
    result = await db.users.update_one(
        {"id": lecturer_id, "role": "lecturer"}, {"$set": {"status": "deleted", "deleted_at": now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    await db.sessions.delete_many({"user_id": lecturer_id})
    background_tasks.add_task(reconcile_lecturer_drive_access, lecturer_id)
    return {"ok": True}


@api_router.get("/whatsapp/settings")
async def whatsapp_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await get_whatsapp_settings(mask=True)


@api_router.put("/whatsapp/settings")
async def update_whatsapp_settings(payload: WhatsAppSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    existing = await get_whatsapp_settings(mask=False)
    doc = payload.model_dump()
    if not doc.get("fonnte_token") and existing.get("fonnte_token"):
        doc["fonnte_token"] = existing.get("fonnte_token", "")
    if not doc.get("waha_api_key") and existing.get("waha_api_key"):
        doc["waha_api_key"] = existing.get("waha_api_key", "")
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.whatsapp_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("whatsapp_settings")
    return await get_whatsapp_settings(mask=True)


@api_router.get("/whatsapp/messages")
async def whatsapp_messages(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await db.whatsapp_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/integrations/settings")
async def get_integrations(_: Dict[str, Any] = Depends(require_campus_admin)):
    """Mengambil konfigurasi seluruh integrasi sistem (Data Sekolah, dan integrasi lain ke depan)."""
    settings = await get_integration_settings(mask=True)
    return {"ok": True, "integrations": settings.get("integrations", {})}


@api_router.put("/integrations/settings")
async def update_integrations(request: Request, user: Dict[str, Any] = Depends(require_campus_admin)):
    """Menyimpan konfigurasi integrasi sistem. Body: {"integrations": {"<nama>": {config...}}}"""
    body = await request.json()
    incoming = body.get("integrations", {}) if isinstance(body, dict) else {}
    if not isinstance(incoming, dict) or not incoming:
        raise HTTPException(status_code=400, detail="Format konfigurasi integrasi tidak valid")

    existing = await get_integration_settings(mask=False)
    integrations = existing.get("integrations", {})
    for name, cfg in incoming.items():
        if not isinstance(cfg, dict):
            continue
        current = integrations.get(name, {})
        if not isinstance(current, dict):
            current = {}
        merged = dict(current)
        for k, v in cfg.items():
            if v is None:
                continue
            if k == "api_key" and not str(v).strip():
                continue  # field API key dikosongkan di UI -> pertahankan key lama
            merged[k] = v
        integrations[name] = merged

    await db.integration_settings.update_one(
        {"id": "main"},
        {"$set": {"integrations": integrations, "updated_at": now_iso(), "updated_by": user["id"]}},
        upsert=True,
    )
    _invalidate_settings_cache("integration_settings")
    settings = await get_integration_settings(mask=True)
    return {"ok": True, "message": "Konfigurasi integrasi berhasil disimpan", "integrations": settings.get("integrations", {})}


@api_router.get("/whatsapp/waha/status")
async def waha_connection_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_whatsapp_settings(mask=False)
    base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
    session = (settings.get("waha_session") or "default").strip() or "default"
    if settings.get("provider") != "waha":
        return {"ok": False, "provider": settings.get("provider", "disabled"), "detail": "Provider WAHA belum dipilih"}
    if not base_url:
        return {"ok": False, "provider": "waha", "session": session, "detail": "WAHA Base URL belum diisi"}
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            session_doc = await fetch_waha_session_status(client, base_url, session, waha_headers(settings.get("waha_api_key", "")))
        if not session_doc:
            return {"ok": False, "provider": "waha", "base_url": base_url, "session": session, "detail": "Sesi WAHA tidak ditemukan"}
        status = str(session_doc.get("status") or "")
        return {
            "ok": status.upper() == "WORKING",
            "provider": "waha",
            "base_url": base_url,
            "session": session,
            "status": status,
            "me": session_doc.get("me"),
            "detail": "Sesi WAHA siap mengirim pesan" if status.upper() == "WORKING" else f"Sesi WAHA belum siap: {status}",
        }
    except Exception as exc:
        return {"ok": False, "provider": "waha", "base_url": base_url, "session": session, "detail": str(exc)}


@api_router.post("/whatsapp/messages/{message_id}/retry")
async def retry_whatsapp_message(message_id: str, background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    msg = await db.whatsapp_messages.find_one({"id": message_id}, {"_id": 0})
    if not msg:
        raise HTTPException(status_code=404, detail="Pesan tidak ditemukan")
    await db.whatsapp_messages.update_one({"id": message_id}, {"$set": {"status": "pending", "error": "", "response": ""}})
    background_tasks.add_task(send_whatsapp_message, message_id)
    return {"ok": True}


@api_router.get("/email/settings")
async def email_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await get_email_settings(mask=True)


@api_router.put("/email/settings")
async def update_email_settings(payload: EmailSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    existing = await get_email_settings(mask=False)
    doc = payload.model_dump()
    if not doc.get("smtp_password") and existing.get("smtp_password"):
        doc["smtp_password"] = existing.get("smtp_password", "")
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.email_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("email_settings")
    return await get_email_settings(mask=True)


@api_router.post("/email/settings/test")
async def test_email_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_email_settings(mask=False)
    if not settings.get("enabled"):
        raise HTTPException(status_code=400, detail="Email belum diaktifkan")
    test_email = settings.get("smtp_user") or settings.get("from_email", "")
    if not test_email:
        raise HTTPException(status_code=400, detail="Tidak ada email tujuan untuk tes. Isi SMTP user atau from email.")
    result = await send_email_message(
        test_email,
        "Tes Konfigurasi Email - E-Learning Dosen",
        "<h3>Email Berhasil Terkirim</h3><p>Konfigurasi SMTP pada aplikasi E-Learning Dosen sudah benar.</p>",
    )
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Gagal mengirim email tes"))
    return {"ok": True, "message": f"Email tes berhasil dikirim ke {test_email}"}


@api_router.get("/sso/settings")
async def get_oidc_admin_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    doc = await db.oidc_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    return oidc_admin_view(oidc_settings(), "admin_ui" if doc else "environment", doc)


@api_router.put("/sso/settings")
async def update_oidc_admin_settings(payload: OidcSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    discovery_url = validate_oidc_url(payload.discovery_url, "Discovery URL", required=payload.enabled)
    issuer = validate_oidc_url(payload.issuer, "Issuer", required=False)
    redirect_uri = validate_oidc_url(payload.redirect_uri, "Redirect URI", required=payload.enabled)
    frontend_url = validate_oidc_url(payload.frontend_url, "Frontend URL", required=payload.enabled)
    client_id = payload.client_id.strip()
    scopes = " ".join(payload.scopes.split())
    if payload.enabled and not payload.local_login_enabled:
        raise HTTPException(status_code=400, detail="Login lokal tidak boleh dinonaktifkan jika SSO aktif agar mencegah terkunci dari sistem.")
    if payload.enabled and not client_id:
        raise HTTPException(status_code=400, detail="Client ID wajib diisi saat SSO aktif")
    if payload.enabled and "openid" not in scopes.split():
        raise HTTPException(status_code=400, detail="Scope openid wajib disertakan")

    existing_secret = oidc_settings().get("client_secret", "")
    client_secret = "" if payload.clear_client_secret else (payload.client_secret.strip() or existing_secret)
    doc = {
        "id": "main",
        "enabled": bool(payload.enabled),
        "discovery_url": discovery_url,
        "issuer": issuer,
        "client_id": client_id,
        "client_secret_encrypted": encrypt_secret(client_secret) if client_secret else "",
        "redirect_uri": redirect_uri,
        "frontend_url": frontend_url,
        "scopes": scopes or "openid profile email roles",
        "local_login_enabled": bool(payload.local_login_enabled),
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.oidc_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    return await load_oidc_runtime_settings()


@api_router.post("/sso/settings/test")
async def test_oidc_admin_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    metadata = await oidc_discovery(force=True)
    settings = oidc_require_settings()
    return {
        "ok": True,
        "message": "Koneksi SCI-ID berhasil",
        "issuer": metadata.get("issuer", ""),
        "authorization_endpoint": metadata.get("authorization_endpoint", ""),
        "client_id": settings.get("client_id", ""),
        "client_secret_configured": bool(settings.get("client_secret")),
    }


@api_router.post("/classes/join-request")
async def request_join_class(payload: JoinRequestInput, user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Hanya mahasiswa yang dapat meminta masuk kelas")
    class_code = clean_code(payload.class_code)
    class_doc = await db.classes.find_one({"class_code": class_code, "status": "active"}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kode kelas tidak ditemukan")
    if class_doc["id"] in user.get("class_ids", []):
        return {"status": "approved", "message": "Mahasiswa sudah terdaftar di kelas ini", "class_id": class_doc["id"]}
    invited = await db.enrollment_requests.find_one(
        {"class_id": class_doc["id"], "student_id": user["id"], "status": "invited"}, {"_id": 0}
    )
    if invited:
        await db.users.update_one({"id": user["id"]}, {"$addToSet": {"class_ids": class_doc["id"]}})
        await db.classes.update_one({"id": class_doc["id"]}, {"$addToSet": {"student_ids": user["id"]}})
        await db.enrollment_requests.update_one(
            {"id": invited["id"]},
            {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": user["id"], "accepted_by_student": True}},
        )
        updated = await db.enrollment_requests.find_one({"id": invited["id"]}, {"_id": 0})
        return public_doc(updated)
    existing = await db.enrollment_requests.find_one(
        {"class_id": class_doc["id"], "student_id": user["id"], "status": "pending"}, {"_id": 0}
    )
    if existing:
        return public_doc(existing)
    request_doc = {
        "id": new_id(),
        "class_id": class_doc["id"],
        "class_name": class_doc["name"],
        "class_code": class_doc.get("class_code", ""),
        "lecturer_id": class_doc.get("lecturer_id", ""),
        "lecturer_name": class_doc.get("lecturer_name", ""),
        "student_id": user["id"],
        "student_name": user["name"],
        "student_nim": user.get("nim", ""),
        "student_email": user.get("email", ""),
        "status": "pending",
        "requested_at": now_iso(),
    }
    await db.enrollment_requests.insert_one(request_doc)
    return public_doc(request_doc)


@api_router.get("/enrollment-requests")
async def list_enrollment_requests(user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") == "admin":
        return await db.enrollment_requests.find({}, {"_id": 0}).sort("requested_at", -1).to_list(1000)
    if user.get("role") == "lecturer":
        return await db.enrollment_requests.find(
            {"class_id": {"$in": await lecturer_class_ids(user)}}, {"_id": 0}
        ).sort("requested_at", -1).to_list(1000)
    return await db.enrollment_requests.find({"student_id": user["id"]}, {"_id": 0}).sort("requested_at", -1).to_list(1000)


@api_router.post("/enrollment-requests/{request_id}/approve")
async def approve_enrollment_request(request_id: str, user: Dict[str, Any] = Depends(require_admin)):
    request_doc = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    await require_class_mutation_access(request_doc.get("class_id", ""), user)
    await db.users.update_one({"id": request_doc["student_id"]}, {"$addToSet": {"class_ids": request_doc["class_id"]}})
    await db.classes.update_one({"id": request_doc["class_id"]}, {"$addToSet": {"student_ids": request_doc["student_id"]}})
    await db.enrollment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": user["id"]}},
    )
    updated = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    return updated


@api_router.post("/enrollment-requests/{request_id}/reject")
async def reject_enrollment_request(request_id: str, user: Dict[str, Any] = Depends(require_admin)):
    request_doc = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    await require_class_mutation_access(request_doc.get("class_id", ""), user)
    await db.enrollment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "rejected", "rejected_at": now_iso(), "rejected_by": user["id"]}},
    )
    updated = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    return updated


def notification_excerpt(value: Any, fallback: str, limit: int = 120) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return fallback
    return text if len(text) <= limit else f"{text[: limit - 1].rstrip()}…"


def notification_time_is_visible(value: Any, cutoff: datetime) -> bool:
    parsed = parse_iso_datetime(str(value or ""))
    return bool(parsed and parsed >= cutoff)


async def notification_class_scope(
    user: Dict[str, Any],
    semester_id: str = "",
) -> tuple[List[str], str]:
    """Return the user's notification class scope for one academic period.

    Notifications are shown in the context of the selected semester. An
    explicit ``all`` keeps the old cross-semester view available, while a
    missing semester defaults to the active period so an older client cannot
    accidentally load every historical class on startup.
    """
    clean_semester_id = str(semester_id or "").strip()
    if not clean_semester_id:
        active_tahun_ajaran = await db.tahun_ajaran.find_one(
            {"is_active": True},
            {"_id": 0, "id": 1},
        )
        clean_semester_id = str((active_tahun_ajaran or {}).get("id") or "").strip()

    all_class_ids = await lecturer_class_ids(user)
    if clean_semester_id == "all":
        return all_class_ids, clean_semester_id
    if not clean_semester_id or not all_class_ids:
        return [], clean_semester_id

    selected_tahun_ajaran = await db.tahun_ajaran.find_one(
        {"id": clean_semester_id},
        {"_id": 0},
    )
    if not selected_tahun_ajaran:
        raise HTTPException(status_code=404, detail="Tahun ajaran yang dipilih tidak ditemukan")

    scoped_classes = await db.classes.find(
        {"id": {"$in": all_class_ids}},
        {
            "_id": 0,
            "id": 1,
            "tahun_ajaran_id": 1,
            "academic_year_id": 1,
            "academic_year": 1,
            "tahun_ajaran": 1,
            "semester": 1,
        },
    ).to_list(5000)
    scoped_class_ids = [
        str(class_doc.get("id") or "")
        for class_doc in scoped_classes
        if class_doc.get("id")
        and class_matches_tahun_ajaran(class_doc, selected_tahun_ajaran, clean_semester_id)
    ]
    return scoped_class_ids, clean_semester_id


async def notification_center_payload(
    user: Dict[str, Any],
    limit: int = 30,
    semester_id: str = "",
) -> Dict[str, Any]:
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=NOTIFICATION_LOOKBACK_DAYS)
    cutoff_iso = cutoff.isoformat()
    class_ids, scoped_semester_id = await notification_class_scope(user, semester_id)
    material_query: Dict[str, Any] = {"class_id": {"$in": class_ids}}
    if user.get("role") == "student":
        material_query["is_active"] = True
    materials = await db.materials.find(
        material_query,
        {
            "_id": 0,
            "id": 1,
            "class_id": 1,
            "title": 1,
            "meeting": 1,
            "course_name": 1,
            "class_name": 1,
        },
    ).to_list(5000)
    material_map = {item["id"]: item for item in materials}
    material_ids = list(material_map)
    events: List[Dict[str, Any]] = []

    if material_ids:
        comments = await db.comments.find(
            {
                "material_id": {"$in": material_ids},
                "author_id": {"$ne": user["id"]},
                "created_at": {"$gte": cutoff_iso},
            },
            {"_id": 0},
        ).sort("created_at", -1).to_list(1000)
        for comment in comments:
            material = material_map.get(comment.get("material_id", ""), {})
            occurred_at = str(comment.get("created_at") or "")
            is_reply = bool(comment.get("parent_id"))
            events.append(
                notification_event(
                    kind="discussion",
                    source_id=str(comment.get("id") or ""),
                    occurred_at=occurred_at,
                    title="Balasan diskusi baru" if is_reply else "Komentar diskusi baru",
                    message=(
                        f"{comment.get('author_name') or 'Pengguna'} pada "
                        f"{material.get('title') or 'materi'}: "
                        f"{notification_excerpt(comment.get('content'), 'Mengirim lampiran diskusi')}"
                    ),
                    actor_name=str(comment.get("author_name") or ""),
                    target={
                        "page": "courses" if user.get("role") == "student" else "materials",
                        "object_type": "comment",
                        "object_id": str(comment.get("id") or ""),
                        "comment_id": str(comment.get("id") or ""),
                        "material_id": str(comment.get("material_id") or ""),
                        "class_id": str(material.get("class_id") or ""),
                    },
                )
            )

    if user.get("role") in {"admin", "lecturer"}:
        submissions = await db.submissions.find(
            {
                "class_id": {"$in": class_ids},
                "submitted_at": {"$gte": cutoff_iso},
            },
            {"_id": 0},
        ).sort("submitted_at", -1).to_list(1000)
        for submission in submissions:
            occurred_at = str(submission.get("submitted_at") or "")
            events.append(
                notification_event(
                    kind="submission",
                    source_id=str(submission.get("id") or ""),
                    occurred_at=occurred_at,
                    title="Submission tugas baru",
                    message=(
                        f"{submission.get('student_name') or 'Mahasiswa'} mengumpulkan "
                        f"{submission.get('assignment_title') or 'tugas'}"
                    ),
                    actor_name=str(submission.get("student_name") or ""),
                    target={
                        "page": "grading",
                        "object_type": "submission",
                        "object_id": str(submission.get("id") or ""),
                        "submission_id": str(submission.get("id") or ""),
                        "assignment_id": str(submission.get("assignment_id") or ""),
                        "class_id": str(submission.get("class_id") or ""),
                    },
                )
            )
        enrollments = await db.enrollment_requests.find(
            {
                "class_id": {"$in": class_ids},
                "status": "pending",
                "requested_at": {"$gte": cutoff_iso},
            },
            {"_id": 0},
        ).sort("requested_at", -1).to_list(1000)
        for enrollment in enrollments:
            occurred_at = str(enrollment.get("requested_at") or "")
            events.append(
                notification_event(
                    kind="enrollment",
                    source_id=str(enrollment.get("id") or ""),
                    occurred_at=occurred_at,
                    title="Permintaan masuk kelas",
                    message=(
                        f"{enrollment.get('student_name') or 'Mahasiswa'} meminta masuk "
                        f"{enrollment.get('class_name') or 'kelas'}"
                    ),
                    actor_name=str(enrollment.get("student_name") or ""),
                    target={
                        "page": "students",
                        "object_type": "enrollment",
                        "object_id": str(enrollment.get("id") or ""),
                        "request_id": str(enrollment.get("id") or ""),
                        "class_id": str(enrollment.get("class_id") or ""),
                    },
                )
            )
    else:
        assignments = await db.assignments.find(
            {
                "class_id": {"$in": class_ids},
                "is_active": True,
            },
            {"_id": 0},
        ).to_list(2000)
        for assignment in assignments:
            occurred_at = str(
                assignment.get("published_notification_sent_at")
                or assignment.get("published_at")
                or assignment.get("created_at")
                or ""
            )
            if (
                not assignment_is_published(assignment, now)
                or not notification_time_is_visible(occurred_at, cutoff)
            ):
                continue
            events.append(
                notification_event(
                    kind="assignment",
                    source_id=str(assignment.get("id") or ""),
                    occurred_at=occurred_at,
                    title="Tugas baru tersedia",
                    message=(
                        f"{assignment.get('title') or 'Tugas'} · "
                        f"{assignment.get('course_name') or assignment.get('class_name') or 'Kelas'}"
                    ),
                    actor_name=str(assignment.get("lecturer_name") or ""),
                    target={
                        "page": "assignments",
                        "object_type": "assignment",
                        "object_id": str(assignment.get("id") or ""),
                        "assignment_id": str(assignment.get("id") or ""),
                        "class_id": str(assignment.get("class_id") or ""),
                    },
                )
            )
        submissions = await db.submissions.find(
            {"student_id": user["id"]},
            {"_id": 0},
        ).to_list(2000)
        for submission in submissions:
            graded_at = str(submission.get("graded_at") or "")
            if notification_time_is_visible(graded_at, cutoff):
                events.append(
                    notification_event(
                        kind="grade",
                        source_id=str(submission.get("id") or ""),
                        occurred_at=graded_at,
                        title="Nilai dan feedback tersedia",
                        message=(
                            f"{submission.get('assignment_title') or 'Tugas'} mendapat nilai "
                            f"{submission.get('grade') if submission.get('grade') is not None else '-'}"
                        ),
                        actor_name=str(submission.get("lecturer_name") or ""),
                        target={
                            "page": "assignments",
                            "object_type": "assignment",
                            "object_id": str(submission.get("assignment_id") or ""),
                            "assignment_id": str(submission.get("assignment_id") or ""),
                            "submission_id": str(submission.get("id") or ""),
                            "class_id": str(submission.get("class_id") or ""),
                        },
                    )
                )
            revision_at = str(submission.get("revision_requested_at") or "")
            if notification_time_is_visible(revision_at, cutoff):
                events.append(
                    notification_event(
                        kind="revision",
                        source_id=str(submission.get("id") or ""),
                        occurred_at=revision_at,
                        title="Revisi tugas diminta",
                        message=(
                            f"{submission.get('assignment_title') or 'Tugas'}: "
                            f"{notification_excerpt(submission.get('revision_note'), 'Periksa arahan revisi dosen')}"
                        ),
                        actor_name=str(submission.get("lecturer_name") or ""),
                        target={
                            "page": "assignments",
                            "object_type": "assignment",
                            "object_id": str(submission.get("assignment_id") or ""),
                            "assignment_id": str(submission.get("assignment_id") or ""),
                            "submission_id": str(submission.get("id") or ""),
                            "class_id": str(submission.get("class_id") or ""),
                        },
                    )
                )

    event_ids = list({item["id"] for item in events})
    read_receipts = (
        await db.notification_reads.find(
            {
                "user_id": user["id"],
                "notification_id": {"$in": event_ids},
            },
            {"_id": 0},
        ).to_list(max(100, len(event_ids)))
        if event_ids
        else []
    )
    payload = finalize_notifications(events, read_receipts, limit)
    payload["lookback_days"] = NOTIFICATION_LOOKBACK_DAYS
    payload["semester_id"] = scoped_semester_id
    return payload


@api_router.get("/notifications")
async def list_notifications(
    limit: int = 30,
    semester_id: str = "",
    user: Dict[str, Any] = Depends(get_current_user),
):
    return await notification_center_payload(user, limit, semester_id)


@api_router.post("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    clean_id = str(notification_id or "").strip().lower()
    if not re.fullmatch(r"[a-f0-9]{32}", clean_id):
        raise HTTPException(status_code=400, detail="ID notifikasi tidak valid")
    read_at = now_iso()
    await db.notification_reads.update_one(
        {"user_id": user["id"], "notification_id": clean_id},
        {
            "$set": {
                "id": f"{user['id']}:{clean_id}",
                "user_id": user["id"],
                "notification_id": clean_id,
                "read_at": read_at,
            }
        },
        upsert=True,
    )
    cutoff = (
        datetime.now(timezone.utc) - timedelta(days=NOTIFICATION_READ_RETENTION_DAYS)
    ).isoformat()
    return {"ok": True, "notification_id": clean_id, "read_at": read_at}


@api_router.post("/notifications/read-all")
async def mark_all_notifications_read(
    semester_id: str = "",
    user: Dict[str, Any] = Depends(get_current_user),
):
    payload = await notification_center_payload(user, limit=500, semester_id=semester_id)
    items = payload.get("items", [])
    unread_items = [item for item in items if not item.get("read")]
    read_at = now_iso()

    if unread_items:
        for item in unread_items:
            clean_id = str(item.get("id") or "").strip().lower()
            if clean_id:
                await db.notification_reads.update_one(
                    {"user_id": user["id"], "notification_id": clean_id},
                    {
                        "$set": {
                            "id": f"{user['id']}:{clean_id}",
                            "user_id": user["id"],
                            "notification_id": clean_id,
                            "read_at": read_at,
                        }
                    },
                    upsert=True,
                )
    return {"ok": True, "count": len(unread_items), "read_at": read_at}


async def user_activity_dashboard_payload(days: int = 14) -> Dict[str, Any]:
    safe_days = max(7, min(int(days or 14), 90))
    now = datetime.now(timezone.utc)
    cutoff = (now - timedelta(days=safe_days + 1)).isoformat()
    logs = await db.user_activity_logs.find(
        {"created_at": {"$gte": cutoff}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(50000)
    payload = aggregate_user_activity(
        logs,
        safe_days,
        now,
        USER_ACTIVITY_TIMEZONE,
    )
    payload["retention_days"] = USER_ACTIVITY_RETENTION_DAYS
    return payload


@api_router.get("/user-activity")
async def get_user_activity(
    days: int = 14,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    return await user_activity_dashboard_payload(days)


def class_matches_tahun_ajaran(
    class_doc: Dict[str, Any],
    tahun_ajaran: Dict[str, Any],
    semester_id: str,
) -> bool:
    """Samakan cakupan kelas dengan selector Tahun Ajaran di frontend."""
    if str(class_doc.get("tahun_ajaran_id") or "") == semester_id:
        return True
    if str(class_doc.get("academic_year_id") or "") == semester_id:
        return True
    target_year = str(tahun_ajaran.get("tahun") or tahun_ajaran.get("academic_year") or "").strip()
    target_semester = str(tahun_ajaran.get("semester") or "").strip().lower()
    class_year = str(class_doc.get("academic_year") or class_doc.get("tahun_ajaran") or "").strip()
    class_semester = str(class_doc.get("semester") or "").strip().lower()
    return bool(target_year and target_semester and class_year == target_year and class_semester == target_semester)


@api_router.get("/dashboard")
async def dashboard(
    include_activity: bool = True,
    semester_id: str = "",
    user: Dict[str, Any] = Depends(get_current_user),
):
    all_class_ids = await lecturer_class_ids(user)
    scoped_classes = await db.classes.find(
        {"id": {"$in": all_class_ids}}, {"_id": 0}
    ).to_list(5000)
    selected_tahun_ajaran: Dict[str, Any] = {}
    clean_semester_id = str(semester_id or "").strip()
    if clean_semester_id and clean_semester_id != "all":
        selected_tahun_ajaran = await db.tahun_ajaran.find_one(
            {"id": clean_semester_id}, {"_id": 0}
        ) or {}
        if not selected_tahun_ajaran:
            raise HTTPException(status_code=404, detail="Tahun ajaran yang dipilih tidak ditemukan")
        scoped_classes = [
            class_doc
            for class_doc in scoped_classes
            if class_matches_tahun_ajaran(class_doc, selected_tahun_ajaran, clean_semester_id)
        ]

    class_ids = [item["id"] for item in scoped_classes if item.get("id")]
    active_classes = [item for item in scoped_classes if item.get("status") == "active"]
    active_courses_count = len({item.get("course_id") for item in active_classes if item.get("course_id")})
    active_classes_count = len(active_classes)
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}, "is_active": True}, {"_id": 0}).to_list(500)
    submissions = await db.submissions.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(2000)
    material_ids = [item["id"] for item in await db.materials.find({"class_id": {"$in": class_ids}}, {"_id": 0, "id": 1}).to_list(5000)]
    comments = await db.comments.find({"material_id": {"$in": material_ids}}, {"_id": 0}).sort("created_at", -1).to_list(5)
    student_ids = list({student_id for item in active_classes for student_id in item.get("student_ids", [])})
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    now = datetime.now(timezone.utc)
    soon = now + timedelta(days=3)
    published_assignments = [assignment for assignment in assignments if assignment_is_published(assignment, now)]
    near_deadline = 0
    for assignment in published_assignments:
        try:
            deadline = datetime.fromisoformat(assignment["deadline"].replace("Z", "+00:00"))
            if now <= deadline <= soon:
                near_deadline += 1
        except Exception:
            continue
    submitted_pairs = {(s.get("assignment_id"), s.get("student_id")) for s in submissions}
    missing = 0
    classes_for_missing = active_classes
    for assignment in published_assignments:
        class_doc = next((c for c in classes_for_missing if c["id"] == assignment.get("class_id")), None)
        if class_doc:
            for student_id in class_doc.get("student_ids", []):
                if (assignment["id"], student_id) not in submitted_pairs:
                    missing += 1
    ungraded = len([s for s in submissions if s.get("status") in ["Sudah Submit", "Terlambat", "Direvisi"]])
    avg_grade_values = [s.get("grade", 0) for s in submissions if isinstance(s.get("grade"), (int, float))]
    avg_grade = round(sum(avg_grade_values) / len(avg_grade_values), 1) if avg_grade_values else 0
    progress_map = await calculate_student_progress_many([s["id"] for s in students], class_ids)
    risk_high = sum(1 for p in progress_map.values() if p.get("risk_label") == "Risiko Tinggi")
    student_progress = []
    for student in students:
        student_doc = public_doc(student.copy()) or {}
        student_doc["progress"] = progress_map.get(student.get("id", ""), {})
        student_progress.append(student_doc)
    user_activity = (
        await user_activity_dashboard_payload(14)
        if include_activity and is_campus_admin(user)
        else None
    )
    return {
        "summary": {
            "semester_id": clean_semester_id,
            "semester_label": (
                f"{selected_tahun_ajaran.get('tahun', '')} {selected_tahun_ajaran.get('semester', '')}".strip()
                if selected_tahun_ajaran
                else "Semua semester"
            ),
            "active_courses": active_courses_count,
            "active_classes": active_classes_count,
            "active_assignments": len(assignments),
            "near_deadline": near_deadline,
            "missing_submissions": missing,
            "ungraded_submissions": ungraded,
            "latest_comments": len(comments),
            "avg_grade": avg_grade,
            "high_risk_students": risk_high,
            **await storage_status_summary(),
        },
        "latest_comments": comments,
        "student_progress": student_progress,
        "user_activity": user_activity,
    }


CLEAN_DATA_MODULES = {
    "academic": {
        "label": "Akademik, mata kuliah & kelas",
        "description": "Menghapus prodi, mata kuliah, kelas, materi, tugas, submission, diskusi, enrollment, dan file akademik terkait.",
    },
    "students": {
        "label": "Mahasiswa",
        "description": "Menghapus akun mahasiswa, sesi mahasiswa, enrollment, submission, komentar, chat, dan file upload mahasiswa.",
    },
    "materials": {
        "label": "Materi & diskusi",
        "description": "Menghapus materi, file materi, komentar diskusi, dan lampiran komentar. Tugas yang terhubung akan dilepas dari materi.",
    },
    "assignments": {
        "label": "Tugas & submission",
        "description": "Menghapus tugas, submission, lampiran soal, file submission, dan reminder tugas.",
    },
    "grades": {
        "label": "Nilai",
        "description": "Mengosongkan nilai, feedback, histori nilai pada submission, serta menghapus range predikat nilai.",
    },
    "chat": {
        "label": "Chat",
        "description": "Menghapus semua pesan chat dan foto yang dikirim lewat chat.",
    },
    "notifications": {
        "label": "Notifikasi & reset password",
        "description": "Menghapus reminder in-app, histori WhatsApp, email, OTP, dan request reset password.",
    },
    "all": {
        "label": "Semua data percobaan",
        "description": "Membersihkan semua modul percobaan tanpa menghapus akun dosen, settings aplikasi, dan konfigurasi Drive.",
    },
}


def remove_local_stored_file(local_path: str) -> bool:
    if not local_path:
        return False
    path = Path(local_path).resolve()
    storage_roots = (STORAGE_ROOT.resolve(), PMB_STORAGE_ROOT.resolve())
    storage_root = next(
        (root for root in storage_roots if root == path or root in path.parents),
        None,
    )
    if storage_root is None:
        raise ValueError(f"Path file di luar storage root: {path}")
    existed = path.exists()
    path.unlink(missing_ok=True)
    parent = path.parent
    while parent != storage_root and storage_root in parent.parents:
        try:
            parent.rmdir()
        except OSError:
            break
        parent = parent.parent
    return existed


async def queue_drive_file_deletion(file_doc: Dict[str, Any], error: str = "") -> None:
    drive_file_id = str(file_doc.get("drive_file_id") or "").strip()
    if not drive_file_id:
        return
    existing = await db.drive_delete_queue.find_one({"id": drive_file_id}, {"_id": 0}) or {}
    await db.drive_delete_queue.update_one(
        {"id": drive_file_id},
        {
            "$set": {
                "id": drive_file_id,
                "drive_file_id": drive_file_id,
                "file_name": file_doc.get("file_name", ""),
                "source_file_id": file_doc.get("id", ""),
                "status": "pending",
                "last_error": error[:500],
                "next_retry_at": "",
                "updated_at": now_iso(),
                "created_at": existing.get("created_at", now_iso()),
                "attempt_date": existing.get("attempt_date", ""),
                "attempts_today": int(existing.get("attempts_today", 0) or 0),
                "attempts_total": int(existing.get("attempts_total", 0) or 0),
            },
        },
        upsert=True,
    )


async def delete_stored_files(query: Dict[str, Any]) -> int:
    files = await db.stored_files.find(
        query,
        {
            "_id": 0,
            "id": 1,
            "file_name": 1,
            "local_path": 1,
            "drive_file_id": 1,
        },
    ).to_list(5000)
    drive_files = [item for item in files if item.get("drive_file_id")]
    drive_settings = await get_google_drive_settings(mask=False) if drive_files else {}
    for file_doc in files:
        drive_file_id = str(file_doc.get("drive_file_id") or "").strip()
        if drive_file_id:
            if google_drive_upload_enabled(drive_settings):
                try:
                    await asyncio.to_thread(
                        delete_drive_file_sync,
                        drive_file_id,
                        drive_settings,
                    )
                    await db.drive_delete_queue.delete_one({"id": drive_file_id})
                except Exception as exc:
                    logger.warning(
                        "Penghapusan file Drive %s ditunda: %s",
                        drive_file_id,
                        exc,
                    )
                    await queue_drive_file_deletion(file_doc, google_drive_error_message(exc))
            else:
                await queue_drive_file_deletion(
                    file_doc,
                    "Google Drive belum aktif saat file lokal dihapus.",
                )
        local_path = str(file_doc.get("local_path") or "")
        if local_path:
            try:
                remove_local_stored_file(local_path)
            except Exception as exc:
                logger.warning("Gagal menghapus file lokal %s: %s", local_path, exc)
    if files:
        await db.stored_files.delete_many({"id": {"$in": [item["id"] for item in files]}})
    return len(files)


async def process_drive_delete_queue(limit: int = 100) -> int:
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        return 0
    now = datetime.now(timezone.utc)
    today = sync_attempt_day(now, STORAGE_POLICY_TIMEZONE)
    processed = 0
    items = await db.drive_delete_queue.find(
        {"status": {"$in": ["pending", "failed"]}},
        {"_id": 0},
    ).sort("created_at", 1).to_list(limit)
    for item in items:
        if not retry_is_due(str(item.get("next_retry_at") or ""), now):
            continue
        attempts_today = (
            int(item.get("attempts_today", 0) or 0)
            if item.get("attempt_date") == today
            else 0
        )
        if attempts_today >= DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY:
            await db.drive_delete_queue.update_one(
                {"id": item["id"]},
                {
                    "$set": {
                        "next_retry_at": next_drive_retry_at(
                            attempts_today,
                            now,
                            STORAGE_POLICY_TIMEZONE,
                        ),
                        "updated_at": now_iso(),
                    }
                },
            )
            continue
        attempts_today += 1
        try:
            await asyncio.to_thread(
                delete_drive_file_sync,
                item.get("drive_file_id", ""),
                settings,
            )
            await db.drive_delete_queue.delete_one({"id": item["id"]})
            processed += 1
        except Exception as exc:
            await db.drive_delete_queue.update_one(
                {"id": item["id"]},
                {
                    "$set": {
                        "status": "failed",
                        "attempt_date": today,
                        "attempts_today": attempts_today,
                        "attempts_total": int(item.get("attempts_total", 0) or 0) + 1,
                        "last_error": google_drive_error_message(exc),
                        "last_attempt_at": now_iso(),
                        "next_retry_at": next_drive_retry_at(
                            attempts_today,
                            now,
                            STORAGE_POLICY_TIMEZONE,
                        ),
                        "updated_at": now_iso(),
                    }
                },
            )
    return processed


async def retry_due_drive_syncs(limit: int = 100) -> int:
    now = datetime.now(timezone.utc)
    files = await db.stored_files.find(
        {"drive_sync_status": {"$in": ["pending", "failed", "not_configured"]}},
        {"_id": 0, "id": 1, "drive_next_retry_at": 1},
    ).sort("uploaded_at", 1).to_list(limit)
    attempted = 0
    for file_doc in files:
        if not retry_is_due(str(file_doc.get("drive_next_retry_at") or ""), now):
            continue
        await sync_stored_file_to_drive(file_doc["id"])
        attempted += 1
    return attempted


async def purge_expired_synced_local_files(limit: int = 500) -> int:
    now = datetime.now(timezone.utc)
    files = await db.stored_files.find(
        {
            "drive_sync_status": "synced",
            "drive_file_id": {"$ne": ""},
            "local_path": {"$ne": ""},
        },
        {
            "_id": 0,
            "id": 1,
            "local_path": 1,
            "drive_uploaded_at": 1,
            "updated_at": 1,
            "uploaded_at": 1,
        },
    ).sort("drive_uploaded_at", 1).to_list(limit)
    purged = 0
    for file_doc in files:
        synced_at = (
            file_doc.get("drive_uploaded_at")
            or file_doc.get("updated_at")
            or file_doc.get("uploaded_at")
            or ""
        )
        if not local_copy_is_expired(
            str(synced_at),
            now,
            DRIVE_LOCAL_RETENTION_DAYS,
        ):
            continue
        try:
            remove_local_stored_file(str(file_doc.get("local_path") or ""))
        except Exception as exc:
            logger.warning(
                "Retensi gagal menghapus file lokal %s: %s",
                file_doc.get("id", ""),
                exc,
            )
            continue
        await db.stored_files.update_one(
            {"id": file_doc["id"]},
            {
                "$set": {
                    "local_path": "",
                    "local_available": False,
                    "local_purged_at": now_iso(),
                    "local_retention_days": DRIVE_LOCAL_RETENTION_DAYS,
                    "updated_at": now_iso(),
                }
            },
        )
        await refresh_embedded_file_references(file_doc["id"])
        purged += 1
    return purged


async def run_storage_maintenance_once() -> Dict[str, int]:
    return {
        "drive_deletions": await process_drive_delete_queue(),
        "drive_sync_attempts": await retry_due_drive_syncs(),
        "local_files_purged": await purge_expired_synced_local_files(),
    }


async def storage_maintenance_scheduler() -> None:
    while True:
        try:
            await run_storage_maintenance_once()
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.exception("Scheduler pemeliharaan file bermasalah: %s", exc)
        await asyncio.sleep(STORAGE_MAINTENANCE_INTERVAL_SECONDS)


async def clean_data_module_counts() -> Dict[str, int]:
    student_ids = [
        item["id"]
        for item in await db.users.find({"role": "student"}, {"_id": 0, "id": 1}).to_list(5000)
    ]
    return {
        "academic": (
            await db.programs.count_documents({})
            + await db.courses.count_documents({})
            + await db.classes.count_documents({})
            + await db.enrollment_requests.count_documents({})
        ),
        "students": (
            await db.users.count_documents({"role": "student"})
            + await db.sessions.count_documents({"user_id": {"$in": student_ids}} if student_ids else {"_never": "_never"})
        ),
        "materials": (
            await db.materials.count_documents({})
            + await db.comments.count_documents({})
            + await db.stored_files.count_documents({"record_type": {"$in": ["material_attachment", "comment_attachment"]}})
        ),
        "assignments": (
            await db.assignments.count_documents({})
            + await db.submissions.count_documents({})
            + await db.stored_files.count_documents({"record_type": {"$in": ["assignment_attachment", "submission"]}})
        ),
        "grades": (
            await db.submissions.count_documents({"grade": {"$ne": None}})
            + await db.grade_predicates.count_documents({})
        ),
        "chat": (
            await db.chat_messages.count_documents({})
            + await db.stored_files.count_documents({"record_type": "chat_image"})
        ),
        "notifications": (
            await db.reminder_logs.count_documents({})
            + await db.whatsapp_messages.count_documents({})
            + await db.email_messages.count_documents({})
            + await db.password_reset_requests.count_documents({})
            + await db.password_reset_otps.count_documents({})
            + await db.notification_reads.count_documents({})
        ),
    }


def deleted_count(result: Any) -> int:
    return int(getattr(result, "deleted_count", 0) or 0)


def modified_count(result: Any) -> int:
    return int(getattr(result, "modified_count", 0) or 0)


def _clean_period_year(document: Dict[str, Any]) -> str:
    raw_values = [
        document.get("academic_year"),
        document.get("tahun_ajaran"),
        document.get("name"),
        document.get("nama"),
        document.get("tahun"),
        document.get("year"),
    ]
    for value in raw_values:
        match = re.search(r"\d{4}\s*/\s*\d{4}", str(value or "").strip())
        if match:
            return re.sub(r"\s+", "", match.group(0))
    raw = next((str(value).strip() for value in raw_values if str(value or "").strip()), "")
    match = re.search(r"\d{4}", raw)
    if match:
        start_year = int(match.group(0))
        code = str(document.get("code") or document.get("kode") or "").strip()
        if re.fullmatch(r"\d{5}", code):
            return f"{start_year}/{start_year + 1}" if code[-1] == "1" else f"{start_year - 1}/{start_year}"
        if re.fullmatch(r"\d{4}", raw):
            semester = str(document.get("semester") or "").strip().lower()
            return f"{start_year - 1}/{start_year}" if semester == "genap" else f"{start_year}/{start_year + 1}"
    return raw


def _clean_period_semester(document: Dict[str, Any]) -> str:
    semester = str(document.get("semester") or "").strip()
    if semester:
        return semester
    raw = str(document.get("name") or "").strip().lower()
    if "genap" in raw:
        return "Genap"
    if "ganjil" in raw:
        return "Ganjil"
    code = str(document.get("code") or document.get("kode") or "").strip()
    if re.fullmatch(r"\d{5}", code):
        return "Ganjil" if code[-1] == "1" else "Genap"
    return ""


def _clean_period_identity(document: Dict[str, Any]) -> tuple[str, str]:
    return (
        _clean_period_year(document),
        _clean_period_semester(document).lower(),
    )


def _clean_period_ids(values: List[Any]) -> List[str]:
    return list(dict.fromkeys(str(value).strip() for value in values if str(value or "").strip()))


def _clean_ids_query(field: str, values: List[str]) -> Dict[str, Any]:
    return {field: {"$in": values or ["__clean_data_no_match__"]}}


def _clean_scope_or(terms: List[tuple[str, List[str]]]) -> Dict[str, Any]:
    clauses = [
        {field: {"$in": values}}
        for field, values in terms
        if values
    ]
    return {"$or": clauses} if clauses else {"_clean_data_no_match__": "1"}


async def resolve_clean_data_period(selector: str) -> Dict[str, Any]:
    """Resolve selector Tahun Ajaran and all IDs used by legacy period data."""
    clean_selector = str(selector or "").strip()
    if not clean_selector or clean_selector == "all":
        raise HTTPException(status_code=400, detail="Semester harus dipilih")

    selected: Optional[Dict[str, Any]] = None
    for collection, fields in (
        (db.tahun_ajaran, ["id", "kode"]),
        (db.academic_periods, ["id", "code"]),
    ):
        query = {"$or": [{field: clean_selector} for field in fields]}
        selected = await collection.find_one(query, {"_id": 0})
        if selected:
            break
    if not selected:
        raise HTTPException(status_code=404, detail="Semester yang dipilih tidak ditemukan")

    year, semester_key = _clean_period_identity(selected)
    display_semester = str(selected.get("semester") or "").strip()
    if not display_semester:
        display_semester = "Genap" if semester_key == "genap" else "Ganjil"
    if not year or not semester_key:
        raise HTTPException(status_code=422, detail="Data semester belum memiliki tahun dan jenis semester yang valid")

    period_ids = _clean_period_ids(
        [selected.get("id"), selected.get("kode"), selected.get("code"), clean_selector]
    )
    for collection in (db.tahun_ajaran, db.academic_periods):
        candidates = await collection.find({}, {"_id": 0}).to_list(5000)
        for candidate in candidates:
            if _clean_period_identity(candidate) != (year, semester_key):
                continue
            period_ids.extend(
                [candidate.get("id"), candidate.get("kode"), candidate.get("code")]
            )
    period_ids = _clean_period_ids(period_ids)
    normalized = {
        **selected,
        "tahun": year,
        "academic_year": year,
        "semester": display_semester,
    }
    return {
        "selector": clean_selector,
        "period": normalized,
        "period_ids": period_ids,
        "label": f"{year} {display_semester}".strip(),
    }


async def collect_clean_data_semester_scope(period_info: Dict[str, Any]) -> Dict[str, Any]:
    """Collect every semester-scoped ID before a destructive operation starts."""
    period = period_info["period"]
    selector = period_info["selector"]
    period_ids = period_info["period_ids"]
    all_classes = await db.classes.find({}, {"_id": 0}).to_list(10000)
    class_docs = [
        item
        for item in all_classes
        if class_matches_tahun_ajaran(item, period, selector)
    ]
    class_ids = _clean_period_ids([item.get("id") for item in class_docs])

    materials = await db.materials.find(
        _clean_ids_query("class_id", class_ids), {"_id": 0, "id": 1}
    ).to_list(10000)
    assignments = await db.assignments.find(
        _clean_ids_query("class_id", class_ids), {"_id": 0, "id": 1}
    ).to_list(10000)
    submissions = await db.submissions.find(
        _clean_ids_query("class_id", class_ids), {"_id": 0, "id": 1}
    ).to_list(20000)
    material_ids = _clean_period_ids([item.get("id") for item in materials])
    assignment_ids = _clean_period_ids([item.get("id") for item in assignments])
    submission_ids = _clean_period_ids([item.get("id") for item in submissions])

    comments = await db.comments.find(
        _clean_ids_query("material_id", material_ids), {"_id": 0, "id": 1}
    ).to_list(20000)
    comment_ids = _clean_period_ids([item.get("id") for item in comments])

    bills = await db.tuition_bills.find(
        _clean_ids_query("academic_period_id", period_ids), {"_id": 0, "id": 1}
    ).to_list(20000)
    bill_ids = _clean_period_ids([item.get("id") for item in bills])

    sk_documents = await db.sk_mengajar.find({}, {"_id": 0, "id": 1, "tahun_ajaran": 1, "semester": 1}).to_list(5000)
    sk_ids = _clean_period_ids([
        item.get("id")
        for item in sk_documents
        if _clean_period_identity(item) == (period["tahun"], period["semester"].lower())
    ])
    sk_validation_documents = await db.sk_validations.find(
        {}, {"_id": 0, "id": 1, "tahun_ajaran": 1, "semester": 1}
    ).to_list(10000)
    sk_validation_ids = _clean_period_ids([
        item.get("id")
        for item in sk_validation_documents
        if _clean_period_identity(item) == (period["tahun"], period["semester"].lower())
    ])

    stored_file_query = _clean_scope_or([
        ("class_id", class_ids),
        ("rps_class_id", class_ids),
        ("material_id", material_ids),
        ("discussion_material_id", material_ids),
        ("assignment_id", assignment_ids),
        ("submission_id", submission_ids),
        ("comment_id", comment_ids),
    ])
    reminder_query = _clean_scope_or([
        ("assignment_id", assignment_ids),
        ("class_id", class_ids),
    ])
    return {
        **period_info,
        "class_docs": class_docs,
        "class_ids": class_ids,
        "material_ids": material_ids,
        "assignment_ids": assignment_ids,
        "submission_ids": submission_ids,
        "comment_ids": comment_ids,
        "bill_ids": bill_ids,
        "sk_ids": sk_ids,
        "sk_validation_ids": sk_validation_ids,
        "stored_file_query": stored_file_query,
        "reminder_query": reminder_query,
    }


async def clean_data_semester_counts(scope: Dict[str, Any]) -> Dict[str, int]:
    period_query = _clean_ids_query("academic_period_id", scope["period_ids"])
    bill_query = _clean_ids_query("id", scope["bill_ids"])
    return {
        "kelas": len(scope["class_ids"]),
        "materi": await db.materials.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "tugas": await db.assignments.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "submission": await db.submissions.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "diskusi": await db.comments.count_documents(_clean_ids_query("material_id", scope["material_ids"])),
        "presensi": await db.attendance_sessions.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "enrollment": await db.enrollment_requests.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "rps": await db.rps.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "predikat_nilai": await db.grade_predicates.count_documents(_clean_ids_query("class_id", scope["class_ids"])),
        "krs": await db.krs.count_documents(period_query),
        "khs": await db.khs.count_documents(period_query),
        "tagihan": await db.tuition_bills.count_documents(period_query),
        "pembayaran": await db.tuition_payments.count_documents(bill_query),
        "kalender": await db.academic_calendar_events.count_documents(_clean_ids_query("academic_year_id", scope["period_ids"])),
        "deadline": await db.academic_deadline_settings.count_documents(_clean_ids_query("academic_year_id", scope["period_ids"])),
        "sk_mengajar": len(scope["sk_ids"]),
        "file": await db.stored_files.count_documents(scope["stored_file_query"]),
    }


async def execute_clean_data_semester(scope: Dict[str, Any]) -> Dict[str, Any]:
    """Delete semester transactions while preserving master data and accounts."""
    affected: Dict[str, int] = {}
    class_ids = scope["class_ids"]
    period_ids = scope["period_ids"]
    bill_ids = scope["bill_ids"]

    affected["file"] = await delete_stored_files(scope["stored_file_query"])
    affected["diskusi"] = deleted_count(await db.comments.delete_many(_clean_ids_query("material_id", scope["material_ids"])))
    affected["submission"] = deleted_count(await db.submissions.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["tugas"] = deleted_count(await db.assignments.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["materi"] = deleted_count(await db.materials.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["presensi"] = deleted_count(await db.attendance_sessions.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["enrollment"] = deleted_count(await db.enrollment_requests.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["rps"] = deleted_count(await db.rps.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["predikat_nilai"] = deleted_count(await db.grade_predicates.delete_many(_clean_ids_query("class_id", class_ids)))
    affected["reminder"] = deleted_count(await db.reminder_logs.delete_many(scope["reminder_query"]))
    affected["kelas"] = deleted_count(await db.classes.delete_many(_clean_ids_query("id", class_ids)))
    if class_ids:
        affected["tautan_kelas_mahasiswa"] = modified_count(
            await db.users.update_many(
                {"class_ids": {"$in": class_ids}},
                {"$pull": {"class_ids": {"$in": class_ids}}},
            )
        )
    else:
        affected["tautan_kelas_mahasiswa"] = 0

    affected["krs"] = deleted_count(await db.krs.delete_many(_clean_ids_query("academic_period_id", period_ids)))
    affected["khs"] = deleted_count(await db.khs.delete_many(_clean_ids_query("academic_period_id", period_ids)))
    affected["pembayaran"] = deleted_count(await db.tuition_payments.delete_many(_clean_ids_query("bill_id", bill_ids)))
    affected["tagihan"] = deleted_count(await db.tuition_bills.delete_many(_clean_ids_query("academic_period_id", period_ids)))
    affected["kalender"] = deleted_count(await db.academic_calendar_events.delete_many(_clean_ids_query("academic_year_id", period_ids)))
    affected["deadline"] = deleted_count(await db.academic_deadline_settings.delete_many(_clean_ids_query("academic_year_id", period_ids)))
    affected["sk_mengajar"] = deleted_count(await db.sk_mengajar.delete_many(_clean_ids_query("id", scope["sk_ids"])))
    affected["validasi_sk"] = deleted_count(await db.sk_validations.delete_many(_clean_ids_query("id", scope["sk_validation_ids"])))
    return {
        "module": "semester",
        "label": scope["label"],
        "period": scope["period"],
        "affected": affected,
    }


async def execute_clean_data_module(module: str) -> Dict[str, Any]:
    if module not in CLEAN_DATA_MODULES:
        raise HTTPException(status_code=404, detail="Modul clean data tidak ditemukan")
    affected: Dict[str, int] = {}

    if module == "all":
        for child in ["chat", "materials", "assignments", "students", "academic", "grades", "notifications"]:
            child_result = await execute_clean_data_module(child)
            for key, value in child_result.get("affected", {}).items():
                affected[f"{child}.{key}"] = value
        affected["stored_files_remaining_deleted"] = await delete_stored_files({})
        return {"module": module, "label": CLEAN_DATA_MODULES[module]["label"], "affected": affected}

    if module == "chat":
        affected["chat_messages"] = deleted_count(await db.chat_messages.delete_many({}))
        affected["stored_files"] = await delete_stored_files({"record_type": "chat_image"})

    elif module == "materials":
        affected["materials"] = deleted_count(await db.materials.delete_many({}))
        affected["comments"] = deleted_count(await db.comments.delete_many({}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["material_attachment", "comment_attachment"]}})
        affected["assignments_unlinked"] = modified_count(await db.assignments.update_many({}, {"$set": {"material_id": ""}}))

    elif module == "assignments":
        affected["assignments"] = deleted_count(await db.assignments.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({"assignment_id": {"$ne": ""}}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["assignment_attachment", "submission"]}})

    elif module == "grades":
        affected["submissions_reset"] = modified_count(
            await db.submissions.update_many(
                {},
                {
                    "$set": {
                        "grade": None,
                        "feedback": "",
                        "grade_predicate": "",
                        "grade_history": [],
                        "review_status": "submitted",
                    },
                    "$unset": {"graded_at": "", "graded_by": "", "rubric_scores": ""},
                },
            )
        )
        affected["grade_predicates"] = deleted_count(await db.grade_predicates.delete_many({}))

    elif module == "notifications":
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({}))
        affected["whatsapp_messages"] = deleted_count(await db.whatsapp_messages.delete_many({}))
        affected["email_messages"] = deleted_count(await db.email_messages.delete_many({}))
        affected["password_reset_requests"] = deleted_count(await db.password_reset_requests.delete_many({}))
        affected["password_reset_otps"] = deleted_count(await db.password_reset_otps.delete_many({}))
        affected["notification_reads"] = deleted_count(await db.notification_reads.delete_many({}))

    elif module == "students":
        student_ids = [
            item["id"]
            for item in await db.users.find({"role": "student"}, {"_id": 0, "id": 1}).to_list(5000)
        ]
        student_query = {"$in": student_ids} if student_ids else {"$in": ["__none__"]}
        affected["student_sessions"] = deleted_count(await db.sessions.delete_many({"user_id": student_query}))
        affected["notification_reads"] = deleted_count(await db.notification_reads.delete_many({"user_id": student_query}))
        affected["student_users"] = deleted_count(await db.users.delete_many({"role": "student"}))
        affected["class_memberships_reset"] = modified_count(await db.classes.update_many({}, {"$set": {"student_ids": []}}))
        affected["enrollment_requests"] = deleted_count(await db.enrollment_requests.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({"student_id": student_query}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({"student_id": student_query}))
        affected["comments"] = deleted_count(await db.comments.delete_many({"author_id": student_query}))
        affected["chat_messages"] = deleted_count(await db.chat_messages.delete_many({"participant_ids": {"$in": student_ids}}))
        affected["stored_files"] = await delete_stored_files({"uploaded_by": student_query})

    elif module == "academic":
        affected["programs"] = deleted_count(await db.programs.delete_many({}))
        affected["courses"] = deleted_count(await db.courses.delete_many({}))
        affected["classes"] = deleted_count(await db.classes.delete_many({}))
        affected["enrollment_requests"] = deleted_count(await db.enrollment_requests.delete_many({}))
        affected["materials"] = deleted_count(await db.materials.delete_many({}))
        affected["comments"] = deleted_count(await db.comments.delete_many({}))
        affected["assignments"] = deleted_count(await db.assignments.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({}))
        affected["student_class_links"] = modified_count(await db.users.update_many({"role": "student"}, {"$set": {"class_ids": []}}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["assignment_attachment", "submission", "material_attachment", "comment_attachment"]}})

    return {"module": module, "label": CLEAN_DATA_MODULES[module]["label"], "affected": affected}


@api_router.get("/clean-data/summary")
async def clean_data_summary(_: Dict[str, Any] = Depends(require_campus_admin)):
    counts = await clean_data_module_counts()
    return [
        {"key": key, **meta, "count": counts.get(key, sum(counts.values()) if key == "all" else 0)}
        for key, meta in CLEAN_DATA_MODULES.items()
    ]


@api_router.get("/clean-data/semester-summary")
async def clean_data_semester_summary(
    semester_id: str = "",
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    period_info = await resolve_clean_data_period(semester_id)
    scope = await collect_clean_data_semester_scope(period_info)
    return {
        "semester": period_info["period"],
        "label": period_info["label"],
        "counts": await clean_data_semester_counts(scope),
        "protected": [
            "Akun dan profil mahasiswa",
            "Program studi, mata kuliah, kurikulum, serta dosen",
            "Master tahun ajaran/semester yang dipilih",
            "Komponen dan skema biaya keuangan",
        ],
    }


@api_router.post("/clean-data/semester/{semester_id}")
async def clean_data_semester(
    semester_id: str,
    payload: CleanDataInput,
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    period_info = await resolve_clean_data_period(semester_id)
    if payload.confirmation != "HAPUS SEMESTER":
        raise HTTPException(status_code=400, detail="Konfirmasi penghapusan semester tidak valid")
    if payload.confirmation_label.strip() != period_info["label"]:
        raise HTTPException(
            status_code=400,
            detail=f"Ketik nama semester persis: {period_info['label']}",
        )
    scope = await collect_clean_data_semester_scope(period_info)
    result = await execute_clean_data_semester(scope)
    result.update(
        {
            "cleaned_at": now_iso(),
            "cleaned_by": user.get("name") or user.get("username") or user.get("id", ""),
        }
    )
    return result


@api_router.post("/clean-data/{module}")
async def clean_data(module: str, payload: CleanDataInput, _: Dict[str, Any] = Depends(require_campus_admin)):
    if payload.confirmation != "HAPUS":
        raise HTTPException(status_code=400, detail="Konfirmasi clean data tidak valid")
    result = await execute_clean_data_module(module)
    result["cleaned_at"] = now_iso()
    return result


@api_router.get("/storage/status")
async def storage_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    return {
        **await storage_status_summary(),
        "policy": {
            "local_retention_days_after_drive_sync": DRIVE_LOCAL_RETENTION_DAYS,
            "max_drive_sync_attempts_per_day": DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY,
            "timezone": str(STORAGE_POLICY_TIMEZONE),
        },
        "files": {
            "total": await db.stored_files.count_documents({}),
            "google_drive": await db.stored_files.count_documents({"storage_provider": "google_drive"}),
            "server_local": await db.stored_files.count_documents({"storage_provider": "server_local"}),
            "drive_pending": await db.stored_files.count_documents({"drive_sync_status": "pending"}),
            "drive_failed": await db.stored_files.count_documents({"drive_sync_status": "failed"}),
            "local_purged": await db.stored_files.count_documents(
                {"drive_sync_status": "synced", "local_available": False}
            ),
            "drive_delete_pending": await db.drive_delete_queue.count_documents({}),
        },
    }


async def drive_sync_overview(limit: int = 50) -> Dict[str, Any]:
    query = {
        "record_type": {
            "$in": [
                "submission",
                "assignment_attachment",
                "material_attachment",
                "physical_document",
            ]
        }
    }
    summary = {
        "total": await db.stored_files.count_documents(query),
        "pending": await db.stored_files.count_documents({**query, "drive_sync_status": "pending"}),
        "synced": await db.stored_files.count_documents({**query, "drive_sync_status": "synced"}),
        "failed": await db.stored_files.count_documents({**query, "drive_sync_status": "failed"}),
        "not_configured": await db.stored_files.count_documents({**query, "drive_sync_status": "not_configured"}),
        "local_purged": await db.stored_files.count_documents(
            {**query, "drive_sync_status": "synced", "local_available": False}
        ),
        "drive_delete_pending": await db.drive_delete_queue.count_documents({}),
        "local_retention_days": DRIVE_LOCAL_RETENTION_DAYS,
        "max_attempts_per_day": DRIVE_SYNC_MAX_ATTEMPTS_PER_DAY,
    }
    files = await db.stored_files.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(limit)
    submission_ids = [item.get("submission_id") for item in files if item.get("submission_id")]
    assignment_ids = [item.get("assignment_id") for item in files if item.get("assignment_id")]
    student_ids = [item.get("student_id") for item in files if item.get("student_id")]
    submissions = await db.submissions.find({"id": {"$in": submission_ids}}, {"_id": 0}).to_list(limit) if submission_ids else []
    assignments = await db.assignments.find({"id": {"$in": assignment_ids}}, {"_id": 0}).to_list(limit) if assignment_ids else []
    students = await db.users.find({"id": {"$in": student_ids}}, {"_id": 0, "id": 1, "name": 1, "nim": 1}).to_list(limit) if student_ids else []
    submissions_by_id = {item["id"]: item for item in submissions}
    assignments_by_id = {item["id"]: item for item in assignments}
    students_by_id = {item["id"]: item for item in students}
    items = []
    for item in files:
        submission = submissions_by_id.get(item.get("submission_id", ""), {})
        assignment = assignments_by_id.get(item.get("assignment_id", ""), {})
        student = students_by_id.get(item.get("student_id", ""), {})
        items.append(
            {
                "id": item.get("id", ""),
                "file_id": item.get("file_id", item.get("id", "")),
                "file_name": item.get("file_name", ""),
                "size": item.get("size", 0),
                "record_type": item.get("record_type", ""),
                "upload_status": item.get("upload_status", ""),
                "drive_sync_status": item.get("drive_sync_status", ""),
                "drive_error": item.get("drive_error", ""),
                "drive_file_url": item.get("drive_file_url", ""),
                "drive_sync_attempt_date": item.get("drive_sync_attempt_date", ""),
                "drive_sync_attempts_today": item.get("drive_sync_attempts_today", 0),
                "drive_sync_attempts_total": item.get("drive_sync_attempts_total", 0),
                "drive_last_attempt_at": item.get("drive_last_attempt_at", ""),
                "drive_next_retry_at": item.get("drive_next_retry_at", ""),
                "local_available": item.get("local_available", bool(item.get("local_path"))),
                "local_purged_at": item.get("local_purged_at", ""),
                "uploaded_at": item.get("uploaded_at", ""),
                "updated_at": item.get("updated_at", ""),
                "assignment_id": item.get("assignment_id", ""),
                "assignment_title": assignment.get("title") or submission.get("assignment_title", ""),
                "course_name": assignment.get("course_name", ""),
                "class_name": assignment.get("class_name", ""),
                "submission_id": item.get("submission_id", ""),
                "student_name": item.get("student_name") or student.get("name") or submission.get("student_name", ""),
                "student_nim": item.get("student_nim") or student.get("nim") or submission.get("student_nim", ""),
                "document_type": item.get("document_type", ""),
                "document_label": item.get("document_label") or PHYSICAL_DOCUMENT_TYPES.get(item.get("document_type", ""), ""),
                "angkatan": item.get("angkatan", ""),
            }
        )
    return {"summary": summary, "items": items}


@api_router.get("/migration/old-siap/summary")
async def get_old_siap_summary(_: Dict[str, Any] = Depends(require_campus_admin)):
    file_path = Path("/Users/syahrulanwar/Documents/Project Web/OLD-SIAP/siap_siakad.json")
    if not file_path.exists():
        return {"exists": False, "message": "File siap_siakad.json tidak ditemukan"}
    
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            raw_data = json.load(f)
        tables = {}
        if isinstance(raw_data, list):
            for item in raw_data:
                if isinstance(item, dict) and item.get("type") == "table":
                    tables[item.get("name")] = len(item.get("data", []))
        
        entities = {
            "fakultas": tables.get("fakultas", 0),
            "prodi": tables.get("prodi", 0),
            "tahun_ajaran": tables.get("tahun", 0),
            "pegawai_dosen": tables.get("pegawai", 0),
            "mahasiswa": tables.get("mhsw", 0),
            "mata_kuliah": tables.get("mk", 0),
            "kelas_jadwal": tables.get("jadwal", 0),
            "krs": tables.get("krs", 0),
            "khs": tables.get("khs", 0),
        }
        return {"exists": True, "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 2), "entities": entities}
    except Exception as e:
        return {"exists": False, "message": str(e)}


migration_progress_state = {
    "status": "idle",
    "step": "",
    "progress_percent": 0,
    "logs": [],
    "result": None,
}

@api_router.get("/migration/old-siap/status")
async def get_old_siap_migration_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    return migration_progress_state

@api_router.post("/migration/old-siap/run")
async def run_old_siap_migration(background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    if migration_progress_state["status"] == "running":
        raise HTTPException(status_code=400, detail="Migrasi sedang berjalan...")

    async def execute_migration_task():
        migration_progress_state["status"] = "running"
        migration_progress_state["progress_percent"] = 5
        migration_progress_state["step"] = "Membaca file data OLD-SIAP..."
        migration_progress_state["logs"] = ["Memulai proses migrasi data OLD-SIAP..."]
        
        try:
            file_path = Path("/Users/syahrulanwar/Documents/Project Web/OLD-SIAP/siap_siakad.json")
            with open(file_path, "r", encoding="utf-8") as f:
                raw_data = json.load(f)
            
            tables = {}
            if isinstance(raw_data, list):
                for item in raw_data:
                    if isinstance(item, dict) and item.get("type") == "table":
                        tables[item.get("name")] = item.get("data", [])
            
            raw_fakultas = tables.get("fakultas", [])
            raw_prodi = tables.get("prodi", [])
            raw_tahun = tables.get("tahun", [])
            raw_pegawai = tables.get("pegawai", [])
            raw_mhsw = tables.get("mhsw", [])
            raw_mk = tables.get("mk", [])
            raw_jadwal = tables.get("jadwal", [])
            raw_kelas = tables.get("kelas", [])
            raw_krs = tables.get("krs", [])

            # 1. Fakultas
            migration_progress_state["step"] = "Memigrasikan Fakultas..."
            migration_progress_state["progress_percent"] = 15
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_fakultas)} fakultas...")
            f_count = 0
            for f in raw_fakultas:
                fid = f.get("FakultasID") or "default-fakultas"
                nama = f.get("Nama") or f.get("NamaIns") or "Fakultas Utama"
                kode = f.get("KodePTI") or f.get("KodeID") or "FT"
                ex = await db.fakultas.find_one({"id": fid}, {"_id": 0})
                doc = {"id": fid, "kode": kode, "nama": nama, "status": "active", "created_at": now_iso()}
                if not ex: await db.fakultas.insert_one(doc)
                else: await db.fakultas.update_one({"id": fid}, {"$set": doc})
                f_count += 1

            # 2. Prodi
            migration_progress_state["step"] = "Memigrasikan Program Studi..."
            migration_progress_state["progress_percent"] = 30
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_prodi)} program studi...")
            prodi_map = {}
            p_count = 0
            for p in raw_prodi:
                pid = p.get("ProdiID")
                nama = p.get("Nama", "")
                kode = pid
                jenjang = p.get("NamaJenjang") or "S1"
                prodi_map[pid] = {"id": pid, "nama": nama, "kode": kode}
                ex = await db.programs.find_one({"id": pid}, {"_id": 0})
                doc = {
                    "id": pid, "code": kode, "kode": kode, "name": nama, "nama": nama,
                    "description": f"Program Studi {nama} ({jenjang})", "status": "active",
                    "jenjang": jenjang, "akreditasi": p.get("Akreditasi") or "B",
                    "kaprodi": p.get("Pejabat") or "", "created_at": now_iso()
                }
                if not ex: await db.programs.insert_one(doc)
                else: await db.programs.update_one({"id": pid}, {"$set": doc})
                p_count += 1

            # 2b. Rombel (Kelas per Angkatan)
            migration_progress_state["step"] = "Memigrasikan Rombel (Kelas per Angkatan)..."
            migration_progress_state["progress_percent"] = 40
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_kelas)} rombel...")
            rombel_map = {}
            mhsw_kelas_map = {}
            for m in raw_mhsw:
                kid = str(m.get("KelasID") or "").strip()
                mhs_id = str(m.get("MhswID") or "").strip().upper()
                if kid and kid != "0" and mhs_id:
                    mhsw_kelas_map.setdefault(kid, set()).add(mhs_id)
            rb_count = 0
            for k in raw_kelas:
                kid = str(k.get("KelasID") or "").strip()
                if not kid: continue
                nama_full = str(k.get("Nama") or "").strip()
                prodi_id = str(k.get("ProdiID") or "").strip()
                prodi_obj = prodi_map.get(prodi_id, {})
                tahun_id = str(k.get("TahunID") or "").strip()
                angkatan = tahun_id[:4] if len(tahun_id) >= 4 else ""
                rombel_letter = nama_full
                if "-" in nama_full:
                    rombel_letter = nama_full.split("-")[-1].strip() or rombel_letter
                rombel_id = f"RLM-{kid}"
                rombel_map[kid] = {"id": rombel_id, "nama": rombel_letter, "prodi_id": prodi_id}
                student_ids = sorted(mhsw_kelas_map.get(kid, set()))
                rb_doc = {
                    "id": rombel_id, "kode": nama_full, "nama": rombel_letter,
                    "prodi_id": prodi_id, "prodi_name": prodi_obj.get("nama", ""),
                    "angkatan": angkatan, "student_ids": student_ids,
                    "status": "active", "created_at": now_iso()
                }
                ex = await db.rombel.find_one({"id": rombel_id}, {"_id": 0})
                if not ex: await db.rombel.insert_one(rb_doc)
                else: await db.rombel.update_one({"id": rombel_id}, {"$set": rb_doc})
                rb_count += 1

            # 3. Dosen / Pegawai
            migration_progress_state["step"] = "Memigrasikan Dosen & Pegawai..."
            migration_progress_state["progress_percent"] = 45
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_pegawai)} dosen & pegawai...")
            default_password_hash = hash_password("Dosen123!")
            default_mhs_password_hash = hash_password("Mahasiswa1231!")
            dosen_map = {}
            d_count = 0

            used_emails = set()
            existing_users = await db.users.find({}, {"_id": 0, "id": 1, "email": 1}).to_list(10000)
            existing_email_to_id = {u.get("email", "").lower(): u.get("id") for u in existing_users if u.get("email")}

            def make_unique_email(raw_email, default_prefix, user_id):
                email = (raw_email or f"{default_prefix}.{user_id.lower()}@demo.id").strip().lower()
                if "@" not in email:
                    email = f"{email}@demo.id"
                owner_id = existing_email_to_id.get(email)
                if (owner_id and owner_id != user_id) or (email in used_emails and owner_id != user_id):
                    parts = email.split("@", 1)
                    email = f"{parts[0]}+{user_id.lower()}@{parts[1]}"
                used_emails.add(email)
                existing_email_to_id[email] = user_id
                return email

            for p in raw_pegawai:
                login = str(p.get("Login") or "").strip()
                if not login: continue
                name = str(p.get("Nama") or login).strip()
                email = make_unique_email(p.get("Email"), "dosen", login)
                wa = str(p.get("WA") or p.get("Handphone") or "").strip()
                prodi_id = str(p.get("ProdiID") or "").strip()
                level_id = str(p.get("LevelID") or "")
                levels = [x.strip() for x in level_id.split(",") if x.strip()]
                role = "admin" if "1" in levels else "lecturer"
                dosen_map[login] = {"id": login, "name": name, "email": email}
                ex = await db.users.find_one({"id": login}, {"_id": 0})
                doc = {
                    "id": login, "role": role, "username": login, "employee_id": str(p.get("NIDN") or login).strip(),
                    "name": name, "email": email, "whatsapp": wa, "password_hash": default_password_hash,
                    "status": "active" if p.get("NA") != "Y" else "inactive", "prodi_id": prodi_id,
                    "created_at": now_iso(), "last_login_at": ""
                }
                if not ex: await db.users.insert_one(doc)
                else: await db.users.update_one({"id": login}, {"$set": doc})
                d_count += 1

            # 4. Mahasiswa
            migration_progress_state["step"] = "Memigrasikan Mahasiswa..."
            migration_progress_state["progress_percent"] = 60
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_mhsw)} mahasiswa...")
            m_count = 0
            for m in raw_mhsw:
                nim = str(m.get("MhswID") or m.get("Login") or "").strip().upper()
                if not nim: continue
                name = str(m.get("Nama") or nim).strip()
                email = make_unique_email(m.get("Email"), nim.lower(), nim)
                wa = str(m.get("WA") or m.get("Handphone") or "").strip()
                prodi_id = str(m.get("ProdiID") or "").strip()
                prodi_obj = prodi_map.get(prodi_id, {})
                tahun_id = str(m.get("TahunID") or "")
                angkatan = tahun_id[:4] if len(tahun_id) >= 4 else "2024"
                pa_id = str(m.get("PenasehatAkademik") or "").strip()
                pa_name = dosen_map.get(pa_id, {}).get("name", "")

                rombel_kid = str(m.get("KelasID") or "").strip()
                rombel_id = rombel_map.get(rombel_kid, {}).get("id", "")

                ex = await db.users.find_one({"nim": nim, "role": "student"}, {"_id": 0})
                doc = {
                    "id": nim, "role": "student", "username": nim.lower(), "nim": nim, "name": name,
                    "email": email, "whatsapp": wa, "password_hash": default_mhs_password_hash,
                    "status": "active" if m.get("StatusMhswID") == "A" else "inactive",
                    "class_ids": [], "rombel_id": rombel_id, "prodi_id": prodi_id, "prodi_name": prodi_obj.get("nama", ""),
                    "prodi_kode": prodi_obj.get("kode", ""), "angkatan": angkatan,
                    "dosen_wali_id": pa_id, "dosen_wali_name": pa_name,
                    "created_at": now_iso(), "last_login_at": ""
                }
                if not ex: await db.users.insert_one(doc)
                else: await db.users.update_one({"id": ex["id"]}, {"$set": doc})
                m_count += 1

            # 4b. Master Kurikulum
            migration_progress_state["step"] = "Memigrasikan Master Kurikulum..."
            migration_progress_state["progress_percent"] = 65
            raw_kurikulum = tables.get("kurikulum", [])
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_kurikulum)} master kurikulum...")
            kurikulum_map = {}
            kur_count = 0
            for k in raw_kurikulum:
                kid = str(k.get("KurikulumID") or "").strip()
                if not kid: continue
                kode = str(k.get("KurikulumKode") or k.get("SKKurikulum") or f"KUR-{kid}").strip()
                nama = str(k.get("Nama") or f"Kurikulum {kid}").strip()
                p_id = str(k.get("ProdiID") or "").strip()
                prodi_obj = prodi_map.get(p_id, {})
                tahun_id = str(k.get("TahunID") or "")
                tahun_mulai = tahun_id[:4] if len(tahun_id) >= 4 else "2023"

                kurikulum_map[kid] = {"id": kid, "kode": kode, "nama": nama, "prodi_id": p_id}
                ex = await db.kurikulum.find_one({"id": kid}, {"_id": 0})
                doc = {
                    "id": kid, "kode": kode, "nama": nama, "prodi_id": p_id,
                    "prodi_nama": prodi_obj.get("nama", ""), "tahun_mulai": tahun_mulai,
                    "status": "active" if k.get("NA") != "Y" else "inactive",
                    "total_sks_wajib": int(k.get("SKSWAjib") or 0),
                    "total_sks_pilihan": int(k.get("SKSPilihan") or 0),
                    "created_at": now_iso()
                }
                if not ex: await db.kurikulum.insert_one(doc)
                else: await db.kurikulum.update_one({"id": kid}, {"$set": doc})
                kur_count += 1

            # 5. Mata Kuliah (Courses)
            migration_progress_state["step"] = "Memigrasikan Mata Kuliah (Courses)..."
            migration_progress_state["progress_percent"] = 75
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_mk)} mata kuliah...")
            mk_count = 0
            for mk in raw_mk:
                mk_id = str(mk.get("MKID") or "").strip()
                nama = str(mk.get("Nama") or "").strip()
                if not nama: continue
                mk_kode = str(mk.get("MKKode") or mk_id).strip()
                sks = int(mk.get("SKS") or 3)
                sks_tm = int(mk.get("SKSTatapMuka") or sks)
                sks_pr = int(mk.get("SKSPraktikum") or 0) + int(mk.get("SKSPraktekLap") or 0)
                p_id = str(mk.get("ProdiID") or "").strip()
                prodi_obj = prodi_map.get(p_id, {})

                kid = str(mk.get("KurikulumID") or "").strip()
                kur_obj = kurikulum_map.get(kid, {})
                kur_kode = kur_obj.get("kode", "")

                try: sem_paket = int(mk.get("Sesi") or 1)
                except ValueError: sem_paket = 1

                wajib = str(mk.get("Wajib") or "Y").upper()
                sifat = "Wajib" if wajib == "Y" else "Pilihan"

                ex = await db.courses.find_one({"id": mk_id}, {"_id": 0})
                doc = {
                    "id": mk_id, "kurikulum_id": kid, "kurikulum_kode": kur_kode,
                    "program_id": p_id, "prodi_id": p_id, "program_name": prodi_obj.get("nama", ""),
                    "code": mk_kode, "kode": mk_kode, "name": nama, "nama": nama,
                    "credits": sks, "sks": sks, "total_sks": sks,
                    "sks_teori": sks_tm, "sks_praktikum": sks_pr,
                    "semester_paket": sem_paket, "semester": sem_paket,
                    "sifat": sifat, "description": str(mk.get("Deskripsi") or "").strip(),
                    "status": "active" if mk.get("NA") != "Y" else "inactive", "created_at": now_iso()
                }
                if not ex: await db.courses.insert_one(doc)
                else: await db.courses.update_one({"id": mk_id}, {"$set": doc})
                mk_count += 1

            # 6. Kelas & Jadwal
            migration_progress_state["step"] = "Memigrasikan Kelas Kuliah..."
            migration_progress_state["progress_percent"] = 90
            migration_progress_state["logs"].append(f"Memigrasikan {len(raw_jadwal)} kelas kuliah...")

            active_semesters = {str(t.get("TahunID") or "").strip() for t in raw_tahun if t.get("NA") == "N"}
            active_tahun_id = max(
                (str(j.get("TahunID") or "").strip() for j in raw_jadwal if str(j.get("TahunID") or "").strip()),
                default="20252",
            ) if raw_jadwal else "20252"

            class_students_map = {}
            for krs in raw_krs:
                jid = str(krs.get("JadwalID") or "").strip()
                mhs_id = str(krs.get("MhswID") or "").strip()
                if jid and mhs_id and jid != "0":
                    class_students_map.setdefault(jid, set()).add(mhs_id)

            j_count = 0
            for j in raw_jadwal:
                jid = str(j.get("JadwalID") or "").strip()
                if not jid: continue
                kelas_kid = str(j.get("KelasID") or "").strip()
                rombel_obj = rombel_map.get(kelas_kid, {})
                rombel_id = rombel_obj.get("id", "")
                if rombel_obj.get("nama"):
                    c_name = f"Kelas {rombel_obj['nama']}"
                else:
                    c_name = str(j.get("NamaKelas_old") or "").strip()
                    if not c_name or c_name == "0" or c_name == "01": c_name = f"Kelas {jid}"
                tahun_id = str(j.get("TahunID") or "20241")
                ay = f"{tahun_id[:4]}/{int(tahun_id[:4])+1}" if len(tahun_id) >= 4 else "2024/2025"
                sem = "Ganjil" if tahun_id.endswith("1") else "Genap"
                d_id = str(j.get("DosenID") or "").strip()
                p_id = str(j.get("ProdiID") or "").strip()
                st_ids = list(class_students_map.get(jid, set()))

                class_status = "active" if tahun_id == active_tahun_id else "ended"

                ex = await db.classes.find_one({"id": jid}, {"_id": 0})
                doc = {
                    "id": jid, "academic_year": ay, "semester": sem, "course_id": str(j.get("MKID") or "").strip(),
                    "course_name": str(j.get("Nama") or "").strip(), "name": c_name,
                    "schedule": f"Hari {j.get('HariID','')} {j.get('JamMulai','')}-{j.get('JamSelesai','')}".strip(),
                    "class_code": f"KLS{jid.zfill(4)}", "lecturer_id": d_id,
                    "lecturer_name": dosen_map.get(d_id, {}).get("name", ""),
                    "status": class_status, "rombel_id": rombel_id,
                    "student_ids": st_ids, "program_id": p_id,
                    "program_name": prodi_map.get(p_id, {}).get("nama", ""), "created_at": now_iso()
                }
                if not ex: await db.classes.insert_one(doc)
                else: await db.classes.update_one({"id": jid}, {"$set": doc})
                j_count += 1
                for st in st_ids:
                    await db.users.update_one({"nim": st}, {"$addToSet": {"class_ids": jid}})

            # 7. Tahun Ajaran
            migration_progress_state["step"] = "Memigrasikan Tahun Ajaran..."
            migration_progress_state["progress_percent"] = 92
            ta_count = 0
            for t in raw_tahun:
                tid = str(t.get("TahunID") or "").strip()
                if not tid: continue
                nama = str(t.get("Nama") or "").strip()
                ex = await db.tahun_ajaran.find_one({"id": tid}, {"_id": 0})
                is_active = (tid == active_tahun_id)
                doc = {
                    "id": tid, "tahun": tid[:4] if len(tid) >= 4 else "2024",
                    "semester": "Ganjil" if tid.endswith("1") else "Genap",
                    "nama": nama or tid, "is_active": is_active,
                    "status": "active" if is_active else "closed",
                    "activated_at": now_iso() if is_active else "",
                    "created_at": now_iso()
                }
                if not ex: await db.tahun_ajaran.insert_one(doc)
                else: await db.tahun_ajaran.update_one({"id": tid}, {"$set": doc})
                ta_count += 1

            # 8. Predikat Nilai
            migration_progress_state["step"] = "Memigrasikan Predikat Nilai..."
            migration_progress_state["progress_percent"] = 95
            raw_nilai = tables.get("nilai", [])
            predicates_list = []
            seen_labels = set()
            for n in raw_nilai:
                label = str(n.get("Nama") or "").strip()
                if not label or label in seen_labels: continue
                seen_labels.add(label)
                try:
                    predicates_list.append({
                        "label": label, "min_score": float(n.get("NilaiMin") or 0),
                        "max_score": float(n.get("NilaiMax") or 100), "gpa": float(n.get("Bobot") or 0)
                    })
                except ValueError: pass
            if predicates_list:
                predicates_list.sort(key=lambda x: x["min_score"], reverse=True)
                await db.grade_predicates.update_one({"class_id": ""}, {"$set": {"class_id": "", "predicates": predicates_list, "updated_at": now_iso()}}, upsert=True)

            # 9. KHS & KRS (Nilai Mahasiswa)
            migration_progress_state["step"] = "Memigrasikan KHS & Nilai KRS..."
            migration_progress_state["progress_percent"] = 98
            raw_khs = tables.get("khs", [])
            krs_by_student_tahun = {}
            krs_grouped = {}
            for krs in raw_krs:
                mhs_id = str(krs.get("MhswID") or "").strip()
                t_id = str(krs.get("TahunID") or "").strip()
                if mhs_id and t_id:
                    key = f"{mhs_id}_{t_id}"
                    item = {
                        "course_code": str(krs.get("MKKode") or "").strip(),
                        "course_name": str(krs.get("Nama") or "").strip(),
                        "sks": int(krs.get("SKS") or 0),
                        "class_id": str(krs.get("JadwalID") or "").strip(),
                        "grade_letter": str(krs.get("GradeNilai") or "").strip(),
                        "grade_point": float(krs.get("BobotNilai") or 0.0),
                        "score": float(krs.get("NilaiAkhir") or 0.0),
                    }
                    krs_by_student_tahun.setdefault(key, []).append(item)
                    krs_grouped.setdefault(key, {
                        "id": f"krs_{mhs_id}_{t_id}", "student_id": mhs_id,
                        "academic_period_id": t_id, "status": "approved",
                        "created_at": now_iso(), "courses": []
                    })["courses"].append(item)

            krs_count = 0
            for key, krs_doc in krs_grouped.items():
                await db.krs.update_one(
                    {"student_id": krs_doc["student_id"], "academic_period_id": krs_doc["academic_period_id"]},
                    {"$set": krs_doc}, upsert=True
                )
                krs_count += 1

            khs_count = 0
            for khs in raw_khs:
                khs_id = str(khs.get("KHSID") or "").strip()
                mhs_id = str(khs.get("MhswID") or "").strip()
                t_id = str(khs.get("TahunID") or "").strip()
                if not khs_id or not mhs_id or not t_id: continue
                try:
                    ips, ipk = float(khs.get("IPS") or 0.0), float(khs.get("IPK") or 0.0)
                    sks_sem, sks_kum = int(khs.get("SKS") or 0), int(khs.get("TotalSKS") or 0)
                except ValueError: ips, ipk, sks_sem, sks_kum = 0.0, 0.0, 0, 0

                khs_doc = {
                    "id": khs_id, "student_id": mhs_id, "academic_period_id": t_id,
                    "period_name": parse_tahun_label(t_id), "ips": ips, "ipk": ipk,
                    "total_sks_semester": sks_sem, "total_sks_kumulatif": sks_kum,
                    "grades": krs_by_student_tahun.get(f"{mhs_id}_{t_id}", []),
                    "updated_at": now_iso()
                }
                await db.khs.update_one({"student_id": mhs_id, "academic_period_id": t_id}, {"$set": khs_doc}, upsert=True)
                khs_count += 1

            # 10. Submissions & Assignments (Nilai Evaluasi & Rekap)
            migration_progress_state["step"] = "Memigrasikan Submissions & Rekap Nilai..."
            migration_progress_state["progress_percent"] = 99
            classes_with_grades = {str(krs.get("JadwalID")).strip() for krs in raw_krs if krs.get("JadwalID") and str(krs.get("JadwalID")).strip() != "0"}
            assignment_map = {}
            asgn_count = 0
            sub_count = 0

            for jid in classes_with_grades:
                class_doc = await db.classes.find_one({"id": jid}, {"_id": 0})
                if not class_doc: continue
                asgn_id = f"asgn_old_{jid}"
                assignment_map[jid] = asgn_id
                asgn_doc = {
                    "id": asgn_id, "class_id": jid,
                    "course_name": class_doc.get("course_name", "Mata Kuliah"),
                    "class_name": class_doc.get("name", "Kelas"),
                    "title": "Evaluasi Akhir Semester (Migrasi OLD-SIAP)",
                    "description": "Nilai akumulasi akhir semester dari database OLD-SIAP",
                    "deadline": now_iso(), "is_active": True, "assessment_category": "uas", "created_at": now_iso()
                }
                await db.assignments.update_one({"id": asgn_id}, {"$set": asgn_doc}, upsert=True)
                asgn_count += 1

            for krs in raw_krs:
                jid = str(krs.get("JadwalID") or "").strip()
                mhs_id = str(krs.get("MhswID") or "").strip()
                if not jid or jid == "0" or not mhs_id: continue
                asgn_id = assignment_map.get(jid)
                if not asgn_id: continue
                try:
                    score = float(krs.get("NilaiAkhir") or 0.0)
                    grade_letter = str(krs.get("GradeNilai") or "").strip()
                except ValueError: score, grade_letter = 0.0, ""

                sub_id = f"sub_old_{jid}_{mhs_id}"
                sub_doc = {
                    "id": sub_id, "assignment_id": asgn_id, "student_id": mhs_id, "class_id": jid,
                    "status": "Sudah Dinilai", "submitted_at": now_iso(), "grade": score,
                    "grade_predicate": grade_letter,
                    "feedback": f"Migrasi SIAP - Grade: {grade_letter}, Bobot: {krs.get('BobotNilai', 0)}",
                    "created_at": now_iso()
                }
                await db.submissions.update_one({"id": sub_id}, {"$set": sub_doc}, upsert=True)
                sub_count += 1

            migration_progress_state["status"] = "success"
            migration_progress_state["progress_percent"] = 100
            migration_progress_state["step"] = "Migrasi selesai"
            migration_progress_state["logs"].append("Seluruh data dari OLD-SIAP (termasuk nilai & KHS) berhasil dimigrasikan!")
            migration_progress_state["result"] = {
                "fakultas": f_count, "prodi": p_count, "kurikulum": kur_count, "dosen": d_count,
                "mahasiswa": m_count, "mk": mk_count, "kelas": j_count, "tahun_ajaran": ta_count,
                "predikat_nilai": len(predicates_list), "khs": khs_count, "krs": krs_count,
                "evaluasi_kelas": asgn_count, "submissions_nilai": sub_count,
            }
        except Exception as err:
            migration_progress_state["status"] = "failed"
            migration_progress_state["step"] = "Terjadi kesalahan"
            migration_progress_state["logs"].append(f"ERROR: {str(err)}")

    background_tasks.add_task(execute_migration_task)
    return {"message": "Proses migrasi dimulai di background"}


@api_router.get("/database-backups")
async def list_database_backups(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_database_backup_settings()
    backups = await db.database_backups.find({}, {"_id": 0, "local_path": 0}).sort("created_at", -1).to_list(100)
    drive_settings = await get_google_drive_settings(mask=True)
    return {
        "settings": settings,
        "backups": backups,
        "drive_ready": google_drive_upload_enabled(drive_settings),
        "drive_account": drive_settings.get("service_account_email", ""),
        "drive_folder": f"{drive_settings.get('root_folder_name', 'E-Learning Dosen')} / Database Backups",
    }


@api_router.put("/database-backups/settings")
async def update_database_backup_settings(
    payload: DatabaseBackupSettingsInput,
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.frequency not in {"daily", "weekly"}:
        raise HTTPException(status_code=400, detail="Frekuensi backup harus harian atau mingguan")
    if not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", payload.run_time):
        raise HTTPException(status_code=400, detail="Waktu backup harus memakai format HH:MM")
    if payload.enabled and payload.upload_to_drive:
        drive_settings = await get_google_drive_settings(mask=True)
        if not google_drive_upload_enabled(drive_settings):
            raise HTTPException(
                status_code=400,
                detail="Aktifkan dan simpan konfigurasi Google Drive sebelum menjadwalkan backup ke Drive",
            )
    doc = {
        "id": "main",
        **payload.model_dump(),
        "timezone": str(BACKUP_TIMEZONE),
        "next_run_at": next_database_backup_at(payload.model_dump()) if payload.enabled else "",
        "updated_at": now_iso(),
        "updated_by": user["id"],
        "running": False,
    }
    await db.database_backup_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    return await get_database_backup_settings()


@api_router.post("/database-backups/run")
async def run_database_backup(user: Dict[str, Any] = Depends(require_campus_admin)):
    running = await db.database_backups.find_one({"status": "running"}, {"_id": 0, "id": 1})
    if running:
        raise HTTPException(status_code=409, detail="Proses backup lain masih berjalan")
    return await create_database_backup("manual", user["id"])


@api_router.get("/database-backups/{backup_id}/download")
async def download_database_backup(
    backup_id: str, _: Dict[str, Any] = Depends(require_campus_admin)
):
    backup = await db.database_backups.find_one({"id": backup_id}, {"_id": 0})
    if not backup or not backup.get("local_available") or not backup.get("local_path"):
        raise HTTPException(status_code=404, detail="File backup lokal tidak tersedia")
    path = Path(backup["local_path"]).resolve()
    if BACKUP_ROOT.resolve() not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="File backup lokal tidak ditemukan")
    return FileResponse(path, filename=backup.get("file_name", path.name), media_type="application/gzip")


@api_router.get("/drive/settings")
async def get_drive_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return {**await get_google_drive_settings(mask=True), **await storage_status_summary(), **await drive_sync_overview()}


@api_router.get("/drive/sync-status")
async def get_drive_sync_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await drive_sync_overview()


@api_router.post("/drive/sync/{file_id}/retry")
async def retry_drive_sync_file(file_id: str, background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    await db.stored_files.update_one(
        {"id": file_id},
        {"$set": {"drive_sync_status": "pending", "upload_status": "stored_on_server", "updated_at": now_iso()}, "$unset": {"drive_error": ""}},
    )
    await refresh_embedded_file_references(file_id)
    background_tasks.add_task(sync_stored_file_to_drive, file_id)
    return {"ok": True, "file_id": file_id, "status": "pending"}


@api_router.post("/drive/sync/retry-failed")
async def retry_failed_drive_sync(background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    failed = await db.stored_files.find({"drive_sync_status": "failed"}, {"_id": 0, "id": 1}).to_list(100)
    for item in failed:
        file_id = item["id"]
        await db.stored_files.update_one(
            {"id": file_id},
            {"$set": {"drive_sync_status": "pending", "upload_status": "stored_on_server", "updated_at": now_iso()}, "$unset": {"drive_error": ""}},
        )
        await refresh_embedded_file_references(file_id)
        background_tasks.add_task(sync_stored_file_to_drive, file_id)
    return {"ok": True, "queued": len(failed)}


@api_router.put("/drive/settings")
async def update_drive_settings(
    payload: GoogleDriveSettingsInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.lecturer_folder_role not in {"reader", "writer"}:
        raise HTTPException(status_code=400, detail="Akses folder dosen harus reader atau writer")
    delegated_user = payload.google_workspace_delegated_user.strip().lower()
    if delegated_user and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", delegated_user):
        raise HTTPException(status_code=400, detail="Email Google Workspace untuk Meet tidak valid")
    existing = await db.google_drive_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    service_account_json_encrypted = existing.get("service_account_json_encrypted", "")
    service_account_email = existing.get("service_account_email", "")
    if payload.clear_service_account:
        service_account_json_encrypted = ""
        service_account_email = ""
    if payload.service_account_json.strip():
        try:
            info = normalize_service_account_payload(payload.service_account_json)
        except Exception:
            raise HTTPException(status_code=400, detail="Service account JSON tidak valid")
        normalized = json.dumps(info, separators=(",", ":"), ensure_ascii=False)
        service_account_json_encrypted = encrypt_secret(normalized)
        service_account_email = info.get("client_email", "")
    doc = {
        "id": "main",
        "enabled": payload.enabled,
        "root_folder_id": payload.root_folder_id.strip(),
        "root_folder_name": safe_path_segment(payload.root_folder_name or "E-Learning Dosen"),
        "require_upload": payload.require_upload,
        "lecturer_folder_sharing_enabled": payload.lecturer_folder_sharing_enabled,
        "lecturer_folder_role": payload.lecturer_folder_role,
        "google_meet_enabled": payload.google_meet_enabled,
        "google_workspace_delegated_user": delegated_user,
        "service_account_json_encrypted": service_account_json_encrypted,
        "service_account_email": service_account_email,
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.google_drive_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("google_drive_settings")
    background_tasks.add_task(reconcile_all_lecturer_drive_access)
    return {**await get_google_drive_settings(mask=True), **await storage_status_summary(), **await drive_sync_overview()}


def test_drive_connection_sync(settings: Dict[str, Any]) -> Dict[str, Any]:
    service = get_drive_service(settings)
    root_folder_id = settings.get("root_folder_id", "")
    folder_name = ""
    if root_folder_id:
        folder = (
            service.files()
            .get(fileId=root_folder_id, fields="id,name,mimeType", supportsAllDrives=True)
            .execute()
        )
        parent_id = folder.get("id", "")
        folder_name = folder.get("name", "")
    else:
        root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
        parent_id = drive_find_or_create_folder(service, root_folder_name, None)
        folder_name = root_folder_name
    test_name = f".nugas-upload-test-{uuid.uuid4().hex[:8]}.txt"
    media = MediaIoBaseUpload(io.BytesIO(b"nugas google drive upload test"), mimetype="text/plain", resumable=True)
    uploaded = (
        service.files()
        .create(
            body={"name": test_name, "parents": [parent_id]},
            media_body=media,
            fields="id,name",
            supportsAllDrives=True,
        )
        .execute()
    )
    uploaded_id = uploaded.get("id", "")
    if uploaded_id:
        service.files().delete(fileId=uploaded_id, supportsAllDrives=True).execute()
    return {"ok": True, "folder_id": parent_id, "folder_name": folder_name, "upload_test": "ok"}


@api_router.post("/drive/settings/test")
async def test_drive_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        raise HTTPException(status_code=400, detail="Google Drive belum aktif atau credential belum tersimpan")
    try:
        return await asyncio.to_thread(test_drive_connection_sync, settings)
    except Exception as exc:
        logger.exception("Tes koneksi Google Drive gagal: %s", exc)
        raise HTTPException(status_code=400, detail=google_drive_error_message(exc))


@api_router.post("/drive/settings/test-meet")
async def test_google_meet_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_google_drive_settings(mask=False)
    if not settings.get("google_meet_enabled"):
        raise HTTPException(status_code=400, detail="Google Meet REST API belum diaktifkan pada konfigurasi admin")
    if not settings.get("service_account_configured"):
        raise HTTPException(status_code=400, detail="Service Account JSON belum tersimpan")
    delegated_user = settings.get("google_workspace_delegated_user", "")
    if not delegated_user:
        raise HTTPException(status_code=400, detail="Email akun Google Workspace untuk pengujian belum diisi")
    try:
        meet = await asyncio.to_thread(create_google_meet_space_sync, settings, delegated_user)
    except Exception as exc:
        logger.warning("Tes koneksi Google Meet gagal: %s", exc)
        raise HTTPException(status_code=400, detail=google_meet_error_message(exc)) from exc
    return {
        "ok": True,
        "message": "Koneksi Google Meet berhasil dan ruang uji telah dibuat",
        "delegated_user": delegated_user,
        **meet,
    }


@api_router.get("/programs")
async def list_programs(_: Dict[str, Any] = Depends(get_current_user)):
    return await db.programs.find({"status": {"$ne": "deleted"}}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/programs")
async def create_program(payload: ProgramInput, _: Dict[str, Any] = Depends(require_admin)):
    doc = payload.model_dump()
    doc.update({"id": new_id(), "status": "active", "created_at": now_iso()})
    await db.programs.insert_one(doc)
    return public_doc(doc)


@api_router.put("/programs/{program_id}")
async def update_program(program_id: str, payload: ProgramInput, _: Dict[str, Any] = Depends(require_admin)):
    existing = await db.programs.find_one({"id": program_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    update = payload.model_dump()
    update["updated_at"] = now_iso()
    await db.programs.update_one({"id": program_id}, {"$set": update})
    await db.courses.update_many({"program_id": program_id}, {"$set": {"program_name": payload.name}})
    await db.classes.update_many({"program_id": program_id}, {"$set": {"program_name": payload.name}})
    doc = await db.programs.find_one({"id": program_id}, {"_id": 0})
    return public_doc(doc)


@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str, _: Dict[str, Any] = Depends(require_admin)):
    linked_course = await db.courses.find_one({"program_id": program_id, "status": {"$ne": "deleted"}}, {"_id": 0, "id": 1})
    if linked_course:
        raise HTTPException(status_code=400, detail="Prodi masih dipakai mata kuliah. Hapus atau pindahkan mata kuliah dulu.")
    result = await db.programs.update_one({"id": program_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    return {"ok": True}


@api_router.get("/courses")
async def list_courses(_: Dict[str, Any] = Depends(get_current_user)):
    courses = await db.courses.find({"status": {"$ne": "deleted"}}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [await enrich_course_payload(item) for item in courses]


@api_router.get("/courses/{course_id}/grade-weights")
async def get_course_grade_weights(course_id: str, user: Dict[str, Any] = Depends(require_admin)):
    course = await require_course_access(course_id, user)
    customized = isinstance(course.get("grade_weights"), dict)
    return {
        "course_id": course_id,
        "course_name": course.get("name", ""),
        "weights": grade_weights_from_document(course.get("grade_weights")),
        "customized": customized,
    }


@api_router.put("/courses/{course_id}/grade-weights")
async def update_course_grade_weights(
    course_id: str,
    payload: GradeWeightsInput,
    user: Dict[str, Any] = Depends(require_admin),
):
    course = await require_course_access(course_id, user)
    weights = validate_grade_weights(payload.model_dump())
    await db.courses.update_one(
        {"id": course_id},
        {"$set": {"grade_weights": weights, "grade_weights_updated_at": now_iso(), "grade_weights_updated_by": user["id"]}},
    )
    return {
        "course_id": course_id,
        "course_name": course.get("name", ""),
        "weights": weights,
        "customized": True,
    }


@api_router.delete("/courses/{course_id}/grade-weights")
async def reset_course_grade_weights(course_id: str, user: Dict[str, Any] = Depends(require_admin)):
    course = await require_course_access(course_id, user)
    await db.courses.update_one(
        {"id": course_id},
        {"$unset": {"grade_weights": "", "grade_weights_updated_at": "", "grade_weights_updated_by": ""}},
    )
    return {
        "course_id": course_id,
        "course_name": course.get("name", ""),
        "weights": dict(DEFAULT_GRADE_WEIGHTS),
        "customized": False,
    }


@api_router.post("/courses")
async def create_course(payload: CourseInput, _: Dict[str, Any] = Depends(require_admin)):
    program_query = {"id": payload.program_id, "status": {"$ne": "deleted"}} if payload.program_id else {"status": {"$ne": "deleted"}}
    program = await db.programs.find_one(program_query, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    doc = payload.model_dump()
    doc["program_id"] = program["id"]
    doc.update({"id": new_id(), "program_name": program["name"], "status": "active", "created_at": now_iso()})
    await db.courses.insert_one(doc)
    return public_doc(doc)


@api_router.put("/courses/{course_id}")
async def update_course(course_id: str, payload: CourseInput, _: Dict[str, Any] = Depends(require_admin)):
    existing = await db.courses.find_one({"id": course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    selected_program_id = payload.program_id or existing.get("program_id", "")
    program = await db.programs.find_one({"id": selected_program_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    update = payload.model_dump()
    update["program_id"] = program["id"]
    update["sks"] = payload.credits
    update.update({"program_name": program["name"], "updated_at": now_iso()})
    await db.courses.update_one({"id": course_id}, {"$set": update})
    await db.classes.update_many(
        {"course_id": course_id},
        {"$set": {"course_name": payload.name, "sks": payload.credits, "program_id": payload.program_id, "program_name": program["name"]}},
    )
    doc = await db.courses.find_one({"id": course_id}, {"_id": 0})
    return await enrich_course_payload(public_doc(doc))


@api_router.delete("/courses/{course_id}")
async def delete_course(course_id: str, _: Dict[str, Any] = Depends(require_admin)):
    linked_class = await db.classes.find_one({"course_id": course_id, "status": {"$ne": "deleted"}}, {"_id": 0, "id": 1})
    if linked_class:
        raise HTTPException(status_code=400, detail="Mata kuliah masih dipakai kelas. Hapus kelas dulu.")
    result = await db.courses.update_one({"id": course_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    return {"ok": True}


@api_router.get("/classes")
async def list_classes(user: Dict[str, Any] = Depends(get_current_user)):
    if user["role"] == "student":
        query = {"id": {"$in": user.get("class_ids", [])}, "status": {"$ne": "deleted"}}
        classes = await db.classes.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    elif user["role"] == "lecturer" and not is_campus_admin(user):
        cids = await lecturer_class_ids(user, include_deleted=False)
        classes = await db.classes.find({"id": {"$in": cids}, "status": {"$ne": "deleted"}}, {"_id": 0}).sort("created_at", -1).to_list(500)
    else:
        query = {"status": {"$ne": "deleted"}}
        classes = await db.classes.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [await enrich_class_payload(item) for item in classes]


@api_router.post("/classes")
async def create_class(payload: ClassInput, user: Dict[str, Any] = Depends(require_admin)):
    course = await db.courses.find_one({"id": payload.course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    dosen_id, dosen_nama = await require_course_lecturer(course)
    class_name = await normalize_new_class_name(course, payload.name)
    code_seed = f"{course.get('code', 'KLS')}{class_name}{uuid.uuid4().hex[:4]}"
    doc = payload.model_dump()
    doc["name"] = class_name
    doc.update(
        {
            "id": new_id(),
            "course_name": course["name"],
            "course_code": course.get("code") or course.get("kode") or "",
            "sks": course.get("sks", course.get("total_sks", 0)),
            "program_id": course.get("program_id") or course.get("prodi_id", ""),
            "program_name": course.get("program_name", ""),
            "class_code": clean_code(code_seed),
            "lecturer_id": dosen_id,
            "lecturer_name": dosen_nama,
            "status": "active",
            "student_ids": [],
            "grade_weights_snapshot": grade_weights_from_document(course.get("grade_weights")),
            "grade_weights_snapshot_customized": isinstance(course.get("grade_weights"), dict),
            "created_at": now_iso(),
        }
    )
    await db.classes.insert_one(doc)
    return await enrich_class_payload(public_doc(doc))


@api_router.post("/classes/{class_id}/duplicate")
async def duplicate_class_for_new_period(
    class_id: str,
    payload: ClassDuplicateInput,
    user: Dict[str, Any] = Depends(require_admin),
):
    if payload.confirmation.strip().upper() != "DUPLIKASI":
        raise HTTPException(status_code=400, detail="Ketik DUPLIKASI untuk membuat kelas periode baru.")
    source = await require_class_access(class_id, user)
    academic_year = payload.academic_year.strip()
    semester = payload.semester.strip()
    name = payload.name.strip()
    if academic_year == str(source.get("academic_year", "")).strip() and semester.lower() == str(source.get("semester", "")).strip().lower():
        raise HTTPException(status_code=409, detail="Periode kelas baru harus berbeda dari kelas sumber.")
    course = await db.courses.find_one(
        {"id": source.get("course_id", ""), "status": {"$ne": "deleted"}},
        {"_id": 0},
    )
    if not course:
        raise HTTPException(status_code=409, detail="Mata kuliah sumber sudah tidak tersedia.")
    lecturer_id, lecturer_name = await require_course_lecturer(course)
    name = await normalize_new_class_name(course, name)
    duplicate = await db.classes.find_one(
        {
            "course_id": source.get("course_id", ""),
            "lecturer_id": lecturer_id,
            "academic_year": academic_year,
            "semester": semester,
            "name": name,
            "status": {"$ne": CLASS_STATUS_DELETED},
        },
        {"_id": 0, "id": 1},
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Kelas dengan mata kuliah, nama, dan periode tersebut sudah ada.")
    code_seed = f"{course.get('code', 'KLS')}{name}{uuid.uuid4().hex[:4]}"
    doc = {
        "id": new_id(),
        "academic_year": academic_year,
        "semester": semester,
        "course_id": course["id"],
        "course_name": course.get("name", source.get("course_name", "")),
        "course_code": course.get("code") or source.get("course_code", ""),
        "sks": course.get("sks", course.get("total_sks", 0)),
        "program_id": course.get("program_id", source.get("program_id", "")),
        "program_name": course.get("program_name", source.get("program_name", "")),
        "name": name,
        "schedule": payload.schedule.strip(),
        "class_code": clean_code(code_seed),
        "lecturer_id": lecturer_id,
        "lecturer_name": lecturer_name,
        "status": CLASS_STATUS_ACTIVE,
        "student_ids": [],
        "grade_weights_snapshot": grade_weights_from_document(course.get("grade_weights")),
        "grade_weights_snapshot_customized": isinstance(course.get("grade_weights"), dict),
        "duplicated_from_class_id": source["id"],
        "duplicated_by": user["id"],
        "created_at": now_iso(),
    }
    await db.classes.insert_one(doc)
    return await enrich_class_payload(public_doc(doc))


@api_router.put("/classes/{class_id}")
async def update_class(class_id: str, payload: ClassInput, user: Dict[str, Any] = Depends(require_admin)):
    existing = await require_class_mutation_access(class_id, user)
    course = await db.courses.find_one({"id": payload.course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    await require_course_lecturer(course)
    update = payload.model_dump()
    update.update(
        {
            "course_name": course["name"],
            "course_code": course.get("code") or course.get("kode", ""),
            "sks": course.get("sks", course.get("total_sks", 0)),
            "program_id": course.get("program_id", ""),
            "program_name": course.get("program_name", ""),
            "updated_at": now_iso(),
        }
    )
    await db.classes.update_one({"id": class_id}, {"$set": update})
    doc = await db.classes.find_one({"id": class_id}, {"_id": 0})
    return await enrich_class_payload(public_doc(doc))


@api_router.delete("/classes/{class_id}")
async def delete_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    if not class_allows_learning(class_doc):
        raise HTTPException(status_code=409, detail="Kelas yang sudah berakhir/final tidak dapat dihapus. Gunakan arsip untuk menyimpan histori.")
    result = await db.classes.update_one({"id": class_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    await db.users.update_many({"class_ids": class_id}, {"$pull": {"class_ids": class_id}})
    return {"ok": True}


@api_router.post("/classes/{class_id}/archive")
async def archive_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    if class_doc.get("status") != CLASS_STATUS_FINALIZED:
        raise HTTPException(status_code=409, detail="Finalisasi nilai terlebih dahulu sebelum mengarsipkan kelas.")
    await db.classes.update_one(
        {"id": class_id},
        {"$set": {"status": CLASS_STATUS_ARCHIVED, "archived_at": now_iso()}},
    )
    return {"ok": True}


@api_router.post("/classes/{class_id}/end")
async def end_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    if class_doc.get("status") != CLASS_STATUS_ACTIVE:
        raise HTTPException(status_code=409, detail=f"Kelas tidak dapat diakhiri karena statusnya {class_status_label(class_doc.get('status', ''))}.")
    course = await db.courses.find_one({"id": class_doc.get("course_id")}, {"_id": 0}) or {}
    await db.classes.update_one(
        {"id": class_id},
        {
            "$set": {
                "status": CLASS_STATUS_ENDED,
                "ended_at": now_iso(),
                "grade_weights_snapshot": grade_weights_from_document(course.get("grade_weights")),
                "grade_weights_snapshot_customized": isinstance(course.get("grade_weights"), dict),
            }
        },
    )
    return {"ok": True}


@api_router.post("/classes/{class_id}/finalize")
async def finalize_class(
    class_id: str,
    payload: FinalizationInput,
    user: Dict[str, Any] = Depends(require_admin),
):
    if payload.confirmation.strip().upper() != "FINALISASI":
        raise HTTPException(status_code=400, detail="Ketik FINALISASI untuk mengunci nilai kelas.")
    class_doc = await require_class_access(class_id, user)
    if class_doc.get("status") != CLASS_STATUS_ENDED:
        raise HTTPException(status_code=409, detail="Kelas harus berstatus Berakhir sebelum nilai difinalisasi.")
    recaps = await build_grade_recap(user, class_id, use_snapshots=False)
    recap = recaps[0] if recaps else None
    incomplete = [
        f"{student.get('student_name', 'Mahasiswa')} ({student.get('student_nim', '-')})"
        for student in (recap or {}).get("students", [])
        if not student.get("grade_complete")
    ]
    if incomplete:
        raise HTTPException(
            status_code=409,
            detail="Nilai belum lengkap untuk: " + ", ".join(incomplete[:10]),
        )
    finalized_at = now_iso()
    snapshot = recap or {
        "class_id": class_id,
        "class_name": class_doc.get("name", ""),
        "students": [],
        "student_count": 0,
        "total_assignments": 0,
        "class_average": 0,
        "grade_distribution": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0},
        "grade_weights": grade_weights_from_document(class_doc.get("grade_weights_snapshot")),
        "grade_weights_customized": bool(class_doc.get("grade_weights_snapshot_customized", False)),
        "assignments": [],
    }
    snapshot["finalized_at"] = finalized_at
    snapshot["finalized_by"] = user["id"]
    snapshot["class_status"] = CLASS_STATUS_FINALIZED
    snapshot["class_status_label"] = class_status_label(CLASS_STATUS_FINALIZED)
    await db.classes.update_one(
        {"id": class_id},
        {
            "$set": {
                "status": CLASS_STATUS_FINALIZED,
                "finalized_at": finalized_at,
                "finalized_by": user["id"],
                "final_grade_snapshot": snapshot,
                "grade_weights_snapshot": snapshot.get("grade_weights", DEFAULT_GRADE_WEIGHTS),
            }
        },
    )
    return {"ok": True, "status": CLASS_STATUS_FINALIZED, "snapshot": snapshot}


@api_router.get("/classes/{class_id}/students")
async def class_students(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    students = await db.users.find(
        {"role": "student", "class_ids": class_id}, {"_id": 0, "password_hash": 0}
    ).sort("name", 1).to_list(1000)
    progress_map = await calculate_student_progress_many([s["id"] for s in students], [class_id])
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return {"class": await enrich_class_payload(class_doc), "students": students}


@api_router.post("/classes/{class_id}/students/{student_id}/remove")
async def remove_student_from_class(class_id: str, student_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_mutation_access(class_id, user)
    await db.classes.update_one({"id": class_id}, {"$pull": {"student_ids": student_id}})
    await db.users.update_one({"id": student_id}, {"$pull": {"class_ids": class_id}})
    return {"ok": True}


@api_router.post("/classes/{class_id}/students/{student_id}/add")
async def add_existing_student_to_class(class_id: str, student_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    student = await get_active_student(student_id)
    result = await add_student_to_class_record(class_doc, student, user["id"])
    return {"ok": True, **result}


@api_router.post("/classes/{class_id}/students/bulk-add")
async def bulk_add_existing_students_to_class(class_id: str, payload: StudentIdsInput, user: Dict[str, Any] = Depends(require_admin)):
    student_ids = unique_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal satu mahasiswa")
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    results = []
    for student_id in student_ids:
        try:
            student = await get_active_student(student_id)
            results.append(await add_student_to_class_record(class_doc, student, user["id"]))
        except HTTPException as exc:
            results.append({"student_id": student_id, "status": "skipped", "detail": exc.detail})
    added = len([item for item in results if item.get("status") == "approved"])
    already_joined = len([item for item in results if item.get("status") == "already_joined"])
    skipped = len([item for item in results if item.get("status") == "skipped"])
    return {"ok": True, "added": added, "already_joined": already_joined, "skipped": skipped, "results": results}


@api_router.post("/classes/{class_id}/students/{student_id}/invite")
async def invite_student_to_class(
    class_id: str,
    student_id: str,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    student = await get_active_student(student_id)
    result = await invite_student_to_class_record(class_doc, student, background_tasks, user["id"])
    return {"ok": True, **result}


@api_router.post("/classes/{class_id}/students/bulk-invite")
async def bulk_invite_students_to_class(
    class_id: str,
    payload: StudentIdsInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    student_ids = unique_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal satu mahasiswa")
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    results = []
    for student_id in student_ids:
        try:
            student = await get_active_student(student_id)
            results.append(await invite_student_to_class_record(class_doc, student, background_tasks, user["id"]))
        except HTTPException as exc:
            results.append({"student_id": student_id, "status": "skipped", "detail": exc.detail, "delivery_status": ""})
    invited = len([item for item in results if item.get("status") in ["invited", "pending"]])
    already_joined = len([item for item in results if item.get("status") == "already_joined"])
    skipped = len([item for item in results if item.get("status") == "skipped"])
    queued = len([item for item in results if item.get("delivery_status") == "pending"])
    pending_config = len([item for item in results if item.get("delivery_status") == "pending_config"])
    no_whatsapp = len([item for item in results if item.get("delivery_status") == "no_whatsapp"])
    return {
        "ok": True,
        "invited": invited,
        "already_joined": already_joined,
        "skipped": skipped,
        "queued": queued,
        "pending_config": pending_config,
        "no_whatsapp": no_whatsapp,
        "results": results,
    }


@api_router.post("/students/{student_id}/status")
async def update_student_status(
    student_id: str,
    payload: Dict[str, str],
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    status = payload.get("status")
    if status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="Status harus active atau inactive")
    await db.users.update_one({"id": student_id, "role": "student"}, {"$set": {"status": status, "status_updated_at": now_iso()}})
    return {"ok": True, "status": status}


@api_router.post("/students/{student_id}/reset-password")
async def reset_student_password(
    student_id: str,
    payload: ResetPasswordInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    student = await db.users.find_one({"id": student_id, "role": "student"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Mahasiswa tidak ditemukan")
    new_password = payload.password.strip() if payload.password.strip() else (student.get("nim") or "Mahasiswa1231!")
    if len(new_password) < 3:
        raise HTTPException(status_code=400, detail="Password minimal 3 karakter")
    await db.users.update_one(
        {"id": student_id},
        {"$set": {"password_hash": hash_password(new_password), "password_reset_at": now_iso()}},
    )
    return {"ok": True, "temporary_password": new_password}


def split_program_scope_values(*sources: Any) -> List[str]:
    """Normalisasi scope prodi lama yang kadang tersimpan sebagai CSV."""
    values: List[str] = []
    for source in sources:
        source_values = source if isinstance(source, (list, tuple, set)) else [source]
        for value in source_values:
            for item in re.split(r"[,;|\n]+", str(value or "")):
                clean_value = item.strip()
                if clean_value and clean_value not in values:
                    values.append(clean_value)
    return values


def preferred_program_scope_values(
    user: Dict[str, Any],
    structural_scope_values: Optional[List[str]] = None,
) -> List[str]:
    """Penunjukan jabatan aktif mengalahkan field prodi lama pada profil."""
    active_structural_scope = split_program_scope_values(structural_scope_values or [])
    if active_structural_scope:
        return active_structural_scope

    derived_scope = split_program_scope_values(
        user.get("access_scope_prodi_ids"),
        user.get("kaprodi_prodi_id"),
    )
    if derived_scope:
        return derived_scope

    # Profil dari OLD-SIAKAD dapat berisi beberapa kode prodi sekaligus. Nilai
    # ini hanya dipakai sementara bila akun belum memiliki penunjukan struktural.
    return split_program_scope_values(user.get("prodi_id"))


async def active_program_manager_scope_values(user: Dict[str, Any]) -> List[str]:
    """Ambil scope Kaprodi/Sekprodi aktif langsung dari penunjukan jabatan."""
    assignments = await db.jabatan_assignments.find(
        {"user_id": user.get("id", "")},
        {"_id": 0, "jabatan_kode": 1, "prodi_id": 1, "status": 1},
    ).to_list(100)
    return split_program_scope_values([
        assignment.get("prodi_id")
        for assignment in assignments
        if str(assignment.get("jabatan_kode") or "").upper() in {"KAPRODI", "SEKPRODI"}
        and str(assignment.get("status") or "active").lower() not in {"inactive", "revoked"}
    ])


async def resolved_program_scope_values(
    user: Dict[str, Any],
    structural_scope_values: Optional[List[str]] = None,
) -> List[str]:
    """Kembangkan ID/kode prodi scope ke seluruh alias yang dipakai data mahasiswa."""
    raw_values = preferred_program_scope_values(
        user,
        structural_scope_values,
    )
    return await resolve_program_identifiers(db, raw_values)


def _analysis_period_year(value: Any) -> str:
    match = re.search(r"(?:19|20)\d{2}", str(value or ""))
    return match.group(0) if match else ""


def _analysis_period_semester(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if raw in {"1", "ganjil", "gasal", "odd"} or "ganjil" in raw or "gasal" in raw:
        return "ganjil"
    if raw in {"2", "genap", "even"} or "genap" in raw:
        return "genap"
    return raw


def _analysis_class_matches_period(
    class_doc: Dict[str, Any],
    selector: str,
    period_doc: Optional[Dict[str, Any]],
) -> bool:
    raw_selector = str(selector or "").strip()
    if not raw_selector or raw_selector == "all":
        return True
    direct_period_ids = (
        class_doc.get("tahun_ajaran_id"),
        class_doc.get("academic_year_id"),
        class_doc.get("academic_period_id"),
        class_doc.get("period_id"),
    )
    if any(str(value or "").strip() == raw_selector for value in direct_period_ids):
        return True

    target_year = _analysis_period_year(
        (period_doc or {}).get("tahun")
        or (period_doc or {}).get("academic_year")
        or (period_doc or {}).get("nama")
        or raw_selector
    )
    target_semester = _analysis_period_semester(
        (period_doc or {}).get("semester")
        or ("Ganjil" if raw_selector[-1:] == "1" and raw_selector.isdigit() else "Genap" if raw_selector[-1:] == "2" and raw_selector.isdigit() else "")
    )
    class_year = _analysis_period_year(
        class_doc.get("academic_year")
        or class_doc.get("tahun_ajaran")
        or class_doc.get("academic_year_label")
    )
    class_semester = _analysis_period_semester(class_doc.get("semester") or class_doc.get("term"))
    if not target_year or not class_year or target_year != class_year:
        return False
    return bool(target_semester and class_semester and target_semester == class_semester)


def _analysis_program_fields(document: Dict[str, Any]) -> List[str]:
    return [
        str(document.get(key) or "").strip()
        for key in ("prodi_id", "prodi_kode", "program_id", "prodi_name", "program_name", "nama_prodi")
        if str(document.get(key) or "").strip()
    ]


def _analysis_is_program_manager(
    user: Dict[str, Any],
    structural_scope: Optional[List[str]] = None,
) -> bool:
    return (
        user_is_admin_or_access_role(user, "academic_operator")
        or bool(structural_scope)
        or user_is_program_manager(user)
    )


def _analysis_has_global_program_scope(user: Dict[str, Any]) -> bool:
    """Admin dan BAAK dapat memilih satu Prodi atau melihat seluruh Prodi."""
    return user_is_admin_or_access_role(user, "academic_operator")


async def require_program_analysis_user(
    user: Dict[str, Any] = Depends(get_current_user),
) -> Dict[str, Any]:
    if _analysis_is_program_manager(user):
        return user
    raise HTTPException(
        status_code=403,
        detail="Halaman ini hanya tersedia untuk Kaprodi, operator akademik, atau Admin Kampus",
    )


async def _analysis_resolve_program_scope(
    user: Dict[str, Any],
    requested_prodi: str = "",
) -> tuple[List[str], List[str], List[str]]:
    structural_scope = await active_program_manager_scope_values(user)
    if not _analysis_is_program_manager(user, structural_scope):
        raise HTTPException(
            status_code=403,
            detail="Halaman ini hanya tersedia untuk Kaprodi, operator akademik, atau Admin Kampus",
        )

    own_scope_values = await resolved_program_scope_values(user, structural_scope)
    has_global_scope = _analysis_has_global_program_scope(user)
    requested_value = str(requested_prodi or "").strip()
    if requested_value:
        own_lookup = {value.upper() for value in own_scope_values}
        if not has_global_scope and requested_value.upper() not in own_lookup:
            raise HTTPException(status_code=403, detail="Prodi di luar kewenangan Anda")
        scope_values = await resolved_program_scope_values(
            {"prodi_id": requested_value},
            [requested_value],
        )
    else:
        scope_values = [] if has_global_scope else own_scope_values
    return structural_scope, own_scope_values, scope_values


async def _analysis_program_options() -> List[Dict[str, Any]]:
    programs = await db.programs.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "nama": 1, "code": 1, "kode": 1, "status": 1},
    ).sort("name", 1).to_list(500)
    return [item for item in programs if str(item.get("status") or "active").lower() != "deleted"]


async def _analysis_program_label(
    requested_prodi: str,
    user: Dict[str, Any],
    class_docs: Optional[List[Dict[str, Any]]] = None,
) -> str:
    requested_value = str(requested_prodi or "").strip()
    if requested_value:
        program = await db.programs.find_one(
            {
                "$or": [
                    {"id": requested_value}, {"code": requested_value},
                    {"kode": requested_value}, {"name": requested_value},
                    {"nama": requested_value},
                ]
            },
            {"_id": 0, "name": 1, "nama": 1},
        )
        if program:
            return str(program.get("name") or program.get("nama") or requested_value)
    if _analysis_has_global_program_scope(user):
        return "Semua Program Studi"
    return str(
        user.get("prodi_nama")
        or user.get("prodi_name")
        or next(
            (
                item.get("program_name") or item.get("prodi_name")
                for item in (class_docs or [])
                if item.get("program_name") or item.get("prodi_name")
            ),
            "",
        )
        or "Program Studi"
    )


def _analysis_student_identifiers(student: Dict[str, Any]) -> set[str]:
    return {
        str(student.get(key) or "").strip()
        for key in ("id", "username", "nim")
        if str(student.get(key) or "").strip()
    }


def _analysis_parse_inactive_days(value: Any) -> Optional[int]:
    if not value:
        return None
    parsed = parse_iso_datetime(str(value))
    if not parsed:
        return None
    return max(0, (datetime.now(timezone.utc) - parsed).days)


def _analysis_is_late_submission(item: Dict[str, Any]) -> bool:
    if str(item.get("status") or "").strip().lower() in {"terlambat", "late"}:
        return True
    try:
        return float(item.get("late_days") or 0) > 0
    except (TypeError, ValueError):
        return False


def _analysis_risk(
    *,
    attendance_percentage: Optional[float],
    average_grade: float,
    missing_assignments: int,
    late_submissions: int,
    inactive_days: Optional[int],
    status: str,
) -> Dict[str, Any]:
    score = 0
    reasons: List[str] = []
    if attendance_percentage is not None:
        if attendance_percentage < 60:
            score += 4
            reasons.append("Kehadiran di bawah 60%")
        elif attendance_percentage < 75:
            score += 2
            reasons.append("Kehadiran di bawah 75%")
    if average_grade > 0:
        if average_grade < 60:
            score += 3
            reasons.append("Rata-rata nilai di bawah 60")
        elif average_grade < 70:
            score += 1
            reasons.append("Rata-rata nilai di bawah 70")
    if missing_assignments >= 3:
        score += 3
        reasons.append(f"{missing_assignments} tugas belum dikumpulkan")
    elif missing_assignments > 0:
        score += 1
        reasons.append(f"{missing_assignments} tugas belum dikumpulkan")
    if late_submissions >= 2:
        score += 2
        reasons.append(f"{late_submissions} tugas terlambat")
    if inactive_days is None:
        score += 1
        reasons.append("Belum ada aktivitas login")
    elif inactive_days > 14:
        score += 2
        reasons.append(f"Tidak aktif {inactive_days} hari")
    if str(status or "active").lower() != "active":
        score += 2
        reasons.append("Status mahasiswa tidak aktif")
    if score >= 8:
        label = "Risiko Tinggi"
    elif score >= 5:
        label = "Perlu Perhatian"
    elif score >= 2:
        label = "Risiko Rendah"
    else:
        label = "Aman"
    return {"score": score, "label": label, "reasons": reasons[:5]}


@api_router.get("/prodi/analisis-mahasiswa")
async def prodi_student_analysis(
    semester_id: str = "",
    prodi_id: str = "",
    user: Dict[str, Any] = Depends(require_program_analysis_user),
):
    """Analisis akademik dan risiko seluruh mahasiswa dalam scope Prodi."""
    requested_prodi = str(prodi_id or "").strip()
    structural_scope, own_scope_values, scope_values = await _analysis_resolve_program_scope(
        user,
        requested_prodi,
    )

    raw_semester_id = str(semester_id or "").strip()
    period_doc = None
    if raw_semester_id and raw_semester_id != "all":
        period_doc = await db.tahun_ajaran.find_one(
            {"$or": [{"id": raw_semester_id}, {"kode": raw_semester_id}, {"code": raw_semester_id}]},
            {"_id": 0},
        )
        if not period_doc:
            raise HTTPException(status_code=404, detail="Tahun ajaran yang dipilih tidak ditemukan")

    class_docs = await db.classes.find({"status": {"$ne": "deleted"}}, {"_id": 0}).to_list(5000)
    if scope_values:
        scope_lookup = {str(value).upper() for value in scope_values}
        class_docs = [
            item for item in class_docs
            if any(value.upper() in scope_lookup for value in _analysis_program_fields(item))
        ]
    elif not _analysis_has_global_program_scope(user):
        # Fallback aman untuk akun struktural lama yang belum memiliki prodi_id.
        # Jangan membuka semua kelas hanya karena field scope belum termigrasi.
        assigned_class_ids = set(await lecturer_class_ids(user))
        class_docs = [item for item in class_docs if str(item.get("id")) in assigned_class_ids]
    if raw_semester_id and raw_semester_id != "all":
        class_docs = [item for item in class_docs if _analysis_class_matches_period(item, raw_semester_id, period_doc)]

    class_by_id = {str(item.get("id")): item for item in class_docs if item.get("id")}
    class_ids = list(class_by_id)
    all_students = await db.users.find(
        {"role": "student"},
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(10000)
    student_by_identifier: Dict[str, Dict[str, Any]] = {}
    for student in all_students:
        for identifier in _analysis_student_identifiers(student):
            student_by_identifier.setdefault(identifier, student)

    class_member_ids: Dict[str, set[str]] = {class_id: set() for class_id in class_ids}
    student_class_ids: Dict[str, set[str]] = {}
    for class_id, class_doc in class_by_id.items():
        for raw_student_id in class_doc.get("student_ids", []):
            student = student_by_identifier.get(str(raw_student_id).strip())
            if student:
                sid = str(student.get("id"))
                class_member_ids[class_id].add(sid)
                student_class_ids.setdefault(sid, set()).add(class_id)

    if scope_values:
        scope_lookup = {str(value).upper() for value in scope_values}
        selected_students = [
            student for student in all_students
            if any(value.upper() in scope_lookup for value in _analysis_program_fields(student))
            or str(student.get("id")) in student_class_ids
        ]
    else:
        selected_students = list(all_students)
    selected_students_by_id = {str(student.get("id")): student for student in selected_students if student.get("id")}
    selected_student_ids = set(selected_students_by_id)

    for student in selected_students:
        sid = str(student.get("id"))
        for raw_class_id in student.get("class_ids", []) or []:
            class_id = str(raw_class_id).strip()
            if class_id in class_by_id:
                student_class_ids.setdefault(sid, set()).add(class_id)
                class_member_ids.setdefault(class_id, set()).add(sid)

    assignments = await db.assignments.find(
        {"class_id": {"$in": class_ids}, "is_active": True}, {"_id": 0}
    ).to_list(20000) if class_ids else []
    assignments = [item for item in assignments if assignment_is_published(item)]
    assignments_by_class: Dict[str, List[Dict[str, Any]]] = {}
    for assignment in assignments:
        assignments_by_class.setdefault(str(assignment.get("class_id")), []).append(assignment)

    submissions = await db.submissions.find(
        {"class_id": {"$in": class_ids}}, {"_id": 0}
    ).to_list(50000) if class_ids else []
    submissions_by_student: Dict[str, List[Dict[str, Any]]] = {}
    for submission in submissions:
        student = student_by_identifier.get(str(submission.get("student_id") or "").strip())
        if student and str(student.get("id")) in selected_student_ids:
            submissions_by_student.setdefault(str(student.get("id")), []).append(submission)

    attendance_sessions = await db.attendance_sessions.find(
        {"class_id": {"$in": class_ids}}, {"_id": 0}
    ).to_list(10000) if class_ids else []
    open_sessions_by_class: Dict[str, int] = {}
    attendance_by_student: Dict[str, Dict[str, int]] = {}
    for session in attendance_sessions:
        if session.get("status") not in {"open", "closed"}:
            continue
        class_id = str(session.get("class_id") or "")
        open_sessions_by_class[class_id] = open_sessions_by_class.get(class_id, 0) + 1
        for record in session.get("records", []) or []:
            student = student_by_identifier.get(str(record.get("student_id") or "").strip())
            if not student or str(student.get("id")) not in selected_student_ids:
                continue
            sid = str(student.get("id"))
            if class_id not in student_class_ids.get(sid, set()):
                continue
            stats = attendance_by_student.setdefault(sid, {"hadir": 0, "izin": 0, "sakit": 0, "alpa": 0})
            status = str(record.get("status") or "Alpa").lower()
            if status == "hadir":
                stats["hadir"] += 1
            elif status == "izin":
                stats["izin"] += 1
            elif status == "sakit":
                stats["sakit"] += 1
            else:
                stats["alpa"] += 1

    snapshot_grades_by_student: Dict[str, List[float]] = {}
    for class_doc in class_docs:
        snapshot = class_doc.get("final_grade_snapshot") or {}
        for snapshot_student in snapshot.get("students", []) or []:
            student = student_by_identifier.get(str(snapshot_student.get("student_id") or "").strip())
            if not student:
                student = student_by_identifier.get(str(snapshot_student.get("student_nim") or "").strip())
            grade = snapshot_student.get("weighted_grade", snapshot_student.get("average"))
            if student and isinstance(grade, (int, float)):
                snapshot_grades_by_student.setdefault(str(student.get("id")), []).append(float(grade))

    student_results: List[Dict[str, Any]] = []
    class_metric_map: Dict[str, List[Dict[str, Any]]] = {}
    for student in selected_students:
        sid = str(student.get("id"))
        student_class_set = student_class_ids.get(sid, set())
        student_assignments = [
            assignment
            for class_id in student_class_set
            for assignment in assignments_by_class.get(class_id, [])
        ]
        student_submissions = [
            submission for submission in submissions_by_student.get(sid, [])
            if str(submission.get("class_id") or "") in student_class_set
        ]
        submitted_assignment_ids = {str(item.get("assignment_id")) for item in student_submissions if item.get("assignment_id")}
        graded_values = [
            float(item.get("grade")) for item in student_submissions
            if isinstance(item.get("grade"), (int, float))
        ]
        if not graded_values:
            graded_values = snapshot_grades_by_student.get(sid, [])
        average_grade = round(sum(graded_values) / len(graded_values), 1) if graded_values else 0.0
        missing_assignments = len([
            item for item in student_assignments
            if str(item.get("id")) not in submitted_assignment_ids
        ])
        late_submissions = len([
            item for item in student_submissions
            if _analysis_is_late_submission(item)
        ])
        attendance_stats = attendance_by_student.get(sid, {"hadir": 0, "izin": 0, "sakit": 0, "alpa": 0})
        total_open_sessions = sum(open_sessions_by_class.get(class_id, 0) for class_id in student_class_set)
        attendance_percentage = round(attendance_stats["hadir"] / total_open_sessions * 100, 1) if total_open_sessions else None
        submission_rate = round(len(submitted_assignment_ids) / len(student_assignments) * 100, 1) if student_assignments else None
        inactive_days = _analysis_parse_inactive_days(student.get("last_login_at"))
        risk = _analysis_risk(
            attendance_percentage=attendance_percentage,
            average_grade=average_grade,
            missing_assignments=missing_assignments,
            late_submissions=late_submissions,
            inactive_days=inactive_days,
            status=str(student.get("status") or "active"),
        )
        class_items = [
            {
                "id": class_id,
                "name": class_by_id[class_id].get("name", ""),
                "course_name": class_by_id[class_id].get("course_name", ""),
                "course_code": class_by_id[class_id].get("course_code", ""),
                "academic_year": class_by_id[class_id].get("academic_year", ""),
                "semester": class_by_id[class_id].get("semester", ""),
            }
            for class_id in sorted(student_class_set)
            if class_id in class_by_id
        ]
        result = {
            "id": sid,
            "name": student.get("name") or student.get("nama") or "Mahasiswa",
            "nim": student.get("nim") or student.get("username") or "-",
            "email": student.get("email", ""),
            "prodi_id": student.get("prodi_id") or student.get("program_id") or "",
            "prodi_name": student.get("prodi_name") or student.get("program_name") or "",
            "angkatan": student.get("angkatan") or student.get("tahun_masuk") or "-",
            "status": student.get("status") or "active",
            "last_login_at": student.get("last_login_at", ""),
            "inactive_days": inactive_days,
            "class_count": len(class_items),
            "classes": class_items,
            "attendance": {
                **attendance_stats,
                "total_open": total_open_sessions,
                "percentage": attendance_percentage,
                "is_eligible_exam": attendance_percentage is None or attendance_percentage >= 75,
            },
            "grades": {
                "average": average_grade,
                "graded_count": len(graded_values),
                "low_grade_count": len([value for value in graded_values if value < 60]),
            },
            "learning": {
                "assignments_total": len(student_assignments),
                "submitted": len(submitted_assignment_ids),
                "missing": missing_assignments,
                "late": late_submissions,
                "submission_rate": submission_rate,
            },
            "risk": risk,
        }
        student_results.append(result)
        for class_id in student_class_set:
            class_metric_map.setdefault(class_id, []).append(result)

    student_results.sort(key=lambda item: (-item["risk"]["score"], item["grades"]["average"], item["name"]))
    average_attendance_values = [item["attendance"]["percentage"] for item in student_results if item["attendance"]["percentage"] is not None]
    average_grade_values = [item["grades"]["average"] for item in student_results if item["grades"]["graded_count"]]
    submission_rates = [item["learning"]["submission_rate"] for item in student_results if item["learning"]["submission_rate"] is not None]
    risk_distribution = {label: len([item for item in student_results if item["risk"]["label"] == label]) for label in ("Risiko Tinggi", "Perlu Perhatian", "Risiko Rendah", "Aman")}
    attendance_buckets = {
        "<60%": len([item for item in student_results if item["attendance"]["percentage"] is not None and item["attendance"]["percentage"] < 60]),
        "60-74%": len([item for item in student_results if item["attendance"]["percentage"] is not None and 60 <= item["attendance"]["percentage"] < 75]),
        "≥75%": len([item for item in student_results if item["attendance"]["percentage"] is not None and item["attendance"]["percentage"] >= 75]),
        "Belum ada data": len([item for item in student_results if item["attendance"]["percentage"] is None]),
    }
    class_summary = []
    for class_id, class_doc in class_by_id.items():
        metrics = class_metric_map.get(class_id, [])
        class_grades = [item["grades"]["average"] for item in metrics if item["grades"]["graded_count"]]
        class_attendance = [item["attendance"]["percentage"] for item in metrics if item["attendance"]["percentage"] is not None]
        class_summary.append({
            "id": class_id,
            "name": class_doc.get("name", ""),
            "course_name": class_doc.get("course_name", ""),
            "course_code": class_doc.get("course_code", ""),
            "student_count": len(metrics),
            "average_grade": round(sum(class_grades) / len(class_grades), 1) if class_grades else 0,
            "average_attendance": round(sum(class_attendance) / len(class_attendance), 1) if class_attendance else None,
            "high_risk_count": len([item for item in metrics if item["risk"]["label"] == "Risiko Tinggi"]),
        })
    class_summary.sort(key=lambda item: (-item["high_risk_count"], item["average_grade"], item["course_name"]))

    period_label = "Semua semester"
    if period_doc:
        period_label = period_doc.get("nama") or f"{period_doc.get('tahun', '')} {period_doc.get('semester', '')}".strip()
    prodi_name = await _analysis_program_label(requested_prodi, user, class_docs)
    return {
        "ok": True,
        "scope": {
            "is_admin": user.get("role") == "admin",
            "is_global": _analysis_has_global_program_scope(user),
            "can_select_prodi": _analysis_has_global_program_scope(user),
            "prodi_id": requested_prodi or (own_scope_values[0] if own_scope_values else ""),
            "prodi_name": prodi_name,
        },
        "prodi_list": await _analysis_program_options(),
        "period": {"id": raw_semester_id or "all", "label": period_label},
        "summary": {
            "total_students": len(student_results),
            "active_students": len([item for item in student_results if str(item["status"]).lower() == "active"]),
            "high_risk": risk_distribution["Risiko Tinggi"],
            "needs_attention": risk_distribution["Perlu Perhatian"],
            "low_risk": risk_distribution["Risiko Rendah"],
            "safe": risk_distribution["Aman"],
            "average_attendance": round(sum(average_attendance_values) / len(average_attendance_values), 1) if average_attendance_values else None,
            "average_grade": round(sum(average_grade_values) / len(average_grade_values), 1) if average_grade_values else 0,
            "average_submission_rate": round(sum(submission_rates) / len(submission_rates), 1) if submission_rates else None,
            "no_login_activity": len([item for item in student_results if item["inactive_days"] is None]),
        },
        "risk_distribution": risk_distribution,
        "attendance_buckets": attendance_buckets,
        "class_summary": class_summary,
        "students": student_results,
    }


@api_router.get("/students")
async def list_students(user: Dict[str, Any] = Depends(require_student_records_reader)):
    query: Dict[str, Any] = {"role": "student"}
    progress_class_ids: Optional[List[str]] = None
    structural_kaprodi_scope = await active_program_manager_scope_values(user)

    is_kaprodi = user_is_program_manager(user) or bool(structural_kaprodi_scope)
    kaprodi_scope_values = await resolved_program_scope_values(
        user,
        structural_kaprodi_scope,
    ) if is_kaprodi else []

    if user.get("role") != "admin" and is_kaprodi and kaprodi_scope_values:
        query["$or"] = [
            {"prodi_id": {"$in": kaprodi_scope_values}},
            {"prodi_kode": {"$in": kaprodi_scope_values}},
            {"program_id": {"$in": kaprodi_scope_values}},
            {"prodi_name": {"$in": kaprodi_scope_values}},
            {"program_name": {"$in": kaprodi_scope_values}},
        ]
    elif user_has_access_role(user, "finance_officer"):
        # Finance staff need the student master for billing and verification,
        # but this does not grant them academic mutation or administrator UI.
        progress_class_ids = None
    elif user.get("role") == "lecturer":
        class_ids = await lecturer_class_ids(user)
        progress_class_ids = class_ids
        homebase_scope_values = await resolved_program_scope_values(user)
        if not homebase_scope_values:
            return []
        query["$or"] = [
            {"prodi_id": {"$in": homebase_scope_values}},
            {"prodi_kode": {"$in": homebase_scope_values}},
            {"program_id": {"$in": homebase_scope_values}},
            {"prodi_name": {"$in": homebase_scope_values}},
            {"program_name": {"$in": homebase_scope_values}},
        ]
    students = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("name", 1).to_list(2000)
    progress_map = await calculate_student_progress_many(
        [s["id"] for s in students], progress_class_ids
    )
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return students


STUDENT_PROFILE_TEXT_FIELDS = (
    "parent_email", "parent_rt", "parent_rw", "parent_kota", "parent_provinsi",
    "parent_kode_pos", "parent_negara", "kewarganegaraan", "rt", "rw", "dusun",
    "kelurahan", "kecamatan", "kode_wilayah", "jenis_tinggal_id", "jenis_tinggal",
    "transportasi_id", "transportasi", "asal_sekolah", "status_sipil", "no_kk",
    "npwp", "no_kip", "no_kps", "kebutuhan_khusus", "tinggi_badan", "berat_badan",
    "semester_masuk", "tanggal_masuk", "jenis_pendaftaran_id", "jenis_pendaftaran",
    "jalur_masuk_id", "jalur_masuk", "jenis_pembiayaan_id", "jenis_pembiayaan",
    "status_mahasiswa_id", "feeder_student_id", "feeder_registration_id", "foto_url",
)
STUDENT_REGISTRATION_FIELDS = (
    "semester_masuk", "tanggal_masuk", "jenis_pendaftaran_id", "jenis_pendaftaran",
    "jalur_masuk_id", "jalur_masuk", "jenis_pembiayaan_id", "jenis_pembiayaan",
    "status_mahasiswa_id",
)


def _student_text(value: Any) -> str:
    return str(value or "").strip()


def _student_profile_fields(
    payload: StudentInput | StudentUpdateInput,
    existing: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Keep flat admin fields and the richer Old SIAKAD-shaped structures together."""
    result: Dict[str, Any] = {}
    for field in STUDENT_PROFILE_TEXT_FIELDS:
        value = getattr(payload, field, None)
        if value is not None:
            result[field] = _student_text(value)

    existing = existing or {}
    raw_parents = getattr(payload, "orang_tua", None)
    if raw_parents is not None:
        result["orang_tua"] = raw_parents or {}
    elif existing.get("orang_tua") is not None:
        result["orang_tua"] = existing.get("orang_tua")

    # The generic fields remain useful for the existing UI, while `orang_tua`
    # gives imports and future integrations the same shape as Old SIAKAD.
    generic_parent_fields = (
        "parent_name", "parent_phone", "parent_email", "parent_job", "parent_address",
        "parent_rt", "parent_rw", "parent_kota", "parent_provinsi", "parent_kode_pos",
        "parent_negara",
    )
    if any(getattr(payload, field, None) is not None for field in generic_parent_fields):
        parent_source = result.get("orang_tua", existing.get("orang_tua") or {})
        parents = dict(parent_source or {})
        wali = dict(parents.get("wali") or {})
        parent_map = {
            "parent_name": "nama", "parent_phone": "telepon", "parent_email": "email",
            "parent_job": "pekerjaan", "parent_address": "alamat", "parent_rt": "rt",
            "parent_rw": "rw", "parent_kota": "kota", "parent_provinsi": "provinsi",
            "parent_kode_pos": "kode_pos", "parent_negara": "negara",
        }
        for source, target in parent_map.items():
            value = getattr(payload, source, None)
            if value is not None:
                wali[target] = _student_text(value)
        parents["wali"] = wali
        result["orang_tua"] = parents

    raw_registration = getattr(payload, "registration", None)
    if raw_registration is not None:
        registration = dict(raw_registration or {})
    else:
        registration = dict(existing.get("registration") or {})
    for field in STUDENT_REGISTRATION_FIELDS:
        value = getattr(payload, field, None)
        if value is not None:
            registration[field] = _student_text(value)
    if registration or raw_registration is not None:
        result["registration"] = registration

    raw_pddikti = getattr(payload, "pddikti_ids", None)
    if raw_pddikti is not None:
        result["pddikti_ids"] = raw_pddikti or {}
    elif existing.get("pddikti_ids") is not None:
        result["pddikti_ids"] = existing.get("pddikti_ids")
    return result


@api_router.post("/students")
async def create_student(payload: StudentInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    if payload.class_id:
        await require_class_mutation_access(payload.class_id, user)
    if payload.status not in {"active", "inactive", "lulus", "cuti", "do"}:
        raise HTTPException(status_code=400, detail="Status mahasiswa tidak valid")
    identity = student_identity_values(payload.email, payload.nim, payload.nim, payload.whatsapp)
    existing = await db.users.find_one(
        student_identity_conflict_query(
            identity["email"],
            identity["nim"],
            identity["username"],
            identity["whatsapp"],
        ),
        {"_id": 0},
    )
    if existing:
        raise HTTPException(status_code=409, detail="Email, username, NIM, atau WhatsApp mahasiswa sudah digunakan")
    
    prodi_name = ""
    prodi_kode = ""
    if payload.prodi_id:
        p = await db.programs.find_one({"id": payload.prodi_id}, {"_id": 0, "nama": 1, "name": 1, "kode": 1, "code": 1})
        if p:
            prodi_name = p.get("nama") or p.get("name", "")
            prodi_kode = p.get("kode") or p.get("code", "")

    dosen_wali_name = ""
    if payload.dosen_wali_id:
        dw = await db.users.find_one({"id": payload.dosen_wali_id}, {"_id": 0, "name": 1})
        if dw:
            dosen_wali_name = dw.get("name", "")

    student_id = new_id()
    doc = {
        "id": student_id,
        "role": "student",
        "username": identity["username"],
        "nim": identity["nim"],
        "name": payload.name,
        "email": identity["email"],
        "whatsapp": identity["whatsapp"],
        "nik": _student_text(payload.nik),
        "nisn": _student_text(payload.nisn),
        "gender": _student_text(payload.gender) or "L",
        "agama": _student_text(payload.agama) or "Islam",
        "tempat_lahir": _student_text(payload.tempat_lahir),
        "tanggal_lahir": _student_text(payload.tanggal_lahir),
        "alamat": _student_text(payload.alamat),
        "kota": _student_text(payload.kota),
        "provinsi": _student_text(payload.provinsi),
        "kode_pos": _student_text(payload.kode_pos),
        "password_hash": hash_password(payload.password),
        "status": payload.status,
        "class_ids": [payload.class_id] if payload.class_id else [],
        "prodi_id": payload.prodi_id or "",
        "prodi_name": prodi_name,
        "prodi_kode": prodi_kode,
        "angkatan": payload.angkatan or "2024",
        "dosen_wali_id": payload.dosen_wali_id or "",
        "dosen_wali_name": dosen_wali_name,
        "parent_name": _student_text(payload.parent_name),
        "parent_phone": _student_text(payload.parent_phone),
        "parent_job": _student_text(payload.parent_job),
        "parent_address": _student_text(payload.parent_address),
        "created_at": now_iso(),
        "last_login_at": "",
    }
    doc.update(_student_profile_fields(payload))
    await db.users.insert_one(doc)
    if payload.class_id:
        await db.classes.update_one({"id": payload.class_id}, {"$addToSet": {"student_ids": student_id}})
    return public_doc(doc)


@api_router.put("/students/{student_id}")
@api_router.post("/students/{student_id}")
async def update_student(
    student_id: str,
    payload: StudentUpdateInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    existing = await db.users.find_one(
        {"id": student_id, "role": "student"},
        {"_id": 0},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Mahasiswa tidak ditemukan")
    status = payload.status if payload.status is not None else existing.get("status") or "active"
    if status not in {"active", "inactive", "lulus", "cuti", "do"}:
        raise HTTPException(status_code=400, detail="Status mahasiswa tidak valid")

    current_whatsapp = existing.get("whatsapp", "")
    identity = student_identity_values(
        payload.email,
        payload.nim,
        payload.nim,
        payload.whatsapp if payload.whatsapp is not None else current_whatsapp,
    )
    duplicate = await db.users.find_one(
        student_identity_conflict_query(
            identity["email"],
            identity["nim"],
            identity["username"],
            identity["whatsapp"],
            exclude_user_id=student_id,
        ),
        {"_id": 0, "id": 1},
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Email, username, NIM, atau WhatsApp mahasiswa sudah digunakan")

    target_prodi_id = payload.prodi_id if payload.prodi_id is not None else existing.get("prodi_id", "")
    prodi_name = existing.get("prodi_name", "")
    prodi_kode = existing.get("prodi_kode", "")
    if target_prodi_id:
        program = await db.programs.find_one(
            {"id": target_prodi_id},
            {"_id": 0, "nama": 1, "name": 1, "kode": 1, "code": 1},
        )
        if not program:
            raise HTTPException(status_code=400, detail="Program studi tidak ditemukan")
        prodi_name = program.get("nama") or program.get("name", "")
        prodi_kode = program.get("kode") or program.get("code", "")

    target_dosen_wali_id = payload.dosen_wali_id if payload.dosen_wali_id is not None else existing.get("dosen_wali_id", "")
    dosen_wali_name = existing.get("dosen_wali_name", "")
    if target_dosen_wali_id:
        dosen_wali = await db.users.find_one(
            {"id": target_dosen_wali_id, "role": "lecturer"},
            {"_id": 0, "name": 1},
        )
        if not dosen_wali:
            raise HTTPException(status_code=400, detail="Dosen wali tidak ditemukan")
        dosen_wali_name = dosen_wali.get("name", "")

    update = {
        "username": identity["username"],
        "nim": identity["nim"],
        "name": payload.name.strip(),
        "email": identity["email"],
        "whatsapp": identity["whatsapp"],
        "status": status,
        "prodi_id": target_prodi_id or "",
        "prodi_name": prodi_name,
        "prodi_kode": prodi_kode,
        "dosen_wali_id": target_dosen_wali_id or "",
        "dosen_wali_name": dosen_wali_name,
        "updated_at": now_iso(),
    }
    for field in (
        "whatsapp", "nik", "nisn", "gender", "agama", "tempat_lahir", "tanggal_lahir",
        "alamat", "kota", "provinsi", "kode_pos", "angkatan", "parent_name", "parent_phone",
        "parent_job", "parent_address",
    ):
        value = getattr(payload, field, None)
        if value is not None:
            update[field] = _student_text(value)
    if not update.get("gender"):
        update["gender"] = existing.get("gender", "L")
    if not update.get("agama"):
        update["agama"] = existing.get("agama", "Islam")
    update.update(_student_profile_fields(payload, existing))
    await db.users.update_one({"id": student_id}, {"$set": update})

    # Keep denormalized enrollment/request labels in sync with the account.
    if existing.get("name") != update["name"] or existing.get("nim") != update["nim"]:
        await db.enrollment_requests.update_many(
            {"student_id": student_id},
            {"$set": {"student_name": update["name"], "student_nim": update["nim"]}},
        )

    return public_doc(await db.users.find_one({"id": student_id}, {"_id": 0}))


@api_router.post("/classes/{class_id}/students/import")
async def import_students(
    class_id: str,
    file: UploadFile = File(...),
    default_password: str = Form(""),
    prodi_id: str = Form(""),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    class_doc = await require_class_mutation_access(class_id, user)

    prodi_name = ""
    prodi_kode = ""
    if prodi_id:
        p = await db.programs.find_one({"id": prodi_id}, {"_id": 0, "nama": 1, "name": 1, "kode": 1, "code": 1})
        if p:
            prodi_name = p.get("nama") or p.get("name", "")
            prodi_kode = p.get("kode") or p.get("code", "")

    content = await file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    mapping = {name: idx for idx, name in enumerate(headers)}
    default_import_password = default_password.strip()

    def row_value(row: tuple[Any, ...], keys: List[str], fallback_index: Optional[int] = None) -> str:
        index = next((mapping[key] for key in keys if key in mapping), fallback_index)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    created = 0
    skipped = 0
    skipped_reasons = {
        "invalid_row": 0,
        "identity_conflict": 0,
        "invalid_password": 0,
    }
    conflicts: List[Dict[str, Any]] = []
    password_from_column = 0
    password_from_default = 0
    password_from_nim = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        nim = row_value(row, ["nim"], 0)
        name = row_value(row, ["nama", "name"], 1)
        email = row_value(row, ["email"], 2).lower()
        whatsapp = row_value(row, ["whatsapp", "wa", "no hp", "nomor hp"], 3)
        if not nim or not name or not email:
            skipped += 1
            skipped_reasons["invalid_row"] += 1
            continue
        identity = student_identity_values(email, nim, nim, whatsapp)
        conflict = await db.users.find_one(
            student_identity_conflict_query(
                identity["email"],
                identity["nim"],
                identity["username"],
                identity["whatsapp"],
            ),
            {"_id": 0, "id": 1},
        )
        if conflict:
            skipped += 1
            skipped_reasons["identity_conflict"] += 1
            conflicts.append(
                {
                    "row": row_number,
                    "nim": identity["nim"],
                    "reason": "Email, username, NIM, atau WhatsApp sudah terdaftar",
                }
            )
            continue
        row_password = row_value(row, ["password", "pass", "sandi", "kata sandi", "kata_sandi"])
        password = row_password or default_import_password or identity["nim"]
        if len(password) < 3:
            skipped += 1
            skipped_reasons["invalid_password"] += 1
            continue
        if row_password:
            password_from_column += 1
        elif default_import_password:
            password_from_default += 1
        else:
            password_from_nim += 1
        student_id = new_id()
        await db.users.insert_one(
            {
                "id": student_id,
                "role": "student",
                "username": identity["username"],
                "nim": identity["nim"],
                "name": name,
                "email": identity["email"],
                "whatsapp": identity["whatsapp"],
                "password_hash": hash_password(password),
                "status": "active",
                "class_ids": [class_id],
                "prodi_id": prodi_id,
                "prodi_name": prodi_name,
                "prodi_kode": prodi_kode,
                "created_at": now_iso(),
                "last_login_at": "",
            }
        )
        await db.classes.update_one({"id": class_id}, {"$addToSet": {"student_ids": student_id}})
        created += 1
    workbook.close()
    return {
        "created": created,
        "skipped": skipped,
        "skipped_reasons": skipped_reasons,
        "conflicts": conflicts[:100],
        "password_from_column": password_from_column,
        "password_from_default": password_from_default,
        "password_from_nim": password_from_nim,
    }


@api_router.get("/materials")
async def list_materials(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"class_id": {"$in": user.get("class_ids", [])}, "is_active": True}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    elif user.get("role") == "staff":
        query = {"id": "__staff_learning_access_disabled__"}
    materials = await db.materials.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    class_ids = None if user["role"] != "student" else [material.get("class_id", "") for material in materials]
    meeting_labels = await material_meeting_label_map(class_ids)
    for material in materials:
        material["meeting"] = meeting_labels.get(material.get("id", ""), material.get("meeting", ""))
    return await enrich_materials_batch(materials)


@api_router.post("/materials")
async def create_material(payload: MaterialInput, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_mutation_access(payload.class_id, user)
    await require_rps_complete(payload.class_id)
    async with _material_creation_lock:
        meeting_number = await db.materials.count_documents({"class_id": payload.class_id}) + 1
        doc = normalized_material_payload(payload)
        rps_meeting_number = payload.rps_meeting_number or None
        doc.update(
            {
                "id": new_id(),
                "meeting": f"Pertemuan {rps_meeting_number}" if rps_meeting_number else f"Pertemuan {meeting_number}",
                "rps_meeting_number": rps_meeting_number,
                "created_at": now_iso(),
                "created_by": user["id"],
                "lecturer_id": class_doc.get("lecturer_id", user["id"]),
                "lecturer_name": class_doc.get("lecturer_name", user.get("name", "Dosen")),
            }
        )
        await db.materials.insert_one(doc)
    return public_doc(await enrich_material_payload(doc))


@api_router.post("/materials/google-meet")
async def generate_material_google_meet(payload: GoogleMeetInput, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_mutation_access(payload.class_id, user)
    settings = await get_google_drive_settings(mask=False)
    if not settings.get("google_meet_enabled"):
        raise HTTPException(status_code=503, detail="Google Meet belum diaktifkan oleh Admin Kampus")
    try:
        meet = await asyncio.to_thread(create_google_meet_for_app_user_sync, settings, user.get("email", ""))
    except Exception as exc:
        logger.warning("Pembuatan Google Meet gagal: %s", exc)
        raise HTTPException(status_code=503, detail=google_meet_error_message(exc)) from exc
    return {"ok": True, **meet}


@api_router.put("/materials/{material_id}")
async def update_material(material_id: str, payload: MaterialInput, user: Dict[str, Any] = Depends(require_admin)):
    existing = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    await require_class_mutation_access(existing.get("class_id", ""), user)
    class_doc = await require_class_mutation_access(payload.class_id, user)
    await require_rps_complete(payload.class_id)
    existing_attachment = existing.get("attachment") if isinstance(existing.get("attachment"), dict) else {}
    attachment_id = existing_attachment.get("file_id") or existing_attachment.get("id")
    attachment_url = local_file_urls(attachment_id)["file_url"] if attachment_id else ""
    remove_attachment = bool(attachment_id and payload.file_url != attachment_url)
    update = normalized_material_payload(payload)
    update.pop("meeting", None)
    update["updated_at"] = now_iso()
    operation: Dict[str, Any] = {"$set": update}
    if remove_attachment:
        operation["$unset"] = {"attachment": ""}
    await db.materials.update_one({"id": material_id}, operation)
    if remove_attachment:
        await delete_stored_files({"id": attachment_id, "record_type": "material_attachment"})
    elif attachment_id:
        await db.stored_files.update_one(
            {"id": attachment_id, "record_type": "material_attachment"},
            {"$set": {"material_class_id": payload.class_id}},
        )
    updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
    meeting_labels = await material_meeting_label_map([updated.get("class_id", "")])
    updated["meeting"] = meeting_labels.get(material_id, updated.get("meeting", ""))
    return public_doc(await enrich_material_payload(updated))


@api_router.post("/materials/{material_id}/attachment")
async def upload_material_attachment(
    material_id: str,
    background_tasks: BackgroundTasks,
    attachment: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_admin),
):
    material = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    await require_class_mutation_access(material.get("class_id", ""), user)
    ext = attachment.filename.rsplit(".", 1)[-1].lower() if attachment.filename and "." in attachment.filename else ""
    allowed = {"pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx", "csv", "txt", "zip", "png", "jpg", "jpeg", "webp"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="File materi hanya dokumen, gambar, atau ZIP")
    content = await attachment.read(25 * 1024 * 1024 + 1)
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran file materi maksimal 25 MB")
    await attachment.seek(0)
    class_doc = await db.classes.find_one({"id": material.get("class_id", "")}, {"_id": 0}) or {}
    class_doc = await enrich_class_payload(class_doc) if class_doc else {}
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        class_doc.get("course_name", "Mata Kuliah"),
        class_doc.get("name", "Kelas"),
        material.get("title", "Materi"),
        "Lampiran Materi",
    ]
    file_doc = await save_uploaded_file_record(
        attachment,
        hierarchy,
        "DOSEN",
        user.get("name", "Dosen"),
        user["id"],
        record_type="material_attachment",
        background_tasks=background_tasks,
        async_drive=True,
        lecturer_id=class_doc.get("lecturer_id", ""),
    )
    await db.stored_files.update_one(
        {"id": file_doc["id"]},
        {"$set": {"material_id": material_id, "material_class_id": material.get("class_id", "")}},
    )
    old_attachment = material.get("attachment") if isinstance(material.get("attachment"), dict) else {}
    old_file_id = old_attachment.get("file_id") or old_attachment.get("id")
    await db.materials.update_one(
        {"id": material_id},
        {"$set": {"attachment": file_doc, "file_url": file_doc["file_url"], "updated_at": now_iso()}},
    )
    if old_file_id and old_file_id != file_doc["id"]:
        await delete_stored_files({"id": old_file_id, "record_type": "material_attachment"})
    updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
    meeting_labels = await material_meeting_label_map([updated.get("class_id", "")])
    updated["meeting"] = meeting_labels.get(material_id, updated.get("meeting", ""))
    return public_doc(await enrich_material_payload(updated))


@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str, user: Dict[str, Any] = Depends(require_admin)):
    existing = await db.materials.find_one({"id": material_id}, {"_id": 0, "id": 1, "attachment": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    full_material = await db.materials.find_one({"id": material_id}, {"_id": 0, "class_id": 1}) or {}
    await require_class_mutation_access(full_material.get("class_id", ""), user)
    comments = await db.comments.find({"material_id": material_id}, {"_id": 0, "attachment.id": 1, "attachment.file_id": 1}).to_list(5000)
    attachment_ids = []
    for comment in comments:
        attachment = comment.get("attachment") or {}
        file_id = attachment.get("id") or attachment.get("file_id")
        if file_id:
            attachment_ids.append(file_id)
    stored_file_query: Dict[str, Any] = {"record_type": "comment_attachment", "discussion_material_id": material_id}
    if attachment_ids:
        stored_file_query = {
            "record_type": "comment_attachment",
            "$or": [{"discussion_material_id": material_id}, {"id": {"$in": attachment_ids}}],
        }
    deleted_files = await delete_stored_files(stored_file_query)
    material_attachment = existing.get("attachment") if isinstance(existing.get("attachment"), dict) else {}
    material_attachment_id = material_attachment.get("file_id") or material_attachment.get("id")
    material_file_query: Dict[str, Any] = {"record_type": "material_attachment", "material_id": material_id}
    if material_attachment_id:
        material_file_query = {
            "record_type": "material_attachment",
            "$or": [{"material_id": material_id}, {"id": material_attachment_id}],
        }
    deleted_files += await delete_stored_files(material_file_query)
    deleted_comments = deleted_count(await db.comments.delete_many({"material_id": material_id}))
    unlinked_assignments = modified_count(
        await db.assignments.update_many(
            {"material_id": material_id},
            {"$set": {"material_id": "", "material_link_removed_at": now_iso()}},
        )
    )
    await db.materials.delete_one({"id": material_id})
    return {
        "ok": True,
        "comments_deleted": deleted_comments,
        "attachments_deleted": deleted_files,
        "assignments_unlinked": unlinked_assignments,
    }


async def discussion_material_for_user(material_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    material = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Materi tidak ditemukan")
    if user.get("role") == "student":
        if material.get("class_id") not in user.get("class_ids", []) or not material.get("is_active", True):
            raise HTTPException(status_code=403, detail="Diskusi hanya dapat diakses anggota kelas")
    elif user.get("role") == "lecturer":
        await require_class_access(material.get("class_id", ""), user)
    return material


async def validate_comment_parent(parent_id: str, material_id: str) -> None:
    if not parent_id:
        return
    parent = await db.comments.find_one({"id": parent_id, "material_id": material_id}, {"_id": 0, "id": 1})
    if not parent:
        raise HTTPException(status_code=400, detail="Komentar balasan tidak valid")


@api_router.get("/materials/{material_id}/comments")
async def list_comments(material_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    await discussion_material_for_user(material_id, user)
    comments = await db.comments.find({"material_id": material_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    admin_author_ids = list(set(
        c.get("author_id", "") for c in comments
        if c.get("author_role") in {"admin", "lecturer"} and c.get("author_id")
    ))
    admin_users_map: Dict[str, Dict[str, Any]] = {}
    if admin_author_ids:
        admin_docs = await db.users.find({"id": {"$in": admin_author_ids}}, {"_id": 0}).to_list(100)
        admin_users_map = {u["id"]: u for u in admin_docs}
    displayed_admin_names: Dict[str, str] = {}
    for comment in comments:
        if comment.get("author_role") not in {"admin", "lecturer"}:
            continue
        author_id = comment.get("author_id", "")
        if author_id not in displayed_admin_names:
            author = admin_users_map.get(author_id) or {
                "id": author_id,
                "role": "admin",
                "name": comment.get("author_name", ""),
            }
            displayed_admin_names[author_id] = (await chat_contact_payload(author))["name"]
        comment["author_name"] = displayed_admin_names[author_id]
    return comments


@api_router.post("/comments")
async def create_comment(payload: CommentInput, user: Dict[str, Any] = Depends(get_current_user)):
    await discussion_material_for_user(payload.material_id, user)
    material = await db.materials.find_one({"id": payload.material_id}, {"_id": 0, "class_id": 1}) or {}
    await require_class_mutation_access(material.get("class_id", ""), user)
    await validate_comment_parent(payload.parent_id, payload.material_id)
    if not payload.content.strip():
        raise HTTPException(status_code=400, detail="Isi komentar diperlukan")
    author = await chat_contact_payload(user)
    doc = payload.model_dump()
    doc["content"] = doc["content"].strip()
    doc.update(
        {
            "id": new_id(),
            "author_id": user["id"],
            "author_name": author["name"],
            "author_role": user["role"],
            "is_pinned": False,
            "created_at": now_iso(),
        }
    )
    await db.comments.insert_one(doc)
    return public_doc(doc)


@api_router.post("/materials/{material_id}/comments")
async def create_material_comment_with_image(
    material_id: str,
    content: str = Form(""),
    parent_id: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
    image: Optional[UploadFile] = File(None),
    user: Dict[str, Any] = Depends(get_current_user),
):
    material = await discussion_material_for_user(material_id, user)
    await require_class_mutation_access(material.get("class_id", ""), user)
    await validate_comment_parent(parent_id, material_id)
    upload = attachment or image
    text = content.strip()
    if not text and not upload:
        raise HTTPException(status_code=400, detail="Isi komentar atau lampiran diperlukan")
    author = await chat_contact_payload(user)
    attachment_doc = None
    if upload:
        ext = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        allowed = {"jpg", "jpeg", "png", "webp", "gif", "pdf", "doc", "docx", "xls", "xlsx", "txt", "csv"}
        if ext not in allowed:
            raise HTTPException(status_code=400, detail="Lampiran diskusi hanya gambar atau dokumen")
        content_bytes = await upload.read(10 * 1024 * 1024 + 1)
        if len(content_bytes) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran lampiran diskusi maksimal 10 MB")
        await upload.seek(0)
        hierarchy = ["Diskusi", safe_path_segment(material.get("title", "Materi"))]
        attachment_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            user.get("nim", user.get("role", "user")),
            user.get("name", "Pengguna"),
            user["id"],
            record_type="comment_attachment",
            sync_drive=False,
            lecturer_id=material.get("lecturer_id", ""),
        )
        await db.stored_files.update_one(
            {"id": attachment_doc["id"]},
            {"$set": {"discussion_class_id": material.get("class_id", ""), "discussion_material_id": material_id}},
        )
    doc = {
        "id": new_id(),
        "material_id": material_id,
        "content": text,
        "parent_id": parent_id,
        "author_id": user["id"],
        "author_name": author["name"],
        "author_role": user["role"],
        "is_pinned": False,
        "attachment": attachment_doc,
        "created_at": now_iso(),
    }
    await db.comments.insert_one(doc)
    return public_doc(doc)


@api_router.post("/comments/{comment_id}/pin")
async def pin_comment(comment_id: str, user: Dict[str, Any] = Depends(require_admin)):
    comment = await db.comments.find_one({"id": comment_id}, {"_id": 0, "material_id": 1}) or {}
    await discussion_material_for_user(comment.get("material_id", ""), user)
    material = await db.materials.find_one({"id": comment.get("material_id", "")}, {"_id": 0, "class_id": 1}) or {}
    await require_class_mutation_access(material.get("class_id", ""), user)
    await db.comments.update_one({"id": comment_id}, {"$set": {"is_pinned": True}})
    return {"ok": True}


@api_router.get("/assignments")
async def list_assignments(background_tasks: BackgroundTasks, user: Dict[str, Any] = Depends(get_current_user)):
    await dispatch_due_assignment_notifications(background_tasks)
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"class_id": {"$in": user.get("class_ids", [])}, "is_active": True}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    elif user.get("role") == "staff":
        query = {"id": "__staff_learning_access_disabled__"}
    assignments = await db.assignments.find(query, {"_id": 0}).sort("deadline", 1).to_list(1000)
    class_ids = list({item.get("class_id", "") for item in assignments if item.get("class_id")})
    class_docs = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0}).to_list(1000) if class_ids else []
    classes_by_id = {item["id"]: item for item in class_docs}
    for class_doc in classes_by_id.values():
        class_doc["status_label"] = class_status_label(class_doc.get("status", ""))
    settings = await get_app_settings_cached()
    creator_ids = [item.get("created_by") for item in assignments if item.get("created_by")]
    creators = await db.users.find({"id": {"$in": creator_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(200) if creator_ids else []
    creator_names = {item["id"]: item.get("name", "") for item in creators}
    if user["role"] == "student":
        assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
        assignment_ids = [a["id"] for a in assignments]
        my_submissions = await db.submissions.find(
            {"assignment_id": {"$in": assignment_ids}, "student_id": user["id"]}, {"_id": 0}
        ).to_list(1000) if assignment_ids else []
        subs_by_assignment: Dict[str, Dict[str, Any]] = {}
        for s in my_submissions:
            enrich_submission_file_urls(s)
            subs_by_assignment[s["assignment_id"]] = s
        for assignment in assignments:
            assignment["my_submission"] = subs_by_assignment.get(assignment["id"])
    for assignment in assignments:
        class_doc = classes_by_id.get(assignment.get("class_id", ""), {})
        ay = class_doc.get("academic_year") or ""
        sem = class_doc.get("semester") or ""
        assignment["academic_year"] = ay
        assignment["semester"] = sem
        if ay and sem:
            assignment["period_name"] = f"{ay} {sem}"
        elif ay:
            assignment["period_name"] = ay
        else:
            assignment["period_name"] = "Dokumen Evaluasi / SIAP"

        assignment["class_status"] = class_doc.get("status", "")
        assignment["class_status_label"] = class_doc.get("status_label", class_status_label(class_doc.get("status", "")))
        assignment["class_read_only"] = class_is_read_only(class_doc)
        assignment["class_allows_learning"] = class_allows_learning(class_doc)
        assignment["class_allows_grading"] = class_allows_grading(class_doc)
        assignment["assessment_category"] = normalize_assessment_category(assignment.get("assessment_category", "tugas"))
        assignment["publish_status"] = assignment_publish_status(assignment)
        assignment["max_file_size_mb"] = assignment_max_file_size_mb(assignment)
        assignment["max_submission_size_mb"] = assignment["max_file_size_mb"]
        lecturer_name = str(assignment.get("lecturer_name") or "").strip()
        if not lecturer_name or lecturer_name.lower() == "dosen admin":
            lecturer_name = creator_names.get(assignment.get("created_by", ""), "") or settings.get("lecturer_name", "") or "Dosen"
        if lecturer_name.lower() == "dosen admin":
            lecturer_name = "Dosen"
        assignment["lecturer_name"] = lecturer_name
    return assignments


@api_router.post("/assignments")
async def create_assignment(payload: AssignmentInput, background_tasks: BackgroundTasks, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_mutation_access(payload.class_id, user)
    await require_rps_complete(payload.class_id)
    doc = payload.model_dump()
    doc["assessment_category"] = normalize_assessment_category(doc.get("assessment_category"))
    doc["attachment_link"] = str(doc.get("attachment_link") or "").strip()
    doc["max_file_size_mb"] = assignment_max_file_size_mb(doc)
    doc["deadline"] = normalize_optional_datetime(doc.get("deadline", ""), "Deadline")
    if not doc["deadline"]:
        raise HTTPException(status_code=400, detail="Deadline wajib diisi")
    doc["published_at"] = normalize_optional_datetime(doc.get("published_at", ""), "Jadwal tayang")
    publish_at = parse_iso_datetime(doc["published_at"])
    deadline = parse_iso_datetime(doc["deadline"])
    if publish_at and deadline and publish_at > deadline:
        raise HTTPException(status_code=400, detail="Jadwal tayang tidak boleh setelah deadline")
    lecturer_name = str(user.get("name") or class_doc.get("lecturer_name") or "Dosen").strip()
    if lecturer_name.lower() == "dosen admin":
        lecturer_name = str(user.get("name") or "Dosen").strip()
    if lecturer_name.lower() == "dosen admin":
        lecturer_name = "Dosen"
    doc.update(
        {
            "id": new_id(),
            "course_id": class_doc["course_id"],
            "course_name": class_doc.get("course_name", ""),
            "class_name": class_doc["name"],
            "lecturer_id": class_doc.get("lecturer_id", user["id"]),
            "lecturer_name": lecturer_name,
            "created_at": now_iso(),
            "created_by": user["id"],
        }
    )
    await db.assignments.insert_one(doc)
    if assignment_is_published(doc):
        sent_at = now_iso()
        await db.assignments.update_one({"id": doc["id"]}, {"$set": {"published_notification_sent_at": sent_at}})
        doc["published_notification_sent_at"] = sent_at
        await send_assignment_publication_notifications(doc, class_doc, background_tasks)
    response = doc.copy()
    response["publish_status"] = assignment_publish_status(response)
    return public_doc(response)


@api_router.put("/assignments/{assignment_id}")
async def update_assignment(
    assignment_id: str,
    payload: AssignmentInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    existing = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    await require_class_mutation_access(existing.get("class_id", ""), user)
    class_doc = await require_class_mutation_access(payload.class_id, user)
    await require_rps_complete(payload.class_id)
    doc = payload.model_dump()
    doc["assessment_category"] = normalize_assessment_category(doc.get("assessment_category"))
    doc["attachment_link"] = str(doc.get("attachment_link") or "").strip()
    doc["max_file_size_mb"] = assignment_max_file_size_mb(doc)
    doc["deadline"] = normalize_optional_datetime(doc.get("deadline", ""), "Deadline")
    if not doc["deadline"]:
        raise HTTPException(status_code=400, detail="Deadline wajib diisi")
    doc["published_at"] = normalize_optional_datetime(doc.get("published_at", ""), "Jadwal tayang")
    publish_at = parse_iso_datetime(doc["published_at"])
    deadline = parse_iso_datetime(doc["deadline"])
    if publish_at and deadline and publish_at > deadline:
        raise HTTPException(status_code=400, detail="Jadwal tayang tidak boleh setelah deadline")

    if existing.get("deadline"):
        await db.submissions.update_many(
            {"assignment_id": assignment_id, "assignment_deadline": {"$exists": False}},
            {"$set": {"assignment_deadline": existing["deadline"]}},
        )
    await db.submissions.update_many(
        {"assignment_id": assignment_id, "assignment_late_penalty_per_day": {"$exists": False}},
        {"$set": {"assignment_late_penalty_per_day": existing.get("late_penalty_per_day", 0)}},
    )

    doc.update(
        {
            "course_id": class_doc["course_id"],
            "course_name": class_doc.get("course_name", ""),
            "class_name": class_doc["name"],
            "updated_at": now_iso(),
            "updated_by": user["id"],
        }
    )
    await db.assignments.update_one({"id": assignment_id}, {"$set": doc})
    updated = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if (
        updated
        and not assignment_is_published(existing)
        and assignment_is_published(updated)
        and not existing.get("published_notification_sent_at")
    ):
        sent_at = now_iso()
        await db.assignments.update_one({"id": assignment_id}, {"$set": {"published_notification_sent_at": sent_at}})
        updated["published_notification_sent_at"] = sent_at
        await send_assignment_publication_notifications(updated, class_doc, background_tasks)
    response = updated or doc
    response["publish_status"] = assignment_publish_status(response)
    response["max_file_size_mb"] = assignment_max_file_size_mb(response)
    response["max_submission_size_mb"] = response["max_file_size_mb"]
    return public_doc(response)


@api_router.post("/assignments/{assignment_id}/attachments")
async def upload_assignment_attachments(
    assignment_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    assignment = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    await require_class_mutation_access(assignment.get("class_id", ""), user)
    _, files = await multipart_uploads(request, ["files", "file", "files[]"])
    if not files:
        raise HTTPException(status_code=400, detail="Minimal satu lampiran soal harus diunggah")
    class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
    allowed = {"pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx", "csv", "txt", "zip", "png", "jpg", "jpeg", "webp"}
    saved = []
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        assignment.get("course_name", "Mata Kuliah"),
        assignment.get("class_name", "Kelas"),
        assignment.get("title", "Tugas"),
        "Lampiran Soal",
    ]
    for upload in files:
        ext = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        if ext not in allowed:
            raise HTTPException(status_code=400, detail="Lampiran soal hanya dokumen, gambar, atau ZIP")
        content = await upload.read(25 * 1024 * 1024 + 1)
        if len(content) > 25 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran setiap lampiran soal maksimal 25 MB")
        await upload.seek(0)
    for upload in files:
        file_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            "DOSEN",
            user.get("name", "Dosen"),
            user["id"],
            assignment_id=assignment_id,
            record_type="assignment_attachment",
            background_tasks=background_tasks,
            async_drive=True,
            lecturer_id=assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        )
        saved.append(file_doc)
    await db.assignments.update_one({"id": assignment_id}, {"$push": {"attachments": {"$each": saved}}})
    updated = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    return {"attachments": saved, "assignment": updated}


@api_router.get("/submissions")
async def list_submissions(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"student_id": user["id"]}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    elif user.get("role") == "staff":
        query = {"id": "__staff_learning_access_disabled__"}
    submissions = await db.submissions.find(query, {"_id": 0}).sort("submitted_at", -1).to_list(2000)
    for item in submissions:
        if isinstance(item.get("grade"), (int, float)) and not item.get("grade_predicate"):
            item["grade_predicate"] = await calculate_grade_predicate(item["grade"], item.get("class_id", ""))
        enrich_submission_file_urls(item)
    return submissions


@api_router.get("/chat/contacts")
async def chat_contacts(q: str = "", user: Dict[str, Any] = Depends(get_current_user)):
    query = q.strip().lower()
    users: List[Dict[str, Any]] = []
    if query:
        candidates = await db.users.find(
            {
                "id": {"$ne": user["id"]},
                "status": {"$ne": "deleted"},
            },
            {"_id": 0, "password_hash": 0},
        ).sort("name", 1).to_list(5000)
        users = [item for item in candidates if chat_contact_matches_query(item, query)][:20]
    else:
        messages = await db.chat_messages.find(
            {"participant_ids": user["id"]}, {"_id": 0, "participant_ids": 1}
        ).sort("created_at", -1).to_list(100)
        recent_ids: List[str] = []
        for message in messages:
            for participant_id in message.get("participant_ids", []):
                if participant_id != user["id"] and participant_id not in recent_ids:
                    recent_ids.append(participant_id)
        if recent_ids:
            user_docs = await db.users.find(
                {"id": {"$in": recent_ids}}, {"_id": 0, "password_hash": 0}
            ).to_list(50)
            by_id = {item["id"]: item for item in user_docs}
            users = [by_id[item_id] for item_id in recent_ids if item_id in by_id]
    return [
        await chat_contact_view_payload(item, user)
        for item in users
    ]


@api_router.get("/chat/lecturers")
async def chat_lecturers(user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        return []
    class_docs = await db.classes.find(
        {"id": {"$in": user.get("class_ids", [])}, "status": {"$ne": "deleted"}}, {"_id": 0, "lecturer_id": 1}
    ).to_list(1000)
    lecturer_ids = list({item.get("lecturer_id") for item in class_docs if item.get("lecturer_id")})
    lecturers = await db.users.find(
        {"id": {"$in": lecturer_ids}, "role": {"$in": ["admin", "lecturer"]}, "status": {"$ne": "deleted"}},
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(100)
    return [
        await chat_contact_view_payload(item, user)
        for item in lecturers
    ]


@api_router.post("/chat/users/{other_user_id}/read")
async def mark_chat_read_endpoint(
    other_user_id: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    other = await db.users.find_one({"id": other_user_id}, {"_id": 0, "id": 1})
    if not other or other_user_id == user["id"]:
        raise HTTPException(status_code=404, detail="Pengguna chat tidak ditemukan")
    await mark_chat_read(user["id"], other_user_id)
    return {"unread_count": 0}


@api_router.get("/chat/users/{other_user_id}/messages")
async def list_chat_messages(other_user_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    other = await db.users.find_one({"id": other_user_id}, {"_id": 0, "password_hash": 0})
    if not other or other_user_id == user["id"]:
        raise HTTPException(status_code=404, detail="Pengguna chat tidak ditemukan")
    await mark_chat_read(user["id"], other_user_id)
    conversation_id = chat_conversation_id(user["id"], other_user_id)
    messages = await db.chat_messages.find(
        {"conversation_id": conversation_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(500)
    return {
        "contact": {
            **await chat_contact_view_payload(other, user),
        },
        "messages": messages,
    }


@api_router.post("/chat/messages")
async def send_chat_message(
    recipient_id: str = Form(...),
    content: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
    user: Dict[str, Any] = Depends(get_current_user),
):
    recipient = await db.users.find_one({"id": recipient_id}, {"_id": 0, "password_hash": 0})
    if not recipient or recipient_id == user["id"]:
        raise HTTPException(status_code=404, detail="Penerima chat tidak ditemukan")
    text = content.strip()
    if len(text) > 4000:
        raise HTTPException(status_code=400, detail="Pesan maksimal 4000 karakter")
    if not text and not attachment:
        raise HTTPException(status_code=400, detail="Isi pesan atau foto diperlukan")
    participant_ids = sorted([user["id"], recipient_id])
    conversation_id = chat_conversation_id(*participant_ids)
    attachment_doc: Optional[Dict[str, Any]] = None
    if attachment:
        if not (attachment.content_type or "").lower().startswith("image/"):
            raise HTTPException(status_code=400, detail="Lampiran chat hanya berupa foto")
        attachment_content = await attachment.read(5 * 1024 * 1024 + 1)
        if len(attachment_content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran foto maksimal 5 MB")
        await attachment.seek(0)
        attachment_doc = await save_uploaded_file_record(
            attachment,
            ["Chat", conversation_id],
            user.get("username") or user.get("nim", "USER"),
            user.get("name", "Pengguna"),
            user["id"],
            record_type="chat_image",
            sync_drive=False,
        )
        await db.stored_files.update_one(
            {"id": attachment_doc["id"]}, {"$set": {"chat_participant_ids": participant_ids}}
        )
    message = {
        "id": new_id(),
        "conversation_id": conversation_id,
        "participant_ids": participant_ids,
        "sender_id": user["id"],
        "recipient_id": recipient_id,
        "content": text,
        "attachment": attachment_doc,
        "created_at": now_iso(),
    }
    await db.chat_messages.insert_one(message)
    response = public_doc(message.copy())
    await chat_connections.send_to_users(set(participant_ids), {"type": "message", "message": response})
    return response


async def stored_file_context(
    file_id: str,
    token: str = "",
    authorization: Optional[str] = Header(None),
) -> tuple[Dict[str, Any], Dict[str, Any], Path, bool]:
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        file_doc = await db.drive_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")

    auth_token = token.strip()
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization.replace("Bearer ", "", 1).strip()

    is_public = file_doc.get("record_type") in {
        "avatar",
        "app_logo",
        "campus_logo",
        "branding_logo",
        "pmb_landing_logo",
        "kop_header",
        "kop_footer",
    }
    user: Dict[str, Any] = {}
    if auth_token:
        session = await db.sessions.find_one({"token": auth_token}, {"_id": 0})
        if session:
            user = await find_user(session["user_id"]) or {}
        elif not is_public:
            raise HTTPException(status_code=401, detail="Sesi tidak ditemukan")
    elif not is_public:
        raise HTTPException(status_code=401, detail="Token diperlukan")
    if user.get("role") == "student" and file_doc.get("submission_id") and file_doc.get("uploaded_by") != user.get("id"):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke file ini")
    if (
        file_doc.get("record_type") == "physical_document"
        and user.get("role") == "student"
        and str(file_doc.get("student_id") or file_doc.get("uploaded_by") or "") != str(user.get("id") or "")
    ):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke dokumen fisik ini")
    if file_doc.get("record_type") == "rps_document" and user.get("role") == "student":
        rps_class_id = str(file_doc.get("class_id") or "").strip()
        if not rps_class_id or rps_class_id not in user.get("class_ids", []):
            raise HTTPException(status_code=403, detail="Dokumen RPS hanya dapat diakses oleh anggota kelas")
    if file_doc.get("record_type") == "chat_image" and user.get("id") not in file_doc.get("chat_participant_ids", []):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke foto chat ini")
    if file_doc.get("record_type") == "comment_attachment" and user.get("role") == "student":
        discussion_class_id = file_doc.get("discussion_class_id", "")
        if not discussion_class_id:
            comment = await db.comments.find_one({"attachment.file_id": file_id}, {"_id": 0, "material_id": 1}) or {}
            material = await db.materials.find_one({"id": comment.get("material_id", "")}, {"_id": 0, "class_id": 1}) or {}
            discussion_class_id = material.get("class_id", "")
        if not discussion_class_id or discussion_class_id not in user.get("class_ids", []):
            raise HTTPException(status_code=403, detail="Lampiran diskusi hanya untuk anggota kelas")
    if file_doc.get("record_type") == "material_attachment" and user.get("role") == "student":
        material_class_id = file_doc.get("material_class_id", "")
        if not material_class_id:
            material = await db.materials.find_one({"attachment.file_id": file_id}, {"_id": 0, "class_id": 1}) or {}
            material_class_id = material.get("class_id", "")
        if not material_class_id or material_class_id not in user.get("class_ids", []):
            raise HTTPException(status_code=403, detail="Lampiran materi hanya untuk anggota kelas")
    if user.get("role") == "lecturer" and file_doc.get("record_type") != "chat_image":
        academic_class_id = file_doc.get("material_class_id") or file_doc.get("discussion_class_id") or file_doc.get("class_id", "")
        if not academic_class_id and file_doc.get("submission_id"):
            submission = await db.submissions.find_one({"id": file_doc["submission_id"]}, {"_id": 0, "class_id": 1}) or {}
            academic_class_id = submission.get("class_id", "")
        if not academic_class_id and file_doc.get("assignment_id"):
            assignment = await db.assignments.find_one({"id": file_doc["assignment_id"]}, {"_id": 0, "class_id": 1}) or {}
            academic_class_id = assignment.get("class_id", "")
        if academic_class_id:
            await require_class_access(academic_class_id, user)
    path = resolved_stored_file_path(file_doc)
    if path:
        if str(file_doc.get("local_path") or "") != str(path):
            await db.stored_files.update_one(
                {"id": file_doc.get("id")},
                {"$set": {"local_path": str(path), "local_available": True}},
            )
            file_doc["local_path"] = str(path)
            file_doc["local_available"] = True
        return user, file_doc, path, False
    drive_file_id = str(file_doc.get("drive_file_id") or "").strip()
    if drive_file_id:
        settings = await get_google_drive_settings(mask=False)
        if not google_drive_upload_enabled(settings):
            raise HTTPException(
                status_code=503,
                detail="Salinan lokal sudah tidak tersedia dan Google Drive belum aktif.",
            )
        try:
            path = await asyncio.to_thread(
                download_drive_file_to_temp_sync,
                drive_file_id,
                file_doc.get("file_name", ""),
                settings,
            )
        except Exception as exc:
            logger.warning("File %s gagal dibaca dari Google Drive: %s", file_id, exc)
            raise HTTPException(
                status_code=503,
                detail=f"File gagal diambil dari Google Drive: {google_drive_error_message(exc)}",
            ) from exc
        return user, file_doc, path, True
    raise HTTPException(status_code=404, detail="File tidak tersedia di server atau Google Drive")


@api_router.get("/files/{file_id}/download")
async def download_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path, temporary = await stored_file_context(file_id, token, authorization)
    return FileResponse(
        path,
        media_type=file_doc.get("mime_type") or "application/octet-stream",
        filename=file_doc.get("file_name") or path.name,
        background=BackgroundTask(remove_temporary_file, path) if temporary else None,
    )


@api_router.get("/files/{file_id}/inline")
async def inline_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path, temporary = await stored_file_context(file_id, token, authorization)
    return FileResponse(
        path,
        media_type=file_doc.get("mime_type") or "application/octet-stream",
        filename=file_doc.get("file_name") or path.name,
        content_disposition_type="inline",
        background=BackgroundTask(remove_temporary_file, path) if temporary else None,
    )


@api_router.get("/files/{file_id}/preview")
async def preview_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path, temporary = await stored_file_context(file_id, token, authorization)
    try:
        kind = preview_kind(file_doc, path)
        response: Dict[str, Any] = {
            "id": file_id,
            "file_name": file_doc.get("file_name") or path.name,
            "mime_type": file_doc.get("mime_type") or "application/octet-stream",
            "size": file_doc.get("size", 0),
            "kind": kind,
            **local_file_urls(file_id),
        }
        if kind in {"pdf", "image"}:
            response["render"] = "inline"
            return response
        try:
            if kind == "docx":
                response.update({"render": "html", "html": preview_docx_html(path)})
                return response
            if kind == "xlsx":
                response.update({"render": "html", "html": preview_xlsx_html(path)})
                return response
            if kind == "text":
                response.update({"render": "html", "html": preview_text_html(path)})
                return response
        except Exception as exc:
            logger.warning("Preview dokumen gagal untuk %s: %s", file_id, exc)
            response.update({"render": "unsupported", "message": "Preview dokumen gagal dibaca. File mungkin rusak atau formatnya tidak valid."})
            return response
        response.update(
            {
                "render": "unsupported",
                "message": "Format file ini belum bisa dipreview langsung. Gunakan file PDF, DOCX, XLSX, TXT, CSV, atau gambar.",
            }
        )
        return response
    finally:
        if temporary:
            remove_temporary_file(path)


@api_router.post("/assignments/{assignment_id}/submit")
async def submit_assignment(
    assignment_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Pengumpulan hanya untuk mahasiswa")
    assignment = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    if assignment["class_id"] not in user.get("class_ids", []):
        raise HTTPException(status_code=403, detail="Mahasiswa tidak terdaftar pada kelas tugas ini")
    class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
    if not class_allows_learning(class_doc):
        raise HTTPException(status_code=409, detail="Kelas sudah berakhir. Submission baru tidak dapat dikirim.")
    if not assignment_is_published(assignment):
        raise HTTPException(status_code=403, detail="Tugas belum tayang")
    submission = await db.submissions.find_one({"assignment_id": assignment_id, "student_id": user["id"]}, {"_id": 0})
    revision_open = bool(submission and (submission.get("status") == "Direvisi" or submission.get("review_status") == "revision_requested"))
    if submission and not revision_open:
        raise HTTPException(
            status_code=409,
            detail="Tugas sudah dikumpulkan. Mahasiswa hanya bisa mengirim ulang jika dosen mengembalikan tugas sebagai revisi.",
        )
    form, upload_files = await multipart_uploads(request, ["files", "file", "files[]"])
    note = str(form.get("note") or "")
    if not upload_files:
        raise HTTPException(status_code=400, detail="Minimal satu file tugas harus dilampirkan")
    allowed = [item.lower().replace(".", "") for item in assignment.get("allowed_formats", [])]
    allowed = sorted(set(allowed + ["jpg", "jpeg", "png", "webp", "pdf", "doc", "docx", "xls", "xlsx", "zip"]))
    for upload in upload_files:
        extension = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        if extension and extension not in allowed:
            raise HTTPException(status_code=400, detail=f"Format file {upload.filename} tidak diizinkan")
    max_file_size_mb = assignment_max_file_size_mb(assignment)
    await validate_upload_file_sizes(upload_files, max_file_size_mb, "file jawaban")
    deadline = datetime.fromisoformat(assignment["deadline"].replace("Z", "+00:00"))
    submit_time = datetime.now(timezone.utc)
    if submit_time > deadline and assignment.get("close_after_deadline") and not revision_open:
        raise HTTPException(status_code=400, detail="Deadline sudah ditutup oleh dosen")
    late_delta = submit_time - deadline if submit_time > deadline else timedelta(0)
    late_hours = round(late_delta.total_seconds() / 3600, 2) if late_delta.total_seconds() > 0 else 0
    late_days = late_delta.days if late_delta.total_seconds() > 0 else 0
    status = "Terlambat" if late_hours > 0 else "Sudah Submit"
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        assignment.get("course_name", "Mata Kuliah"),
        assignment.get("class_name", "Kelas"),
        assignment.get("title", "Tugas"),
    ]
    submission_id = submission["id"] if submission else new_id()
    previous_file_ids: List[str] = []
    if submission:
        previous_files = submission.get("files") if isinstance(submission.get("files"), list) else []
        if isinstance(submission.get("file"), dict):
            previous_files = [*previous_files, submission["file"]]
        previous_file_ids = list(
            {
                str(item.get("file_id") or item.get("id") or "")
                for item in previous_files
                if isinstance(item, dict) and (item.get("file_id") or item.get("id"))
            }
        )
    saved_files = []
    for upload in upload_files:
        file_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            user.get("nim", ""),
            user.get("name", "Mahasiswa"),
            user["id"],
            submission_id,
            assignment_id,
            "submission",
            background_tasks=background_tasks,
            async_drive=True,
            lecturer_id=assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        )
        saved_files.append(file_doc)
    revision_count = int(submission.get("revision_count", 0)) + 1 if submission else 0
    doc = {
        "id": submission_id,
        "assignment_id": assignment_id,
        "assignment_title": assignment["title"],
        "assignment_deadline": assignment.get("deadline", ""),
        "assignment_late_penalty_per_day": assignment.get("late_penalty_per_day", 0),
        "student_id": user["id"],
        "student_name": user["name"],
        "student_nim": user.get("nim", ""),
        "class_id": assignment["class_id"],
        "lecturer_id": assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        "lecturer_name": assignment.get("lecturer_name", class_doc.get("lecturer_name", "")),
        "status": status,
        "review_status": "submitted",
        "note": note,
        "file": saved_files[0],
        "files": saved_files,
        "submitted_at": now_iso(),
        "late_hours": late_hours,
        "late_days": late_days,
        "late_text": f"Terlambat {late_days} hari {round(late_hours % 24, 2)} jam" if late_hours > 0 else "Tepat waktu",
        "revision_count": revision_count,
        "grade": submission.get("grade") if submission else None,
        "feedback": submission.get("feedback", "") if submission else "",
        "revision_note": submission.get("revision_note", "") if submission else "",
        "grade_history": submission.get("grade_history", []) if submission else [],
    }
    await db.submissions.update_one({"id": submission_id}, {"$set": doc}, upsert=True)
    if previous_file_ids:
        await delete_stored_files(
            {
                "id": {"$in": previous_file_ids},
                "record_type": "submission",
            }
        )
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": assignment_id,
            "class_id": assignment["class_id"],
            "lecturer_id": assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
            "student_id": user["id"],
            "reminder_type": "konfirmasi_submit",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": f"Tugas diterima dengan status {status}",
        }
    )
    return public_doc(doc)


@api_router.post("/submissions/{submission_id}/grade")
async def grade_submission(
    submission_id: str,
    payload: GradeInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    submission = await require_submission_access(submission_id, user)
    assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
    await require_rps_complete(submission.get("class_id") or assignment.get("class_id", ""))
    weighted = sum((item.score * item.weight) / 100 for item in payload.rubric_scores)
    penalty = 0.0
    penalty_rate = submission.get("assignment_late_penalty_per_day", assignment.get("late_penalty_per_day", 0))
    if submission.get("status") == "Terlambat" and penalty_rate:
        submitted_at = parse_iso_datetime(submission.get("submitted_at", ""))
        deadline = parse_iso_datetime(submission.get("assignment_deadline", "") or assignment.get("deadline", ""))
        if submitted_at and deadline:
            days_late = max(1, (submitted_at - deadline).days + 1)
            penalty = min(100, float(penalty_rate) * days_late)
    final_grade = min(100, max(0, round(weighted - penalty, 2)))
    grade_predicate = await calculate_grade_predicate(final_grade, submission.get("class_id", assignment.get("class_id", "")))
    history = submission.get("grade_history", [])
    history.append(
        {
            "grade": final_grade,
            "grade_predicate": grade_predicate,
            "rubric_scores": [item.model_dump() for item in payload.rubric_scores],
            "feedback": payload.feedback,
            "revision_note": payload.revision_note,
            "penalty": penalty,
            "graded_by": user["id"],
            "graded_at": now_iso(),
        }
    )
    await db.submissions.update_one(
        {"id": submission_id},
        {
            "$set": {
                "grade": final_grade,
                "grade_predicate": grade_predicate,
                "feedback": payload.feedback,
                "revision_note": payload.revision_note,
                "rubric_scores": [item.model_dump() for item in payload.rubric_scores],
                "late_penalty": penalty,
                "status": payload.status,
                "review_status": "graded",
                "graded_at": now_iso(),
                "grade_history": history,
            }
        },
    )
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": submission["assignment_id"],
            "student_id": submission["student_id"],
            "reminder_type": "nilai_tersedia",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": "Nilai dan feedback sudah tersedia",
        }
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    await notify_submission_status_whatsapp(updated or submission, assignment, "nilai_tersedia", background_tasks)
    return updated


@api_router.post("/submissions/{submission_id}/review")
async def review_submission(submission_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_submission_access(submission_id, user)
    await db.submissions.update_one(
        {"id": submission_id},
        {"$set": {"review_status": "reviewed", "reviewed_at": now_iso(), "reviewed_by": user["id"]}},
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Submission tidak ditemukan")
    return updated


@api_router.post("/submissions/{submission_id}/request-revision")
async def request_revision(
    submission_id: str,
    payload: Dict[str, str],
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    submission = await require_submission_access(submission_id, user)
    await db.submissions.update_one(
        {"id": submission_id},
        {
            "$set": {
                "status": "Direvisi",
                "review_status": "revision_requested",
                "revision_note": payload.get("revision_note", "Perlu revisi"),
                "revision_requested_at": now_iso(),
                "revision_requested_by": user["id"],
            }
        },
    )
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": submission["assignment_id"],
            "student_id": submission["student_id"],
            "reminder_type": "revisi_tugas",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": "Dosen meminta revisi tugas",
        }
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
    await notify_submission_status_whatsapp(updated or submission, assignment, "revisi_tugas", background_tasks)
    return updated


@api_router.post("/submissions/bulk-grade")
async def bulk_grade_submissions(
    payload: BulkGradeInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    results = []
    for item in payload.grades:
        try:
            submission = await require_submission_access(item.submission_id, user)
        except HTTPException:
            results.append({"submission_id": item.submission_id, "status": "not_found"})
            continue
        await require_rps_complete(submission.get("class_id", ""))
        assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
        rubric = assignment.get("rubric") or [{"criterion": "Nilai total", "weight": 100}]
        rubric_scores = [
            {"criterion": r.get("criterion", "Nilai"), "weight": float(r.get("weight", 0)), "score": float(item.score)}
            for r in rubric
        ]
        weighted = sum((score["score"] * score["weight"]) / 100 for score in rubric_scores)
        penalty = 0.0
        penalty_rate = submission.get("assignment_late_penalty_per_day", assignment.get("late_penalty_per_day", 0))
        if submission.get("status") == "Terlambat" and penalty_rate:
            penalty = min(100, float(penalty_rate) * max(1, int(submission.get("late_days", 0)) + 1))
        final_grade = min(100, max(0, round(weighted - penalty, 2)))
        grade_predicate = await calculate_grade_predicate(final_grade, submission.get("class_id", assignment.get("class_id", "")))
        history = submission.get("grade_history", [])
        history.append(
            {
                "grade": final_grade,
                "grade_predicate": grade_predicate,
                "rubric_scores": rubric_scores,
                "feedback": item.feedback,
                "revision_note": item.revision_note,
                "penalty": penalty,
                "graded_by": user["id"],
                "graded_at": now_iso(),
                "bulk": True,
            }
        )
        await db.submissions.update_one(
            {"id": item.submission_id},
            {
                "$set": {
                    "grade": final_grade,
                    "grade_predicate": grade_predicate,
                    "feedback": item.feedback,
                    "revision_note": item.revision_note,
                    "rubric_scores": rubric_scores,
                    "late_penalty": penalty,
                    "status": "Dinilai",
                    "review_status": "graded",
                    "graded_at": now_iso(),
                    "grade_history": history,
                }
            },
        )
        await db.reminder_logs.insert_one(
            {
                "id": new_id(),
                "assignment_id": submission["assignment_id"],
                "student_id": submission["student_id"],
                "reminder_type": "nilai_tersedia",
                "sent_at": now_iso(),
                "status": "in_app",
                "response": "Nilai dan feedback sudah tersedia",
            }
        )
        updated = await db.submissions.find_one({"id": item.submission_id}, {"_id": 0}) or submission
        await notify_submission_status_whatsapp(updated, assignment, "nilai_tersedia", background_tasks)
        results.append({"submission_id": item.submission_id, "status": "graded", "grade": final_grade, "grade_predicate": grade_predicate})
    return {"updated": len([r for r in results if r["status"] == "graded"]), "results": results}


async def calculate_student_progress(student_id: str) -> Dict[str, Any]:
    student = await db.users.find_one({"id": student_id}, {"_id": 0}) or {}
    class_ids = student.get("class_ids", [])
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}, "is_active": True}, {"_id": 0}).to_list(1000)
    assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
    submissions = await db.submissions.find({"student_id": student_id}, {"_id": 0}).to_list(1000)
    submitted_ids = {item["assignment_id"] for item in submissions}
    graded = [item for item in submissions if isinstance(item.get("grade"), (int, float))]
    late = len([item for item in submissions if item.get("status") == "Terlambat"])
    missing = len([assignment for assignment in assignments if assignment["id"] not in submitted_ids])
    avg_grade = round(sum(item["grade"] for item in graded) / len(graded), 1) if graded else 0
    inactive_days = 0
    if student.get("last_login_at"):
        try:
            inactive_days = (datetime.now(timezone.utc) - datetime.fromisoformat(student["last_login_at"].replace("Z", "+00:00"))).days
        except Exception:
            inactive_days = 0
    risk_score = 0
    risk_score += min(4, missing) * 2
    risk_score += min(3, late) * 2
    if graded and avg_grade < 60:
        risk_score += 3
    if inactive_days > 14:
        risk_score += 2
    if risk_score >= 8:
        label = "Risiko Tinggi"
    elif risk_score >= 5:
        label = "Perlu Perhatian"
    elif risk_score >= 2:
        label = "Risiko Rendah"
    else:
        label = "Aman"
    return {
        "submitted": len(submitted_ids),
        "missing": missing,
        "late": late,
        "avg_grade": avg_grade,
        "last_login_at": student.get("last_login_at", ""),
        "status": student.get("status", "active"),
        "risk_label": label,
    }


def _compute_student_progress(
    student: Dict[str, Any],
    published_assignments: List[Dict[str, Any]],
    student_submissions: List[Dict[str, Any]],
) -> Dict[str, Any]:
    submitted_ids = {item["assignment_id"] for item in student_submissions}
    graded = [item for item in student_submissions if isinstance(item.get("grade"), (int, float))]
    late = len([item for item in student_submissions if item.get("status") == "Terlambat"])
    missing = len([assignment for assignment in published_assignments if assignment["id"] not in submitted_ids])
    avg_grade = round(sum(item["grade"] for item in graded) / len(graded), 1) if graded else 0
    inactive_days = 0
    if student.get("last_login_at"):
        try:
            inactive_days = (datetime.now(timezone.utc) - datetime.fromisoformat(student["last_login_at"].replace("Z", "+00:00"))).days
        except Exception:
            inactive_days = 0
    risk_score = 0
    risk_score += min(4, missing) * 2
    risk_score += min(3, late) * 2
    if graded and avg_grade < 60:
        risk_score += 3
    if inactive_days > 14:
        risk_score += 2
    if risk_score >= 8:
        label = "Risiko Tinggi"
    elif risk_score >= 5:
        label = "Perlu Perhatian"
    elif risk_score >= 2:
        label = "Risiko Rendah"
    else:
        label = "Aman"
    return {
        "submitted": len(submitted_ids),
        "missing": missing,
        "late": late,
        "avg_grade": avg_grade,
        "last_login_at": student.get("last_login_at", ""),
        "status": student.get("status", "active"),
        "risk_label": label,
    }


async def calculate_student_progress_many(
    student_ids: List[str], scoped_class_ids: Optional[List[str]] = None
) -> Dict[str, Dict[str, Any]]:
    if not student_ids:
        return {}
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0}).to_list(len(student_ids) * 2)
    students_map: Dict[str, Dict[str, Any]] = {s["id"]: s for s in students}
    all_class_ids: set = set()
    for s in students:
        student_classes = s.get("class_ids", [])
        if scoped_class_ids is not None:
            student_classes = [class_id for class_id in student_classes if class_id in scoped_class_ids]
        all_class_ids.update(student_classes)
    all_assignments: List[Dict[str, Any]] = []
    if all_class_ids:
        all_assignments = await db.assignments.find(
            {"class_id": {"$in": list(all_class_ids)}, "is_active": True}, {"_id": 0}
        ).to_list(5000)
    published_assignments = [a for a in all_assignments if assignment_is_published(a)]
    class_to_assignments: Dict[str, List[Dict[str, Any]]] = {}
    for a in published_assignments:
        cid = a.get("class_id", "")
        class_to_assignments.setdefault(cid, []).append(a)
    all_submissions: List[Dict[str, Any]] = []
    if student_ids:
        all_submissions = await db.submissions.find(
            {"student_id": {"$in": student_ids}}, {"_id": 0}
        ).to_list(5000)
    subs_by_student: Dict[str, List[Dict[str, Any]]] = {}
    for s in all_submissions:
        sid = s.get("student_id", "")
        subs_by_student.setdefault(sid, []).append(s)
    result: Dict[str, Dict[str, Any]] = {}
    for sid in student_ids:
        student = students_map.get(sid, {})
        student_class_ids = student.get("class_ids", [])
        if scoped_class_ids is not None:
            student_class_ids = [class_id for class_id in student_class_ids if class_id in scoped_class_ids]
        student_assignments = []
        for cid in student_class_ids:
            student_assignments.extend(class_to_assignments.get(cid, []))
        student_subs = subs_by_student.get(sid, [])
        if scoped_class_ids is not None:
            student_subs = [
                submission
                for submission in student_subs
                if submission.get("class_id") in scoped_class_ids
            ]
        result[sid] = _compute_student_progress(student, student_assignments, student_subs)
    return result


@api_router.get("/progress")
async def progress(user: Dict[str, Any] = Depends(get_current_user)):
    if user["role"] == "student":
        return {"student": public_doc(user.copy()), "progress": await calculate_student_progress(user["id"])}
    if user.get("role") == "staff" and not user_has_access_role(user, "academic_operator"):
        return []
    class_ids = await lecturer_class_ids(user)
    student_ids = list({
        student_id
        for item in await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "student_ids": 1}).to_list(5000)
        for student_id in item.get("student_ids", [])
    })
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    progress_map = await calculate_student_progress_many([s["id"] for s in students], class_ids)
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return students


@api_router.get("/grade-predicates")
async def get_grade_predicates(class_id: str = "", user: Dict[str, Any] = Depends(require_lecturer_or_academic_manager)):
    if class_id:
        await require_class_access(class_id, user)
    elif user.get("role") == "lecturer":
        return {"class_id": "", "predicates": DEFAULT_GRADE_PREDICATES}
    predicates = await get_grade_predicates_for_class(class_id)
    return {"class_id": class_id, "predicates": predicates}


@api_router.put("/grade-predicates")
async def save_grade_predicates(payload: GradePredicateInput, user: Dict[str, Any] = Depends(require_admin)):
    if payload.class_id:
        await require_class_mutation_access(payload.class_id, user)
        await require_rps_complete(payload.class_id)
    elif user.get("role") == "lecturer":
        raise HTTPException(status_code=400, detail="Dosen harus memilih kelas untuk menyimpan predikat")
    predicates = validate_grade_predicates([item.model_dump() for item in payload.predicates])
    doc = {
        "id": payload.class_id or "global",
        "class_id": payload.class_id,
        "predicates": predicates,
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.grade_predicates.update_one({"class_id": payload.class_id}, {"$set": doc}, upsert=True)
    return doc


@api_router.get("/calendar")
async def calendar(user: Dict[str, Any] = Depends(get_current_user)):
    class_filter: Dict[str, Any] = {}
    if user["role"] == "student":
        class_filter = {"class_id": {"$in": user.get("class_ids", [])}}
    elif user["role"] == "lecturer":
        class_filter = {"class_id": {"$in": await lecturer_class_ids(user)}}
    assignments = await db.assignments.find({**class_filter, "is_active": True}, {"_id": 0}).to_list(1000)
    if user["role"] == "student":
        assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
    materials = await db.materials.find(class_filter, {"_id": 0}).to_list(1000)
    published_calendar_events = await db.academic_calendar_events.find(
        {"status": "published"}, {"_id": 0}
    ).to_list(2000)
    deadline_settings_documents = await db.academic_deadline_settings.find(
        {}, {"_id": 0}
    ).to_list(1000)
    events = [
        calendar_event_payload(event)
        for event in published_calendar_events
        if calendar_event_visible_to_user(event, user)
    ]
    for settings_document in deadline_settings_documents:
        settings = academic_deadline_settings_payload(settings_document)
        for deadline_type, item in settings["deadlines"].items():
            if (
                item.get("enabled")
                and item.get("deadline_at")
                and academic_deadline_visible_to_user(deadline_type, user)
            ):
                events.append(
                    academic_deadline_event_payload(
                        deadline_type,
                        item,
                        settings["academic_year_id"],
                    )
                )
    for assignment in assignments:
        if user["role"] in {"admin", "lecturer"} and assignment.get("published_at") and not assignment_is_published(assignment):
            events.append(
                {
                    "id": f"{assignment['id']}-publish",
                    "source": "assignment",
                    "type": "tayang",
                    "title": f"Tayang: {assignment['title']}",
                    "date": assignment["published_at"],
                    "class_name": assignment.get("class_name", ""),
                    "class_id": assignment.get("class_id", ""),
                }
            )
        events.append(
            {
                "id": assignment["id"],
                "source": "assignment",
                "type": "deadline",
                "title": assignment["title"],
                "date": assignment["deadline"],
                "class_name": assignment.get("class_name", ""),
                "class_id": assignment.get("class_id", ""),
            }
        )
    for material in materials:
        if material.get("locked_until"):
            events.append(
                {
                    "id": material["id"],
                    "source": "material",
                    "type": "materi",
                    "title": material["title"],
                    "date": material["locked_until"],
                    "class_id": material.get("class_id", ""),
                }
            )
    return sorted(events, key=lambda item: item.get("date", ""))


@api_router.get("/calendar/deadlines")
async def get_academic_deadline_settings(
    academic_year_id: str = "",
    user: Dict[str, Any] = Depends(get_current_user),
):
    if not can_manage_academic_calendar(user):
        raise HTTPException(
            status_code=403,
            detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola deadline",
        )
    normalized_academic_year_id = str(academic_year_id or "").strip()
    document = await db.academic_deadline_settings.find_one(
        {"academic_year_id": normalized_academic_year_id},
        {"_id": 0},
    ) or {
        "id": f"academic-deadlines-{normalized_academic_year_id or 'global'}",
        "academic_year_id": normalized_academic_year_id,
        "deadlines": {},
    }
    return academic_deadline_settings_payload(document)


@api_router.put("/calendar/deadlines")
async def save_academic_deadline_settings(
    payload: AcademicDeadlineSettingsInput,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if not can_manage_academic_calendar(user):
        raise HTTPException(
            status_code=403,
            detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola deadline",
        )
    validated = validate_academic_deadline_settings(payload)
    academic_year_id = validated["academic_year_id"]
    document = {
        "id": f"academic-deadlines-{academic_year_id or 'global'}",
        **validated,
        "updated_by": user["id"],
        "updated_at": now_iso(),
    }
    await db.academic_deadline_settings.update_one(
        {"academic_year_id": academic_year_id},
        {"$set": document},
        upsert=True,
    )
    return academic_deadline_settings_payload(document)


@api_router.get("/calendar/academic")
async def list_academic_calendar_events(
    academic_year_id: str = "",
    include_archived: bool = False,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Daftar penuh kalender institusi untuk halaman kontrol akademik."""
    if not can_manage_academic_calendar(user):
        raise HTTPException(status_code=403, detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola kalender")
    query: Dict[str, Any] = {}
    if academic_year_id:
        query["academic_year_id"] = academic_year_id
    if not include_archived:
        query["status"] = {"$ne": "archived"}
    return await db.academic_calendar_events.find(query, {"_id": 0}).sort("start_at", 1).to_list(2000)


@api_router.post("/calendar/academic")
async def create_academic_calendar_event(
    payload: AcademicCalendarEventInput,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if not can_manage_academic_calendar(user):
        raise HTTPException(status_code=403, detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola kalender")
    doc = {
        "id": new_id(),
        **validate_academic_calendar_event(payload),
        "created_by": user["id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.academic_calendar_events.insert_one(doc)
    return public_doc(doc.copy())


@api_router.put("/calendar/academic/{event_id}")
async def update_academic_calendar_event(
    event_id: str,
    payload: AcademicCalendarEventInput,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if not can_manage_academic_calendar(user):
        raise HTTPException(status_code=403, detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola kalender")
    existing = await db.academic_calendar_events.find_one({"id": event_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kegiatan kalender tidak ditemukan")
    updates = {
        **validate_academic_calendar_event(payload),
        "updated_by": user["id"],
        "updated_at": now_iso(),
    }
    await db.academic_calendar_events.update_one({"id": event_id}, {"$set": updates})
    return {**existing, **updates}


@api_router.delete("/calendar/academic/{event_id}")
async def archive_academic_calendar_event(
    event_id: str,
    user: Dict[str, Any] = Depends(get_current_user),
):
    """Arsipkan kegiatan agar riwayat tetap dapat diaudit tanpa tampil ke pengguna."""
    if not can_manage_academic_calendar(user):
        raise HTTPException(status_code=403, detail="Hanya Admin Kampus atau Operator Akademik yang dapat mengelola kalender")
    existing = await db.academic_calendar_events.find_one({"id": event_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kegiatan kalender tidak ditemukan")
    updates = {"status": "archived", "archived_by": user["id"], "archived_at": now_iso(), "updated_at": now_iso()}
    await db.academic_calendar_events.update_one({"id": event_id}, {"$set": updates})
    return {"ok": True, "event": {**existing, **updates}}


@api_router.post("/reminders/send")
async def send_reminder(payload: ReminderInput, user: Dict[str, Any] = Depends(require_admin)):
    assignment = await db.assignments.find_one({"id": payload.assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    class_doc = await require_class_mutation_access(assignment.get("class_id", ""), user)
    doc = payload.model_dump()
    doc.update({
        "id": new_id(),
        "class_id": assignment.get("class_id", ""),
        "lecturer_id": class_doc.get("lecturer_id", user["id"]),
        "sent_at": now_iso(),
        "status": "in_app",
        "response": "Reminder tersimpan di aplikasi",
    })
    await db.reminder_logs.insert_one(doc)
    return public_doc(doc)


@api_router.get("/reminders")
async def list_reminders(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"student_id": user["id"]}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    reminders = await db.reminder_logs.find(query, {"_id": 0}).sort("sent_at", -1).to_list(1000)
    if user["role"] == "student":
        assignment_ids = [item.get("assignment_id") for item in reminders if item.get("assignment_id")]
        if assignment_ids:
            assignments = await db.assignments.find({"id": {"$in": assignment_ids}}, {"_id": 0}).to_list(1000)
            visible_ids = {item["id"] for item in assignments if assignment_is_published(item)}
            reminders = [item for item in reminders if not item.get("assignment_id") or item.get("assignment_id") in visible_ids]
    return reminders


async def build_grade_recap(
    user: Optional[Dict[str, Any]] = None,
    class_id: Optional[str] = None,
    use_snapshots: bool = True,
) -> List[Dict[str, Any]]:
    if user and not is_campus_admin(user):
        class_ids = await lecturer_class_ids(user)
        if class_id and class_id not in class_ids:
            raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
        scoped_ids = [class_id] if class_id else class_ids
    else:
        if class_id:
            scoped_ids = [class_id]
        elif user:
            class_ids = await lecturer_class_ids(user)
            scoped_ids = class_ids
        else:
            scoped_ids = []
    classes = await db.classes.find({"id": {"$in": scoped_ids}}, {"_id": 0}).to_list(500)
    if not classes:
        return []
    course_ids = list({item.get("course_id") for item in classes if item.get("course_id")})
    courses = await db.courses.find({"id": {"$in": course_ids}}, {"_id": 0}).to_list(500)
    courses_by_id = {item["id"]: item for item in courses}
    assignments = await db.assignments.find({"class_id": {"$in": scoped_ids}, "is_active": True}, {"_id": 0}).to_list(5000)
    assignments_by_class: Dict[str, List[Dict[str, Any]]] = {}
    for item in assignments:
        assignments_by_class.setdefault(item.get("class_id", ""), []).append(item)
    submissions = await db.submissions.find({"class_id": {"$in": scoped_ids}, "grade": {"$ne": None}}, {"_id": 0}).to_list(10000)
    submissions_by_class_student: Dict[tuple[str, str], List[Dict[str, Any]]] = {}
    for item in submissions:
        submissions_by_class_student.setdefault((item.get("class_id", ""), item.get("student_id", "")), []).append(item)

    result: List[Dict[str, Any]] = []
    for class_doc in classes:
        cid = class_doc["id"]
        course = courses_by_id.get(class_doc.get("course_id"), {})
        if use_snapshots and class_doc.get("status") in CLASS_STATUSES_READ_ONLY and class_doc.get("final_grade_snapshot"):
            result.append(class_doc["final_grade_snapshot"])
            continue
        snapshot_weights = class_doc.get("grade_weights_snapshot") if class_doc.get("status") != CLASS_STATUS_ACTIVE else None
        weights = grade_weights_from_document(snapshot_weights or course.get("grade_weights"))
        enrolled_ids = list(dict.fromkeys(class_doc.get("student_ids", [])))
        students = await db.users.find(
            {"id": {"$in": enrolled_ids}, "role": "student"},
            {"_id": 0, "password_hash": 0},
        ).to_list(max(1000, len(enrolled_ids) * 2 or 1))
        student_map = {item["id"]: item for item in students}
        student_ids = list(dict.fromkeys(enrolled_ids + list(student_map.keys())))
        class_assignments = assignments_by_class.get(cid, [])
        class_students: List[Dict[str, Any]] = []
        scored_totals: List[float] = []
        for sid in student_ids:
            student = student_map.get(sid, {})
            student_submissions = submissions_by_class_student.get((cid, sid), [])
            component_values: Dict[str, List[float]] = {component: [] for component in GRADE_WEIGHT_COMPONENTS}
            scores = []
            for submission in student_submissions:
                assignment = next((item for item in class_assignments if item.get("id") == submission.get("assignment_id")), {})
                component = normalize_assessment_category(assignment.get("assessment_category", "tugas"))
                score = float(submission.get("grade", 0))
                component_values[component].append(score)
                scores.append({
                    "assignment_id": submission.get("assignment_id", ""),
                    "assignment_title": submission.get("assignment_title", assignment.get("title", "")),
                    "assessment_category": component,
                    "grade": score,
                    "grade_predicate": submission.get("grade_predicate", ""),
                })
            component_scores = {
                component: round(sum(values) / len(values), 2) if values else None
                for component, values in component_values.items()
            }
            available_components = [
                component for component in GRADE_WEIGHT_COMPONENTS if component_scores[component] is not None and weights[component] > 0
            ]
            available_weight = sum(weights[component] for component in available_components)
            weighted_grade = round(
                sum(component_scores[component] * weights[component] for component in available_components) / available_weight,
                2,
            ) if available_weight else 0
            if scores:
                scored_totals.append(weighted_grade)
            class_students.append({
                "student_id": sid,
                "student_name": student.get("name", "Mahasiswa"),
                "student_nim": student.get("nim", ""),
                "scores": sorted(scores, key=lambda item: item.get("assignment_title", "")),
                "component_scores": component_scores,
                "weighted_grade": weighted_grade,
                "average": weighted_grade,
                "total_graded": len(scores),
                "grade_complete": len(available_components) == len(GRADE_WEIGHT_COMPONENTS),
                "grade_predicate": await calculate_grade_predicate(weighted_grade, cid) if scores else "-",
            })
        distribution = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}
        for score in scored_totals:
            if score >= 85:
                distribution["A"] += 1
            elif score >= 70:
                distribution["B"] += 1
            elif score >= 60:
                distribution["C"] += 1
            elif score >= 50:
                distribution["D"] += 1
            else:
                distribution["E"] += 1
        result.append({
            "class_id": cid,
            "course_id": class_doc.get("course_id", ""),
            "course_name": course.get("name") or class_doc.get("course_name", ""),
            "class_name": class_doc.get("name", ""),
            "class_status": class_doc.get("status", CLASS_STATUS_ACTIVE),
            "class_status_label": class_status_label(class_doc.get("status", CLASS_STATUS_ACTIVE)),
            "ended_at": class_doc.get("ended_at", ""),
            "finalized_at": class_doc.get("finalized_at", ""),
            "student_count": len(class_students),
            "total_assignments": len(class_assignments),
            "class_average": round(sum(scored_totals) / len(scored_totals), 2) if scored_totals else 0,
            "grade_distribution": distribution,
            "grade_weights": weights,
            "grade_weights_customized": (
                bool(class_doc.get("grade_weights_snapshot_customized", False))
                if snapshot_weights
                else isinstance(course.get("grade_weights"), dict)
            ),
            "assignments": [
                {
                    "id": item.get("id", ""),
                    "title": item.get("title", "Tugas"),
                    "assessment_category": normalize_assessment_category(item.get("assessment_category", "tugas")),
                }
                for item in class_assignments
            ],
            "students": sorted(class_students, key=lambda item: (item.get("student_name", ""), item.get("student_nim", ""))),
        })
    return sorted(result, key=lambda item: (item.get("course_name", ""), item.get("class_name", "")))


@api_router.get("/reports/grades.xlsx")
async def export_grades(class_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_admin)):
    recaps = await build_grade_recap(user, class_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rekap Nilai"
    sheet.append(["NIM", "Nama", "Mata Kuliah", "Kelas", "Nilai Akhir", "Predikat", "Status", "Nilai Tugas", "Nilai UTS", "Nilai UAS", "Bobot Tugas", "Bobot UTS", "Bobot UAS"])
    for recap in recaps:
        weights = recap.get("grade_weights", DEFAULT_GRADE_WEIGHTS)
        for student in recap.get("students", []):
            components = student.get("component_scores", {})
            sheet.append([
                student.get("student_nim", ""),
                student.get("student_name", ""),
                recap.get("course_name", ""),
                recap.get("class_name", ""),
                student.get("weighted_grade", 0),
                student.get("grade_predicate", "-"),
                "Lengkap" if student.get("grade_complete") else "Sementara",
                components.get("tugas", ""),
                components.get("uts", ""),
                components.get("uas", ""),
                weights.get("tugas", 25),
                weights.get("uts", 35),
                weights.get("uas", 40),
            ])
    for column in sheet.columns:
        column_letter = column[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 32)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = "rekap-nilai.xlsx" if not class_id else f"rekap-nilai-{class_id}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _pdf_text(value: Any) -> str:
    text = str(value or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    return text.encode("latin-1", "replace").decode("latin-1")


def build_grade_recap_pdf(recaps: List[Dict[str, Any]]) -> bytes:
    lines = ["REKAP NILAI MAHASISWA", ""]
    for recap in recaps:
        weights = recap.get("grade_weights", DEFAULT_GRADE_WEIGHTS)
        lines.extend([
            f"Mata Kuliah: {recap.get('course_name', '')} | Kelas: {recap.get('class_name', '')}",
            f"Bobot: Tugas {weights.get('tugas', 25)}% | UTS {weights.get('uts', 35)}% | UAS {weights.get('uas', 40)}%",
            "NIM | Nama | Tugas | UTS | UAS | Nilai Akhir | Status | Predikat",
        ])
        for student in recap.get("students", []):
            components = student.get("component_scores", {})
            lines.append(
                f"{student.get('student_nim', '')} | {student.get('student_name', '')} | "
                f"{components.get('tugas', '-')} | {components.get('uts', '-')} | {components.get('uas', '-')} | "
                f"{student.get('weighted_grade', 0)} | "
                f"{'Lengkap' if student.get('grade_complete') else 'Sementara'} | {student.get('grade_predicate', '-') }"
            )
        lines.append("")
    if not recaps:
        lines.append("Belum ada data rekap nilai.")

    chunks = [lines[index:index + 52] for index in range(0, len(lines), 52)] or [[]]
    objects: List[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    page_ids = [3 + index * 2 for index in range(len(chunks))]
    content_ids = [page_id + 1 for page_id in page_ids]
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("ascii"))
    for page_id, content_id, page_lines in zip(page_ids, content_ids, chunks):
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 {3 + len(chunks) * 2} 0 R >> >> /Contents {content_id} 0 R >>".encode("ascii"))
        content_lines = ["BT", "/F1 8 Tf", "40 760 Td", "11 TL"]
        for line in page_lines:
            content_lines.append(f"({_pdf_text(line[:130])}) Tj")
            content_lines.append("T*")
        content_lines.append("ET")
        content = "\n".join(content_lines).encode("latin-1", "replace")
        objects.append(f"<< /Length {len(content)} >>\nstream\n".encode("ascii") + content + b"\nendstream")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii"))
    return bytes(output)


@api_router.get("/reports/grades.pdf")
async def export_grades_pdf(class_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_admin)):
    pdf = build_grade_recap_pdf(await build_grade_recap(user, class_id))
    filename = "rekap-nilai.pdf" if not class_id else f"rekap-nilai-{class_id}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/reports/summary")
async def report_summary(user: Dict[str, Any] = Depends(require_lecturer_or_academic_manager)):
    class_ids = await lecturer_class_ids(user)
    class_docs = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "student_ids": 1}).to_list(5000)
    student_ids = list({student_id for item in class_docs for student_id in item.get("student_ids", [])})
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    submissions = await db.submissions.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(5000)
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(1000)
    return {
        "total_students": len(students),
        "total_assignments": len(assignments),
        "total_submissions": len(submissions),
        "graded_submissions": len([item for item in submissions if item.get("status") == "Dinilai"]),
        "late_submissions": len([item for item in submissions if item.get("status") == "Terlambat"]),
    }


@api_router.get("/reports/grade-recap")
async def grade_recap(class_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_lecturer_or_academic_manager)):
    return await build_grade_recap(user, class_id)


@api_router.get("/settings")
async def get_settings(_: Dict[str, Any] = Depends(get_current_user)):
    stored_settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    settings = {**default_app_settings(), **stored_settings}
    active_ta = await db.tahun_ajaran.find_one({"is_active": True}, {"_id": 0})
    if active_ta:
        settings["active_academic_year"] = active_ta.get("tahun") or settings.get("active_academic_year")
        settings["active_semester"] = active_ta.get("semester") or settings.get("active_semester")
    return settings


@api_router.get("/settings/public")
async def get_public_settings():
    defaults = default_app_settings()
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or defaults
    return {
        # Hanya expose identitas dan kontak kampus yang memang ditampilkan
        # pada halaman publik. Field admin, integrasi, dan konfigurasi internal
        # tetap tidak ikut dikirim ke browser.
        "app_name": settings.get("app_name") or defaults["app_name"],
        "meta_description": settings.get("meta_description") or defaults["meta_description"],
        "campus_name": settings.get("campus_name", defaults["campus_name"]),
        "campus_code": settings.get("campus_code", defaults["campus_code"]),
        "institution_type": settings.get("institution_type", defaults["institution_type"]),
        "accreditation": settings.get("accreditation", defaults["accreditation"]),
        "accreditation_sk": settings.get("accreditation_sk", defaults["accreditation_sk"]),
        "campus_motto": settings.get("campus_motto", defaults["campus_motto"]),
        "campus_phone": settings.get("campus_phone", defaults["campus_phone"]),
        "campus_whatsapp": settings.get("campus_whatsapp", defaults["campus_whatsapp"]),
        "campus_email": settings.get("campus_email", defaults["campus_email"]),
        "campus_website": settings.get("campus_website", defaults["campus_website"]),
        "campus_address": settings.get("campus_address", defaults["campus_address"]),
        "app_logo_url": settings.get("app_logo_url", defaults["app_logo_url"]),
        "campus_logo_url": settings.get("campus_logo_url", defaults["campus_logo_url"]),
    }


@api_router.put("/settings")
async def update_settings(payload: AppSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    doc = payload.model_dump()
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.app_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return await db.app_settings.find_one({"id": "main"}, {"_id": 0})


@api_router.post("/academic-years/rollover-preview")
async def rollover_preview(user: Dict[str, Any] = Depends(require_admin)):
    classes = await db.classes.find(
        {"id": {"$in": await lecturer_class_ids(user)}, "status": "active"}, {"_id": 0}
    ).to_list(500)
    return {
        "recommended_flow": [
            "Export rekap nilai dan arsip kelas semester lama.",
            "Ubah active_academic_year dan active_semester pada Settings.",
            "Duplikasi kelas yang masih dipakai ke tahun ajaran baru dengan kode kelas baru.",
            "Mahasiswa lama tidak otomatis ikut kelas baru; mereka mengajukan kode kelas dan dosen approve.",
            "Materi/tugas semester lama tetap bisa dilihat di arsip, sedangkan submission baru mengikuti deadline kelas baru.",
        ],
        "active_classes_to_archive": len(classes),
        "classes": classes,
    }


@app.websocket("/api/chat/ws")
async def chat_websocket(websocket: WebSocket, token: str = ""):
    session = await db.sessions.find_one({"token": token.strip()}, {"_id": 0}) if token.strip() else None
    if not session:
        await websocket.close(code=1008)
        return
    user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        await websocket.close(code=1008)
        return
    if not await chat_connections.connect(user["id"], websocket):
        return
    try:
        while True:
            payload = json.loads(await websocket.receive_text())
            if payload.get("type") == "viewing":
                target_id = str(payload.get("user_id", ""))
                if target_id and target_id != user["id"]:
                    await chat_connections.set_viewing(user["id"], websocket, target_id)
                else:
                    await chat_connections.set_viewing(user["id"], websocket, "")
            elif payload.get("type") == "ping":
                await websocket.send_json({"type": "pong"})
    except (WebSocketDisconnect, json.JSONDecodeError):
        await chat_connections.disconnect(user["id"], websocket)


@app.middleware("http")
async def user_activity_logging_middleware(request: Request, call_next):
    started_at = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        user = getattr(request.state, "current_user", None)
        if (
            user
            and request.url.path.startswith("/api/")
            and should_log_user_activity(request.method, request.url.path)
        ):
            queue_user_activity(
                user,
                request.method,
                request.url.path,
                500,
                round((time.perf_counter() - started_at) * 1000),
                request,
            )
        raise
    user = getattr(request.state, "current_user", None)
    if (
        user
        and request.url.path.startswith("/api/")
        and should_log_user_activity(request.method, request.url.path)
    ):
        queue_user_activity(
            user,
            request.method,
            request.url.path,
            response.status_code,
            round((time.perf_counter() - started_at) * 1000),
            request,
        )
    return response

_cors_raw = os.environ.get("CORS_ORIGINS", "*")
if _cors_raw == "*":
    _cors_origins = ["*"]
else:
    _cors_origins = [o.strip() for o in _cors_raw.split(",") if o.strip()]
    for _def_orig in ["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:3001", "http://127.0.0.1:3001"]:
        if _def_orig not in _cors_origins:
            _cors_origins.append(_def_orig)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)


# ==========================================
# RPS & ATTENDANCE (PRESENSI 16 SESI) APIs
# ==========================================

def generate_default_16_meetings(course_name: str) -> List[Dict[str, Any]]:
    meetings = []
    for i in range(1, 17):
        if i == 8:
            topic = "Evaluasi Tengah Semester (UTS)"
            sub_topic = "Ujian Tertulis / Praktek Tengah Semester"
            is_exam = True
        elif i == 16:
            topic = "Evaluasi Akhir Semester (UAS)"
            sub_topic = "Ujian Akhir Semester / Proyek Akhir"
            is_exam = True
        else:
            week_num = i if i < 8 else i - 1
            topic = f"Pertemuan {i}: Pengenalan & Pokok Bahasan Minggu ke-{week_num}"
            sub_topic = f"Pembahasan konsep dasar & studi kasus {course_name} Bagian {week_num}"
            is_exam = False
        if is_exam:
            meetings.append({
                "meeting_number": i,
                "topic": topic,
                "sub_topic": sub_topic,
                "learning_outcome": "Menunjukkan penguasaan keseluruhan capaian pembelajaran",
                "method": "Evaluasi Mandiri",
                "materials": "Kisi-kisi & modul seluruh pertemuan",
                "assignments": "Ujian Evaluasi",
                "waktu": "90 menit",
                "penilaian_teknik": "Ujian tertulis / praktik",
                "penilaian_indikator": "Ketepatan jawaban terhadap kisi-kisi",
                "penilaian_kriteria": "Nilai ujian",
                "penilaian_bobot": "15%" if i == 8 else "20%",
                "is_exam": True,
            })
        else:
            meetings.append({
                "meeting_number": i,
                "topic": topic,
                "sub_topic": sub_topic,
                "learning_outcome": f"Mahasiswa mampu memahami dan menguasai konsep Sesi {i}",
                "method": "Presentasi, diskusi, dan penugasan",
                "materials": f"Slide & Modul Pertemuan {i}",
                "assignments": "Kuis singkat & Latihan mandiri",
                "waktu": "KPB 3x50",
                "penilaian_teknik": "FGD / penugasan",
                "penilaian_indikator": "Pemahaman terhadap materi",
                "penilaian_kriteria": "Kehadiran, penguasaan materi",
                "penilaian_bobot": "",
                "is_exam": False,
            })
    return meetings


class RPSMeetingItem(BaseModel):
    meeting_number: int
    topic: str
    sub_topic: str = ""
    learning_outcome: str = ""
    method: str = "Tatap Muka / Blended Learning"
    materials: str = ""
    assignments: str = ""
    waktu: str = ""
    penilaian_teknik: str = ""
    penilaian_indikator: str = ""
    penilaian_kriteria: str = ""
    penilaian_bobot: str = ""
    is_exam: bool = False


class RPSInput(BaseModel):
    course_code: str = ""
    semester: str = ""
    sks: str = ""
    program_name: str = ""
    lecturer_name: str = ""
    compiled_at: str = ""
    cpl_sikap: str = ""
    cpl_keterampilan_umum: str = ""
    cpl_pengetahuan: str = ""
    cpl_keterampilan_khusus: str = ""
    keterangan: str = ""
    cpmk: str = ""
    description: str = ""
    references: str = ""
    document_url: str = ""
    meetings: List[RPSMeetingItem] = []


class RPSApprovalInput(BaseModel):
    action: str = Field(..., description="approve atau reject")
    note: str = ""


RPS_UPLOAD_MAX_BYTES = 20 * 1024 * 1024


def rps_meeting_label(meeting: Dict[str, Any]) -> str:
    return f"sesi {meeting.get('meeting_number', '?')}"


def rps_completeness_from_document(
    rps: Optional[Dict[str, Any]],
) -> tuple[bool, List[str]]:
    """Cek kelengkapan RPS dari dokumen yang sudah dimuat."""
    if not rps:
        return False, ["RPS belum disusun"]

    # Lengkap = dokumen RPS tersimpan, kolom header wajib terisi (CPMK,
    # deskripsi, daftar referensi), dan seluruh 16 pertemuan terisi kolom
    # topik + materi pembelajaran; pertemuan non-ujian juga wajib terisi
    # kemampuan yang diharapkan, metode, dan waktu.
    missing: List[str] = []
    for field, label in (
        ("cpmk", "CPMK"),
        ("description", "Deskripsi Mata Kuliah"),
        ("references", "Daftar Referensi"),
    ):
        if not str(rps.get(field) or "").strip():
            missing.append(label)

    meetings = rps.get("meetings") or []
    if len(meetings) < 16:
        missing.append(f"16 pertemuan (baru terisi {len(meetings)})")
    else:
        for m in meetings:
            num = m.get("meeting_number", "?")
            if not str(m.get("topic") or "").strip():
                missing.append(f"topik {rps_meeting_label(m)}")
            if not str(m.get("materials") or "").strip():
                missing.append(f"materi pembelajaran {rps_meeting_label(m)}")
            if not m.get("is_exam"):
                for field, label in (
                    ("learning_outcome", "kemampuan yang diharapkan"),
                    ("method", "metode pembelajaran"),
                    ("waktu", "waktu (menit)"),
                ):
                    if not str(m.get(field) or "").strip():
                        missing.append(f"{label} {rps_meeting_label(m)}")

    return (not missing), sorted(set(missing))


async def class_rps_complete(class_id: str) -> tuple[bool, List[str]]:
    """Cek apakah RPS kelas sudah disusun lengkap sesuai template."""
    rps = await db.rps.find_one({"class_id": class_id}, {"_id": 0})
    return rps_completeness_from_document(rps)


def rps_review_status(
    rps: Optional[Dict[str, Any]],
    complete: bool,
) -> str:
    if not rps:
        return "not_started"
    raw_status = str(rps.get("approval_status") or "").strip().lower()
    if raw_status in {"approved", "rejected", "pending"}:
        return raw_status
    return "pending" if complete else "draft"


def _rps_course_group_key(document: Dict[str, Any]) -> str:
    return str(
        document.get("course_id")
        or document.get("course_code")
        or document.get("code")
        or document.get("course_name")
        or document.get("name")
        or document.get("id")
        or "unknown"
    ).strip().lower()


def _rps_group_seed(document: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "course_id": document.get("course_id") or document.get("id", ""),
        "course_code": document.get("course_code") or document.get("code") or document.get("kode", ""),
        "course_name": document.get("course_name") or document.get("name") or document.get("nama", "Mata Kuliah"),
        "program_name": document.get("program_name") or document.get("prodi_name") or document.get("nama_prodi", ""),
        "class_count": 0,
        "rps_count": 0,
        "complete_count": 0,
        "not_started_count": 0,
        "draft_count": 0,
        "pending_count": 0,
        "approved_count": 0,
        "rejected_count": 0,
        "classes": [],
    }


@api_router.get("/prodi/analisis-rps")
async def prodi_rps_analysis(
    semester_id: str = "",
    prodi_id: str = "",
    user: Dict[str, Any] = Depends(require_program_analysis_user),
):
    """Analisis kelengkapan dan status approval RPS seluruh mata kuliah Prodi."""
    requested_prodi = str(prodi_id or "").strip()
    structural_scope, own_scope_values, scope_values = await _analysis_resolve_program_scope(
        user,
        requested_prodi,
    )

    raw_semester_id = str(semester_id or "").strip()
    period_doc = None
    if raw_semester_id and raw_semester_id != "all":
        period_doc = await db.tahun_ajaran.find_one(
            {"$or": [{"id": raw_semester_id}, {"kode": raw_semester_id}, {"code": raw_semester_id}]},
            {"_id": 0},
        )
        if not period_doc:
            raise HTTPException(status_code=404, detail="Tahun ajaran yang dipilih tidak ditemukan")

    class_docs = await db.classes.find({"status": {"$ne": "deleted"}}, {"_id": 0}).to_list(5000)
    if scope_values:
        scope_lookup = {str(value).upper() for value in scope_values}
        class_docs = [
            item for item in class_docs
            if any(value.upper() in scope_lookup for value in _analysis_program_fields(item))
        ]
    elif not _analysis_has_global_program_scope(user):
        assigned_class_ids = set(await lecturer_class_ids(user))
        class_docs = [item for item in class_docs if str(item.get("id")) in assigned_class_ids]
    if raw_semester_id and raw_semester_id != "all":
        class_docs = [
            item for item in class_docs
            if _analysis_class_matches_period(item, raw_semester_id, period_doc)
        ]

    class_ids = [str(item.get("id")) for item in class_docs if item.get("id")]
    rps_docs = await db.rps.find(
        {"class_id": {"$in": class_ids}},
        {"_id": 0},
    ).to_list(10000) if class_ids else []
    rps_by_class = {str(item.get("class_id")): item for item in rps_docs if item.get("class_id")}

    # Mata kuliah aktif ditentukan dari kelas yang sudah lolos scope Prodi dan
    # filter semester di atas. Jangan memulai grup dari master `courses`,
    # karena master tersebut berisi seluruh riwayat mata kuliah lintas semester.
    groups: Dict[str, Dict[str, Any]] = {}
    for class_doc in class_docs:
        groups.setdefault(_rps_course_group_key(class_doc), _rps_group_seed(class_doc))

    class_results: List[Dict[str, Any]] = []
    status_counts = {status: 0 for status in ("not_started", "draft", "pending", "approved", "rejected")}
    complete_count = 0
    rps_count = 0
    total_meetings = 0
    total_classes = len(class_docs)
    for class_doc in class_docs:
        class_id = str(class_doc.get("id"))
        rps = rps_by_class.get(class_id)
        complete, missing = rps_completeness_from_document(rps)
        status = rps_review_status(rps, complete)
        course_key = _rps_course_group_key(class_doc)
        group = groups.setdefault(course_key, _rps_group_seed(class_doc))
        meetings_count = len((rps or {}).get("meetings") or [])
        group["class_count"] += 1
        group["rps_count"] += int(bool(rps))
        group["complete_count"] += int(complete)
        group[f"{status}_count"] += 1
        group["classes"].append(class_id)
        status_counts[status] += 1
        complete_count += int(complete)
        rps_count += int(bool(rps))
        total_meetings += meetings_count
        class_results.append({
            "id": class_id,
            "course_id": class_doc.get("course_id", ""),
            "course_code": class_doc.get("course_code", ""),
            "course_name": class_doc.get("course_name", "Mata Kuliah"),
            "class_name": class_doc.get("name") or class_doc.get("class_code", ""),
            "program_name": class_doc.get("program_name", ""),
            "lecturer_name": class_doc.get("lecturer_name", ""),
            "academic_year": class_doc.get("academic_year", ""),
            "semester": class_doc.get("semester", ""),
            "student_count": len(class_doc.get("student_ids", []) or []),
            "rps_id": (rps or {}).get("id", ""),
            "has_rps": bool(rps),
            "is_complete": complete,
            "missing_fields": missing[:12],
            "meetings_count": meetings_count,
            "meetings_target": 16,
            "meeting_coverage_percent": round(min(meetings_count, 16) / 16 * 100, 1),
            "approval_status": status,
            "approval_note": (rps or {}).get("approval_note", ""),
            "approved_at": (rps or {}).get("approved_at", ""),
            "approved_by": (rps or {}).get("approved_by", ""),
            "updated_at": (rps or {}).get("updated_at", ""),
            "document_file_name": ((rps or {}).get("document_file") or {}).get("file_name", ""),
            "document_url": (rps or {}).get("document_url", ""),
            "rps_preview": {
                "cpmk": (rps or {}).get("cpmk", ""),
                "description": (rps or {}).get("description", ""),
                "references": (rps or {}).get("references", ""),
                "meetings": [
                    {
                        "meeting_number": meeting.get("meeting_number"),
                        "topic": meeting.get("topic", ""),
                        "materials": meeting.get("materials", ""),
                        "learning_outcome": meeting.get("learning_outcome", ""),
                        "is_exam": bool(meeting.get("is_exam")),
                    }
                    for meeting in ((rps or {}).get("meetings") or [])
                ],
            },
        })

    course_results = []
    for group in groups.values():
        group["course_count_status"] = "not_offered" if group["class_count"] == 0 else "offered"
        group["class_ids"] = group.pop("classes")
        course_results.append(group)
    course_results.sort(key=lambda item: (str(item.get("course_name") or "").lower(), str(item.get("course_code") or "")))
    class_results.sort(key=lambda item: (str(item.get("course_name") or "").lower(), str(item.get("class_name") or "").lower()))

    period_label = "Semua semester"
    if period_doc:
        period_label = period_doc.get("nama") or f"{period_doc.get('tahun', '')} {period_doc.get('semester', '')}".strip()
    prodi_name = await _analysis_program_label(requested_prodi, user, class_docs)
    return {
        "ok": True,
        "scope": {
            "is_admin": user.get("role") == "admin",
            "is_global": _analysis_has_global_program_scope(user),
            "can_select_prodi": _analysis_has_global_program_scope(user),
            "prodi_id": requested_prodi or (own_scope_values[0] if own_scope_values else ""),
            "prodi_name": prodi_name,
            "scope_values": scope_values,
            "structural_scope": structural_scope,
        },
        "prodi_list": await _analysis_program_options(),
        "period": {"id": raw_semester_id or "all", "label": period_label},
        "summary": {
            "total_courses": len(course_results),
            "offered_courses": len([item for item in course_results if item["class_count"] > 0]),
            "total_classes": total_classes,
            "rps_count": rps_count,
            "not_started": status_counts["not_started"],
            "draft": status_counts["draft"],
            "pending_approval": status_counts["pending"],
            "approved": status_counts["approved"],
            "rejected": status_counts["rejected"],
            "complete_rps": complete_count,
            "completeness_percent": round(complete_count / total_classes * 100, 1) if total_classes else 0,
            "approval_percent": round(status_counts["approved"] / rps_count * 100, 1) if rps_count else 0,
            "meeting_coverage_percent": round(total_meetings / (total_classes * 16) * 100, 1) if total_classes else 0,
        },
        "status_distribution": status_counts,
        "courses": course_results,
        "classes": class_results,
    }


@api_router.post("/prodi/rps/{class_id}/approval")
async def approve_prodi_rps(
    class_id: str,
    payload: RPSApprovalInput,
    user: Dict[str, Any] = Depends(require_program_analysis_user),
):
    """Approve atau reject RPS kelas oleh Ketua Prodi/Admin."""
    _, _, scope_values = await _analysis_resolve_program_scope(user)
    class_doc = await db.classes.find_one(
        {"id": class_id, "status": {"$ne": "deleted"}},
        {"_id": 0},
    )
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas mata kuliah tidak ditemukan")
    if not _analysis_has_global_program_scope(user):
        scope_lookup = {str(value).upper() for value in scope_values}
        class_in_scope = any(value.upper() in scope_lookup for value in _analysis_program_fields(class_doc))
        if not class_in_scope and not scope_values:
            class_in_scope = class_id in set(await lecturer_class_ids(user))
        if not class_in_scope:
            raise HTTPException(status_code=403, detail="Kelas berada di luar kewenangan Prodi Anda")

    rps = await db.rps.find_one({"class_id": class_id}, {"_id": 0})
    if not rps:
        raise HTTPException(status_code=404, detail="RPS untuk kelas ini belum tersedia")
    complete, missing = rps_completeness_from_document(rps)
    action = str(payload.action or "").strip().lower()
    if action in {"approve", "approved"}:
        if not complete:
            raise HTTPException(
                status_code=400,
                detail=f"RPS belum lengkap dan belum dapat disetujui. Kolom: {', '.join(missing[:8])}",
            )
        status = "approved"
    elif action in {"reject", "rejected"}:
        status = "rejected"
    else:
        raise HTTPException(status_code=400, detail="Aksi approval harus approve atau reject")

    reviewed_at = now_iso()
    note = str(payload.note or "").strip()
    update_doc = {
        "approval_status": status,
        "approval_note": note,
        "reviewed_at": reviewed_at,
        "reviewed_by": user.get("id", ""),
    }
    if status == "approved":
        update_doc.update({"approved_at": reviewed_at, "approved_by": user.get("id", "")})
    else:
        update_doc.update({"approved_at": "", "approved_by": ""})
    await db.rps.update_one({"class_id": class_id}, {"$set": update_doc})
    return {
        "ok": True,
        "class_id": class_id,
        "approval_status": status,
        "approval_note": note,
        "reviewed_at": reviewed_at,
        "reviewed_by": user.get("id", ""),
    }


async def require_rps_complete(class_id: str) -> None:
    complete, missing = await class_rps_complete(class_id)
    if not complete:
        detail = (
            "RPS belum disusun lengkap untuk kelas ini. "
            "Susun dan lengkapi RPS terlebih dahulu sebelum membuat materi, tugas, atau penilaian."
        )
        if missing:
            detail += f" Kolom yang belum lengkap: {', '.join(missing[:8])}."
        raise HTTPException(status_code=400, detail=detail)


class StudentAttendanceRecord(BaseModel):
    student_id: str
    status: str = "Hadir"
    note: str = ""


class AttendanceSessionInput(BaseModel):
    meeting_number: int
    date: str = ""
    topic: str = ""
    status: str = "open"
    records: List[StudentAttendanceRecord] = []


class SubmitPinInput(BaseModel):
    meeting_number: int
    pin_code: str


class SubmitQrInput(BaseModel):
    qr_content: str


class LockAttendanceInput(BaseModel):
    meeting_number: int
    locked: bool = True


def _attendance_qr_content(class_id: str, meeting_number: int, secret: str) -> str:
    return f"QRATT:{class_id}:{meeting_number}:{secret}"


def _qr_png_data_url(content: str) -> str:
    qr = segno.make(content, error="m")
    buf = io.BytesIO()
    qr.save(buf, kind="png", scale=6, border=2)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _attendance_expired(exp_str: str) -> bool:
    if not exp_str:
        return False
    exp_dt = parse_iso_datetime(exp_str)
    return bool(exp_dt) and datetime.now(timezone.utc) > exp_dt


def _mark_student_present(session: Dict[str, Any], student_id: str, note: str) -> None:
    records = session.get("records") or []
    rec = next((r for r in records if r.get("student_id") == student_id), None)
    if rec:
        rec["status"] = "Hadir"
        rec["note"] = note
        rec["updated_at"] = now_iso()
    else:
        records.append({
            "student_id": student_id,
            "status": "Hadir",
            "note": note,
            "updated_at": now_iso()
        })
    session["records"] = records
    session["updated_at"] = now_iso()


async def _load_attendance_session(class_id: str, meeting_number: int) -> Dict[str, Any]:
    return await db.attendance_sessions.find_one(
        {"class_id": class_id, "meeting_number": meeting_number}, {"_id": 0}
    ) or {
        "class_id": class_id,
        "meeting_number": meeting_number,
        "date": now_iso()[:10],
        "topic": f"Pertemuan {meeting_number}",
        "status": "draft",
        "pin_code": "",
        "pin_expires_at": "",
        "qr_code": "",
        "qr_content": "",
        "qr_expires_at": "",
        "expires_minutes": 15,
        "method": "",
        "locked": False,
        "records": []
    }


async def _load_rps_doc(class_doc: Dict[str, Any]) -> Dict[str, Any]:
    class_id = class_doc.get("id", "")
    rps = await db.rps.find_one({"class_id": class_id}, {"_id": 0})
    if not rps:
        rps = {
            "class_id": class_id,
            "cpmk": f"Memahami konsep dasar dan terapan dari mata kuliah {class_doc.get('course_name', 'ini')}.",
            "description": f"Rencana Pembelajaran Semester untuk {class_doc.get('course_name', '')}.",
            "references": "1. Buku Referensi Utama\n2. Modul Perkuliahan LMS",
            "document_url": "",
            "meetings": generate_default_16_meetings(class_doc.get("course_name", "Mata Kuliah")),
            "updated_at": now_iso()
        }
    rps.setdefault("course_code", class_doc.get("course_code", ""))
    rps.setdefault("semester", class_doc.get("semester", ""))
    rps.setdefault("sks", str(class_doc.get("sks") or ""))
    rps.setdefault("program_name", class_doc.get("program_name", ""))
    rps.setdefault("lecturer_name", class_doc.get("lecturer_name", ""))
    rps.setdefault("compiled_at", "")
    rps.setdefault("cpl_sikap", "")
    rps.setdefault("cpl_keterampilan_umum", "")
    rps.setdefault("cpl_pengetahuan", "")
    rps.setdefault("cpl_keterampilan_khusus", "")
    rps.setdefault("keterangan", "Kegiatan Proses Belajar (KPB); Kegiatan Penanganan Terstruktur (KPT); dan Kegiatan Mandiri (KM); Seminar (S); Praktikum/Praktik Lapangan (P/PL).")
    return rps


@api_router.get("/classes/{class_id}/rps")
async def get_class_rps(class_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    await require_class_access(class_id, user)
    class_doc = await db.classes.find_one({"id": class_id}, {"_id": 0}) or {}
    class_doc = await enrich_class_payload(class_doc) if class_doc else {}
    rps = await _load_rps_doc(class_doc)
    complete, missing = await class_rps_complete(class_id)
    rps["is_complete"] = complete
    rps["missing_fields"] = missing
    return rps


@api_router.post("/classes/{class_id}/rps/upload")
async def upload_class_rps_pdf(
    class_id: str,
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_admin),
):
    """Store an official RPS document and return a reviewable extraction draft."""

    class_doc = await require_class_access(class_id, user)
    class_doc = await enrich_class_payload(class_doc)
    filename = str(file.filename or "").strip()
    extension = Path(filename).suffix.lower()
    if not filename or extension not in {".pdf", ".docx", ".doc"}:
        raise HTTPException(status_code=400, detail="Dokumen RPS harus berupa PDF, DOCX, atau Word .doc")

    content = await file.read(RPS_UPLOAD_MAX_BYTES + 1)
    if len(content) > RPS_UPLOAD_MAX_BYTES:
        raise HTTPException(status_code=400, detail="Ukuran dokumen RPS maksimal 20 MB")
    if not content:
        raise HTTPException(status_code=400, detail="File dokumen RPS kosong")

    try:
        extraction = parse_rps_document(content, extension, class_doc)
    except RPSPdfDependencyError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except RPSPdfParseError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        await file.seek(0)

    file_doc = await save_uploaded_file_record(
        file,
        [
            "RPS",
            class_doc.get("academic_year", "Tahun Akademik"),
            class_doc.get("semester", "Semester"),
            class_doc.get("course_name", "Mata Kuliah"),
            class_doc.get("name", "Kelas"),
        ],
        class_doc.get("course_code") or class_id,
        class_doc.get("name") or class_doc.get("course_name", "Kelas"),
        user["id"],
        record_type="rps_document",
        sync_drive=False,
        lecturer_id=class_doc.get("lecturer_id", ""),
    )
    await db.stored_files.update_one(
        {"id": file_doc["id"]},
        {"$set": {"class_id": class_id, "rps_class_id": class_id}},
    )

    document_file = {
        key: file_doc.get(key, "")
        for key in (
            "id",
            "file_id",
            "file_name",
            "mime_type",
            "size",
            "file_url",
            "preview_url",
            "inline_url",
        )
    }
    uploaded_at = now_iso()
    await db.rps.update_one(
        {"class_id": class_id},
        {
            "$set": {
                "class_id": class_id,
                "document_url": file_doc.get("file_url", ""),
                "document_file": document_file,
                "document_uploaded_at": uploaded_at,
                "document_extraction": {
                    "uploaded_at": uploaded_at,
                    "stats": extraction.get("stats", {}),
                    "warnings": extraction.get("warnings", []),
                },
                "approval_status": "pending",
                "approval_note": "",
                "reviewed_at": "",
                "reviewed_by": "",
                "approved_at": "",
                "approved_by": "",
            }
        },
        upsert=True,
    )
    return {
        "document_url": file_doc.get("file_url", ""),
        "document_file": document_file,
        **extraction,
    }


@api_router.post("/classes/{class_id}/rps")
async def save_class_rps(class_id: str, payload: RPSInput, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    class_doc = await enrich_class_payload(class_doc)
    doc = payload.model_dump()
    doc.update({
        "class_id": class_id,
        "course_name": class_doc.get("course_name", ""),
        "course_code": doc.get("course_code") or class_doc.get("course_code", ""),
        "semester": doc.get("semester") or class_doc.get("semester", ""),
        "sks": doc.get("sks") or str(class_doc.get("sks") or ""),
        "program_name": doc.get("program_name") or class_doc.get("program_name", ""),
        "lecturer_name": doc.get("lecturer_name") or class_doc.get("lecturer_name", ""),
        "updated_at": now_iso(),
        "updated_by": user["id"],
        # Setiap perubahan dosen harus masuk review ulang oleh Kaprodi.
        "approval_status": "pending",
        "approval_note": "",
        "reviewed_at": "",
        "reviewed_by": "",
        "approved_at": "",
        "approved_by": "",
    })
    await db.rps.update_one({"class_id": class_id}, {"$set": doc}, upsert=True)
    complete, missing = await class_rps_complete(class_id)
    doc["is_complete"] = complete
    doc["missing_fields"] = missing
    return doc


@api_router.post("/classes/{class_id}/rps/generate-meetings")
async def generate_class_rps_meetings(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    meetings = generate_default_16_meetings(class_doc.get("course_name", "Mata Kuliah"))
    rps = await db.rps.find_one({"class_id": class_id}, {"_id": 0}) or {
        "class_id": class_id,
        "cpmk": f"Memahami konsep dasar dan terapan dari mata kuliah {class_doc.get('course_name', 'ini')}.",
        "description": f"Rencana Pembelajaran Semester untuk {class_doc.get('course_name', '')}.",
        "references": "1. Buku Referensi Utama\n2. Modul Perkuliahan LMS",
        "document_url": "",
    }
    rps["meetings"] = meetings
    rps["updated_at"] = now_iso()
    rps["updated_by"] = user["id"]
    rps.update({
        "approval_status": "pending",
        "approval_note": "",
        "reviewed_at": "",
        "reviewed_by": "",
        "approved_at": "",
        "approved_by": "",
    })
    await db.rps.update_one({"class_id": class_id}, {"$set": rps}, upsert=True)
    return rps


# ---------------------------------------------------------------------------
# Export RPS ke Excel / Word / PDF
# ---------------------------------------------------------------------------

_RPS_MEETING_COLUMNS = [
    ("Minggu/Pertemuan ke", 900),
    ("Kemampuan yang Diharapkan", 2600),
    ("Materi Pembelajaran", 2200),
    ("Bentuk/Metode/Pengalaman Belajar", 2000),
    ("Waktu", 900),
    ("Penilaian (Teknik)", 1400),
    ("Penilaian (Bobot)", 800),
    ("Penilaian (Indikator)", 1700),
    ("Penilaian (Kriteria)", 1700),
    ("Tugas/Aktivitas", 1700),
]


def _rps_export_value(rps: Dict[str, Any], key: str) -> str:
    return str(rps.get(key) or "").strip()


def _rps_export_filename(class_doc: Dict[str, Any], ext: str) -> str:
    import re as _re
    base = class_doc.get("course_code") or class_doc.get("course_name") or "RPS"
    base = _re.sub(r"[^A-Za-z0-9\-_]+", "-", str(base)).strip("-")
    return f"RPS-{base or 'Kelas'}.{ext}"


def build_rps_xlsx(rps: Dict[str, Any], class_doc: Dict[str, Any], settings: Dict[str, Any], stream) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RPS"
    sheet.sheet_view.showGridLines = False

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    label_fill = PatternFill("solid", fgColor="F2F2F2")
    title_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)
    normal = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header = (settings.get("kop_letterhead") or "").splitlines() or ["RENCANA PEMBELAJARAN SEMESTER (RPS)"]
    for line in header[:2]:
        sheet.append([line])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
        cell = sheet.cell(row=sheet.max_row, column=1)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center
        sheet.row_dimensions[sheet.max_row].height = 20

    sheet.append(["RENCANA PEMBELAJARAN SEMESTER (RPS)"])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
    sheet.cell(row=sheet.max_row, column=1).font = title_font
    sheet.cell(row=sheet.max_row, column=1).alignment = center
    sheet.append([f"{_rps_export_value(rps, 'course_code')} - {_rps_export_value(rps, 'program_name') or class_doc.get('program_name', '')} - {class_doc.get('course_name', '')}"])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
    sheet.cell(row=sheet.max_row, column=1).alignment = center
    sheet.append([])

    identity = [
        ("Kode Mata Kuliah", _rps_export_value(rps, "course_code")),
        ("Semester / SKS", f"{_rps_export_value(rps, 'semester')} / {_rps_export_value(rps, 'sks')} SKS"),
        ("Program Studi", _rps_export_value(rps, "program_name") or class_doc.get("program_name", "")),
        ("Dosen Pengampu", _rps_export_value(rps, "lecturer_name") or class_doc.get("lecturer_name", "")),
        ("Tanggal Penyusunan", _rps_export_value(rps, "compiled_at")),
        ("Rombongan Belajar", f"{class_doc.get('course_name', '')} - {class_doc.get('name', '')}"),
    ]
    for label, value in identity:
        sheet.append([label, value, "", "", "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=2, end_row=sheet.max_row, end_column=10)
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=1).fill = label_fill
        sheet.cell(row=sheet.max_row, column=1).alignment = normal
        sheet.cell(row=sheet.max_row, column=2).alignment = normal
        sheet.cell(row=sheet.max_row, column=2).border = border
    sheet.append([])

    def append_section(title: str, text: str) -> None:
        sheet.append([title, "", "", "", "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
        sheet.cell(row=sheet.max_row, column=1).font = sub_font
        sheet.cell(row=sheet.max_row, column=1).fill = header_fill
        sheet.append([text, "", "", "", "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
        sheet.cell(row=sheet.max_row, column=1).alignment = normal
        sheet.cell(row=sheet.max_row, column=1).border = border

    append_section("CAPAIAN PEMBELAJARAN LULUSAN (CPL) - Sikap", _rps_export_value(rps, "cpl_sikap"))
    append_section("CPL - Keterampilan Umum", _rps_export_value(rps, "cpl_keterampilan_umum"))
    append_section("CPL - Pengetahuan", _rps_export_value(rps, "cpl_pengetahuan"))
    append_section("CPL - Keterampilan Khusus", _rps_export_value(rps, "cpl_keterampilan_khusus"))
    append_section("CPMK", _rps_export_value(rps, "cpmk"))
    append_section("Deskripsi Mata Kuliah", _rps_export_value(rps, "description"))
    append_section("Daftar Referensi", _rps_export_value(rps, "references"))

    sheet.append([])
    sheet.append(["TABEL PEMBELAJARAN (16 PERTEMUAN)", "", "", "", "", "", "", "", "", ""])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
    sheet.cell(row=sheet.max_row, column=1).font = sub_font
    sheet.cell(row=sheet.max_row, column=1).fill = header_fill

    header_row = sheet.max_row + 1
    for col_idx, (col_name, _width) in enumerate(_RPS_MEETING_COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    sheet.row_dimensions[header_row].height = 30

    for meeting in (rps.get("meetings") or []):
        row_vals = [
            meeting.get("meeting_number", ""),
            meeting.get("learning_outcome") or meeting.get("sub_cpmk") or "",
            meeting.get("materials") or "",
            meeting.get("method") or "",
            meeting.get("waktu") or "",
            meeting.get("penilaian_teknik") or "",
            meeting.get("penilaian_bobot") or "",
            meeting.get("penilaian_indikator") or "",
            meeting.get("penilaian_kriteria") or "",
            meeting.get("assignments") or "",
        ]
        sheet.append(row_vals)
        for col_idx in range(1, 11):
            cell = sheet.cell(row=sheet.max_row, column=col_idx)
            cell.font = Font(size=9)
            cell.alignment = normal
            cell.border = border

    if (rps.get("keterangan") or "").strip():
        sheet.append([])
        sheet.append(["KETERANGAN", "", "", "", "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
        sheet.cell(row=sheet.max_row, column=1).font = sub_font
        sheet.cell(row=sheet.max_row, column=1).fill = header_fill
        sheet.append([str(rps.get("keterangan") or ""), "", "", "", "", "", "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=10)
        sheet.cell(row=sheet.max_row, column=1).alignment = normal

    widths = [18, 32, 30, 28, 12, 20, 12, 24, 24, 22]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx)].width = width
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    workbook.save(stream)


def _docx_escape(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _docx_par(text: str, *, bold: bool = False, size: int = 22, center: bool = False, after: int = 80) -> str:
    ppr = ""
    if center:
        ppr += '<w:jc w:val="center"/>'
    if after:
        ppr += f'<w:spacing w:after="{after}"/>'
    rpr = f"<w:b/>" if bold else ""
    if size:
        rpr += f'<w:sz w:val="{size}"/><w:szCs w:val="{size}"/>'
    return (
        f'<w:p><w:pPr>{ppr}</w:pPr>'
        f'<w:r><w:rPr>{rpr}</w:rPr><w:t xml:space="preserve">{_docx_escape(text)}</w:t></w:r></w:p>'
    )


def _docx_label_block(label: str, value: str) -> str:
    rpr = "<w:b/>"
    return (
        f'<w:p><w:pPr><w:spacing w:after="40"/></w:pPr>'
        f'<w:r><w:rPr>{rpr}</w:rPr><w:t>{_docx_escape(label)}</w:t></w:r>'
        f'<w:r><w:t xml:space="preserve"> : {_docx_escape(value)}</w:t></w:r></w:p>'
    )


def _docx_section(title: str, value: str) -> str:
    return (
        _docx_par(title, bold=True, size=24, after=40)
        + _docx_par(value if value else "-", size=20, after=80)
    )


def _docx_table(rows: List[List[str]]) -> str:
    borders = (
        '<w:tblBorders>'
        '<w:top w:val="single" w:sz="4" w:color="999999"/>'
        '<w:left w:val="single" w:sz="4" w:color="999999"/>'
        '<w:bottom w:val="single" w:sz="4" w:color="999999"/>'
        '<w:right w:val="single" w:sz="4" w:color="999999"/>'
        '<w:insideH w:val="single" w:sz="4" w:color="999999"/>'
        '<w:insideV w:val="single" w:sz="4" w:color="999999"/>'
        '</w:tblBorders>'
    )
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for _, width in _RPS_MEETING_COLUMNS)
    trs = []
    for row_idx, row in enumerate(rows):
        tcs = []
        for col_idx, (text, _width) in enumerate(zip(row, _RPS_MEETING_COLUMNS)):
            width = _RPS_MEETING_COLUMNS[col_idx][1]
            shade = '<w:shd w:val="clear" w:fill="D9E2F3"/>' if row_idx == 0 else ""
            rpr = "<w:b/>" if row_idx == 0 else ""
            tc = (
                f'<w:tc><w:tcPr><w:tcW w:w="{width}" w:type="dxa"/>{shade}'
                f'<w:vAlign w:val="top"/></w:tcPr>'
                f'<w:p><w:r><w:rPr>{rpr}<w:sz w:val="16"/><w:szCs w:val="16"/></w:rPr>'
                f'<w:t xml:space="preserve">{_docx_escape(text)}</w:t></w:r></w:p></w:tc>'
            )
            tcs.append(tc)
        trs.append(f"<w:tr>{''.join(tcs)}</w:tr>")
    return (
        f'<w:tbl><w:tblPr><w:tblW w:w="15900" w:type="dxa"/>{borders}'
        f'<w:tblLayout w:type="fixed"/></w:tblPr>'
        f'<w:tblGrid>{grid}</w:tblGrid>{"".join(trs)}</w:tbl>'
    )


def build_rps_docx(rps: Dict[str, Any], class_doc: Dict[str, Any], settings: Dict[str, Any]) -> bytes:
    header_lines = (settings.get("kop_letterhead") or "").splitlines() or ["RENCANA PEMBELAJARAN SEMESTER (RPS)"]
    body = []
    for line in header_lines[:2]:
        body.append(_docx_par(line, bold=True, size=26, center=True, after=20))
    body.append(_docx_par("RENCANA PEMBELAJARAN SEMESTER (RPS)", bold=True, size=30, center=True, after=40))
    body.append(_docx_par(
        f"{_rps_export_value(rps, 'course_code')} - {_rps_export_value(rps, 'program_name') or class_doc.get('program_name', '')} - {class_doc.get('course_name', '')}",
        bold=True, size=22, center=True, after=200,
    ))

    identity = [
        ("Kode Mata Kuliah", _rps_export_value(rps, "course_code")),
        ("Semester / SKS", f"{_rps_export_value(rps, 'semester')} / {_rps_export_value(rps, 'sks')} SKS"),
        ("Program Studi", _rps_export_value(rps, "program_name") or class_doc.get("program_name", "")),
        ("Dosen Pengampu", _rps_export_value(rps, "lecturer_name") or class_doc.get("lecturer_name", "")),
        ("Tanggal Penyusunan", _rps_export_value(rps, "compiled_at")),
        ("Rombongan Belajar", f"{class_doc.get('course_name', '')} - {class_doc.get('name', '')}"),
    ]
    for label, value in identity:
        body.append(_docx_label_block(label, value))
    body.append(_docx_par("", after=120))

    body.append(_docx_section("CPL - Sikap", _rps_export_value(rps, "cpl_sikap")))
    body.append(_docx_section("CPL - Keterampilan Umum", _rps_export_value(rps, "cpl_keterampilan_umum")))
    body.append(_docx_section("CPL - Pengetahuan", _rps_export_value(rps, "cpl_pengetahuan")))
    body.append(_docx_section("CPL - Keterampilan Khusus", _rps_export_value(rps, "cpl_keterampilan_khusus")))
    body.append(_docx_section("CPMK", _rps_export_value(rps, "cpmk")))
    body.append(_docx_section("Deskripsi Mata Kuliah", _rps_export_value(rps, "description")))
    body.append(_docx_section("Daftar Referensi", _rps_export_value(rps, "references")))

    body.append(_docx_par("TABEL PEMBELAJARAN (16 PERTEMUAN)", bold=True, size=24, after=80))
    rows = [[name for name, _width in _RPS_MEETING_COLUMNS]]
    for meeting in (rps.get("meetings") or []):
        rows.append([
            str(meeting.get("meeting_number", "")),
            meeting.get("learning_outcome") or meeting.get("sub_cpmk") or "",
            meeting.get("materials") or "",
            meeting.get("method") or "",
            meeting.get("waktu") or "",
            meeting.get("penilaian_teknik") or "",
            meeting.get("penilaian_bobot") or "",
            meeting.get("penilaian_indikator") or "",
            meeting.get("penilaian_kriteria") or "",
            meeting.get("assignments") or "",
        ])
    body.append(_docx_table(rows))

    if (rps.get("keterangan") or "").strip():
        body.append(_docx_par("", after=60))
        body.append(_docx_section("KETERANGAN", _rps_export_value(rps, "keterangan")))

    body_xml = "".join(body)
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body_xml}"
        '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        '<w:pgMar w:top="720" w:right="720" w:bottom="720" w:left="720" w:header="360" w:footer="360" w:gutter="0"/>'
        "</w:sectPr></w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        "</Relationships>"
    )
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("word/document.xml", document_xml)
    return stream.getvalue()


def _pdf_split_line(text: str, width: int) -> List[str]:
    text = _pdf_text(text)
    if len(text) <= width:
        return [text] if text else [""]
    out = []
    for chunk in text.split(" "):
        if not chunk:
            continue
        if out and len(out[-1]) + len(chunk) + 1 <= width:
            out[-1] += " " + chunk
        elif len(chunk) > width:
            start = 0
            while start < len(chunk):
                out.append(chunk[start:start + width])
                start += width
        else:
            out.append(chunk)
    return out or [""]


def build_rps_pdf(rps: Dict[str, Any], class_doc: Dict[str, Any], settings: Dict[str, Any]) -> bytes:
    lines: List[str] = []
    header_lines = (settings.get("kop_letterhead") or "").splitlines() or ["RENCANA PEMBELAJARAN SEMESTER (RPS)"]
    lines.extend(header_lines[:2])
    lines.append("RENCANA PEMBELAJARAN SEMESTER (RPS)")
    lines.append(
        f"{_rps_export_value(rps, 'course_code')} - "
        f"{_rps_export_value(rps, 'program_name') or class_doc.get('program_name', '')} - "
        f"{class_doc.get('course_name', '')}"
    )
    lines.append("")
    identity = [
        ("Kode MK", _rps_export_value(rps, "course_code")),
        ("Semester/SKS", f"{_rps_export_value(rps, 'semester')} / {_rps_export_value(rps, 'sks')} SKS"),
        ("Prodi", _rps_export_value(rps, "program_name") or class_doc.get("program_name", "")),
        ("Dosen Pengampu", _rps_export_value(rps, "lecturer_name") or class_doc.get("lecturer_name", "")),
        ("Tanggal Penyusunan", _rps_export_value(rps, "compiled_at")),
        ("Rombongan Belajar", f"{class_doc.get('course_name', '')} - {class_doc.get('name', '')}"),
    ]
    for label, value in identity:
        lines.append(f"{label}: {value}")
    lines.append("")
    sections = [
        ("CPL - Sikap", _rps_export_value(rps, "cpl_sikap")),
        ("CPL - Keterampilan Umum", _rps_export_value(rps, "cpl_keterampilan_umum")),
        ("CPL - Pengetahuan", _rps_export_value(rps, "cpl_pengetahuan")),
        ("CPL - Keterampilan Khusus", _rps_export_value(rps, "cpl_keterampilan_khusus")),
        ("CPMK", _rps_export_value(rps, "cpmk")),
        ("Deskripsi Mata Kuliah", _rps_export_value(rps, "description")),
        ("Daftar Referensi", _rps_export_value(rps, "references")),
    ]
    for title, value in sections:
        lines.append(f"[{title}]")
        for wrapped in _pdf_split_line(value, 100):
            lines.append(f"   {wrapped}")
        lines.append("")
    lines.append("=" * 100)
    lines.append("TABEL PEMBELAJARAN (16 PERTEMUAN)")
    lines.append("=" * 100)
    for meeting in (rps.get("meetings") or []):
        lines.append("")
        lines.append(f"PERTEMUAN {meeting.get('meeting_number', '?')} - {meeting.get('topic') or ''}")
        pairs = [
            ("Kemampuan yang diharapkan", meeting.get("learning_outcome") or meeting.get("sub_cpmk") or ""),
            ("Materi Pembelajaran", meeting.get("materials") or ""),
            ("Bentuk/Metode/Pengalaman Belajar", meeting.get("method") or ""),
            ("Waktu", meeting.get("waktu") or ""),
            ("Penilaian Teknik", meeting.get("penilaian_teknik") or ""),
            ("Penilaian Bobot", meeting.get("penilaian_bobot") or ""),
            ("Penilaian Indikator", meeting.get("penilaian_indikator") or ""),
            ("Penilaian Kriteria", meeting.get("penilaian_kriteria") or ""),
            ("Tugas/Aktivitas", meeting.get("assignments") or ""),
        ]
        for label, value in pairs:
            if not str(value or "").strip():
                continue
            for idx, wrapped in enumerate(_pdf_split_line(value, 90)):
                lines.append(f"   {label}: {wrapped}" if idx == 0 else f"   {' ' * (len(label) + 2)}{wrapped}")
    if (rps.get("keterangan") or "").strip():
        lines.append("")
        lines.append("[KETERANGAN]")
        for wrapped in _pdf_split_line(_rps_export_value(rps, "keterangan"), 100):
            lines.append(f"   {wrapped}")

    chunks = [lines[index:index + 55] for index in range(0, len(lines), 55)] or [[]]
    objects: List[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    page_ids = [3 + index * 2 for index in range(len(chunks))]
    content_ids = [page_id + 1 for page_id in page_ids]
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("ascii"))
    for page_id, content_id, page_lines in zip(page_ids, content_ids, chunks):
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 {3 + len(chunks) * 2} 0 R >> >> /Contents {content_id} 0 R >>".encode("ascii"))
        content_lines = ["BT", "/F1 8 Tf", "30 760 Td", "10 TL"]
        for line in page_lines:
            content_lines.append(f"({_pdf_text(line[:160])}) Tj")
            content_lines.append("T*")
        content_lines.append("ET")
        content = "\n".join(content_lines).encode("latin-1", "replace")
        objects.append(f"<< /Length {len(content)} >>\nstream\n".encode("ascii") + content + b"\nendstream")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii"))
    return bytes(output)


@api_router.get("/classes/{class_id}/rps/export.xlsx")
async def export_class_rps_xlsx(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    class_doc = await enrich_class_payload(class_doc)
    rps = await _load_rps_doc(class_doc)
    settings = await get_app_settings_cached()
    stream = io.BytesIO()
    build_rps_xlsx(rps, class_doc, settings, stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={_rps_export_filename(class_doc, 'xlsx')}"},
    )


@api_router.get("/classes/{class_id}/rps/export.docx")
async def export_class_rps_docx(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    class_doc = await enrich_class_payload(class_doc)
    rps = await _load_rps_doc(class_doc)
    settings = await get_app_settings_cached()
    content = build_rps_docx(rps, class_doc, settings)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={_rps_export_filename(class_doc, 'docx')}"},
    )


@api_router.get("/classes/{class_id}/rps/export.pdf")
async def export_class_rps_pdf(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    class_doc = await enrich_class_payload(class_doc)
    rps = await _load_rps_doc(class_doc)
    settings = await get_app_settings_cached()
    content = build_rps_pdf(rps, class_doc, settings)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={_rps_export_filename(class_doc, 'pdf')}"},
    )


@api_router.get("/classes/{class_id}/attendance")
async def get_class_attendance(class_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    class_doc = await require_class_access(class_id, user)
    enrolled_ids = list(dict.fromkeys(class_doc.get("student_ids", [])))
    students_docs = await db.users.find({"$or": [{"id": {"$in": enrolled_ids}}, {"username": {"$in": enrolled_ids}}, {"nim": {"$in": enrolled_ids}}]}, {"_id": 0}).to_list(1000)
    
    sessions = await db.attendance_sessions.find({"class_id": class_id}, {"_id": 0}).to_list(100)
    sessions_by_meeting = {s["meeting_number"]: s for s in sessions}

    rps = await db.rps.find_one({"class_id": class_id}, {"_id": 0})
    rps_meetings = {m["meeting_number"]: m for m in (rps.get("meetings") if rps else [])}

    meeting_slots = []
    for i in range(1, 17):
        session = sessions_by_meeting.get(i)
        rps_m = rps_meetings.get(i, {})
        default_topic = rps_m.get("topic") or ("Evaluasi Tengah Semester (UTS)" if i == 8 else ("Evaluasi Akhir Semester (UAS)" if i == 16 else f"Pertemuan {i}"))
        
        if not session:
            session_data = {
                "class_id": class_id,
                "meeting_number": i,
                "date": now_iso()[:10],
                "topic": default_topic,
                "status": "draft",
                "pin_code": "",
                "pin_expires_at": "",
                "qr_code": "",
                "qr_content": "",
                "qr_expires_at": "",
                "expires_minutes": 15,
                "method": "",
                "locked": False,
                "records": []
            }
        else:
            session_data = {
                **session,
                "pin_code": session.get("pin_code", ""),
                "pin_expires_at": session.get("pin_expires_at", ""),
                "qr_code": session.get("qr_code", ""),
                "qr_content": session.get("qr_content", ""),
                "qr_expires_at": session.get("qr_expires_at", ""),
                "expires_minutes": session.get("expires_minutes", 15),
                "method": session.get("method", ""),
                "locked": session.get("locked", False),
            }

        meeting_slots.append(session_data)

    recap = []
    for sid in enrolled_ids:
        s_doc = next((u for u in students_docs if u.get("id") == sid or u.get("username") == sid or u.get("nim") == sid), {})
        s_name = s_doc.get("name") or s_doc.get("nama") or f"Mahasiswa ({sid})"
        s_nim = s_doc.get("nim") or s_doc.get("username") or sid
        
        hadir_count = 0
        izin_count = 0
        sakit_count = 0
        alpa_count = 0
        total_open_sessions = 0

        for m in meeting_slots:
            if m.get("status") in ["open", "closed"]:
                total_open_sessions += 1
                rec = next((r for r in (m.get("records") or []) if r.get("student_id") == sid), None)
                st = rec.get("status") if rec else "Alpa"
                if st == "Hadir": hadir_count += 1
                elif st == "Izin": izin_count += 1
                elif st == "Sakit": sakit_count += 1
                else: alpa_count += 1

        pct = round((hadir_count / total_open_sessions * 100), 1) if total_open_sessions > 0 else 100.0
        recap.append({
            "student_id": sid,
            "student_name": s_name,
            "student_nim": s_nim,
            "hadir": hadir_count,
            "izin": izin_count,
            "sakit": sakit_count,
            "alpa": alpa_count,
            "total_open": total_open_sessions,
            "percentage": pct,
            "is_eligible_exam": pct >= 75.0
        })

    return {
        "class_id": class_id,
        "meetings": meeting_slots,
        "recap": recap,
        "total_students": len(enrolled_ids)
    }


@api_router.post("/classes/{class_id}/attendance/session")
async def save_attendance_session(class_id: str, payload: AttendanceSessionInput, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    m_no = payload.meeting_number
    if m_no < 1 or m_no > 16:
        raise HTTPException(status_code=400, detail="Nomor pertemuan harus antara 1 sampai 16")
    
    doc = payload.model_dump()
    doc.update({
        "class_id": class_id,
        "meeting_number": m_no,
        "updated_at": now_iso(),
        "updated_by": user["id"]
    })
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": m_no}, {"$set": doc}, upsert=True)
    return doc


@api_router.post("/classes/{class_id}/attendance/lock")
async def lock_attendance_session(class_id: str, payload: LockAttendanceInput, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    m_no = payload.meeting_number
    if m_no < 1 or m_no > 16:
        raise HTTPException(status_code=400, detail="Nomor pertemuan harus antara 1 sampai 16")

    session = await _load_attendance_session(class_id, m_no)
    session["locked"] = bool(payload.locked)
    session["status"] = "closed" if session["locked"] else "draft"
    session["updated_at"] = now_iso()
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": m_no}, {"$set": session}, upsert=True)
    return {"meeting_number": m_no, "locked": session["locked"], "status": session["status"]}


@api_router.post("/classes/{class_id}/attendance/generate-pin")
async def generate_attendance_pin(class_id: str, meeting_number: int, minutes: int = 15, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    if meeting_number < 1 or meeting_number > 16:
        raise HTTPException(status_code=400, detail="Nomor pertemuan harus antara 1 sampai 16")
    session = await _load_attendance_session(class_id, meeting_number)
    if session.get("locked"):
        raise HTTPException(status_code=400, detail="Sesi presensi pertemuan ini dikunci oleh dosen. Buka kunci terlebih dahulu untuk generate PIN.")
    minutes = max(1, min(int(minutes), 1440))
    import random
    pin = f"{random.randint(1000, 9999)}"
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=minutes)).isoformat()

    session.update({
        "status": "open",
        "method": "pin",
        "expires_minutes": minutes,
        "pin_code": pin,
        "pin_expires_at": expires_at,
        "updated_at": now_iso(),
    })
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": meeting_number}, {"$set": session}, upsert=True)
    return {"pin_code": pin, "pin_expires_at": expires_at, "meeting_number": meeting_number, "minutes": minutes}


@api_router.post("/classes/{class_id}/attendance/generate-qr")
async def generate_attendance_qr(class_id: str, meeting_number: int, minutes: int = 15, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    if meeting_number < 1 or meeting_number > 16:
        raise HTTPException(status_code=400, detail="Nomor pertemuan harus antara 1 sampai 16")
    session = await _load_attendance_session(class_id, meeting_number)
    if session.get("locked"):
        raise HTTPException(status_code=400, detail="Sesi presensi pertemuan ini dikunci oleh dosen. Buka kunci terlebih dahulu untuk generate QR.")
    minutes = max(1, min(int(minutes), 1440))
    secret = secrets.token_hex(16)
    content = _attendance_qr_content(class_id, meeting_number, secret)
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=minutes)).isoformat()

    session.update({
        "status": "open",
        "method": "qr",
        "expires_minutes": minutes,
        "qr_code": secret,
        "qr_content": content,
        "qr_expires_at": expires_at,
        "updated_at": now_iso(),
    })
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": meeting_number}, {"$set": session}, upsert=True)
    return {
        "meeting_number": meeting_number,
        "minutes": minutes,
        "qr_code": secret,
        "qr_content": content,
        "qr_expires_at": expires_at,
        "qr_image_data_url": _qr_png_data_url(content),
    }


@api_router.post("/classes/{class_id}/attendance/submit-qr")
async def submit_attendance_qr(class_id: str, payload: SubmitQrInput, user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Hanya mahasiswa yang dapat melakukan presensi dengan QR")

    parts = str(payload.qr_content or "").strip().split(":")
    if len(parts) != 4 or parts[0] != "QRATT":
        raise HTTPException(status_code=400, detail="Kode QR presensi tidak dikenali")
    qr_class_id, meeting_str, secret = parts[1], parts[2], parts[3]
    if qr_class_id != class_id:
        raise HTTPException(status_code=400, detail="Kode QR presensi bukan untuk kelas ini")
    try:
        meeting_number = int(meeting_str)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Kode QR presensi tidak valid")

    session = await _load_attendance_session(class_id, meeting_number)
    if session.get("locked"):
        raise HTTPException(status_code=400, detail="Sesi presensi pertemuan ini dikunci oleh dosen")
    if session.get("status") != "open":
        raise HTTPException(status_code=400, detail="Sesi presensi belum dibuka oleh dosen")
    if str(session.get("qr_code", "")).strip() != str(secret).strip():
        raise HTTPException(status_code=400, detail="Kode QR presensi tidak valid")
    if _attendance_expired(str(session.get("qr_expires_at", ""))):
        raise HTTPException(status_code=400, detail="Kode QR presensi telah kedaluwarsa")

    _mark_student_present(session, user["id"], "Hadir via QR")
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": meeting_number}, {"$set": session})
    return {"ok": True, "message": f"Presensi Pertemuan {meeting_number} Berhasil! Status: Hadir"}


@api_router.post("/classes/{class_id}/attendance/submit-pin")
async def submit_attendance_pin(class_id: str, payload: SubmitPinInput, user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Hanya mahasiswa yang dapat melakukan presensi dengan PIN")
    
    session = await _load_attendance_session(class_id, payload.meeting_number)
    if session.get("locked"):
        raise HTTPException(status_code=400, detail="Sesi presensi pertemuan ini dikunci oleh dosen")
    if session.get("status") != "open":
        raise HTTPException(status_code=400, detail="Sesi presensi belum dibuka oleh dosen")
    
    if str(session.get("pin_code", "")).strip() != str(payload.pin_code).strip():
        raise HTTPException(status_code=400, detail="Kode PIN presensi salah")
    
    if _attendance_expired(str(session.get("pin_expires_at", ""))):
        raise HTTPException(status_code=400, detail="Kode PIN presensi telah kedaluwarsa")
    
    _mark_student_present(session, user["id"], "Hadir via PIN")
    await db.attendance_sessions.update_one({"class_id": class_id, "meeting_number": payload.meeting_number}, {"$set": session})
    return {"ok": True, "message": f"Presensi Pertemuan {payload.meeting_number} Berhasil! Status: Hadir"}


# ==========================================
# BKD LAPORAN DOSEN & BUNDLE PORTOFOLIO APIs
# ==========================================

@api_router.get("/reports/lecturer/summary")
async def get_lecturer_reports_summary(
    semester_id: Optional[str] = None,
    user: Dict[str, Any] = Depends(get_current_user)
):
    class_ids = await lecturer_class_ids(user)
    classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0}).to_list(500)
    
    if semester_id and semester_id != "all":
        target_ta = await db.tahun_ajaran.find_one({"id": semester_id}, {"_id": 0})
        filtered_classes = []
        for c in classes:
            if c.get("tahun_ajaran_id") == semester_id:
                filtered_classes.append(c)
            elif target_ta:
                ta_year = target_ta.get("tahun") or target_ta.get("nama") or ""
                ta_sem = (target_ta.get("semester") or "").lower()
                c_year = c.get("academic_year") or ""
                c_sem = (c.get("semester") or "").lower()
                if ta_year and ta_sem and ta_year in c_year and ta_sem in c_sem:
                    filtered_classes.append(c)
        classes = filtered_classes

    summaries = []
    for c in classes:
        cid = c.get("id")
        rps_doc = await db.rps.find_one({"class_id": cid}, {"_id": 0})
        att_count = await db.attendance_sessions.count_documents({"class_id": cid})
        
        has_rps = bool(rps_doc and len(rps_doc.get("meetings", [])) > 0)
        has_att = att_count >= 1
        has_grades = True
        has_sk = True
        
        comp_count = sum([has_rps, has_att, has_grades, has_sk])
        pct = int((comp_count / 4.0) * 100)
        
        summaries.append({
            "class_id": cid,
            "course_name": c.get("course_name", "Mata Kuliah"),
            "class_code": c.get("name") or c.get("class_code", "01"),
            "program_name": c.get("program_name", "Program Studi"),
            "academic_year": c.get("academic_year", ""),
            "semester": c.get("semester", ""),
            "student_count": len(c.get("student_ids", [])),
            "rps_complete": has_rps,
            "attendance_complete": has_att,
            "grades_complete": has_grades,
            "sk_complete": has_sk,
            "completion_percentage": pct,
            "status": "Lengkap (100%)" if pct == 100 else "Dalam Proses"
        })
        
    return {"classes": summaries}


@api_router.get("/public/reports/bkd-bundle/{class_id}")
async def get_public_bkd_bundle(class_id: str):
    class_doc = await db.classes.find_one({"id": class_id}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
        
    lecturer_id = class_doc.get("lecturer_id")
    lecturer_doc = await db.users.find_one({"id": lecturer_id}, {"_id": 0, "password_hash": 0}) or {
        "name": class_doc.get("lecturer_name", "Syahrul Anwar"),
        "nidn": "0402030001",
        "email": "dosen@kampus.ac.id"
    }
    
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or default_app_settings()
    
    rps_doc = await db.rps.find_one({"class_id": class_id}, {"_id": 0})
    if not rps_doc:
        rps_doc = {
            "class_id": class_id,
            "cpmk": f"Memahami konsep dasar dan terapan dari mata kuliah {class_doc.get('course_name')}.",
            "description": "Rencana Pembelajaran Semester standar perguruan tinggi.",
            "references": "Buku Ajar, Jurnal Ilmiah, Dokumentasi Resmi.",
            "document_url": "",
            "meetings": [
                {
                    "meeting_number": i + 1,
                    "topic": f"Pertemuan {i + 1}: UTS" if i + 1 == 8 else (f"Pertemuan {i + 1}: UAS" if i + 1 == 16 else f"Pertemuan {i + 1}: Topik Bahasan Ke-{i + 1}"),
                    "sub_topic": "Pembahasan konsep dasar & studi kasus terapan",
                    "method": "Tatap Muka / Blended Learning",
                    "materials": "Slide Presentasi, Modul PDF",
                    "is_exam": i + 1 in [8, 16]
                }
                for i in range(16)
            ]
        }
        
    admin_user = {"id": lecturer_doc.get("id") or "admin", "role": "admin"}
    try:
        att_data = await get_class_attendance(class_id, admin_user)
    except Exception:
        att_data = {"meetings": [], "recap": []}
    try:
        recap_list = await build_grade_recap(user=admin_user, class_id=class_id, use_snapshots=False)
        if recap_list and len(recap_list) > 0:
            raw_students = recap_list[0].get("students", [])
            grade_data = []
            for st in raw_students:
                comp = st.get("component_scores", {})
                tugas_score = comp.get("tugas")
                presensi_score = comp.get("presensi")
                uts_score = comp.get("uts")
                uas_score = comp.get("uas")
                
                grade_data.append({
                    "student_id": st.get("student_id"),
                    "student_nim": st.get("student_nim") or st.get("nim") or "-",
                    "student_name": st.get("student_name") or st.get("name") or "-",
                    "assignment_score": round(tugas_score, 1) if tugas_score is not None else 0,
                    "attendance_score": round(presensi_score, 1) if presensi_score is not None else (round(tugas_score, 1) if tugas_score is not None else 0),
                    "uts_score": round(uts_score, 1) if uts_score is not None else 0,
                    "uas_score": round(uas_score, 1) if uas_score is not None else 0,
                    "final_score": st.get("weighted_grade") if st.get("weighted_grade") is not None else (st.get("average") or 0),
                    "grade_letter": st.get("grade_predicate") or "A"
                })
        else:
            grade_data = []
    except Exception as e:
        logger.error(f"Error building public grade recap for class {class_id}: {e}")
        grade_data = []
        
    wadir = None
    direktur = None
    kaprodi_pejabat = None
    try:
        wadir = await _active_pejabat(db, jabatan_kode="WADIR1")
        direktur = await _active_pejabat(db, jabatan_kode="DIREKTUR")
        kaprodi_pejabat = await _active_pejabat(
            db, jabatan_kode="KAPRODI", prodi_id=class_doc.get("program_id") or ""
        )
    except Exception:
        pass

    penetap = wadir or direktur
    if penetap:
        ident = f"NIP {penetap['nip']}" if penetap["nip"] else (f"NIDN {penetap['nidn']}" if penetap["nidn"] else "")
        signatory = f"{penetap['nama']} — {penetap['jabatan']}"
        signatory_detail = ident or ""
    else:
        signatory = "Dekan / Wakil Rektor Bidang Akademik"
        signatory_detail = ""

    kap_nama = (kaprodi_pejabat or {}).get("nama", "") or ""
    if not kap_nama and class_doc.get("program_id"):
        kap_nama = await _pejabat_name_fallback(db, class_doc.get("program_id", ""))
    kap_detail = ""
    if kaprodi_pejabat:
        kap_detail = (
            f"NIP {kaprodi_pejabat['nip']}"
            if kaprodi_pejabat["nip"]
            else (f"NIDN {kaprodi_pejabat['nidn']}" if kaprodi_pejabat["nidn"] else "")
        )

    return {
        "campus": {
            "name": settings.get("campus_name", "UNIVERSITAS NUGAS LAGI"),
            "logo_url": settings.get("campus_logo_url", ""),
            "kop_header_url": settings.get("kop_header_url", ""),
            "kop_footer_url": settings.get("kop_footer_url", ""),
            "academic_year": class_doc.get("academic_year", "2025/2026"),
            "semester": class_doc.get("semester", "Genap")
        },
        "lecturer": {
            "name": lecturer_doc.get("name", "Syahrul Anwar"),
            "nidn": lecturer_doc.get("nidn") or lecturer_doc.get("employee_id") or "0402030001",
            "email": lecturer_doc.get("email", "syahrul@kampus.ac.id")
        },
        "class_info": {
            "id": class_id,
            "course_name": class_doc.get("course_name"),
            "class_code": class_doc.get("name") or class_doc.get("class_code"),
            "program_name": class_doc.get("program_name"),
            "student_count": len(class_doc.get("student_ids", []))
        },
        "rps": rps_doc,
        "attendance": att_data,
        "grades": grade_data,
        "sk_info": {
            "sk_number": f"SK/AKAD/{class_doc.get('academic_year', '2025-2026').replace('/', '-')}/{class_id}",
            "issue_date": _fmt_date_id(datetime.now().astimezone().date()),
            "signatory": signatory,
            "signatory_detail": signatory_detail,
            "kaprodi_name": kap_nama,
            "kaprodi_detail": kap_detail,
        }
    }


async def _pejabat_name_fallback(db: PostgresDatabase, prodi_id: str) -> str:
    try:
        prodi = await db.programs.find_one({"id": prodi_id}, {"_id": 0, "kaprodi": 1})
        return str((prodi or {}).get("kaprodi") or "")
    except Exception:
        return ""


def _fmt_date_id(d: "date") -> str:
    bulan = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ]
    return f"{d.day:02d} {bulan[d.month - 1]} {d.year}"


@api_router.post("/settings/upload-logo")
async def upload_campus_logo(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File logo tidak valid")
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".svg"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Format logo harus berupa JPG, PNG, WEBP, atau SVG")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran logo maksimal 5 MB")

    file_token = secrets.token_hex(8)
    safe_name = safe_path_segment(file.filename) or f"logo{ext}"
    filename = f"logo_campus_{file_token[:8]}_{safe_name}"

    branding_dir = STORAGE_ROOT / "Branding"
    branding_dir.mkdir(parents=True, exist_ok=True)
    file_path = branding_dir / filename
    file_path.write_bytes(content)

    file_id = f"logo-campus-{file_token[:8]}"
    file_doc = {
        "id": file_id,
        "record_type": "campus_logo",
        "owner_user_id": user["id"],
        "file_name": safe_name,
        "original_name": file.filename,
        "mime_type": file.content_type or f"image/{ext.replace('.', '')}",
        "size": len(content),
        "storage_path": portable_storage_path_from_local_path(str(file_path)),
        "local_path": str(file_path),
        "local_available": True,
        "created_at": now_iso(),
    }
    await db.stored_files.update_one({"id": file_id}, {"$set": file_doc}, upsert=True)
    logo_url = f"/api/files/{file_id}/inline"
    await db.app_settings.update_one({"id": "main"}, {"$set": {"campus_logo_url": logo_url}}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return {"logo_url": logo_url}


@api_router.post("/settings/upload-app-logo")
async def upload_app_logo(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File logo aplikasi tidak valid")
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".svg"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Format logo aplikasi harus berupa JPG, PNG, WEBP, atau SVG")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran logo aplikasi maksimal 5 MB")

    file_token = secrets.token_hex(8)
    safe_name = safe_path_segment(file.filename) or f"app_logo{ext}"
    filename = f"logo_aplikasi_{file_token[:8]}_{safe_name}"

    branding_dir = STORAGE_ROOT / "Branding"
    branding_dir.mkdir(parents=True, exist_ok=True)
    file_path = branding_dir / filename
    file_path.write_bytes(content)

    file_id = f"app-logo-{file_token[:8]}"
    file_doc = {
        "id": file_id,
        "record_type": "app_logo",
        "owner_user_id": user["id"],
        "file_name": safe_name,
        "original_name": file.filename,
        "mime_type": file.content_type or f"image/{ext.replace('.', '')}",
        "size": len(content),
        "storage_path": portable_storage_path_from_local_path(str(file_path)),
        "local_path": str(file_path),
        "local_available": True,
        "created_at": now_iso(),
    }
    await db.stored_files.update_one({"id": file_id}, {"$set": file_doc}, upsert=True)
    app_logo_url = f"/api/files/{file_id}/inline"
    await db.app_settings.update_one({"id": "main"}, {"$set": {"app_logo_url": app_logo_url}}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return {"app_logo_url": app_logo_url}


@api_router.post("/settings/upload-kop-header")
async def upload_kop_header(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File KOP Header tidak valid")
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".svg"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Format header KOP harus berupa JPG, PNG, WEBP, atau SVG")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran header KOP maksimal 10 MB")

    file_token = secrets.token_hex(8)
    safe_name = safe_path_segment(file.filename) or f"kop_header{ext}"
    filename = f"kop_header_{file_token[:8]}_{safe_name}"

    branding_dir = STORAGE_ROOT / "Branding"
    branding_dir.mkdir(parents=True, exist_ok=True)
    file_path = branding_dir / filename
    file_path.write_bytes(content)

    file_id = f"kop-header-{file_token[:8]}"
    file_doc = {
        "id": file_id,
        "record_type": "kop_header",
        "owner_user_id": user["id"],
        "file_name": safe_name,
        "original_name": file.filename,
        "mime_type": file.content_type or f"image/{ext.replace('.', '')}",
        "size": len(content),
        "storage_path": portable_storage_path_from_local_path(str(file_path)),
        "local_path": str(file_path),
        "local_available": True,
        "created_at": now_iso(),
    }
    await db.stored_files.update_one({"id": file_id}, {"$set": file_doc}, upsert=True)
    kop_header_url = f"/api/files/{file_id}/inline"
    await db.app_settings.update_one({"id": "main"}, {"$set": {"kop_header_url": kop_header_url}}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return {"kop_header_url": kop_header_url}


@api_router.post("/settings/upload-kop-footer")
async def upload_kop_footer(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File KOP Footer tidak valid")
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".svg"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Format footer KOP harus berupa JPG, PNG, WEBP, atau SVG")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran footer KOP maksimal 10 MB")

    file_token = secrets.token_hex(8)
    safe_name = safe_path_segment(file.filename) or f"kop_footer{ext}"
    filename = f"kop_footer_{file_token[:8]}_{safe_name}"

    branding_dir = STORAGE_ROOT / "Branding"
    branding_dir.mkdir(parents=True, exist_ok=True)
    file_path = branding_dir / filename
    file_path.write_bytes(content)

    file_id = f"kop-footer-{file_token[:8]}"
    file_doc = {
        "id": file_id,
        "record_type": "kop_footer",
        "owner_user_id": user["id"],
        "file_name": safe_name,
        "original_name": file.filename,
        "mime_type": file.content_type or f"image/{ext.replace('.', '')}",
        "size": len(content),
        "storage_path": portable_storage_path_from_local_path(str(file_path)),
        "local_path": str(file_path),
        "local_available": True,
        "created_at": now_iso(),
    }
    await db.stored_files.update_one({"id": file_id}, {"$set": file_doc}, upsert=True)
    kop_footer_url = f"/api/files/{file_id}/inline"
    await db.app_settings.update_one({"id": "main"}, {"$set": {"kop_footer_url": kop_footer_url}}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return {"kop_footer_url": kop_footer_url}


app.include_router(api_router)

_FRONTEND_BUILD = Path(__file__).parent.parent / "frontend" / "build"
if _FRONTEND_BUILD.exists():
    from starlette.staticfiles import StaticFiles
    from starlette.responses import FileResponse

    class CachedStaticFiles(StaticFiles):
        def __init__(self, *args, cache_control: str, **kwargs):
            super().__init__(*args, **kwargs)
            self.cache_control = cache_control

        async def get_response(self, path: str, scope):
            response = await super().get_response(path, scope)
            if response.status_code == 200:
                response.headers["Cache-Control"] = self.cache_control
            return response

    if (_FRONTEND_BUILD / "static").exists():
        app.mount(
            "/static",
            CachedStaticFiles(
                directory=str(_FRONTEND_BUILD / "static"),
                cache_control="public, max-age=31536000, immutable",
            ),
            name="static-assets",
        )
    if (_FRONTEND_BUILD / "campus").exists():
        app.mount(
            "/campus",
            CachedStaticFiles(
                directory=str(_FRONTEND_BUILD / "campus"),
                cache_control="public, max-age=604800",
            ),
            name="campus-assets",
        )
    if (_FRONTEND_BUILD / "templates").exists():
        app.mount(
            "/templates",
            CachedStaticFiles(
                directory=str(_FRONTEND_BUILD / "templates"),
                cache_control="public, max-age=3600",
            ),
            name="download-templates",
        )

    @app.get("/manifest.json")
    async def serve_manifest():
        return FileResponse(
            str(_FRONTEND_BUILD / "manifest.json"),
            headers={"Cache-Control": "no-cache"},
        )

    @app.get("/app-icon.svg")
    async def serve_icon():
        return FileResponse(
            str(_FRONTEND_BUILD / "app-icon.svg"),
            headers={"Cache-Control": "public, max-age=86400"},
        )

    @app.get("/service-worker.js")
    async def serve_sw():
        return FileResponse(
            str(_FRONTEND_BUILD / "service-worker.js"),
            headers={"Cache-Control": "no-cache, must-revalidate"},
        )

    @app.api_route("/{full_path:path}", methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"])
    async def serve_spa(full_path: str, request: Request):
        """Fallback to index.html for React SPA routing with live share metadata."""
        if full_path.startswith("api/") or full_path.startswith("api"):
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail=f"API route not found: /{full_path}")
        if request.method != "GET":
            from fastapi import HTTPException
            raise HTTPException(status_code=405, detail="Method Not Allowed")
        index_file = _FRONTEND_BUILD / "index.html"
        if index_file.exists():
            content = index_file.read_text(encoding="utf-8")
            defaults = default_app_settings()
            settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or {}
            app_name = str(settings.get("app_name") or defaults["app_name"]).strip()
            description = str(settings.get("meta_description") or defaults["meta_description"]).strip()
            safe_app_name = html.escape(app_name, quote=True)
            safe_description = html.escape(description, quote=True)

            def replace_meta(pattern: str, value: str, source: str) -> str:
                return re.sub(
                    pattern,
                    lambda match: f"{match.group(1)}{value}{match.group(2)}",
                    source,
                    count=1,
                    flags=re.IGNORECASE,
                )

            content = replace_meta(r'(<title>).*?(</title>)', safe_app_name, content)
            content = replace_meta(r'(<meta name="description" content=").*?("[^>]*>)', safe_description, content)
            content = replace_meta(r'(<meta property="og:description" content=").*?("[^>]*>)', safe_description, content)
            content = replace_meta(r'(<meta name="twitter:description" content=").*?("[^>]*>)', safe_description, content)
            content = replace_meta(r'(<meta property="og:title" content=").*?("[^>]*>)', safe_app_name, content)
            content = replace_meta(r'(<meta name="twitter:title" content=").*?("[^>]*>)', safe_app_name, content)
            return HTMLResponse(
                content=content,
                headers={"Cache-Control": "no-cache, must-revalidate"},
            )
        return {"detail": "Frontend build not found"}


@app.on_event("startup")
async def on_startup():
    global _database_backup_scheduler_task, _storage_maintenance_scheduler_task
    await db.connect()
    app.state.db = db
    await seed_data()
    storage_reconciliation = await reconcile_local_storage_paths()
    if storage_reconciliation["rebased"] or storage_reconciliation["missing"]:
        logger.info(
            "Rekonsiliasi path storage: %s diperbarui, %s file lokal belum tersedia",
            storage_reconciliation["rebased"],
            storage_reconciliation["missing"],
        )
    # Penugasan yang telah ada sebelum modul Hak Akses diperbarui juga harus
    # menurunkan scope terbaru saat server menyala atau setelah migrasi data.
    assignment_users = await db.jabatan_assignments.find(
        {"user_id": {"$nin": ["", None]}},
        {"_id": 0, "user_id": 1},
    ).to_list(5000)
    for user_id in {item.get("user_id") for item in assignment_users if item.get("user_id")}:
        try:
            await rebuild_user_position_access(db, str(user_id))
        except Exception as error:
            logger.warning("Gagal menyinkronkan hak akses jabatan %s: %s", user_id, error)
    await load_oidc_runtime_settings()
    await ensure_program_course_links()
    await ensure_multi_lecturer_schema()
    await ensure_class_lifecycle_schema()
    # Init tabel master data SIAKAD
    for _col in ["fakultas", "tahun_ajaran", "academic_config", "kurikulum", "gedung", "ruangan"]:
        try:
            await getattr(db, _col).create_index("id", unique=True, sparse=True)
        except Exception:
            pass
    await db.users.update_one({"email": "dosen@demo.id", "username": {"$exists": False}}, {"$set": {"username": "dosenadmin", "whatsapp": "628000000001"}})
    async for student in db.users.find({"role": "student", "username": {"$exists": False}}, {"_id": 0, "id": 1, "nim": 1}):
        if student.get("nim"):
            await db.users.update_one({"id": student["id"]}, {"$set": {"username": str(student["nim"]).lower()}})
    async for assignment in db.assignments.find(
        {
            "$or": [{"material_id": {"$exists": False}}, {"material_id": ""}, {"material_id": None}],
            "material_link_removed_at": {"$exists": False},
        },
        {"_id": 0, "id": 1, "class_id": 1},
    ):
        material = await db.materials.find_one({"class_id": assignment.get("class_id")}, {"_id": 0})
        if material:
            await db.assignments.update_one({"id": assignment["id"]}, {"$set": {"material_id": material["id"]}})
    await db.users.create_index("email", unique=True)
    await ensure_unique_identity_index("username")
    await ensure_unique_identity_index("nim")
    await db.users.create_index("name")
    await db.users.create_index([("sso_issuer", 1), ("sso_subject", 1)], unique=True, sparse=True)
    await db.sessions.create_index("token", unique=True)
    await db.user_activity_logs.create_index("created_at")
    await db.user_activity_logs.create_index("user_id")
    await db.user_activity_logs.create_index("category")
    await db.user_activity_logs.create_index("action")
    await db.notification_reads.create_index(
        [("user_id", 1), ("notification_id", 1)],
        unique=True,
    )
    await db.notification_reads.create_index("read_at")
    await db.database_backups.create_index("created_at")
    await db.oidc_flows.create_index("state_hash", unique=True)
    await db.oidc_flows.create_index("expires_at", expireAfterSeconds=0)
    await db.oidc_login_tickets.create_index("ticket_hash", unique=True)
    await db.oidc_login_tickets.create_index("expires_at", expireAfterSeconds=0)
    await db.chat_messages.create_index([("conversation_id", 1), ("created_at", 1)])
    await db.chat_read_receipts.create_index(
        [("user_id", 1), ("contact_id", 1)],
        unique=True,
    )
    await db.chat_read_receipts.create_index("last_read_at")
    await db.assignments.create_index("class_id")
    await db.assignments.create_index([("is_active", 1), ("published_at", 1)])
    await db.assignments.create_index("deadline")
    await db.academic_calendar_events.create_index("id", unique=True)
    await db.academic_calendar_events.create_index([("status", 1), ("start_at", 1)])
    await db.academic_calendar_events.create_index("academic_year_id")
    await db.academic_deadline_settings.create_index("id", unique=True)
    await db.academic_deadline_settings.create_index("academic_year_id", unique=True)
    await db.submissions.create_index([("assignment_id", 1), ("student_id", 1)])
    await db.submissions.create_index("student_id")
    await db.submissions.create_index("submitted_at")
    await db.submissions.create_index("student_name")
    await db.materials.create_index("class_id")
    await db.materials.create_index("created_at")
    await db.classes.create_index("class_code")
    await db.classes.create_index("status")
    await db.classes.create_index([("lecturer_id", 1), ("status", 1)])
    await db.classes.create_index("created_at")
    await db.programs.create_index("created_at")
    await db.courses.create_index("created_at")
    await db.materials.create_index("lecturer_id")
    await db.submissions.create_index("lecturer_id")
    await db.stored_files.create_index("drive_sync_status")
    await db.stored_files.create_index("record_type")
    await db.stored_files.create_index("lecturer_id")
    await db.stored_files.create_index("uploaded_at")
    await db.stored_files.create_index("drive_next_retry_at")
    await db.drive_delete_queue.create_index("status")
    await db.drive_delete_queue.create_index("next_retry_at")
    await db.comments.create_index("material_id")
    await db.comments.create_index("created_at")
    await db.enrollment_requests.create_index([("class_id", 1), ("student_id", 1)])
    await db.enrollment_requests.create_index("requested_at")
    await db.reminder_logs.create_index("assignment_id")
    await db.reminder_logs.create_index("sent_at")
    await db.whatsapp_messages.create_index("status")
    await db.whatsapp_messages.create_index("created_at")
    await db.password_reset_requests.create_index("requested_at")
    # SIAKAD Core Indexes & Collections
    await db.academic_periods.create_index("code", unique=True)
    await db.academic_periods.create_index("is_active")
    await db.curriculums.create_index("program_id")
    await db.student_profiles.create_index("student_id", unique=True)
    await db.student_profiles.create_index("nim", unique=True, sparse=True)
    await db.student_profiles.create_index("pa_dosen_id")
    await db.lecturer_profiles.create_index("lecturer_id", unique=True)
    await db.lecturer_profiles.create_index("nidn", unique=True, sparse=True)
    await db.krs.create_index([("student_id", 1), ("academic_period_id", 1)], unique=True)
    await db.krs.create_index("status")
    await db.khs.create_index([("student_id", 1), ("academic_period_id", 1)], unique=True)
    await db.tuition_bills.create_index([("student_id", 1), ("academic_period_id", 1)])
    await db.tuition_bills.create_index("status")
    await db.tuition_bills.create_index("scheme_id")
    await db.tuition_bills.create_index(
        [("source", 1), ("pmb_applicant_id", 1)],
        unique=True,
        sparse=True,
    )
    await db.finance_components.create_index("code", unique=True, sparse=True)
    await db.finance_schemes.create_index("code", unique=True, sparse=True)
    await db.finance_scheme_rules.create_index([("scheme_id", 1), ("component_id", 1)])
    await db.tuition_payments.create_index("bill_id")
    await db.tuition_payments.create_index("status")
    await db.payment_accounts.create_index("payment_method")
    await ensure_default_finance_components(db)

    active_period = await db.academic_periods.find_one({"is_active": True}, {"_id": 0})
    if not active_period:
        now_str = now_iso()
        default_period = {
            "id": f"period_{uuid.uuid4().hex[:12]}",
            "code": "20251",
            "name": "Tahun Akademik 2025/2026 Ganjil",
            "year": "2025/2026",
            "semester": "Ganjil",
            "is_active": True,
            "krs_start_at": now_str,
            "krs_end_at": now_str,
            "status": "active",
            "created_at": now_str,
            "updated_at": now_str,
        }
        await db.academic_periods.insert_one(default_period)
        logger.info("Periode akademik default 2025/2026 Ganjil telah diinisialisasi.")

    try:
        await cleanup_old_user_activity_logs()
    except Exception as exc:
        logger.warning("Retensi log aktivitas belum dapat dijalankan: %s", exc)
    _database_backup_scheduler_task = asyncio.create_task(database_backup_scheduler())
    _storage_maintenance_scheduler_task = asyncio.create_task(storage_maintenance_scheduler())


@app.on_event("shutdown")
async def on_shutdown():
    global _database_backup_scheduler_task, _storage_maintenance_scheduler_task
    if _storage_maintenance_scheduler_task:
        _storage_maintenance_scheduler_task.cancel()
        try:
            await _storage_maintenance_scheduler_task
        except asyncio.CancelledError:
            pass
        _storage_maintenance_scheduler_task = None
    if _database_backup_scheduler_task:
        _database_backup_scheduler_task.cancel()
        try:
            await _database_backup_scheduler_task
        except asyncio.CancelledError:
            pass
        _database_backup_scheduler_task = None


@app.on_event("shutdown")
async def shutdown_db_client():
    await db.close()
