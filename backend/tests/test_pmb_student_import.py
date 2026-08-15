import io
from pathlib import Path
from types import SimpleNamespace

from fastapi import UploadFile
from openpyxl import Workbook

from routers.pmb import (
    _nim_program_code,
    _prepare_student_import,
    extract_academic_year_prefix,
    import_admin_students_from_excel,
    sync_imported_students_to_pmb,
)


class FakeCursor:
    def __init__(self, items):
        self.items = items

    async def to_list(self, _limit=None):
        return list(self.items)


class FakeCollection:
    def __init__(self, items=None):
        self.items = list(items or [])

    def find(self, *_args, **_kwargs):
        return FakeCursor(self.items)

    async def find_one(self, *_args, **_kwargs):
        return self.items[0] if self.items else None

    async def insert_one(self, item):
        self.items.append(item)

    async def update_one(self, *_args, **_kwargs):
        return None


class FakeDb:
    def __init__(self):
        self.programs = FakeCollection([
            {"id": "BD-D4", "kode": "BD-D4", "nama": "Bisnis Digital"},
            {"id": "RKJ-D4", "kode": "RKJ-D4", "nama": "Rekayasa Komputer Jaringan"},
        ])
        self.classes = FakeCollection([])
        self.users = FakeCollection([
            {"nim": "2627010001", "email": "existing@example.com", "username": "2627010001"},
        ])
        self.pmb_applicants = FakeCollection([])
        self.pmb_settings = FakeCollection([{
            "id": "pmb_global_settings",
            "nim_prefix": "2627",
            "active_period_name": "2026/2027 Ganjil",
        }])


def make_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["nama", "email", "prodi_kode", "whatsapp", "nim"])
    sheet.append(["Mahasiswa Baru", "new@example.com", "BD-D4", "081234567890", ""])
    sheet.append(["Prodi Tidak Ada", "bad@example.com", "UNKNOWN", "", ""])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_template_style_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Mahasiswa"
    sheet.append(["Template Import Mahasiswa Baru PMB"])
    sheet.append(["Isi satu mahasiswa per baris mulai baris 5"])
    sheet.append([])
    sheet.append(["nama", "email", "prodi_kode", "whatsapp", "nim"])
    sheet.append(["Mahasiswa Template", "template@example.com", "BD-D4", "081234567891", ""])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_second_program_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["nama", "email", "prodi_kode"])
    sheet.append(["Mahasiswa Prodi Dua", "prodi2@example.com", "RKJ-D4"])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_import_preview_generates_nim_and_reports_invalid_program():
    result = __import__("asyncio").run(_prepare_student_import(FakeDb(), make_workbook_bytes()))

    assert result["year_prefix"] == "2627"
    assert len(result["valid_rows"]) == 1
    assert result["valid_rows"][0]["nim"] == "2627010002"
    assert result["rows"][1]["status"] == "error"
    assert "Prodi tidak ditemukan" in result["rows"][1]["message"]


def test_import_finds_header_below_template_title_rows():
    result = __import__("asyncio").run(_prepare_student_import(FakeDb(), make_template_style_workbook_bytes()))

    assert result["sheet_name"] == "Data Mahasiswa"
    assert result["header_row"] == 4
    assert len(result["valid_rows"]) == 1
    assert result["valid_rows"][0]["row"] == 5
    assert result["valid_rows"][0]["nama"] == "Mahasiswa Template"


def test_downloadable_student_template_header_is_detected():
    template_path = Path(__file__).resolve().parents[2] / "frontend" / "public" / "templates" / "template-import-mahasiswa-baru.xlsx"
    result = __import__("asyncio").run(_prepare_student_import(FakeDb(), template_path.read_bytes()))

    assert result["sheet_name"] == "Data Mahasiswa"
    assert result["header_row"] == 4
    assert result["headers"][0] == "nama"


def test_numeric_program_code_is_preserved_for_nim_segment():
    assert _nim_program_code({"kode": "07"}, 1) == "07"
    assert _nim_program_code({"kode": "07-TI"}, 1) == "07"


def test_calendar_year_prefix_is_converted_to_academic_year_prefix():
    assert extract_academic_year_prefix({"nim_prefix": "2026"}) == "2627"
    assert extract_academic_year_prefix({"nim_prefix": "2026/2027"}) == "2627"
    assert extract_academic_year_prefix({"nim_prefix": "2627"}) == "2627"


def test_generated_nim_uses_academic_year_program_code_and_sequence():
    db = FakeDb()
    db.users = FakeCollection([])
    db.pmb_settings = FakeCollection([{
        "id": "pmb_global_settings",
        "nim_prefix": "2026",
        "active_period_name": "2026/2027 Ganjil",
    }])

    result = __import__("asyncio").run(_prepare_student_import(db, make_second_program_workbook_bytes()))

    assert result["year_prefix"] == "2627"
    assert result["valid_rows"][0]["nim"] == "2627020001"


def test_import_endpoint_creates_student_after_valid_preview():
    db = FakeDb()
    db.users = FakeCollection([])
    request = SimpleNamespace(app=SimpleNamespace(state=SimpleNamespace(db=db)))
    upload = UploadFile(filename="mahasiswa-baru.xlsx", file=io.BytesIO(make_template_style_workbook_bytes()))

    result = __import__("asyncio").run(
        import_admin_students_from_excel(
            request=request,
            file=upload,
            default_prodi_id="",
            default_password="Mahasiswa123!",
            user={"id": "admin-test"},
        )
    )

    assert result["created"] == 1
    assert result["skipped"] == 0
    assert len(db.users.items) == 1
    assert db.users.items[0]["name"] == "Mahasiswa Template"
    assert db.users.items[0]["created_by"] == "admin-test"
    assert db.users.items[0]["id"]
    assert db.users.items[0]["password_hash"] != "Mahasiswa123!"
    assert len(db.pmb_applicants.items) == 1
    assert db.pmb_applicants.items[0]["source"] == "pmb_excel_import"


def test_imported_student_is_normalized_into_manual_pmb_review_record():
    db = FakeDb()
    db.users = FakeCollection([{
        "id": "student-import-1",
        "role": "student",
        "source": "pmb_excel_import",
        "nim": "2627010001",
        "name": "Mahasiswa Import",
        "email": "import@example.com",
        "prodi_id": "BD-D4",
        "prodi_name": "Bisnis Digital",
        "prodi_kode": "BD-D4",
    }])

    created = __import__("asyncio").run(sync_imported_students_to_pmb(db))

    assert created == 1
    assert len(db.pmb_applicants.items) == 1
    record = db.pmb_applicants.items[0]
    assert record["source"] == "pmb_excel_import"
    assert record["manual_completion_required"] is True
    assert record["manual_payment_status"] == "pending"
    assert record["is_converted_to_student"] is True
    assert record["student_user_id"] == "student-import-1"
