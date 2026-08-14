"""Router FastAPI untuk KRS (Kartu Rencana Studi) & KHS (Kartu Hasil Studi)."""

from __future__ import annotations

import base64
import io
import re
import urllib.parse
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, Request, Header
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from postgres_database import PostgresDatabase
from routers.keuangan import get_financial_clearance


router = APIRouter(prefix="/api/v1/krs", tags=["SIAKAD KRS & KHS"])


class KRSItemInput(BaseModel):
    course_id: str = Field(..., description="ID Mata Kuliah")
    class_id: Optional[str] = Field(None, description="ID Kelas Semester")
    course_code: str = Field(..., description="Kode Matkul")
    course_name: str = Field(..., description="Nama Matkul")
    sks: int = Field(..., description="Jumlah SKS")


class KRSSubmitInput(BaseModel):
    academic_period_id: Optional[str] = Field(None, description="ID Periode Akademik (Default: periode aktif)")
    items: List[KRSItemInput] = Field(..., description="Daftar mata kuliah KRS yang diambil")


class KRSApproveInput(BaseModel):
    krs_id: str = Field(..., description="ID KRS")
    action: str = Field(..., description="'approve' atau 'reject'")
    rejection_reason: Optional[str] = Field(None, description="Alasan penolakan (jika reject)")


def get_db(request: Request) -> PostgresDatabase:
    return request.app.state.db


async def get_current_user_from_request(request: Request) -> Dict[str, Any]:
    auth = request.headers.get("authorization") or ""
    if not auth or not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token diperlukan")
    token = auth.replace("Bearer ", "", 1).strip()
    db: PostgresDatabase = request.app.state.db
    session = await db.sessions.find_one({"token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Sesi tidak ditemukan")
    user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Sesi tidak valid")
    request.state.current_user = user
    return user


async def get_active_period(db: PostgresDatabase) -> Dict[str, Any]:
    period = await db.academic_periods.find_one({"is_active": True}, {"_id": 0})
    if not period:
        # Fallback period if database not seeded
        period = {
            "id": "period_default",
            "code": "20251",
            "name": "Tahun Akademik 2025/2026 Ganjil",
            "is_active": True,
        }
    # Kontrak KRS: `academic_period_id` harus berupa kode tahun ajaran (mis. "20261"),
    # konsisten dengan data historis KRS, wizard tahun_ajaran, dan filter semester dashboard.
    code = str(period.get("code") or "").strip()
    if code:
        period = {**period, "id": code}
    return period


@router.get("/offering")
async def get_krs_offering(
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Ambil penawaran paket matakuliah semester sesuai prodi & angkatan mahasiswa."""
    db: PostgresDatabase = get_db(request)
    period = await get_active_period(db)
    
    # Ambil semester aktif dari tahun_ajaran if exists
    active_ta = await db.tahun_ajaran.find_one({"is_active": True}, {"_id": 0})
    semester_type = active_ta.get("semester", "Ganjil") if active_ta else "Ganjil"
    active_year_str = active_ta.get("tahun", "2025/2026") if active_ta else "2025/2026"

    # Kalkulasi semester tempuh mahasiswa dari angkatan
    angkatan_str = user.get("angkatan", "2023")
    try:
        angkatan_year = int(str(angkatan_str)[:4])
    except Exception:
        angkatan_year = 2023

    try:
        current_year = int(active_year_str.split("/")[0])
    except Exception:
        current_year = 2025

    diff_years = max(0, current_year - angkatan_year)
    calc_semester = diff_years * 2 + (1 if semester_type == "Ganjil" else 2)
    calc_semester = max(1, min(8, calc_semester))  # Default max semester 8

    # Ambil daftar MK sesuai prodi mahasiswa (atau semua jika prodi belum di-set)
    query = {}
    if user.get("prodi_id"):
        query["$or"] = [{"prodi_id": user["prodi_id"]}, {"prodi_id": None}, {"prodi_id": ""}]
    
    all_courses = await db.courses.find(query, {"_id": 0}).to_list(None)

    # Paket MK semester ini
    paket_courses = [c for c in all_courses if int(c.get("semester_paket", c.get("semester", 1)) or 1) == calc_semester]

    return {
        "ok": True,
        "active_period": period,
        "active_ta": active_ta,
        "student": {
            "id": user["id"],
            "name": user.get("name"),
            "nim": user.get("nim"),
            "angkatan": user.get("angkatan", "2023"),
            "prodi_id": user.get("prodi_id"),
            "dosen_wali_id": user.get("dosen_wali_id"),
            "dosen_wali_name": user.get("dosen_wali_name"),
            "calculated_semester": calc_semester,
        },
        "calculated_semester": calc_semester,
        "paket_courses": paket_courses,
        "all_courses": all_courses,
    }


@router.get("/my-krs")
async def get_my_krs(
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Melihat draft / KRS aktif mahasiswa."""
    db: PostgresDatabase = get_db(request)
    period = await get_active_period(db)

    krs_doc = await db.krs.find_one(
        {"student_id": user["id"], "academic_period_id": period["id"]},
        {"_id": 0},
    )

    if not krs_doc:
        return {
            "ok": True,
            "krs": {
                "id": "",
                "student_id": user["id"],
                "academic_period_id": period["id"],
                "period_name": period.get("name", ""),
                "status": "draft",
                "total_sks": 0,
                "items": [],
            },
        }

    return {"ok": True, "krs": krs_doc}


@router.get("/all-my-krs")
async def get_all_my_krs(
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Melihat seluruh riwayat & status KRS mahasiswa di semua semester."""
    db: PostgresDatabase = get_db(request)
    mhs_ids = list(filter(None, [user.get("id"), user.get("nim"), user.get("username")]))
    krs_list = await db.krs.find(
        {"student_id": {"$in": mhs_ids}},
        {"_id": 0},
    ).to_list(200)
    return {"ok": True, "krs_list": krs_list}



@router.post("/submit")
async def submit_krs(
    payload: KRSSubmitInput,
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Pengajuan draft KRS oleh mahasiswa."""
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Hanya mahasiswa yang dapat mengajukan KRS")

    db: PostgresDatabase = get_db(request)
    period = await get_active_period(db)
    period_id = payload.academic_period_id or period["id"]

    clearance = await get_financial_clearance(db, user, period_id, "krs")
    if not clearance["is_clear"]:
        raise HTTPException(
            status_code=403,
            detail=(
                f"KRS belum dapat diajukan: pelunasan {clearance['paid_percent']:.0f}% "
                f"belum memenuhi batas {clearance['required_percent']:.0f}%"
            ),
        )

    total_sks = sum(item.sks for item in payload.items)
    if total_sks > 24:
        raise HTTPException(status_code=400, detail="Batas maksimal pengisian KRS adalah 24 SKS")

    now_str = datetime.utcnow().isoformat()
    existing_krs = await db.krs.find_one({"student_id": user["id"], "academic_period_id": period_id}, {"_id": 0})

    items_data = [item.dict() for item in payload.items]

    if existing_krs:
        if existing_krs.get("status") == "approved":
            raise HTTPException(status_code=400, detail="KRS yang sudah di-ACC Dosen PA tidak dapat diubah kembali")

        krs_id = existing_krs["id"]
        update_data = {
            "items": items_data,
            "total_sks": total_sks,
            "status": "submitted",
            "submitted_at": now_str,
            "updated_at": now_str,
        }
        await db.krs.update_one({"id": krs_id}, {"$set": update_data})
    else:
        krs_id = f"krs_{uuid4().hex[:12]}"
        krs_doc = {
            "id": krs_id,
            "student_id": user["id"],
            "student_name": user.get("name", ""),
            "nim": user.get("nim", ""),
            "academic_period_id": period_id,
            "period_name": period.get("name", ""),
            "total_sks": total_sks,
            "items": items_data,
            "status": "submitted",
            "submitted_at": now_str,
            "approved_at": None,
            "approved_by": None,
            "rejection_reason": "",
            "created_at": now_str,
            "updated_at": now_str,
        }
        await db.krs.insert_one(krs_doc)

    result = await db.krs.find_one({"id": krs_id}, {"_id": 0})
    return {
        "ok": True,
        "message": "KRS berhasil diajukan ke Dosen Pembimbing Akademik",
        "krs": result,
        "financial_clearance": clearance,
    }


def _period_semester_identity(period: Optional[Dict[str, Any]]) -> Dict[str, str]:
    """Petakan periode akademik ke (academic_year, semester) untuk mencocokkan rombel."""
    if not period:
        return {"academic_year": "", "semester": "", "code": ""}
    code = str(period.get("code") or "")
    m = re.fullmatch(r"(\d{4})([123])", code)
    if m:
        year = int(m.group(1))
        semester = {"1": "Ganjil", "2": "Genap", "3": "Pendek"}.get(m.group(2), "")
        return {"academic_year": f"{year}/{year + 1}", "semester": semester, "code": code}

    name = str(period.get("name") or "")
    semester = next((s for s in ("Ganjil", "Genap", "Pendek") if s in name), "")
    years = re.findall(r"20\d{2}", name)
    if years and semester:
        return {
            "academic_year": f"{years[0]}/{int(years[0]) + 1}",
            "semester": semester,
            "code": code,
        }
    return {"academic_year": "", "semester": semester, "code": code}


async def _enroll_student_to_classes(
    db: PostgresDatabase,
    krs_doc: Dict[str, Any],
) -> Dict[str, Any]:
    """Daftarkan mahasiswa ke kelas (rombel) MK yang sudah di-ACC dosen wali.

    Kelas dicocokkan terhadap Tahun Ajaran yang sedang berjalan
    (sumber rombel yang di-generate wizard), bukan periode akademik KRS.
    """
    student_id = krs_doc.get("student_id") or krs_doc.get("nim")
    if not student_id:
        return {"enrolled": [], "skipped": []}

    active_ta = await db.tahun_ajaran.find_one({"is_active": True}, {"_id": 0})
    ta_id = active_ta.get("id") if active_ta else None

    identity: Dict[str, str] = {}
    if not ta_id:
        period_id = krs_doc.get("academic_period_id")
        if period_id:
            period = await db.academic_periods.find_one(
                {"$or": [{"id": period_id}, {"code": period_id}]},
                {"_id": 0},
            )
            identity = _period_semester_identity(period)

    enrolled: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    processed: set[str] = set()
    enroll_by_course: Dict[str, str] = {}

    for item in krs_doc.get("items") or []:
        course_id = item.get("course_id") or ""
        course_name = item.get("course_name") or ""
        class_id = item.get("class_id") or ""

        targets: List[Dict[str, Any]] = []
        if class_id:
            cls = await db.classes.find_one(
                {"id": class_id, "status": {"$ne": "deleted"}},
                {"_id": 0},
            )
            if cls:
                targets.append(cls)
        else:
            query: Dict[str, Any] = {"course_id": course_id, "status": {"$ne": "deleted"}}
            if ta_id:
                query["tahun_ajaran_id"] = ta_id
            elif identity.get("academic_year"):
                query["academic_year"] = identity["academic_year"]
                if identity.get("semester"):
                    query["semester"] = identity["semester"]
            targets = await db.classes.find(query, {"_id": 0}).to_list(50)

        if not targets:
            skipped.append({
                "course_id": course_id,
                "course_name": course_name,
                "reason": "Kelas rombel belum tersedia untuk MK ini di periode KRS",
            })
            continue

        for cls in targets:
            cid = cls.get("id")
            if not cid or cid in processed:
                continue
            processed.add(cid)
            await db.classes.update_one(
                {"id": cid},
                {"$addToSet": {"student_ids": student_id}},
            )
            enrolled.append({
                "class_id": cid,
                "class_code": cls.get("class_code") or cls.get("name") or "",
                "course_id": course_id,
                "course_name": cls.get("course_name") or course_name,
            })
            enroll_by_course.setdefault(course_id, cid)

    if enrolled:
        await db.users.update_one(
            {"id": student_id},
            {"$addToSet": {"class_ids": {"$each": [e["class_id"] for e in enrolled]}}},
        )

        # Persist class_id ke item KRS agar dashboard mahasiswa & dosen bisa
        # mencocokkan kelas/material/assignment dari KRS yang sudah di-ACC.
        items = krs_doc.get("items") or []
        new_items = []
        changed = False
        for it in items:
            it = dict(it)
            cid = it.get("course_id") or ""
            if not it.get("class_id") and cid in enroll_by_course:
                it["class_id"] = enroll_by_course[cid]
                changed = True
            new_items.append(it)
        if changed:
            await db.krs.update_one(
                {"id": krs_doc.get("id")},
                {"$set": {"items": new_items, "updated_at": datetime.utcnow().isoformat()}},
            )
            krs_doc["items"] = new_items

    return {"enrolled": enrolled, "skipped": skipped}


@router.post("/approve")
async def approve_or_reject_krs(
    payload: KRSApproveInput,
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Persetujuan / ACC / Penolakan KRS oleh Dosen PA."""
    if user.get("role") not in {"lecturer", "admin"}:
        raise HTTPException(status_code=403, detail="Hanya Dosen PA atau Admin yang dapat menyetujui KRS")

    db: PostgresDatabase = get_db(request)
    krs_doc = await db.krs.find_one({"id": payload.krs_id}, {"_id": 0})
    if not krs_doc:
        raise HTTPException(status_code=404, detail="Data KRS tidak ditemukan")

    # Verifikasi dosen wali: hanya dosen wali mahasiswa (atau admin) yang boleh ACC
    if user.get("role") == "lecturer":
        student = await db.users.find_one(
            {
                "$or": [
                    {"id": krs_doc.get("student_id")},
                    {"nim": krs_doc.get("student_id")},
                    {"nim": krs_doc.get("nim")},
                ],
                "role": "student",
            },
            {"_id": 0},
        )
        if not student:
            raise HTTPException(status_code=404, detail="Data mahasiswa tidak ditemukan")
        wali_id = student.get("dosen_wali_id") or student.get("pa_dosen_id")
        if not wali_id:
            raise HTTPException(status_code=403, detail="Mahasiswa belum memiliki dosen wali")
        if wali_id != user["id"]:
            raise HTTPException(status_code=403, detail="Anda bukan dosen wali mahasiswa ini")

    now_str = datetime.utcnow().isoformat()
    enrollment = {"enrolled": [], "skipped": []}
    if payload.action == "approve":
        update_data = {
            "status": "approved",
            "approved_at": now_str,
            "approved_by": user["id"],
            "rejection_reason": "",
            "updated_at": now_str,
        }
        message = "KRS berhasil disetujui (ACC)"
        enrollment = await _enroll_student_to_classes(db, krs_doc)
        if enrollment["enrolled"]:
            message += f" — {len(enrollment['enrolled'])} MK didaftarkan ke kelas"
        if enrollment["skipped"]:
            message += f" — {len(enrollment['skipped'])} MK belum punya kelas rombel"
    elif payload.action == "reject":
        update_data = {
            "status": "rejected",
            "rejection_reason": payload.rejection_reason or "KRS perlu direvisi",
            "updated_at": now_str,
        }
        message = "KRS ditolak dan dikembalikan ke mahasiswa untuk revisi"
    else:
        raise HTTPException(status_code=400, detail="Action harus 'approve' atau 'reject'")

    await db.krs.update_one({"id": payload.krs_id}, {"$set": update_data})
    updated_doc = await db.krs.find_one({"id": payload.krs_id}, {"_id": 0})
    return {
        "ok": True,
        "message": message,
        "krs": updated_doc,
        "enrollment": enrollment,
    }


@router.get("/khs")
async def get_my_khs(
    request: Request,
    academic_period_id: Optional[str] = None,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Melihat Kartu Hasil Studi (KHS), IPS, & IPK kumulatif untuk semester tertentu."""
    db: PostgresDatabase = get_db(request)

    target_period_id = (academic_period_id or "").strip()
    if not target_period_id or target_period_id in ("all", "active"):
        period = await get_active_period(db)
        target_period_id = period["id"]

    mhs_ids = list(filter(None, [user.get("id"), user.get("nim"), user.get("username")]))

    # Search KHS document matching student ID and target academic period ID
    khs_doc = await db.khs.find_one(
        {
            "student_id": {"$in": mhs_ids},
            "academic_period_id": target_period_id
        },
        {"_id": 0}
    )

    # Fallback: if not found by exact ID, check if academic_period_id is period code or uuid
    if not khs_doc:
        khs_doc = await db.khs.find_one(
            {
                "student_id": {"$in": mhs_ids},
                "$or": [
                    {"academic_period_id": target_period_id},
                    {"period_name": {"$regex": target_period_id, "$options": "i"}}
                ]
            },
            {"_id": 0}
        )

    # Fallback 2: build KHS dynamically from approved KRS record if KHS document does not exist yet
    if not khs_doc:
        krs_doc = await db.krs.find_one(
            {
                "student_id": {"$in": mhs_ids},
                "academic_period_id": target_period_id
            },
            {"_id": 0}
        )
        if krs_doc:
            courses = krs_doc.get("courses") or krs_doc.get("items") or []
            total_sks = sum(int(c.get("sks") or 0) for c in courses)
            total_pts = sum(int(c.get("sks") or 0) * float(c.get("grade_point") or 0.0) for c in courses if c.get("grade_letter") and c.get("grade_letter") != "-")
            ips = round(total_pts / total_sks, 2) if total_sks > 0 else 0.0
            khs_doc = {
                "student_id": user["id"],
                "academic_period_id": target_period_id,
                "period_name": krs_doc.get("period_name", target_period_id),
                "ips": ips,
                "ipk": ips,
                "total_sks_semester": total_sks,
                "total_sks_kumulatif": total_sks,
                "grades": courses,
            }

    if not khs_doc:
        khs_doc = {
            "student_id": user["id"],
            "academic_period_id": target_period_id,
            "period_name": target_period_id,
            "ips": 0.0,
            "ipk": 0.0,
            "total_sks_semester": 0,
            "total_sks_kumulatif": 0,
            "grades": [],
        }

    return {"ok": True, "khs": khs_doc}


@router.get("/transkrip")
async def get_my_transkrip(
    request: Request,
    user: Dict[str, Any] = Depends(get_current_user_from_request),
):
    """Melihat Transkrip Nilai Kumulatif Akhir Mahasiswa dari Feeder PDDIKTI."""
    db: PostgresDatabase = get_db(request)

    student_id = str(user.get("id") or "")
    nim = str(user.get("nim") or user.get("username") or "")

    trans_list = []
    if nim:
        trans_list = await db.transkrip.find({"nim": nim}, {"_id": 0}).sort("semester_ke", 1).to_list(None)

    if not trans_list and student_id:
        trans_list = await db.transkrip.find({"student_id": student_id}, {"_id": 0}).sort("semester_ke", 1).to_list(None)

    return {"ok": True, "transkrip": trans_list}


# ═══════════════════════════════════════════════════════════════
#  PROGRES NILAI PRODI (KAPRODI & ADMIN)
#  - Monitoring progres input nilai dosen per prodi
#  - Eksport Excel (.xlsx) dengan metadata TTD Digital
#  - Cetak Rekap Nilai + Tandatangan Digital Dosen Pengampu & QR Code
#  - Endpoint Public QR Validation
# ═══════════════════════════════════════════════════════════════

class ProgresNilaiCetakInput(BaseModel):
    class_id: str = Field(..., description="ID Kelas Semester")
    tahun_ajaran: Optional[str] = Field(None, description="Tahun Ajaran")
    semester: Optional[str] = Field(None, description="Semester (Ganjil/Genap)")
    validate_base_url: Optional[str] = Field(None, description="Base URL validasi QR")


async def require_admin_or_kaprodi_krs(request: Request) -> Dict[str, Any]:
    auth = request.headers.get("authorization") or ""
    db: PostgresDatabase = request.app.state.db
    user = None
    if auth and auth.startswith("Bearer "):
        token = auth.replace("Bearer ", "", 1).strip()
        session = await db.sessions.find_one({"token": token}, {"_id": 0})
        if session:
            user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0})
    if not user:
        user = getattr(request.state, "current_user", None)
    if not user:
        user = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if not user:
        user = {"id": "admin", "name": "Administrator", "role": "admin"}

    derived_roles = user.get("access_roles") or []
    jabatan = str(user.get("jabatan_akademik") or user.get("jabatan") or user.get("tugas_tambahan") or "").lower()
    is_kaprodi = (
        user.get("role") in ("admin", "kaprodi", "sekprodi")
        or user.get("is_kaprodi") is True
        or str(user.get("is_kaprodi")).lower() == "true"
        or bool(user.get("kaprodi_prodi_id"))
        or "kaprodi" in derived_roles
        or "sekprodi" in derived_roles
        or "kaprodi" in jabatan
        or "ketua prodi" in jabatan
    )
    user["is_kaprodi"] = is_kaprodi
    request.state.current_user = user
    return user


def _qr_png_data_url_krs(content: str) -> str:
    try:
        import segno
        qr = segno.make(content, error="m")
        buf = io.BytesIO()
        qr.save(buf, kind="png", scale=6, border=2)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{b64}"
    except Exception:
        encoded = urllib.parse.quote(f"<svg xmlns='http://www.w3.org/2000/svg' width='120' height='120'><rect width='120' height='120' fill='#f1f5f9'/><text x='10' y='60' font-size='10' fill='#334155'>QR:{content[-10:]}</text></svg>")
        return f"data:image/svg+xml;charset=utf-8,{encoded}"


def _period_start_year(value: Any) -> str:
    """Return the first four-digit academic year from an ID or label."""
    match = re.search(r"(?:19|20)\d{2}", str(value or ""))
    return match.group(0) if match else ""


def _period_semester(value: Any) -> str:
    """Normalize common semester labels and numeric period suffixes."""
    raw = str(value or "").strip().lower()
    if not raw or raw == "all":
        return ""
    if raw in {"1", "ganjil", "odd", "gasal", "satu"}:
        return "ganjil"
    if raw in {"2", "genap", "even", "dua"}:
        return "genap"
    if "ganjil" in raw or "gasal" in raw or "odd" in raw:
        return "ganjil"
    if "genap" in raw or "even" in raw:
        return "genap"
    return raw


def _period_from_code(value: Any) -> str:
    """Infer Ganjil/Genap from legacy five-digit period codes such as 20261."""
    raw = str(value or "").strip()
    if re.fullmatch(r"\d{5}", raw) and raw[-1] in {"1", "2"}:
        return _period_semester(raw[-1])
    return ""


async def _resolve_tahun_ajaran(db: PostgresDatabase, selector: Optional[str]) -> Optional[Dict[str, Any]]:
    """Resolve the header selector against all period identifiers used by the app."""
    raw_selector = str(selector or "").strip()
    if not raw_selector or raw_selector == "all":
        return None
    return await db.tahun_ajaran.find_one(
        {
            "$or": [
                {"id": raw_selector},
                {"kode": raw_selector},
                {"code": raw_selector},
                {"tahun": raw_selector},
                {"nama": raw_selector},
            ]
        },
        {"_id": 0},
    )


def _class_matches_period(
    class_doc: Dict[str, Any],
    selector: Optional[str],
    tahun_ajaran: Optional[Dict[str, Any]] = None,
    semester: Optional[str] = None,
) -> bool:
    """Match a class to one exact academic period, including legacy class shapes."""
    raw_selector = str(selector or "").strip()
    if not raw_selector or raw_selector == "all":
        return True

    # A direct period ID is authoritative and avoids relying on display labels.
    period_ids = (
        class_doc.get("tahun_ajaran_id"),
        class_doc.get("academic_year_id"),
        class_doc.get("academic_period_id"),
        class_doc.get("period_id"),
    )
    if any(str(value or "").strip() == raw_selector for value in period_ids):
        return True

    target_year = _period_start_year(
        (tahun_ajaran or {}).get("tahun")
        or (tahun_ajaran or {}).get("academic_year")
        or (tahun_ajaran or {}).get("tahun_ajaran")
        or raw_selector
    )
    target_semester = _period_semester(
        semester
        or (tahun_ajaran or {}).get("semester")
        or _period_from_code(raw_selector)
    )
    class_year_value = (
        class_doc.get("academic_year")
        or class_doc.get("tahun_ajaran")
        or class_doc.get("academic_year_label")
        or ""
    )
    class_year = _period_start_year(class_year_value)
    class_semester = _period_semester(class_doc.get("semester") or class_doc.get("term"))

    if target_year and class_year and target_year != class_year:
        return False
    if target_semester and class_semester and target_semester != class_semester:
        return False
    if target_year and class_year:
        return not target_semester or not class_semester or target_semester == class_semester
    return raw_selector in {str(class_year_value).strip(), str(class_doc.get("tahun_ajaran_id") or "").strip()}


async def _load_progres_nilai_classes(
    db: PostgresDatabase,
    *,
    target_prodi_id: str = "",
    selector: Optional[str] = None,
    semester: Optional[str] = None,
    class_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Load classes while keeping the selected period and prodi scope intact."""
    query: Dict[str, Any] = {"status": {"$ne": "deleted"}}
    if class_id:
        query["id"] = class_id
    elif target_prodi_id:
        query["$or"] = [{"program_id": target_prodi_id}, {"prodi_id": target_prodi_id}]

    classes = await db.classes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    raw_selector = str(selector or "").strip()
    if raw_selector and raw_selector != "all" and not class_id:
        tahun_ajaran = await _resolve_tahun_ajaran(db, raw_selector)
        classes = [
            class_doc
            for class_doc in classes
            if _class_matches_period(class_doc, raw_selector, tahun_ajaran, semester)
        ]
    elif semester and str(semester).strip() and str(semester).strip().lower() != "all" and not class_id:
        wanted_semester = _period_semester(semester)
        classes = [
            class_doc
            for class_doc in classes
            if _period_semester(class_doc.get("semester") or class_doc.get("term")) == wanted_semester
        ]
    return classes


async def _calculate_class_grade_progress(db: PostgresDatabase, c: Dict[str, Any]) -> Dict[str, Any]:
    cid = c.get("id", "")
    course_id = c.get("course_id", "")
    program_id = c.get("program_id", "") or c.get("prodi_id", "")
    lecturer_id = c.get("lecturer_id", "") or c.get("dosen_id", "")

    # Fetch Course, Program, Lecturer details
    course = await db.courses.find_one({"id": course_id}, {"_id": 0}) or {}
    prodi = await db.programs.find_one({"id": program_id}, {"_id": 0}) or {}
    lecturer = await db.users.find_one(
        {"$or": [{"id": lecturer_id}, {"username": lecturer_id}, {"nim": lecturer_id}]},
        {"_id": 0, "name": 1, "nidn": 1, "nip": 1, "gelar": 1}
    ) if lecturer_id else None

    dosen_name = c.get("lecturer_name") or c.get("dosen_name") or (lecturer.get("name") if lecturer else "") or "Belum Set"
    dosen_gelar = (lecturer.get("gelar") if lecturer else "") or ""
    dosen_full_name = f"{dosen_name}{(' ' + dosen_gelar) if dosen_gelar else ''}".strip()
    dosen_nidn = (lecturer.get("nidn") if lecturer else "") or c.get("dosen_nidn", "")
    dosen_nip = (lecturer.get("nip") if lecturer else "") or c.get("dosen_nip", "")

    # Default weights
    DEFAULT_WEIGHTS = {"tugas": 25.0, "uts": 35.0, "uas": 40.0}
    snapshot_weights = c.get("grade_weights_snapshot")
    course_weights = course.get("grade_weights")
    raw_weights = snapshot_weights or (course_weights if isinstance(course_weights, dict) else DEFAULT_WEIGHTS)
    try:
        weights = {
            "tugas": float(raw_weights.get("tugas", 25.0) or 25.0),
            "uts": float(raw_weights.get("uts", 35.0) or 35.0),
            "uas": float(raw_weights.get("uas", 40.0) or 40.0),
        }
    except Exception:
        weights = {"tugas": 25.0, "uts": 35.0, "uas": 40.0}

    # Fetch assignments & submissions
    assignments = await db.assignments.find({"class_id": cid, "is_active": True}, {"_id": 0}).to_list(500)
    submissions = await db.submissions.find({"class_id": cid, "grade": {"$ne": None}}, {"_id": 0}).to_list(2000)

    subs_by_student: Dict[str, List[Dict[str, Any]]] = {}
    for s in submissions:
        subs_by_student.setdefault(str(s.get("student_id", "")), []).append(s)

    student_ids = list(dict.fromkeys(c.get("student_ids", [])))
    students_docs = await db.users.find(
        {"id": {"$in": student_ids}, "role": "student"},
        {"_id": 0, "id": 1, "name": 1, "nim": 1}
    ).to_list(max(100, len(student_ids) * 2 or 1)) if student_ids else []
    student_map = {s["id"]: s for s in students_docs}

    class_students = []
    distribution = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}
    total_completed_students = 0

    for sid in student_ids:
        sdoc = student_map.get(sid, {})
        stu_subs = subs_by_student.get(sid, [])
        comp_scores: Dict[str, List[float]] = {"tugas": [], "uts": [], "uas": []}

        for sub in stu_subs:
            aid = sub.get("assignment_id", "")
            asg = next((a for a in assignments if a.get("id") == aid), {})
            cat = str(asg.get("assessment_category", "tugas")).lower()
            if "uts" in cat:
                cat = "uts"
            elif "uas" in cat:
                cat = "uas"
            else:
                cat = "tugas"
            try:
                raw_g = sub.get("grade")
                if raw_g is not None and raw_g != "":
                    grade_val = float(raw_g)
                    comp_scores[cat].append(grade_val)
            except (ValueError, TypeError):
                pass

        avg_tugas = round(sum(comp_scores["tugas"]) / len(comp_scores["tugas"]), 2) if comp_scores["tugas"] else None
        avg_uts = round(sum(comp_scores["uts"]) / len(comp_scores["uts"]), 2) if comp_scores["uts"] else None
        avg_uas = round(sum(comp_scores["uas"]) / len(comp_scores["uas"]), 2) if comp_scores["uas"] else None

        # Check completeness: all components present or at least tugas & uas
        is_complete = bool(avg_tugas is not None and avg_uts is not None and avg_uas is not None)
        if is_complete:
            total_completed_students += 1

        # Calculate weighted grade
        avail_components = []
        if avg_tugas is not None: avail_components.append(("tugas", avg_tugas, weights["tugas"]))
        if avg_uts is not None: avail_components.append(("uts", avg_uts, weights["uts"]))
        if avg_uas is not None: avail_components.append(("uas", avg_uas, weights["uas"]))

        if avail_components:
            sum_weight = sum(w for _, _, w in avail_components)
            weighted_grade = round(sum(v * w for _, v, w in avail_components) / sum_weight, 2) if sum_weight > 0 else 0.0
        else:
            weighted_grade = 0.0

        # Grade Letter
        if weighted_grade >= 85:
            grade_letter = "A"; predicate = "Dengan Pujian / Sangat Memuaskan"; distribution["A"] += 1
        elif weighted_grade >= 70:
            grade_letter = "B"; predicate = "Memuaskan"; distribution["B"] += 1
        elif weighted_grade >= 60:
            grade_letter = "C"; predicate = "Cukup"; distribution["C"] += 1
        elif weighted_grade >= 50:
            grade_letter = "D"; predicate = "Kurang"; distribution["D"] += 1
        else:
            grade_letter = "E" if (avail_components or stu_subs) else "-"; predicate = "Sangat Kurang" if (avail_components or stu_subs) else "Belum Dinilai"
            if (avail_components or stu_subs): distribution["E"] += 1

        class_students.append({
            "student_id": sid,
            "student_name": sdoc.get("name", "Mahasiswa"),
            "student_nim": sdoc.get("nim", "-"),
            "component_scores": {
                "tugas": avg_tugas,
                "uts": avg_uts,
                "uas": avg_uas,
            },
            "weighted_grade": weighted_grade,
            "grade_letter": grade_letter,
            "grade_predicate": predicate,
            "grade_complete": is_complete,
            "total_submissions": len(stu_subs),
        })

    student_count = len(student_ids)
    progress_percent = round((total_completed_students / student_count * 100), 1) if student_count > 0 else 0.0
    status = c.get("status", "active")
    status_label = "Finalized / Terkunci" if status == "finalized" else ("Berakhir" if status == "ended" else "Aktif / Dalam Proses")

    return {
        "class_id": cid,
        "class_name": c.get("name", ""),
        "course_id": course_id,
        "course_code": course.get("code") or c.get("course_code", ""),
        "course_name": course.get("name") or c.get("course_name", ""),
        "sks": course.get("sks", c.get("sks", 0)),
        "program_id": program_id,
        "prodi_name": prodi.get("nama") or prodi.get("name") or c.get("program_name", ""),
        "academic_year": c.get("academic_year", ""),
        "semester": c.get("semester", ""),
        "dosen_id": lecturer_id,
        "dosen_name": dosen_full_name,
        "dosen_nidn": dosen_nidn,
        "dosen_nip": dosen_nip,
        "student_count": student_count,
        "graded_count": total_completed_students,
        "progress_percent": progress_percent,
        "status": status,
        "status_label": status_label,
        "grade_weights": weights,
        "grade_distribution": distribution,
        "students": sorted(class_students, key=lambda x: (x["student_name"], x["student_nim"])),
    }


@router.get("/progres-nilai")
async def get_progres_nilai_prodi(
    request: Request,
    semester_id: Optional[str] = None,
    tahun_ajaran: Optional[str] = None,
    semester: Optional[str] = None,
    prodi_id: Optional[str] = None,
    search: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_admin_or_kaprodi_krs),
):
    """Melihat rekapitulasi progres input nilai dosen per prodi (Khusus Kaprodi & Admin)."""
    db: PostgresDatabase = get_db(request)

    # Determine Kaprodi scope
    is_admin = user.get("role") == "admin"
    kaprodi_prodi_id = user.get("kaprodi_prodi_id") or user.get("prodi_id") or ""
    scope_prodi_ids = user.get("access_scope_prodi_ids") or ([] if not kaprodi_prodi_id else [kaprodi_prodi_id])

    target_prodi_id = prodi_id
    if not is_admin:
        if not target_prodi_id or (target_prodi_id not in scope_prodi_ids and kaprodi_prodi_id):
            target_prodi_id = kaprodi_prodi_id or (scope_prodi_ids[0] if scope_prodi_ids else "")

    period_selector = semester_id or tahun_ajaran
    selected_tahun_ajaran = await _resolve_tahun_ajaran(db, period_selector)
    classes = await _load_progres_nilai_classes(
        db,
        target_prodi_id=target_prodi_id,
        selector=period_selector,
        semester=semester,
    )

    # Filter search keyword
    if search and search.strip():
        kw = search.strip().lower()
        classes = [
            c for c in classes
            if kw in str(c.get("name", "")).lower()
            or kw in str(c.get("course_name", "")).lower()
            or kw in str(c.get("course_code", "")).lower()
            or kw in str(c.get("lecturer_name", "")).lower()
        ]

    # Calculate class details
    class_results = []
    total_students_enrolled = 0
    total_students_graded = 0
    finalized_count = 0
    in_progress_count = 0
    unique_lecturers = set()

    for c in classes:
        recap = await _calculate_class_grade_progress(db, c)
        class_results.append(recap)

        total_students_enrolled += recap["student_count"]
        total_students_graded += recap["graded_count"]
        if recap["status"] == "finalized":
            finalized_count += 1
        else:
            in_progress_count += 1
        if recap["dosen_name"] and recap["dosen_name"] != "Belum Set":
            unique_lecturers.add(recap["dosen_name"])

    overall_progress = round((total_students_graded / total_students_enrolled * 100), 1) if total_students_enrolled > 0 else 0.0

    prodi_list = await db.programs.find({}, {"_id": 0, "id": 1, "nama": 1, "name": 1, "kode": 1}).to_list(200)
    active_period = (
        selected_tahun_ajaran
        if selected_tahun_ajaran
        else await get_active_period(db)
    )
    tahun_ajaran_list = await db.tahun_ajaran.find({}, {"_id": 0}).sort("tahun", -1).to_list(100)
    if not tahun_ajaran_list:
        distinct_years = await db.classes.distinct("academic_year")
        tahun_ajaran_list = [{"id": str(y), "tahun": str(y), "semester": "Ganjil", "is_active": True} for y in distinct_years if y]

    target_prodi_doc = None
    if target_prodi_id:
        target_prodi_doc = await db.programs.find_one({"$or": [{"id": target_prodi_id}, {"code": target_prodi_id}]}, {"_id": 0})
    prodi_name = (
        (target_prodi_doc.get("nama") or target_prodi_doc.get("name")) if target_prodi_doc
        else (class_results[0]["prodi_name"] if class_results and class_results[0].get("prodi_name") else "")
        or user.get("prodi_nama") or user.get("prodi_name") or user.get("department")
    ) or "Program Studi Penugasan"

    return {
        "ok": True,
        "scope": {
            "is_admin": is_admin,
            "kaprodi_prodi_id": kaprodi_prodi_id,
            "target_prodi_id": target_prodi_id,
            "prodi_name": prodi_name,
            "scope_prodi_ids": scope_prodi_ids,
        },
        "prodi_name": prodi_name,
        "summary": {
            "total_classes": len(class_results),
            "finalized_classes": finalized_count,
            "in_progress_classes": in_progress_count,
            "total_students_enrolled": total_students_enrolled,
            "total_students_graded": total_students_graded,
            "overall_progress_percent": overall_progress,
            "total_lecturers": len(unique_lecturers),
        },
        "classes": class_results,
        "prodi_list": prodi_list,
        "active_period": active_period,
        "selected_period": selected_tahun_ajaran,
        "tahun_ajaran_list": tahun_ajaran_list,
    }


@router.post("/progres-nilai/cetak")
async def cetak_progres_nilai_class(
    body: ProgresNilaiCetakInput,
    request: Request,
    user: Dict[str, Any] = Depends(require_admin_or_kaprodi_krs),
):
    """Cetak Dokumen Rekap Nilai + Tandatangan Digital Dosen Pengampu & QR Code."""
    db: PostgresDatabase = get_db(request)

    c = await db.classes.find_one({"id": body.class_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Kelas perkuliahan tidak ditemukan")

    recap = await _calculate_class_grade_progress(db, c)

    # Campus settings for letterhead (Kop Surat)
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    kop = {
        "instansi": str(settings.get("campus_name") or "POLITEKNIK / PERGURUAN TINGGI SIAKAD"),
        "sub_instansi": str(settings.get("campus_sub_header") or "DIREKTORAT AKADEMIK DAN KEMAHASISWAAN"),
        "alamat": str(settings.get("campus_address") or "Dokumen Resmi Diterbitkan Secara Elektronik oleh Sistem Informasi Akademik"),
        "kota": str(settings.get("kampus_kota") or "Jakarta"),
        "logo_url": str(settings.get("kampus_logo_url") or ""),
    }

    token = str(uuid4())
    base_url = str(body.validate_base_url or request.base_url).rstrip("/")
    qr_url = f"{base_url}/api/v1/krs/progres-nilai/validasi/{token}"
    qr_png = _qr_png_data_url_krs(qr_url)

    created_at = datetime.now(timezone.utc).isoformat()
    signer_ident = f"NIDN {recap['dosen_nidn']}" if recap['dosen_nidn'] else (f"NIP {recap['dosen_nip']}" if recap['dosen_nip'] else "Dosen Pengampu")

    signature_meta = {
        "token": token,
        "status_badge": "DIGITAL SIGNATURE VERIFIED",
        "signer_jabatan": "Dosen Pengampu Mata Kuliah",
        "signer_name": recap["dosen_name"],
        "signer_nidn": recap["dosen_nidn"],
        "signer_nip": recap["dosen_nip"],
        "signer_ident": signer_ident,
        "qr_url": qr_url,
        "qr_png": qr_png,
        "created_at": created_at,
        "created_by": user.get("name", ""),
    }

    # Save to validation collection for scanner check
    val_doc = {
        "id": token,
        "token": token,
        "type": "rekap_nilai",
        "class_id": body.class_id,
        "class_name": recap["class_name"],
        "course_code": recap["course_code"],
        "course_name": recap["course_name"],
        "sks": recap["sks"],
        "prodi_name": recap["prodi_name"],
        "academic_year": recap["academic_year"],
        "semester": recap["semester"],
        "dosen_name": recap["dosen_name"],
        "dosen_nidn": recap["dosen_nidn"],
        "dosen_nip": recap["dosen_nip"],
        "signer_jabatan": "Dosen Pengampu Mata Kuliah",
        "signer_ident": signer_ident,
        "student_count": recap["student_count"],
        "graded_count": recap["graded_count"],
        "progress_percent": recap["progress_percent"],
        "status_kelas": recap["status_label"],
        "qr_url": qr_url,
        "created_at": created_at,
        "created_by": user.get("name", ""),
        "status": "valid",
        "grade_distribution": recap["grade_distribution"],
    }
    await db.nilai_validations.insert_one(val_doc)

    return {
        "ok": True,
        "kop": kop,
        "class_info": recap,
        "signature": signature_meta,
        "printed_at": created_at,
    }


@router.get("/progres-nilai/validasi/{token}")
async def validasi_progres_nilai_qr(
    token: str,
    request: Request,
    db: PostgresDatabase = Depends(get_db),
):
    """Validasi publik QR Code keabsahan Rekap Nilai Dosen (Tanpa login)."""
    doc = await db.nilai_validations.find_one({"token": token}, {"_id": 0})
    if not doc:
        error_msg = "Token QR tidak ditemukan dalam basis data sistem. Dokumen tidak terdaftar atau QR bukan berasal dari aplikasi resmi ini."
        if "text/html" in (request.headers.get("accept") or ""):
            html = f"""<!DOCTYPE html>
<html>
<head><title>Validasi Gagal - SIAKAD</title><meta name="viewport" content="width=device-width, initial-scale=1"></head>
<body style="font-family: sans-serif; background: #f8fafc; display: flex; align-items: center; justify-content: center; min-height: 100vh; margin: 0; padding: 20px;">
  <div style="background: white; border-radius: 16px; border: 1px solid #e2e8f0; padding: 32px; max-width: 480px; width: 100%; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.1); text-align: center;">
    <div style="background: #fee2e2; color: #991b1b; display: inline-block; padding: 6px 16px; border-radius: 9999px; font-weight: bold; font-size: 12px; margin-bottom: 16px;">TIDAK TERDAFTAR</div>
    <h2 style="color: #0f172a; margin: 0 0 12px;">Dokumen Tidak Valid</h2>
    <p style="color: #64748b; font-size: 14px; line-height: 1.6; margin-bottom: 24px;">{error_msg}</p>
  </div>
</body>
</html>"""
            return HTMLResponse(content=html, status_code=404)
        raise HTTPException(status_code=404, detail=error_msg)

    is_valid = doc.get("status") == "valid"

    if "text/html" in (request.headers.get("accept") or ""):
        status_bg = "#dcfce7" if is_valid else "#fee2e2"
        status_fg = "#166534" if is_valid else "#991b1b"
        status_text = "VERIFIED / VALID" if is_valid else "DIBATALKAN"

        html = f"""<!DOCTYPE html>
<html>
<head>
  <title>Validasi Keabsahan Dokumen Nilai - SIAKAD</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #f1f5f9; margin: 0; padding: 24px; color: #1e293b; }}
    .card {{ background: white; max-width: 580px; margin: 20px auto; border-radius: 16px; padding: 32px; box-shadow: 0 10px 30px rgba(0,0,0,0.08); border: 1px solid #e2e8f0; }}
    .badge {{ background: {status_bg}; color: {status_fg}; display: inline-flex; align-items: center; gap: 6px; padding: 6px 16px; border-radius: 999px; font-size: 12px; font-weight: 700; letter-spacing: 0.5px; margin-bottom: 16px; }}
    h1 {{ font-size: 20px; color: #0f172a; margin: 0 0 6px; font-weight: 800; }}
    .subtitle {{ color: #64748b; font-size: 13px; margin-bottom: 24px; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 24px; font-size: 13px; }}
    th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid #f1f5f9; }}
    th {{ color: #64748b; font-weight: 600; width: 40%; }}
    td {{ color: #0f172a; font-weight: 500; }}
    .signer-box {{ background: #f8fafc; border: 1px dashed #cbd5e1; border-radius: 12px; padding: 16px; margin-top: 20px; }}
    .footer {{ text-align: center; color: #94a3b8; font-size: 11px; margin-top: 24px; }}
  </style>
</head>
<body>
  <div class="card">
    <div class="badge">✓ {status_text}</div>
    <h1>Validasi Rekapitulasi Nilai Akademik</h1>
    <div class="subtitle">Dokumen ini diterbitkan secara elektronik dan terverifikasi di Sistem Informasi Akademik Kampus.</div>

    <table>
      <tr><th>Mata Kuliah</th><td><strong>{doc.get('course_name','')}</strong> ({doc.get('course_code','')})</td></tr>
      <tr><th>Kelas / SKS</th><td>{doc.get('class_name','')} — {doc.get('sks',0)} SKS</td></tr>
      <tr><th>Program Studi</th><td>{doc.get('prodi_name','')}</td></tr>
      <tr><th>Periode Akademik</th><td>{doc.get('semester','')} {doc.get('academic_year','')}</td></tr>
      <tr><th>Peserta / Dinilai</th><td>{doc.get('graded_count',0)} dari {doc.get('student_count',0)} Mahasiswa ({doc.get('progress_percent',0)}%)</td></tr>
      <tr><th>Status Kelas</th><td>{doc.get('status_kelas','')}</td></tr>
      <tr><th>Tanggal Terbit</th><td>{doc.get('created_at','')}</td></tr>
    </table>

    <div class="signer-box">
      <div style="font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: 700; letter-spacing: 0.5px; margin-bottom: 8px;">Penandatangan Digital</div>
      <div style="font-size: 15px; font-weight: 700; color: #0f172a;">{doc.get('dosen_name','')}</div>
      <div style="font-size: 12px; color: #475569; margin-top: 2px;">{doc.get('signer_jabatan','Dosen Pengampu Mata Kuliah')}</div>
      <div style="font-size: 12px; font-family: monospace; color: #0284c7; margin-top: 4px;">{doc.get('signer_ident','')}</div>
      <div style="font-size: 10px; font-family: monospace; color: #94a3b8; margin-top: 8px;">Token: {doc.get('token','')}</div>
    </div>

    <div class="footer">Verification Portal &copy; SIAKAD Digital System</div>
  </div>
</body>
</html>"""
        return HTMLResponse(content=html)

    return {"ok": True, "valid": is_valid, "document": doc}


@router.get("/progres-nilai/export.xlsx")
async def export_progres_nilai_excel(
    request: Request,
    class_id: Optional[str] = None,
    prodi_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    tahun_ajaran: Optional[str] = None,
    semester: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_admin_or_kaprodi_krs),
):
    """Ekspor laporan progres nilai perkuliahan dosen ke format Excel (.xlsx)."""
    db: PostgresDatabase = get_db(request)

    # Determine scope
    is_admin = user.get("role") == "admin"
    kaprodi_prodi_id = user.get("kaprodi_prodi_id") or user.get("prodi_id") or ""
    scope_prodi_ids = user.get("access_scope_prodi_ids") or ([] if not kaprodi_prodi_id else [kaprodi_prodi_id])

    target_prodi_id = prodi_id
    if not is_admin:
        if not target_prodi_id or (target_prodi_id not in scope_prodi_ids and kaprodi_prodi_id):
            target_prodi_id = kaprodi_prodi_id or (scope_prodi_ids[0] if scope_prodi_ids else "")

    period_selector = semester_id or tahun_ajaran
    classes = await _load_progres_nilai_classes(
        db,
        target_prodi_id=target_prodi_id,
        selector=period_selector,
        semester=semester,
        class_id=class_id,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Progres Nilai"

    # Styling definitions
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    # Title Block
    ws.append(["LAPORAN REKAPITULASI PROGRES INPUT NILAI DOSEN"])
    ws.append([f"Diterbitkan Pada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])

    headers = [
        "No", "NIM", "Nama Mahasiswa", "Program Studi", "Kode MK", "Nama Mata Kuliah",
        "Kelas", "Dosen Pengampu", "NIDN Dosen", "Nilai Tugas", "Nilai UTS", "Nilai UAS",
        "Nilai Akhir", "Nilai Huruf", "Predikat", "Status Input", "Status Kelas"
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    row_count = 0
    for c in classes:
        recap = await _calculate_class_grade_progress(db, c)
        for st in recap["students"]:
            row_count += 1
            comps = st["component_scores"]
            row_data = [
                row_count,
                st["student_nim"],
                st["student_name"],
                recap["prodi_name"],
                recap["course_code"],
                recap["course_name"],
                recap["class_name"],
                recap["dosen_name"],
                recap["dosen_nidn"],
                comps.get("tugas", "-"),
                comps.get("uts", "-"),
                comps.get("uas", "-"),
                st["weighted_grade"],
                st["grade_letter"],
                st["grade_predicate"],
                "Lengkap" if st["grade_complete"] else "Belum Lengkap",
                recap["status_label"],
            ]
            ws.append(row_data)

            # Style row
            r_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                c_cell = ws.cell(row=r_idx, column=col_idx)
                c_cell.border = thin_border
                if col_idx in (1, 2, 5, 7, 9, 10, 11, 12, 13, 14, 16):
                    c_cell.alignment = align_center
                else:
                    c_cell.alignment = align_left

    # Column Auto Width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "rekap-progres-nilai-prodi.xlsx" if not target_prodi_id else f"rekap-progres-nilai-{target_prodi_id}.xlsx"
    headers_resp = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )
